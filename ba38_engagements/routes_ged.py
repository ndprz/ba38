"""
routes_ged.py
=============

GED (Gestion Électronique des Documents) des engagements de dépenses.

Vue orientée révision comptable (Commissaire aux Comptes).

Fonctionnalités :
- Tableau de tous les engagements avec indicateur de complétude documentaire
- Filtres avancés : période, pôle, rubrique, statut, financement, complétude, EBP
- Indicateur de complétude par dossier (Complet / Incomplet / En cours)
- Accès direct aux pièces jointes depuis le tableau
- Export Excel avec colonne N° EBP

Accès :
- Réservé aux administrateurs (engagement_parametres / lecture)
  → trésorier, président, commissaire aux comptes

Champ clé :
- engagements.numero_ecriture_ebp : référence EBP saisie par le trésorier
  au moment du règlement. Permet le rapprochement avec le grand livre.
"""

from flask import (
    render_template,
    request,
    send_file,
    current_app,
)

from flask_login import login_required, current_user

from ba38_utilitaires.core import (
    get_db_path,
    has_access,
    require_access,
    write_log,
)

from ba38_engagements import engagements_bp

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

from io import BytesIO
from datetime import datetime
from decimal import Decimal

import sqlite3


# ============================================================
# CONSTANTES
# ============================================================

STATUTS_LABELS = {
    "validation_pole":       "En attente pôle",
    "validation_presidence": "En attente bureau",
    "valide":                "Validé",
    "a_payer":               "À payer",
    "reglee":                "Réglé",
    "comptabilise":          "Comptabilisé",
    "termine":               "Terminé",
    "refuse":                "Refusé",
}

# Statuts pour lesquels l'argent a été réglé (paiement fournisseur
# ou remboursement bénévole effectué), qu'il soit déjà comptabilisé
# ou non. Utilisé pour les pièces justificatives attendues et les
# montants "réglés".
STATUTS_TERMINES = {"reglee", "comptabilise", "termine"}

# Statuts pour lesquels le N° d'écriture EBP est exigible.
# Depuis le passage en deux phases, le règlement (reglee) ne
# comporte plus le N° EBP : il n'est saisi qu'à la comptabilisation,
# à réception du relevé bancaire.
STATUTS_EBP_EXIGIBLE = {"comptabilise", "termine"}

# Statuts considérés comme "en cours" (workflow actif)
STATUTS_EN_COURS = {
    "validation_pole",
    "validation_presidence",
    "valide",
    "a_payer",
}


# ============================================================
# HELPERS
# ============================================================

def _est_admin_ged():
    """
    Accès GED réservé aux admins engagement_parametres.
    C'est le même rôle que l'admin du reporting.
    """
    return has_access("engagement_parametres", "lecture")


def _trouver_palier(montant, paliers):
    """
    Retrouve le palier applicable à un montant donné.
    Reprend la même logique que le JS de nouvelle_depense.html :
    le premier palier (trié par montant_max croissant) dont le
    montant_max est >= au montant de la dépense.
    """
    for p in paliers:
        if montant <= p["montant_max"]:
            return p
    return None


def _calculer_completude(engagement, fichiers, paliers):
    """
    Calcule l'indicateur de complétude documentaire d'un engagement.

    paliers : liste des paliers actifs (engagements_parametres),
              utilisée pour ne réclamer un devis que lorsque le
              montant de la dépense le rend effectivement obligatoire.

    Retourne un dict :
        statut      : 'complet' | 'incomplet' | 'en_cours' | 'refuse'
        label       : texte affiché
        badge       : classe CSS Bootstrap
        icone       : emoji
        detail      : liste de str décrivant les manques éventuels
    """

    statut_eng = engagement.get("statut", "")
    sous_type   = engagement.get("sous_type_depense", "")
    types_docs  = {f["type_fichier"] for f in fichiers}
    manques     = []

    # ----------------------------------------------------------
    # Refusé → cas particulier, pas de complétude attendue
    # ----------------------------------------------------------
    if statut_eng == "refuse":
        return {
            "statut":  "refuse",
            "label":   "Refusé",
            "badge":   "bg-danger",
            "icone":   "🔴",
            "detail":  [],
        }

    # ----------------------------------------------------------
    # En cours de workflow → pas encore exigible
    # ----------------------------------------------------------
    if statut_eng in STATUTS_EN_COURS:
        return {
            "statut":  "en_cours",
            "label":   "En cours",
            "badge":   "bg-secondary",
            "icone":   "🔵",
            "detail":  ["Workflow en cours — complétude non encore évaluée"],
        }

    # ----------------------------------------------------------
    # Dossier terminé ou réglé → on vérifie la complétude
    # ----------------------------------------------------------

    # 1. Numéro EBP obligatoire uniquement à partir de la
    #    comptabilisation (saisi à réception du relevé bancaire,
    #    pas au moment du règlement lui-même).
    if statut_eng in STATUTS_EBP_EXIGIBLE:
        if not engagement.get("numero_ecriture_ebp"):
            manques.append("N° écriture EBP manquant")

    # 2. Pièces selon le type
    if sous_type == "deplacement":
        # Note de frais générée automatiquement
        if "note_frais" not in types_docs:
            manques.append("Note de frais manquante")

    else:
        # Achat fournisseur ou bénévole
        nb_devis = sum(
            1 for f in fichiers
            if f["type_fichier"] == "devis"
        )
        a_justificatif = (
            "justificatif" in types_docs
            or "facture" in types_docs
        )

        # Devis exigé uniquement si le palier correspondant au
        # montant de la dépense l'impose (cf. nouvelle_depense.html)
        montant = float(engagement.get("montant_total") or 0)
        palier = _trouver_palier(montant, paliers)

        un_devis_obligatoire = bool(palier) and palier["un_devis"] == "o"
        deux_devis_obligatoires = bool(palier) and palier["deux_devis"] == "o"

        if deux_devis_obligatoires and nb_devis < 2:
            manques.append(
                f"Devis incomplet ({nb_devis}/2 requis)"
            )
        elif un_devis_obligatoire and nb_devis == 0:
            manques.append("Aucun devis joint")

        if not a_justificatif:
            manques.append("Facture / justificatif de règlement manquant")

    # ----------------------------------------------------------
    # Résultat
    # ----------------------------------------------------------
    if not manques:

        # Réglé, pièces complètes, mais pas encore comptabilisé :
        # ce n'est pas une anomalie, juste une étape normale en
        # attente de réception du relevé bancaire.
        if statut_eng == "reglee":
            return {
                "statut":  "en_attente_compta",
                "label":   "Réglé (en attente compta)",
                "badge":   "bg-info text-white",
                "icone":   "🔷",
                "detail":  [],
            }

        return {
            "statut":  "complet",
            "label":   "Complet",
            "badge":   "bg-success",
            "icone":   "✅",
            "detail":  [],
        }
    else:
        return {
            "statut":  "incomplet",
            "label":   "Incomplet",
            "badge":   "bg-warning text-dark",
            "icone":   "⚠️",
            "detail":  manques,
        }


def _build_ged_query(filters):
    """
    Construit la requête SQL pour la GED.
    Toujours en mode admin (vision complète).

    Par défaut, les engagements archivés sont inclus : la GED
    sert de registre comptable/audit, ils doivent rester
    consultables sans action particulière.
    """

    where  = ["1=1"]
    params = []

    if filters.get("archive") == "actifs":
        where.append("COALESCE(e.deleted, 0) = 0")
    elif filters.get("archive") == "archives":
        where.append("COALESCE(e.deleted, 0) = 1")

    if filters.get("date_debut"):
        where.append("DATE(e.cree_le) >= ?")
        params.append(filters["date_debut"])

    if filters.get("date_fin"):
        where.append("DATE(e.cree_le) <= ?")
        params.append(filters["date_fin"])

    if filters.get("pole_id"):
        where.append("e.pole_id = ?")
        params.append(filters["pole_id"])

    if filters.get("rubrique"):
        where.append("d.rubrique = ?")
        params.append(filters["rubrique"])

    if filters.get("subvention_id"):
        if filters["subvention_id"] == "budget":
            where.append("d.subvention_id IS NULL")
        else:
            where.append("d.subvention_id = ?")
            params.append(filters["subvention_id"])

    if filters.get("statut"):
        where.append("e.statut = ?")
        params.append(filters["statut"])

    if filters.get("avec_ebp") == "oui":
        where.append(
            "e.numero_ecriture_ebp IS NOT NULL "
            "AND e.numero_ecriture_ebp != ''"
        )
    elif filters.get("avec_ebp") == "non":
        where.append(
            "(e.numero_ecriture_ebp IS NULL "
            "OR e.numero_ecriture_ebp = '')"
        )

    if filters.get("numero_ebp"):
        where.append("e.numero_ecriture_ebp LIKE ?")
        params.append(f"%{filters['numero_ebp']}%")

    # Le filtre completude est appliqué en Python après calcul
    sql = f"""
        SELECT
            e.id,
            e.cree_le,
            e.statut,
            e.demandeur_nom,
            e.demandeur_email,
            e.paye_le,
            e.numero_ecriture_ebp,
            e.deleted,

            p.nom_affiche AS pole,

            d.objet,
            d.rubrique,
            d.precision_rubrique,
            d.montant_total,
            d.type_engagement,
            d.sous_type_depense,
            d.fournisseur_nom,
            d.beneficiaire_nom,
            d.subvention_id,

            s.nom_subvention

        FROM engagements e

        LEFT JOIN engagements_depenses d
            ON d.engagement_id = e.id

        LEFT JOIN engagement_poles p
            ON p.id = e.pole_id

        LEFT JOIN engagement_subventions s
            ON s.id = d.subvention_id

        WHERE {" AND ".join(where)}
        ORDER BY e.cree_le DESC
    """

    return sql, params


def _charger_fichiers_par_engagement(conn, ids_engagements):
    """
    Charge tous les fichiers pour une liste d'engagement_id.
    Retourne un dict { engagement_id: [fichiers...] }
    """
    if not ids_engagements:
        return {}

    placeholders = ",".join("?" * len(ids_engagements))

    rows = conn.execute(
        f"""
        SELECT
            id,
            engagement_id,
            type_fichier,
            nom_original,
            uploaded_le
        FROM engagements_fichiers
        WHERE engagement_id IN ({placeholders})
        ORDER BY uploaded_le ASC
        """,
        ids_engagements
    ).fetchall()

    result = {}
    for row in rows:
        eid = row["engagement_id"]
        result.setdefault(eid, [])
        result[eid].append(dict(row))

    return result


# ============================================================
# PAGE GED
# ============================================================

@engagements_bp.route(
    "/engagements/ged",
    methods=["GET", "POST"]
)
@login_required
@require_access("engagement_parametres", "lecture")
def engagements_ged():
    """
    Page GED principale.
    Accès réservé aux administrateurs / trésorier / CAC.
    """

    db_path = get_db_path()
    today   = datetime.now().date()

    # Valeurs par défaut : exercice en cours
    if request.method == "POST":
        date_debut_defaut = ""
        date_fin_defaut   = ""
    else:
        date_debut_defaut = today.replace(month=1, day=1).isoformat()
        date_fin_defaut   = today.isoformat()

    filters = {
        "date_debut":   request.form.get("date_debut",    date_debut_defaut).strip(),
        "date_fin":     request.form.get("date_fin",      date_fin_defaut).strip(),
        "pole_id":      request.form.get("pole_id",       "").strip(),
        "rubrique":     request.form.get("rubrique",      "").strip(),
        "subvention_id":request.form.get("subvention_id", "").strip(),
        "statut":       request.form.get("statut",        "").strip(),
        "avec_ebp":     request.form.get("avec_ebp",      "").strip(),
        "numero_ebp":   request.form.get("numero_ebp",    "").strip(),
        "completude":   request.form.get("completude",    "").strip(),
        "archive":      request.form.get("archive",       "").strip(),
    }

    rows        = []
    stats       = {}
    has_search  = request.method == "POST"

    with sqlite3.connect(db_path) as conn:

        conn.row_factory = sqlite3.Row

        # --------------------------------------------------
        # Listes pour les filtres
        # --------------------------------------------------

        poles = conn.execute("""
            SELECT id, nom_affiche
            FROM engagement_poles
            ORDER BY nom_affiche
        """).fetchall()

        rubriques = conn.execute("""
            SELECT DISTINCT rubrique
            FROM engagements_depenses
            WHERE rubrique IS NOT NULL AND rubrique != ''
            ORDER BY rubrique
        """).fetchall()

        subventions = conn.execute("""
            SELECT id, nom_subvention
            FROM engagement_subventions
            ORDER BY nom_subvention
        """).fetchall()

        paliers = conn.execute("""
            SELECT *
            FROM engagements_parametres
            WHERE actif = 1
            ORDER BY montant_max
        """).fetchall()

        # --------------------------------------------------
        # Requête principale
        # --------------------------------------------------

        if has_search:

            sql, params = _build_ged_query(filters)
            raw = conn.execute(sql, params).fetchall()
            engagements_raw = [dict(r) for r in raw]

            # Charger tous les fichiers en une seule requête
            ids = [e["id"] for e in engagements_raw]
            fichiers_map = _charger_fichiers_par_engagement(conn, ids)

            # Calculer la complétude pour chaque engagement
            for eng in engagements_raw:
                fichiers_eng = fichiers_map.get(eng["id"], [])
                eng["fichiers"]    = fichiers_eng
                eng["nb_fichiers"] = len(fichiers_eng)
                eng["completude"]  = _calculer_completude(eng, fichiers_eng, paliers)

            # Filtre complétude (appliqué en Python)
            filtre_completude = filters.get("completude")
            if filtre_completude:
                engagements_raw = [
                    e for e in engagements_raw
                    if e["completude"]["statut"] == filtre_completude
                ]

            rows = engagements_raw

            # --------------------------------------------------
            # Statistiques de la sélection
            # --------------------------------------------------
            nb_total      = len(rows)
            nb_complet    = sum(1 for r in rows if r["completude"]["statut"] == "complet")
            nb_incomplet  = sum(1 for r in rows if r["completude"]["statut"] == "incomplet")
            nb_en_cours   = sum(1 for r in rows if r["completude"]["statut"] == "en_cours")
            nb_refuse     = sum(1 for r in rows if r["completude"]["statut"] == "refuse")
            nb_avec_ebp   = sum(
                1 for r in rows
                if r.get("numero_ecriture_ebp")
            )
            nb_sans_ebp_reglee = sum(
                1 for r in rows
                if r["statut"] in STATUTS_EBP_EXIGIBLE
                and not r.get("numero_ecriture_ebp")
            )

            montant_total = sum(
                float(r["montant_total"] or 0)
                for r in rows
            )
            montant_regle = sum(
                float(r["montant_total"] or 0)
                for r in rows
                if r["statut"] in STATUTS_TERMINES
            )

            stats = {
                "nb_total":           nb_total,
                "nb_complet":         nb_complet,
                "nb_incomplet":       nb_incomplet,
                "nb_en_cours":        nb_en_cours,
                "nb_refuse":          nb_refuse,
                "nb_avec_ebp":        nb_avec_ebp,
                "nb_sans_ebp_reglee": nb_sans_ebp_reglee,
                "montant_total":      montant_total,
                "montant_regle":      montant_regle,
                "pct_complet": (
                    round(nb_complet / nb_total * 100, 1)
                    if nb_total else 0
                ),
            }

            write_log(
                f"[GED] Consultation par {current_user.email} "
                f"— {nb_total} engagement(s)"
            )

    return render_template(
        "engagements/ged.html",
        filters=filters,
        poles=poles,
        rubriques=rubriques,
        subventions=subventions,
        rows=rows,
        stats=stats,
        has_search=has_search,
        statuts=STATUTS_LABELS,
    )


# ============================================================
# EXPORT EXCEL GED
# ============================================================

@engagements_bp.route(
    "/engagements/ged/export-excel",
    methods=["POST"]
)
@login_required
@require_access("engagement_parametres", "lecture")
def engagements_ged_excel():
    """
    Export Excel de la GED avec toutes les colonnes utiles au CAC,
    y compris le N° EBP et l'indicateur de complétude.
    """

    db_path = get_db_path()

    filters = {
        "date_debut":   request.form.get("date_debut",    "").strip(),
        "date_fin":     request.form.get("date_fin",      "").strip(),
        "pole_id":      request.form.get("pole_id",       "").strip(),
        "rubrique":     request.form.get("rubrique",      "").strip(),
        "subvention_id":request.form.get("subvention_id", "").strip(),
        "statut":       request.form.get("statut",        "").strip(),
        "avec_ebp":     request.form.get("avec_ebp",      "").strip(),
        "numero_ebp":   request.form.get("numero_ebp",    "").strip(),
        "completude":   request.form.get("completude",    "").strip(),
        "archive":      request.form.get("archive",       "").strip(),
    }

    with sqlite3.connect(db_path) as conn:

        conn.row_factory = sqlite3.Row

        paliers = conn.execute("""
            SELECT *
            FROM engagements_parametres
            WHERE actif = 1
            ORDER BY montant_max
        """).fetchall()

        sql, params = _build_ged_query(filters)
        raw = conn.execute(sql, params).fetchall()
        engagements_raw = [dict(r) for r in raw]

        ids = [e["id"] for e in engagements_raw]
        fichiers_map = _charger_fichiers_par_engagement(conn, ids)

        for eng in engagements_raw:
            fichiers_eng       = fichiers_map.get(eng["id"], [])
            eng["fichiers"]    = fichiers_eng
            eng["nb_fichiers"] = len(fichiers_eng)
            eng["completude"]  = _calculer_completude(eng, fichiers_eng, paliers)

        if filters.get("completude"):
            engagements_raw = [
                e for e in engagements_raw
                if e["completude"]["statut"] == filters["completude"]
            ]

    # ----------------------------------------------------------
    # Construction du classeur Excel
    # ----------------------------------------------------------

    wb = Workbook()
    ws = wb.active
    ws.title = "GED Engagements"

    # Styles
    BLEU       = "1F3864"
    BLEU_L     = "D9E1F2"
    VERT       = "375623"
    VERT_L     = "E2EFDA"
    ORANGE_L   = "FCE4D6"
    ROUGE_L    = "FFDCE1"
    GRIS_L     = "F2F2F2"

    font_titre  = Font(name="Arial", bold=True, size=12, color=BLEU)
    font_header = Font(name="Arial", bold=True, size=9,  color="FFFFFF")
    font_normal = Font(name="Arial", size=9)
    font_total  = Font(name="Arial", bold=True, size=9)

    fill_header  = PatternFill("solid", start_color=BLEU)
    fill_complet = PatternFill("solid", start_color=VERT_L)
    fill_incompl = PatternFill("solid", start_color=ORANGE_L)
    fill_refuse  = PatternFill("solid", start_color=ROUGE_L)
    fill_gris    = PatternFill("solid", start_color=GRIS_L)

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right",  vertical="center")

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Titre
    ws.merge_cells("A1:P1")
    titre_cell = ws["A1"]
    titre_cell.value = (
        f"GED Engagements BA38 — "
        f"Extrait le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    )
    titre_cell.font      = font_titre
    titre_cell.alignment = center
    ws.row_dimensions[1].height = 22

    # Ligne filtres
    filtres_txt = []
    if filters["date_debut"]: filtres_txt.append(f"Du {filters['date_debut']}")
    if filters["date_fin"]:   filtres_txt.append(f"au {filters['date_fin']}")

    ws.merge_cells("A2:P2")
    filtre_cell = ws["A2"]
    filtre_cell.value = (
        "  |  ".join(filtres_txt) if filtres_txt
        else "Tous les engagements"
    )
    filtre_cell.font      = Font(name="Arial", italic=True, size=8, color="595959")
    filtre_cell.alignment = left

    ws.row_dimensions[3].height = 4  # espace

    # En-têtes colonnes
    colonnes = [
        ("ID",                    6),
        ("Date demande",         13),
        ("Demandeur",            22),
        ("Pôle",                 18),
        ("Objet",                35),
        ("Rubrique",             16),
        ("Sous-type",            12),
        ("Bénéficiaire/Four.",   25),
        ("Financement",          22),
        ("Statut",               16),
        ("Montant (€)",          13),
        ("Réglé le",             13),
        ("N° EBP",               18),   # ← clé CAC
        ("Nb fichiers",          10),
        ("Complétude",           14),
        ("Archivé",              10),
    ]

    ROW_H = 4
    for col_idx, (label, width) in enumerate(colonnes, 1):
        cell            = ws.cell(row=ROW_H, column=col_idx, value=label)
        cell.font       = font_header
        cell.fill       = fill_header
        cell.alignment  = center
        cell.border     = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[ROW_H].height = 18

    # Données
    for row_idx, eng in enumerate(engagements_raw, ROW_H + 1):

        beneficiaire = (
            eng.get("fournisseur_nom")
            or eng.get("beneficiaire_nom")
            or eng.get("demandeur_nom")
            or "—"
        )

        date_str    = (eng["cree_le"] or "")[:10]
        paye_le_str = (eng.get("paye_le") or "")[:10]
        montant     = float(eng.get("montant_total") or 0)
        completude  = eng["completude"]

        valeurs = [
            eng["id"],
            date_str,
            eng.get("demandeur_nom") or "",
            eng.get("pole") or "",
            eng.get("objet") or "",
            eng.get("rubrique") or "",
            eng.get("sous_type_depense") or "",
            beneficiaire,
            eng.get("nom_subvention") or "Budget propre",
            STATUTS_LABELS.get(eng["statut"], eng["statut"]),
            montant,
            paye_le_str,
            eng.get("numero_ecriture_ebp") or "",
            eng["nb_fichiers"],
            completude["label"],
            "Oui" if eng.get("deleted") == 1 else "Non",
        ]

        # Couleur de ligne selon complétude
        c_statut = completude["statut"]
        if c_statut == "complet":
            row_fill = fill_complet
        elif c_statut == "incomplet":
            row_fill = fill_incompl
        elif c_statut == "refuse":
            row_fill = fill_refuse
        else:
            row_fill = fill_gris if row_idx % 2 == 0 else None

        for col_idx, val in enumerate(valeurs, 1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = font_normal
            cell.border    = border
            cell.alignment = right if col_idx in (1, 11, 14) else left
            if row_fill:
                cell.fill = row_fill

        # Format monétaire
        ws.cell(row=row_idx, column=11).number_format = '#,##0.00 €'

    # Ligne total
    row_total = ROW_H + 1 + len(engagements_raw)
    total_montant = sum(
        float(e.get("montant_total") or 0)
        for e in engagements_raw
    )

    for col_idx in range(1, 17):
        cell        = ws.cell(row=row_total, column=col_idx)
        cell.font   = font_total
        cell.fill   = PatternFill("solid", start_color=BLEU)
        cell.border = border
        cell.font   = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cell.alignment = right

    ws.cell(row=row_total, column=1, value="TOTAL")
    ws.cell(
        row=row_total, column=11,
        value=total_montant
    ).number_format = '#,##0.00 €'
    ws.cell(row=row_total, column=15, value=f"{len(engagements_raw)} dossier(s)")

    ws.freeze_panes = "A5"

    # Envoi
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = (
        f"ged_engagements_"
        f"{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )

    write_log(
        f"[GED] Export Excel par {current_user.email} "
        f"— {len(engagements_raw)} dossier(s)"
    )

    return send_file(
        output,
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        download_name=filename,
        as_attachment=True,
    )
