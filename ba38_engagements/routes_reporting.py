"""
routes_reporting.py
===================

Reporting des engagements de dépenses.

Filtres disponibles :
- Période (date début / date fin)
- Demandeur
- Pôle
- Rubrique
- Subvention / budget

Sorties :
- Tableau HTML avec agrégats
- Export Excel (.xlsx)
- Export PDF (reportlab)

Accès :
- Tout utilisateur voit ses propres engagements.
- Les admins (engagement_parametres / lecture) voient tout.
"""

from flask import (
    render_template,
    request,
    redirect,
    url_for,
    flash,
    abort,
    send_file,
    current_app,
)

from flask_login import login_required, current_user

from ba38_utilitaires.core import (
    get_db_path,
    has_access,
    require_access,
    write_log,
)

from ba38_engagements import engagements_bp
from ba38_engagements.utils_financement import calculer_montants_utilises

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

from io import BytesIO
from datetime import datetime
from decimal import Decimal, InvalidOperation

import sqlite3


# ============================================================
# HELPERS
# ============================================================

def _est_admin():
    return has_access("engagement_parametres", "lecture")


def _build_query(filters, voir_tout):
    """
    Construit la requête SQL principale + les paramètres
    en fonction des filtres reçus.
    """

    where = ["COALESCE(e.deleted, 0) = 0"]
    params = []

    if not voir_tout:
        where.append("e.demandeur_id = ?")
        params.append(current_user.id)

    if filters.get("date_debut"):
        where.append("DATE(e.cree_le) >= ?")
        params.append(filters["date_debut"])

    if filters.get("date_fin"):
        where.append("DATE(e.cree_le) <= ?")
        params.append(filters["date_fin"])

    if filters.get("demandeur"):
        where.append(
            "LOWER(e.demandeur_nom) LIKE ?"
        )
        params.append(
            f"%{filters['demandeur'].lower()}%"
        )

    if filters.get("fournisseur"):
        where.append(
            "LOWER(d.fournisseur_nom) LIKE ?"
        )
        params.append(
            f"%{filters['fournisseur'].lower()}%"
        )

    if filters.get("numero_ecriture_ebp"):
        where.append(
            "LOWER(e.numero_ecriture_ebp) LIKE ?"
        )
        params.append(
            f"%{filters['numero_ecriture_ebp'].lower()}%"
        )

    if filters.get("pole_id"):
        where.append("e.pole_id = ?")
        params.append(filters["pole_id"])

    if filters.get("rubrique"):
        where.append("d.rubrique = ?")
        params.append(filters["rubrique"])

    if filters.get("subvention_id"):
        if filters["subvention_id"] == "budget":
            where.append("d.subvention_id IS NULL")
        else:
            where.append("d.subvention_id = ?")
            params.append(filters["subvention_id"])

    if filters.get("statut"):
        where.append("e.statut = ?")
        params.append(filters["statut"])

    sql = f"""
        SELECT
            e.id,
            e.cree_le,
            e.statut,
            e.demandeur_nom,
            e.demandeur_email,
            p.nom_affiche AS pole,
            d.objet,
            d.rubrique,
            d.precision_rubrique,
            d.montant_total,
            d.type_engagement,
            d.sous_type_depense,
            d.fournisseur_nom,
            d.beneficiaire_nom,
            d.subvention_id,
            s.nom_subvention,
            e.numero_ecriture_ebp
        FROM engagements e
        LEFT JOIN engagements_depenses d
            ON d.engagement_id = e.id
        LEFT JOIN engagement_poles p
            ON p.id = e.pole_id
        LEFT JOIN engagement_subventions s
            ON s.id = d.subvention_id
        WHERE {" AND ".join(where)}
        ORDER BY e.cree_le DESC
    """

    return sql, params


def _build_agregats(rows):
    """
    Calcule les totaux par pôle et par rubrique
    à partir des lignes brutes.
    """

    par_pole = {}
    par_rubrique = {}
    par_financement = {}
    total_global = Decimal("0")

    for r in rows:

        montant = Decimal(
            str(r["montant_total"] or 0)
        )

        total_global += montant

        pole = r["pole"] or "—"
        par_pole.setdefault(pole, Decimal("0"))
        par_pole[pole] += montant

        rubrique = r["rubrique"] or "—"
        par_rubrique.setdefault(rubrique, Decimal("0"))
        par_rubrique[rubrique] += montant

        if r["subvention_id"]:
            cle_financement = r["subvention_id"]
            nom_financement = r["nom_subvention"] or "—"
        else:
            cle_financement = "budget"
            nom_financement = "Budget propre (sans subvention)"

        par_financement.setdefault(
            cle_financement,
            {"nom": nom_financement, "montant": Decimal("0")}
        )
        par_financement[cle_financement]["montant"] += montant

    par_pole = dict(
        sorted(par_pole.items(), key=lambda x: -x[1])
    )

    par_rubrique = dict(
        sorted(par_rubrique.items(), key=lambda x: -x[1])
    )

    par_financement = dict(
        sorted(
            par_financement.items(),
            key=lambda x: -x[1]["montant"]
        )
    )

    return {
        "total_global": total_global,
        "nb_engagements": len(rows),
        "par_pole": par_pole,
        "par_rubrique": par_rubrique,
        "par_financement": par_financement,
    }


# ============================================================
# STATUTS LISIBLES
# ============================================================

STATUTS_LABELS = {
    "validation_pole":       "En attente pôle",
    "validation_presidence": "En attente bureau",
    "valide":                "Validé",
    "a_payer":               "À payer",
    "reglee":                "Réglé",
    "comptabilise":          "Comptabilisé",
    "termine":               "Terminé",
    "refuse":                "Refusé",
}

# Statuts pour lesquels l'argent a été réglé (paiement fournisseur
# ou remboursement bénévole effectué), comptabilisé ou non.
STATUTS_REGLES = {"reglee", "comptabilise", "termine"}


# ============================================================
# PAGE REPORTING
# ============================================================

@engagements_bp.route(
    "/engagements/reporting",
    methods=["GET", "POST"]
)
@login_required
@require_access("engagements", "lecture")
def engagements_reporting():

    db_path = get_db_path()
    voir_tout = _est_admin()
    today = datetime.now().date()

    if request.method == "POST":
        date_debut_defaut = ""
        date_fin_defaut = ""
    else:
        date_debut_defaut = today.replace(month=1, day=1).isoformat()
        date_fin_defaut = today.isoformat()

    filters = {
        "date_debut":   request.form.get("date_debut", date_debut_defaut).strip(),
        "date_fin":     request.form.get("date_fin", date_fin_defaut).strip(),
        "demandeur":    request.form.get("demandeur", "").strip(),
        "fournisseur":  request.form.get("fournisseur", "").strip(),
        "numero_ecriture_ebp": request.form.get("numero_ecriture_ebp", "").strip(),
        "pole_id":      request.form.get("pole_id", "").strip(),
        "rubrique":     request.form.get("rubrique", "").strip(),
        "subvention_id":request.form.get("subvention_id", "").strip(),
        "statut":       request.form.get("statut", "").strip(),
    }

    rows = []
    agregats = None
    has_search = request.method == "POST"

    with sqlite3.connect(db_path) as conn:

        conn.row_factory = sqlite3.Row

        poles = conn.execute("""
            SELECT id, nom_affiche
            FROM engagement_poles
            ORDER BY nom_affiche
        """).fetchall()

        rubriques = conn.execute("""
            SELECT DISTINCT rubrique
            FROM engagements_depenses
            WHERE rubrique IS NOT NULL
              AND rubrique != ''
            ORDER BY rubrique
        """).fetchall()

        subventions = conn.execute("""
            SELECT id, nom_subvention
            FROM engagement_subventions
            ORDER BY nom_subvention
        """).fetchall()

        financement_selectionne = None

        if has_search:

            sql, params = _build_query(filters, voir_tout)
            raw = conn.execute(sql, params).fetchall()
            rows = [dict(r) for r in raw]
            agregats = _build_agregats(rows)

            ids_financement = [
                cle for cle in agregats["par_financement"]
                if cle != "budget"
            ]

            recus_par_id = {}
            utilises_par_id = {}

            if ids_financement:

                placeholders = ",".join("?" for _ in ids_financement)

                recus = conn.execute(f"""
                    SELECT id, montant_prevu, montant_recu
                    FROM engagement_subventions
                    WHERE id IN ({placeholders})
                """, ids_financement).fetchall()

                recus_par_id = {r["id"]: dict(r) for r in recus}

                utilises_par_id = calculer_montants_utilises(
                    conn, ids_financement
                )

            for cle, info in agregats["par_financement"].items():

                etat = recus_par_id.get(cle, {})
                montant_recu = etat.get("montant_recu") or 0
                montant_prevu = etat.get("montant_prevu") or 0
                montant_utilise = utilises_par_id.get(
                    cle, Decimal("0")
                )

                # Base de calcul : le reçu si dispo, sinon le prévu
                base = montant_recu or montant_prevu

                info["montant_restant"] = (
                    Decimal(str(base)) - montant_utilise
                )

                if base:
                    info["pct_utilisation"] = round(
                        float(montant_utilise) / base * 100,
                        1
                    )
                    info["base_calcul"] = (
                        "reçu" if montant_recu else "prévu"
                    )
                else:
                    info["pct_utilisation"] = None

            if filters["subvention_id"] and filters["subvention_id"] != "budget":

                row_fin = conn.execute("""
                    SELECT id, nom_subvention, nom_organisme,
                           montant_prevu, montant_recu
                    FROM engagement_subventions
                    WHERE id = ?
                """, (filters["subvention_id"],)).fetchone()

                if row_fin:

                    financement_selectionne = dict(row_fin)

                    montant_utilise = calculer_montants_utilises(
                        conn, [row_fin["id"]]
                    ).get(row_fin["id"], Decimal("0"))

                    financement_selectionne["montant_utilise"] = float(
                        montant_utilise
                    )

                    montant_recu = financement_selectionne["montant_recu"] or 0
                    montant_prevu = financement_selectionne["montant_prevu"] or 0

                    # Base de calcul : le reçu si dispo, sinon le prévu
                    base = montant_recu or montant_prevu

                    financement_selectionne["montant_restant"] = float(
                        Decimal(str(base)) - montant_utilise
                    )

                    if base:
                        pct = round(
                            float(montant_utilise) / base * 100,
                            1
                        )
                        financement_selectionne["pct_utilisation"] = pct
                        financement_selectionne["pct_largeur"] = min(pct, 100)
                        financement_selectionne["base_calcul"] = (
                            "reçu" if montant_recu else "prévu"
                        )
                    else:
                        financement_selectionne["pct_utilisation"] = None
                        financement_selectionne["pct_largeur"] = None

    return render_template(
        "engagements/reporting.html",
        filters=filters,
        poles=poles,
        rubriques=rubriques,
        subventions=subventions,
        rows=rows,
        agregats=agregats,
        financement_selectionne=financement_selectionne,
        has_search=has_search,
        statuts=STATUTS_LABELS,
        voir_tout=voir_tout,
    )


# ============================================================
# EXPORT EXCEL
# ============================================================

@engagements_bp.route(
    "/engagements/reporting/export-excel",
    methods=["POST"]
)
@login_required
@require_access("engagements", "lecture")
def engagements_reporting_excel():

    db_path = get_db_path()
    voir_tout = _est_admin()

    filters = {
        "date_debut":    request.form.get("date_debut", "").strip(),
        "date_fin":      request.form.get("date_fin", "").strip(),
        "demandeur":     request.form.get("demandeur", "").strip(),
        "fournisseur":   request.form.get("fournisseur", "").strip(),
        "numero_ecriture_ebp": request.form.get("numero_ecriture_ebp", "").strip(),
        "pole_id":       request.form.get("pole_id", "").strip(),
        "rubrique":      request.form.get("rubrique", "").strip(),
        "subvention_id": request.form.get("subvention_id", "").strip(),
        "statut":        request.form.get("statut", "").strip(),
    }

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        sql, params = _build_query(filters, voir_tout)
        raw = conn.execute(sql, params).fetchall()
        rows = [dict(r) for r in raw]

    agregats = _build_agregats(rows)

    # =========================================================
    # WORKBOOK
    # =========================================================

    wb = Workbook()

    # ---------------------------------------------------------
    # STYLES
    # ---------------------------------------------------------

    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill   = PatternFill("solid", start_color="1F3864")
    section_font  = Font(name="Arial", bold=True, size=10, color="1F3864")
    section_fill  = PatternFill("solid", start_color="D9E1F2")
    normal_font   = Font(name="Arial", size=10)
    total_font    = Font(name="Arial", bold=True, size=10)
    total_fill    = PatternFill("solid", start_color="F2F2F2")
    center        = Alignment(horizontal="center", vertical="center")
    left          = Alignment(horizontal="left",   vertical="center")
    right         = Alignment(horizontal="right",  vertical="center")

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(cell):
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    def style_normal(cell, align=left):
        cell.font      = normal_font
        cell.alignment = align
        cell.border    = border

    def style_total(cell, align=right):
        cell.font      = total_font
        cell.fill      = total_fill
        cell.alignment = align
        cell.border    = border

    # =========================================================
    # FEUILLE 1 — LISTE DÉTAILLÉE
    # =========================================================

    ws1 = wb.active
    ws1.title = "Détail engagements"

    # En-tête titre
    ws1.merge_cells("A1:N1")
    titre = ws1["A1"]
    titre.value = (
        f"Reporting Engagements — Extrait au "
        f"{datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    titre.font = Font(name="Arial", bold=True, size=12, color="1F3864")
    titre.alignment = center
    ws1.row_dimensions[1].height = 22

    # Ligne filtres
    ws1.merge_cells("A2:N2")
    filtres_txt = []
    if filters["date_debut"]:
        filtres_txt.append(f"Du {filters['date_debut']}")
    if filters["date_fin"]:
        filtres_txt.append(f"au {filters['date_fin']}")
    if filters["demandeur"]:
        filtres_txt.append(f"Demandeur : {filters['demandeur']}")
    if filters["fournisseur"]:
        filtres_txt.append(f"Fournisseur : {filters['fournisseur']}")
    filtres_cell = ws1["A2"]
    filtres_cell.value = "  |  ".join(filtres_txt) if filtres_txt else "Tous les engagements"
    filtres_cell.font = Font(name="Arial", italic=True, size=9, color="595959")
    filtres_cell.alignment = left

    ws1.row_dimensions[3].height = 4  # espace

    # Colonnes
    colonnes = [
        ("ID",          8),
        ("Date",       12),
        ("Demandeur",  22),
        ("Pôle",       20),
        ("Objet",      35),
        ("Rubrique",   18),
        ("Sous-type",  14),
        ("Bénéficiaire/Fournisseur", 28),
        ("Subvention", 22),
        ("Statut",     18),
        ("N° écriture EBP", 18),
        ("Montant (€)", 14),
    ]

    ROW_HEADER = 4
    for col_idx, (label, width) in enumerate(colonnes, 1):
        cell = ws1.cell(row=ROW_HEADER, column=col_idx, value=label)
        style_header(cell)
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    ws1.row_dimensions[ROW_HEADER].height = 18

    # Données
    for row_idx, r in enumerate(rows, ROW_HEADER + 1):

        montant = float(r["montant_total"] or 0)

        beneficiaire = (
            r["fournisseur_nom"]
            or r["beneficiaire_nom"]
            or r["demandeur_nom"]
            or "—"
        )

        statut_label = STATUTS_LABELS.get(
            r["statut"], r["statut"]
        )

        date_str = r["cree_le"][:10] if r["cree_le"] else ""

        vals = [
            r["id"],
            date_str,
            r["demandeur_nom"] or "",
            r["pole"] or "",
            r["objet"] or "",
            r["rubrique"] or "",
            r["sous_type_depense"] or "",
            beneficiaire,
            r["nom_subvention"] or "Budget propre",
            statut_label,
            r["numero_ecriture_ebp"] or "",
            montant,
        ]

        for col_idx, val in enumerate(vals, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 12:
                cell.number_format = '#,##0.00 €'
                style_normal(cell, right)
            elif col_idx == 1:
                style_normal(cell, center)
            else:
                style_normal(cell, left)

        # Zébrage
        if row_idx % 2 == 0:
            for col_idx in range(1, 13):
                ws1.cell(row=row_idx, column=col_idx).fill = \
                    PatternFill("solid", start_color="EEF2FA")

    # Ligne total
    row_total = ROW_HEADER + len(rows) + 1
    ws1.merge_cells(
        f"A{row_total}:K{row_total}"
    )
    total_label = ws1[f"A{row_total}"]
    total_label.value = f"TOTAL — {len(rows)} engagement(s)"
    style_total(total_label, left)

    total_montant = ws1.cell(
        row=row_total,
        column=12,
        value=f"=SUM(L{ROW_HEADER + 1}:L{row_total - 1})"
    )
    total_montant.number_format = '#,##0.00 €'
    style_total(total_montant, right)

    # Freeze
    ws1.freeze_panes = f"A{ROW_HEADER + 1}"

    # =========================================================
    # FEUILLE 2 — AGRÉGATS PAR PÔLE
    # =========================================================

    ws2 = wb.create_sheet("Par pôle")

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 14

    ws2.merge_cells("A1:C1")
    t = ws2["A1"]
    t.value = "Répartition par pôle"
    t.font = Font(name="Arial", bold=True, size=12, color="1F3864")
    t.alignment = center
    ws2.row_dimensions[1].height = 22

    for col, label in [(1, "Pôle"), (2, "Montant (€)"), (3, "Nb engagements")]:
        cell = ws2.cell(row=3, column=col, value=label)
        style_header(cell)

    # Nb par pôle depuis rows
    nb_par_pole = {}
    for r in rows:
        pole = r["pole"] or "—"
        nb_par_pole[pole] = nb_par_pole.get(pole, 0) + 1

    for i, (pole, montant) in enumerate(
        agregats["par_pole"].items(), 4
    ):
        c1 = ws2.cell(row=i, column=1, value=pole)
        c2 = ws2.cell(row=i, column=2, value=float(montant))
        c3 = ws2.cell(row=i, column=3, value=nb_par_pole.get(pole, 0))
        c2.number_format = '#,##0.00 €'
        style_normal(c1)
        style_normal(c2, right)
        style_normal(c3, center)
        if i % 2 == 0:
            for c in [c1, c2, c3]:
                c.fill = PatternFill("solid", start_color="EEF2FA")

    row_t = 4 + len(agregats["par_pole"])
    ws2.merge_cells(f"A{row_t}:A{row_t}")
    c_lbl = ws2.cell(row=row_t, column=1, value="TOTAL")
    c_mon = ws2.cell(
        row=row_t, column=2,
        value=f"=SUM(B4:B{row_t - 1})"
    )
    c_nb = ws2.cell(
        row=row_t, column=3,
        value=f"=SUM(C4:C{row_t - 1})"
    )
    c_mon.number_format = '#,##0.00 €'
    for c in [c_lbl, c_mon, c_nb]:
        style_total(c, right)

    # =========================================================
    # FEUILLE 3 — AGRÉGATS PAR RUBRIQUE
    # =========================================================

    ws3 = wb.create_sheet("Par rubrique")

    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 14

    ws3.merge_cells("A1:C1")
    t3 = ws3["A1"]
    t3.value = "Répartition par rubrique"
    t3.font = Font(name="Arial", bold=True, size=12, color="1F3864")
    t3.alignment = center
    ws3.row_dimensions[1].height = 22

    for col, label in [(1, "Rubrique"), (2, "Montant (€)"), (3, "Nb engagements")]:
        cell = ws3.cell(row=3, column=col, value=label)
        style_header(cell)

    nb_par_rubrique = {}
    for r in rows:
        rub = r["rubrique"] or "—"
        nb_par_rubrique[rub] = nb_par_rubrique.get(rub, 0) + 1

    for i, (rub, montant) in enumerate(
        agregats["par_rubrique"].items(), 4
    ):
        c1 = ws3.cell(row=i, column=1, value=rub)
        c2 = ws3.cell(row=i, column=2, value=float(montant))
        c3 = ws3.cell(row=i, column=3, value=nb_par_rubrique.get(rub, 0))
        c2.number_format = '#,##0.00 €'
        style_normal(c1)
        style_normal(c2, right)
        style_normal(c3, center)
        if i % 2 == 0:
            for c in [c1, c2, c3]:
                c.fill = PatternFill("solid", start_color="EEF2FA")

    row_t3 = 4 + len(agregats["par_rubrique"])
    c_lbl3 = ws3.cell(row=row_t3, column=1, value="TOTAL")
    c_mon3 = ws3.cell(
        row=row_t3, column=2,
        value=f"=SUM(B4:B{row_t3 - 1})"
    )
    c_nb3 = ws3.cell(
        row=row_t3, column=3,
        value=f"=SUM(C4:C{row_t3 - 1})"
    )
    c_mon3.number_format = '#,##0.00 €'
    for c in [c_lbl3, c_mon3, c_nb3]:
        style_total(c, right)

    # =========================================================
    # ENVOI
    # =========================================================

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = (
        f"reporting_engagements_"
        f"{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )

    write_log(
        f"[REPORTING] Export Excel par {current_user.email}"
    )

    return send_file(
        output,
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        download_name=filename,
        as_attachment=True,
    )


# ============================================================
# EXPORT PDF
# ============================================================

@engagements_bp.route(
    "/engagements/reporting/export-pdf",
    methods=["POST"]
)
@login_required
@require_access("engagements", "lecture")
def engagements_reporting_pdf():

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
            HRFlowable,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    except ImportError:
        flash(
            "⚠️ La librairie reportlab est requise "
            "pour l'export PDF.",
            "danger"
        )
        return redirect(
            url_for("engagements.engagements_reporting")
        )

    db_path = get_db_path()
    voir_tout = _est_admin()

    filters = {
        "date_debut":    request.form.get("date_debut", "").strip(),
        "date_fin":      request.form.get("date_fin", "").strip(),
        "demandeur":     request.form.get("demandeur", "").strip(),
        "fournisseur":   request.form.get("fournisseur", "").strip(),
        "numero_ecriture_ebp": request.form.get("numero_ecriture_ebp", "").strip(),
        "pole_id":       request.form.get("pole_id", "").strip(),
        "rubrique":      request.form.get("rubrique", "").strip(),
        "subvention_id": request.form.get("subvention_id", "").strip(),
        "statut":        request.form.get("statut", "").strip(),
    }

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        sql, params = _build_query(filters, voir_tout)
        raw = conn.execute(sql, params).fetchall()
        rows = [dict(r) for r in raw]

    agregats = _build_agregats(rows)

    # ---------------------------------------------------------
    # DOCUMENT
    # ---------------------------------------------------------

    output = BytesIO()

    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()

    style_titre = ParagraphStyle(
        "titre",
        parent=styles["Heading1"],
        fontSize=14,
        textColor=colors.HexColor("#1F3864"),
        spaceAfter=4,
    )

    style_sous_titre = ParagraphStyle(
        "sous_titre",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#595959"),
        spaceAfter=8,
        fontName="Helvetica-Oblique",
    )

    style_section = ParagraphStyle(
        "section",
        parent=styles["Heading2"],
        fontSize=11,
        textColor=colors.HexColor("#1F3864"),
        spaceBefore=12,
        spaceAfter=4,
    )

    BLEU   = colors.HexColor("#1F3864")
    BLEU_L = colors.HexColor("#D9E1F2")
    GRIS   = colors.HexColor("#F2F2F2")
    BLANC  = colors.white

    story = []

    # Titre
    story.append(
        Paragraph(
            "Reporting — Engagements de dépenses",
            style_titre
        )
    )

    filtres_txt = []
    if filters["date_debut"]:
        filtres_txt.append(f"Du {filters['date_debut']}")
    if filters["date_fin"]:
        filtres_txt.append(f"au {filters['date_fin']}")
    if filters["demandeur"]:
        filtres_txt.append(f"Demandeur : {filters['demandeur']}")
    if filters["fournisseur"]:
        filtres_txt.append(f"Fournisseur : {filters['fournisseur']}")

    sous_titre_txt = (
        ("  |  ".join(filtres_txt) if filtres_txt else "Tous les engagements")
        + f"  —  Extrait le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    )

    story.append(Paragraph(sous_titre_txt, style_sous_titre))
    story.append(HRFlowable(width="100%", thickness=1, color=BLEU))
    story.append(Spacer(1, 6))

    # -------------------------------------------------------
    # SYNTHÈSE CHIFFRES CLÉS
    # -------------------------------------------------------

    story.append(Paragraph("Synthèse", style_section))

    synthese_data = [
        ["Engagements", "Montant total", "À payer", "Réglés"],
    ]

    nb_a_payer   = sum(1 for r in rows if r["statut"] == "a_payer")
    nb_reglees   = sum(1 for r in rows if r["statut"] in STATUTS_REGLES)
    mt_a_payer   = sum(
        float(r["montant_total"] or 0)
        for r in rows
        if r["statut"] == "a_payer"
    )

    synthese_data.append([
        str(agregats["nb_engagements"]),
        f"{float(agregats['total_global']):,.2f} €".replace(",", " "),
        f"{nb_a_payer} ({mt_a_payer:,.2f} €)".replace(",", " "),
        str(nb_reglees),
    ])

    t_synthese = Table(
        synthese_data,
        colWidths=[60 * mm, 70 * mm, 80 * mm, 50 * mm]
    )
    t_synthese.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), BLEU),
        ("TEXTCOLOR",   (0, 0), (-1, 0), BLANC),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 10),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("BACKGROUND",  (0, 1), (-1, 1), GRIS),
        ("FONTNAME",    (0, 1), (-1, 1), "Helvetica-Bold"),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
        ("TOPPADDING",  (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(t_synthese)
    story.append(Spacer(1, 10))

    # -------------------------------------------------------
    # PAR PÔLE
    # -------------------------------------------------------

    story.append(Paragraph("Répartition par pôle", style_section))

    pole_data = [["Pôle", "Montant (€)", "Nb"]]
    nb_par_pole = {}
    for r in rows:
        pole = r["pole"] or "—"
        nb_par_pole[pole] = nb_par_pole.get(pole, 0) + 1

    for pole, montant in agregats["par_pole"].items():
        pole_data.append([
            pole,
            f"{float(montant):,.2f} €".replace(",", " "),
            str(nb_par_pole.get(pole, 0)),
        ])

    pole_data.append([
        "TOTAL",
        f"{float(agregats['total_global']):,.2f} €".replace(",", " "),
        str(agregats["nb_engagements"]),
    ])

    t_pole = Table(
        pole_data,
        colWidths=[90 * mm, 60 * mm, 30 * mm]
    )
    t_pole.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0),  (-1, 0),  BLEU),
        ("TEXTCOLOR",     (0, 0),  (-1, 0),  BLANC),
        ("FONTNAME",      (0, 0),  (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0),  (-1, -1), 9),
        ("ALIGN",         (1, 0),  (-1, -1), "RIGHT"),
        ("ALIGN",         (0, 0),  (0, -1),  "LEFT"),
        ("BACKGROUND",    (0, -1), (-1, -1), GRIS),
        ("FONTNAME",      (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS",(0, 1), (-1, -2), [BLANC, BLEU_L]),
        ("GRID",          (0, 0),  (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
        ("TOPPADDING",    (0, 0),  (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0),  (-1, -1), 4),
    ]))
    story.append(t_pole)
    story.append(Spacer(1, 10))

    # -------------------------------------------------------
    # PAR RUBRIQUE
    # -------------------------------------------------------

    story.append(Paragraph("Répartition par rubrique", style_section))

    rub_data = [["Rubrique", "Montant (€)", "Nb"]]
    nb_par_rub = {}
    for r in rows:
        rub = r["rubrique"] or "—"
        nb_par_rub[rub] = nb_par_rub.get(rub, 0) + 1

    for rub, montant in agregats["par_rubrique"].items():
        rub_data.append([
            rub,
            f"{float(montant):,.2f} €".replace(",", " "),
            str(nb_par_rub.get(rub, 0)),
        ])

    rub_data.append([
        "TOTAL",
        f"{float(agregats['total_global']):,.2f} €".replace(",", " "),
        str(agregats["nb_engagements"]),
    ])

    t_rub = Table(
        rub_data,
        colWidths=[90 * mm, 60 * mm, 30 * mm]
    )
    t_rub.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0),  (-1, 0),  BLEU),
        ("TEXTCOLOR",     (0, 0),  (-1, 0),  BLANC),
        ("FONTNAME",      (0, 0),  (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0),  (-1, -1), 9),
        ("ALIGN",         (1, 0),  (-1, -1), "RIGHT"),
        ("ALIGN",         (0, 0),  (0, -1),  "LEFT"),
        ("BACKGROUND",    (0, -1), (-1, -1), GRIS),
        ("FONTNAME",      (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS",(0, 1), (-1, -2), [BLANC, BLEU_L]),
        ("GRID",          (0, 0),  (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
        ("TOPPADDING",    (0, 0),  (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0),  (-1, -1), 4),
    ]))
    story.append(t_rub)
    story.append(Spacer(1, 14))

    # -------------------------------------------------------
    # LISTE DÉTAILLÉE
    # -------------------------------------------------------

    story.append(Paragraph("Liste détaillée", style_section))

    detail_headers = [
        "ID", "Date", "Demandeur", "Pôle",
        "Objet", "Rubrique", "Bénéficiaire / Fournisseur",
        "Subvention", "Statut", "N° écriture EBP", "Montant €"
    ]
    detail_data = [detail_headers]

    for r in rows:

        beneficiaire = (
            r["fournisseur_nom"]
            or r["beneficiaire_nom"]
            or r["demandeur_nom"]
            or "—"
        )

        date_str = r["cree_le"][:10] if r["cree_le"] else ""
        montant_str = (
            f"{float(r['montant_total'] or 0):,.2f}"
            .replace(",", " ")
        )

        detail_data.append([
            str(r["id"]),
            date_str,
            (r["demandeur_nom"] or "")[:16],
            (r["pole"] or "")[:14],
            (r["objet"] or "")[:26],
            (r["rubrique"] or "")[:14],
            beneficiaire[:20],
            (r["nom_subvention"] or "Budget propre")[:16],
            STATUTS_LABELS.get(r["statut"], r["statut"]),
            r["numero_ecriture_ebp"] or "—",
            montant_str,
        ])

    # Ligne total
    detail_data.append([
        "", "", "", "", "", "", "", "", "",
        "TOTAL",
        f"{float(agregats['total_global']):,.2f}".replace(",", " "),
    ])

    col_widths = [
        10*mm, 18*mm, 28*mm, 24*mm,
        36*mm, 20*mm, 30*mm, 24*mm,
        20*mm, 20*mm, 20*mm
    ]

    t_detail = Table(detail_data, colWidths=col_widths, repeatRows=1)
    t_detail.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0),  (-1, 0),  BLEU),
        ("TEXTCOLOR",     (0, 0),  (-1, 0),  BLANC),
        ("FONTNAME",      (0, 0),  (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0),  (-1, -1), 7.5),
        ("ALIGN",         (0, 0),  (0, -1),  "CENTER"),
        ("ALIGN",         (10, 0), (10, -1), "RIGHT"),
        ("BACKGROUND",    (0, -1), (-1, -1), GRIS),
        ("FONTNAME",      (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS",(0, 1), (-1, -2), [BLANC, BLEU_L]),
        ("GRID",          (0, 0),  (-1, -1), 0.4, colors.HexColor("#BFBFBF")),
        ("TOPPADDING",    (0, 0),  (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0),  (-1, -1), 3),
        ("VALIGN",        (0, 0),  (-1, -1), "MIDDLE"),
    ]))
    story.append(t_detail)

    # ---------------------------------------------------------
    # GÉNÉRATION
    # ---------------------------------------------------------

    doc.build(story)
    output.seek(0)

    filename = (
        f"reporting_engagements_"
        f"{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    )

    write_log(
        f"[REPORTING] Export PDF par {current_user.email}"
    )

    return send_file(
        output,
        mimetype="application/pdf",
        download_name=filename,
        as_attachment=True,
    )