"""
routes_suivi_budgetaire.py
===========================

Suivi budgétaire des subventions / budgets.

Pour chaque subvention ou budget paramétré (engagement_subventions),
affiche le total dépensé et le total restant sur une période choisie
(année calendaire ou durée totale d'utilisation du budget), avec le
détail des dépenses rattachées.

Sorties :
- Tableau HTML par subvention/budget
- Export Excel (.xlsx)
- Export PDF (reportlab)

Accès : aligné sur la page Reporting (engagements/lecture, restriction
"mes engagements" pour les non-admins).
"""

from flask import (
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
)

from flask_login import login_required, current_user

from utils import (
    get_db_path,
    has_access,
    require_access,
)

from ba38_engagements import engagements_bp
from ba38_engagements.routes_reporting import STATUTS_LABELS

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from io import BytesIO
from datetime import datetime
from decimal import Decimal

import sqlite3


# ============================================================
# HELPERS
# ============================================================

def _est_admin():
    return has_access("engagement_parametres", "lecture")


def _date_fr_vers_iso(valeur):
    """Convertit une date au format jj/mm/aaaa (champs .date-fr) en ISO."""

    if not valeur:
        return None

    try:
        return datetime.strptime(valeur.strip(), "%d/%m/%Y").date().isoformat()
    except ValueError:
        return None


def _charger_donnees(conn, mode, annee, voir_tout):
    """
    Retourne la liste des subventions/budgets avec, pour chacun,
    la période considérée, le total dépensé, le total restant et
    le détail des dépenses rattachées.
    """

    subventions = conn.execute("""
        SELECT *
        FROM engagement_subventions
        ORDER BY nom_subvention
    """).fetchall()

    subventions_data = []

    for s in subventions:

        s = dict(s)

        if mode == "duree":
            periode_debut_affiche = s["utilisation_date_debut"] or None
            periode_fin_affiche = s["utilisation_date_fin"] or None
            date_debut_iso = _date_fr_vers_iso(periode_debut_affiche)
            date_fin_iso = _date_fr_vers_iso(periode_fin_affiche)
        else:
            periode_debut_affiche = f"01/01/{annee}"
            periode_fin_affiche = f"31/12/{annee}"
            date_debut_iso = f"{annee}-01-01"
            date_fin_iso = f"{annee}-12-31"

        where = [
            "d.subvention_id = ?",
            "COALESCE(e.deleted, 0) = 0",
            "e.statut != 'refuse'",
        ]
        params = [s["id"]]

        if not voir_tout:
            where.append("e.demandeur_id = ?")
            params.append(current_user.id)

        if date_debut_iso:
            where.append("DATE(e.cree_le) >= ?")
            params.append(date_debut_iso)

        if date_fin_iso:
            where.append("DATE(e.cree_le) <= ?")
            params.append(date_fin_iso)

        sql = f"""
            SELECT
                e.id AS engagement_id,
                e.cree_le,
                e.statut,
                e.demandeur_nom,
                d.objet,
                d.fournisseur_nom,
                d.beneficiaire_nom,
                d.montant_total
            FROM engagements_depenses d
            JOIN engagements e
                ON e.id = d.engagement_id
            WHERE {" AND ".join(where)}
            ORDER BY e.cree_le DESC
        """

        depenses = [dict(r) for r in conn.execute(sql, params).fetchall()]

        total_depense = sum(
            Decimal(str(d["montant_total"] or 0)) for d in depenses
        )

        base = Decimal(str(s["montant_recu"] or s["montant_prevu"] or 0))
        total_restant = base - total_depense

        subventions_data.append({
            "subvention": s,
            "periode": {
                "debut": periode_debut_affiche,
                "fin": periode_fin_affiche,
            },
            "base": float(base),
            "total_depense": float(total_depense),
            "total_restant": float(total_restant),
            "depenses": depenses,
        })

    return subventions_data


def _lire_filtres(source):
    annee_defaut = datetime.now().year
    mode = source.get("mode", "annee").strip() or "annee"
    try:
        annee = int(source.get("annee", annee_defaut))
    except (TypeError, ValueError):
        annee = annee_defaut
    return mode, annee


# ============================================================
# PAGE SUIVI BUDGÉTAIRE
# ============================================================

@engagements_bp.route(
    "/engagements/suivi-budgetaire",
    methods=["GET", "POST"]
)
@login_required
@require_access("engagements", "lecture")
def suivi_budgetaire():

    db_path = get_db_path()
    voir_tout = _est_admin()

    mode, annee = _lire_filtres(
        request.form if request.method == "POST" else request.args
    )

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        subventions_data = _charger_donnees(conn, mode, annee, voir_tout)

    return render_template(
        "engagements/suivi_budgetaire.html",
        subventions_data=subventions_data,
        mode=mode,
        annee=annee,
        voir_tout=voir_tout,
        statuts_labels=STATUTS_LABELS,
    )


# ============================================================
# EXPORT EXCEL
# ============================================================

@engagements_bp.route(
    "/engagements/suivi-budgetaire/export-excel",
    methods=["POST"]
)
@login_required
@require_access("engagements", "lecture")
def suivi_budgetaire_excel():

    db_path = get_db_path()
    voir_tout = _est_admin()
    mode, annee = _lire_filtres(request.form)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        subventions_data = _charger_donnees(conn, mode, annee, voir_tout)

    wb = Workbook()

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="1F3864")
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", bold=True, size=10)
    total_fill = PatternFill("solid", start_color="F2F2F2")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(cell):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    def style_normal(cell, align=left):
        cell.font = normal_font
        cell.alignment = align
        cell.border = border

    def style_total(cell, align=right):
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = align
        cell.border = border

    periode_txt = (
        f"Année {annee}" if mode == "annee"
        else "Durée totale d'utilisation du budget"
    )

    # ------------------------------------------------------
    # FEUILLE 1 — SYNTHÈSE
    # ------------------------------------------------------

    ws1 = wb.active
    ws1.title = "Synthèse"

    ws1.merge_cells("A1:H1")
    titre = ws1["A1"]
    titre.value = (
        f"Suivi budgétaire — {periode_txt} — Extrait au "
        f"{datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    titre.font = Font(name="Arial", bold=True, size=12, color="1F3864")
    titre.alignment = center
    ws1.row_dimensions[1].height = 22

    colonnes1 = [
        ("Subvention / Budget", 30),
        ("Organisme", 22),
        ("Période", 24),
        ("Prévu (€)", 14),
        ("Reçu (€)", 14),
        ("Dépensé (€)", 14),
        ("Restant (€)", 14),
    ]

    ROW_HEADER = 3
    for col_idx, (label, width) in enumerate(colonnes1, 1):
        cell = ws1.cell(row=ROW_HEADER, column=col_idx, value=label)
        style_header(cell)
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    row_idx = ROW_HEADER + 1
    total_depense_global = Decimal("0")
    total_restant_global = Decimal("0")

    for item in subventions_data:

        s = item["subvention"]
        periode = item["periode"]
        periode_cell = (
            f"{periode['debut'] or '—'} → {periode['fin'] or '—'}"
        )

        vals = [
            s["nom_subvention"],
            s["nom_organisme"] or "",
            periode_cell,
            float(s["montant_prevu"] or 0),
            float(s["montant_recu"] or 0),
            item["total_depense"],
            item["total_restant"],
        ]

        for col_idx, val in enumerate(vals, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            if col_idx >= 4:
                cell.number_format = '#,##0.00 €'
                style_normal(cell, right)
            else:
                style_normal(cell, left)

        if row_idx % 2 == 0:
            for col_idx in range(1, 8):
                ws1.cell(row=row_idx, column=col_idx).fill = \
                    PatternFill("solid", start_color="EEF2FA")

        total_depense_global += Decimal(str(item["total_depense"]))
        total_restant_global += Decimal(str(item["total_restant"]))

        row_idx += 1

    total_row = row_idx
    ws1.cell(row=total_row, column=1, value="TOTAL")
    ws1.merge_cells(
        start_row=total_row, start_column=1,
        end_row=total_row, end_column=5
    )
    for col_idx in range(1, 6):
        style_total(ws1.cell(row=total_row, column=col_idx), left)

    cell = ws1.cell(row=total_row, column=6, value=float(total_depense_global))
    cell.number_format = '#,##0.00 €'
    style_total(cell)

    cell = ws1.cell(row=total_row, column=7, value=float(total_restant_global))
    cell.number_format = '#,##0.00 €'
    style_total(cell)

    ws1.freeze_panes = f"A{ROW_HEADER + 1}"

    # ------------------------------------------------------
    # FEUILLE 2 — DÉTAIL DES DÉPENSES
    # ------------------------------------------------------

    ws2 = wb.create_sheet("Détail dépenses")

    colonnes2 = [
        ("Subvention / Budget", 28),
        ("ID engagement", 12),
        ("Date", 12),
        ("Demandeur", 22),
        ("Objet", 32),
        ("Bénéficiaire / Fournisseur", 26),
        ("Statut", 18),
        ("Montant (€)", 14),
    ]

    for col_idx, (label, width) in enumerate(colonnes2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=label)
        style_header(cell)
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    row_idx = 2
    for item in subventions_data:

        s = item["subvention"]

        for d in item["depenses"]:

            beneficiaire = (
                d["fournisseur_nom"]
                or d["beneficiaire_nom"]
                or d["demandeur_nom"]
                or "—"
            )

            statut_label = STATUTS_LABELS.get(d["statut"], d["statut"])
            date_str = d["cree_le"][:10] if d["cree_le"] else ""

            vals = [
                s["nom_subvention"],
                d["engagement_id"],
                date_str,
                d["demandeur_nom"] or "",
                d["objet"] or "",
                beneficiaire,
                statut_label,
                float(d["montant_total"] or 0),
            ]

            for col_idx, val in enumerate(vals, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                if col_idx == 8:
                    cell.number_format = '#,##0.00 €'
                    style_normal(cell, right)
                elif col_idx == 2:
                    style_normal(cell, center)
                else:
                    style_normal(cell, left)

            if row_idx % 2 == 0:
                for col_idx in range(1, 9):
                    ws2.cell(row=row_idx, column=col_idx).fill = \
                        PatternFill("solid", start_color="EEF2FA")

            row_idx += 1

    ws2.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
        download_name=(
            f"suivi_budgetaire_{datetime.now():%Y%m%d_%H%M}.xlsx"
        ),
        as_attachment=True,
    )


# ============================================================
# EXPORT PDF
# ============================================================

@engagements_bp.route(
    "/engagements/suivi-budgetaire/export-pdf",
    methods=["POST"]
)
@login_required
@require_access("engagements", "lecture")
def suivi_budgetaire_pdf():

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    except ImportError:
        flash(
            "⚠️ La librairie reportlab est requise "
            "pour l'export PDF.",
            "danger"
        )
        return redirect(url_for("engagements.suivi_budgetaire"))

    db_path = get_db_path()
    voir_tout = _est_admin()
    mode, annee = _lire_filtres(request.form)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        subventions_data = _charger_donnees(conn, mode, annee, voir_tout)

    periode_txt = (
        f"Année {annee}" if mode == "annee"
        else "Durée totale d'utilisation du budget"
    )

    output = BytesIO()

    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()

    style_titre = ParagraphStyle(
        "titre",
        parent=styles["Heading1"],
        fontSize=14,
        textColor=colors.HexColor("#1F3864"),
        spaceAfter=4,
    )

    style_sous_titre = ParagraphStyle(
        "sous_titre",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#595959"),
        spaceAfter=10,
        fontName="Helvetica-Oblique",
    )

    style_section = ParagraphStyle(
        "section",
        parent=styles["Heading2"],
        fontSize=11,
        textColor=colors.HexColor("#1F3864"),
        spaceBefore=14,
        spaceAfter=4,
    )

    BLEU = colors.HexColor("#1F3864")
    BLEU_L = colors.HexColor("#D9E1F2")
    GRIS = colors.HexColor("#F2F2F2")
    BLANC = colors.white

    story = []

    story.append(Paragraph("Suivi budgétaire — Subventions et budgets", style_titre))
    story.append(Paragraph(
        f"{periode_txt} — Extrait au "
        f"{datetime.now().strftime('%d/%m/%Y %H:%M')}",
        style_sous_titre
    ))

    for item in subventions_data:

        s = item["subvention"]
        periode = item["periode"]

        story.append(Paragraph(
            f"{s['nom_subvention']}"
            + (f" — {s['nom_organisme']}" if s["nom_organisme"] else ""),
            style_section
        ))

        periode_cell = f"{periode['debut'] or '—'} → {periode['fin'] or '—'}"

        infos_data = [
            ["Période", "Prévu (€)", "Reçu (€)", "Dépensé (€)", "Restant (€)"],
            [
                periode_cell,
                f"{float(s['montant_prevu'] or 0):,.2f}".replace(",", " "),
                f"{float(s['montant_recu'] or 0):,.2f}".replace(",", " "),
                f"{item['total_depense']:,.2f}".replace(",", " "),
                f"{item['total_restant']:,.2f}".replace(",", " "),
            ],
        ]

        infos_table = Table(
            infos_data,
            colWidths=[70 * mm, 45 * mm, 45 * mm, 45 * mm, 45 * mm]
        )
        infos_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BLEU),
            ("TEXTCOLOR", (0, 0), (-1, 0), BLANC),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
            ("BACKGROUND", (0, 1), (-1, 1), GRIS),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))

        story.append(infos_table)
        story.append(Spacer(1, 6))

        if item["depenses"]:

            dep_header = ["Date", "Objet", "Bénéficiaire / Fournisseur", "Statut", "Montant (€)"]
            dep_rows = [dep_header]

            for d in item["depenses"]:

                beneficiaire = (
                    d["fournisseur_nom"]
                    or d["beneficiaire_nom"]
                    or d["demandeur_nom"]
                    or "—"
                )
                statut_label = STATUTS_LABELS.get(d["statut"], d["statut"])
                date_str = d["cree_le"][8:10] + "/" + d["cree_le"][5:7] + "/" + d["cree_le"][0:4] if d["cree_le"] else ""

                dep_rows.append([
                    date_str,
                    d["objet"] or "—",
                    beneficiaire,
                    statut_label,
                    f"{float(d['montant_total'] or 0):,.2f}".replace(",", " "),
                ])

            dep_table = Table(
                dep_rows,
                colWidths=[25 * mm, 90 * mm, 60 * mm, 35 * mm, 40 * mm],
                repeatRows=1,
            )

            dep_style = [
                ("BACKGROUND", (0, 0), (-1, 0), BLEU_L),
                ("TEXTCOLOR", (0, 0), (-1, 0), BLEU),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (4, 0), (4, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D9D9")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]

            for i in range(1, len(dep_rows)):
                if i % 2 == 0:
                    dep_style.append(
                        ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F7F9FC"))
                    )

            dep_table.setStyle(TableStyle(dep_style))
            story.append(dep_table)

        else:
            story.append(Paragraph(
                "Aucune dépense sur la période.",
                style_sous_titre
            ))

        story.append(Spacer(1, 4))

    doc.build(story)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/pdf",
        download_name=(
            f"suivi_budgetaire_{datetime.now():%Y%m%d_%H%M}.pdf"
        ),
        as_attachment=True,
    )
