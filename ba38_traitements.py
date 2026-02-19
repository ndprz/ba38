# ba38_traitements.py
import io, re, os
import pandas as pd
import sqlite3
import json
import tempfile
import uuid
import tempfile
import json

from datetime import datetime
from collections import defaultdict
from pathlib import Path
from flask import Blueprint, request, render_template, flash, redirect, url_for, send_file,abort,current_app, session
from flask_login import login_required
from utils import get_google_services, write_log, envoyer_mail,get_db_path,upload_file_to_drive_path,slugify_filename
from utils import get_drive_folder_id_from_path
from openpyxl.utils import get_column_letter


from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from googleapiclient.discovery_cache.base import Cache
from google.oauth2 import service_account
from googleapiclient.discovery import build


from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from datetime import date
from pathlib import Path

BA380_SHARED_DRIVE_ID = os.getenv("BA380_SHARED_DRIVE_ID")

if not BA380_SHARED_DRIVE_ID:
    raise RuntimeError("BA380_SHARED_DRIVE_ID non défini dans l'environnement")


# ===============================
# 📌 Blueprint
# ===============================
traitements_bp = Blueprint("traitements", __name__)

# Ancien dossier générique (pour tests)
FOLDER_ID_TRAITEMENTS = "1_RiRtqyjwxcgCo9csqL8ckjePmaFwviy"


# ===============================
# 🛠️ Menu utilitaires (accessible à tous les utilisateurs connectés)
# ===============================
@traitements_bp.route("/utilitaires")
@login_required
def utilitaires():
    return render_template("utilitaires.html")



# ===============================
# 📂 Traitement simple Drive (Excel / CSV)
# ===============================
@traitements_bp.route("/traitement_drive", methods=["GET", "POST"])
@login_required
def traitement_drive():
    """
    📂 Liste les fichiers d’un dossier Google Drive et affiche un aperçu avec pandas.
    """
    client, drive_service, creds = get_google_services()
    if drive_service is None:
        flash("❌ Connexion Google Drive impossible", "danger")
        return "Erreur Drive"

    try:
        results = drive_service.files().list(
            q=f"'{FOLDER_ID_TRAITEMENTS}' in parents and trashed=false",
            fields="files(id, name)",
        ).execute()
        fichiers = results.get("files", [])
    except Exception as e:
        write_log(f"❌ Erreur accès Drive : {e}")
        flash("Erreur d’accès au dossier Google Drive", "danger")
        return "Erreur Drive"

    if request.method == "POST":
        file_id = request.form.get("file_id")
        if not file_id:
            flash("❌ Aucun fichier sélectionné", "danger")
        else:
            try:
                request_dl = drive_service.files().get_media(fileId=file_id)
                fh = io.BytesIO()
                downloader = MediaIoBaseDownload(fh, request_dl)
                done = False
                while not done:
                    status, done = downloader.next_chunk()

                fh.seek(0)
                try:
                    df = pd.read_excel(fh)
                except Exception:
                    fh.seek(0)
                    df = pd.read_csv(fh, sep=";")

                flash(f"✅ {len(df)} lignes lues dans {df.shape[1]} colonnes", "success")
                return render_template("traitement_drive.html", fichiers=fichiers, apercu=df.head().to_html())

            except Exception as e:
                write_log(f"❌ Erreur traitement fichier Drive : {e}")
                flash(f"Erreur traitement fichier : {e}", "danger")

    return render_template("traitement_drive.html", fichiers=fichiers)


# ===============================
# 📅 Utilitaire jour de semaine
# ===============================
def jour_semaine(date_str):
    try:
        d = datetime.strptime(date_str, "%d/%m/%Y")
        return d.weekday()  # 0=lundi … 6=dimanche
    except:
        return None

# ===============================================
# 🧹 Supprimer complètement le contenu d’un dossier Drive
# (Drive partagé, pagination, attente réelle)
# ===============================================
def delete_drive_folder_contents(drive_path, wait_until_empty=True, timeout=30):
    """
    Supprime TOUS les fichiers d'un dossier Google Drive existant
    (Drive partagé compatible), avec pagination.
    Optionnellement attend que le dossier soit réellement vide.
    """

    import os
    import time
    from utils import write_log, SERVICE_ACCOUNT_FILE
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    try:
        if not BA380_SHARED_DRIVE_ID:
            write_log("❌ BA380_SHARED_DRIVE_ID non défini")
            return

        credentials = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE,
            scopes=["https://www.googleapis.com/auth/drive"]
        )

        service = build("drive", "v3", credentials=credentials)

        # 🔎 Résolution du dossier (sans création)
        folder_id = get_drive_folder_id_from_path(
            drive_path,
            BA380_SHARED_DRIVE_ID
        )

        if not folder_id:
            write_log(f"⚠️ Dossier Drive inexistant : {drive_path}")
            return

        write_log(f"🧹 Nettoyage dossier Drive : {drive_path}")

        # ============================
        # 🔁 LISTE COMPLÈTE (pagination)
        # ============================
        files = []
        page_token = None

        while True:
            response = service.files().list(
                q=f"'{folder_id}' in parents and trashed=false",
                corpora="drive",
                driveId=BA380_SHARED_DRIVE_ID,
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
                fields="nextPageToken, files(id, name)",
                pageSize=1000,
                pageToken=page_token
            ).execute()

            files.extend(response.get("files", []))
            page_token = response.get("nextPageToken")

            if not page_token:
                break

        if not files:
            write_log("ℹ️ Aucun fichier à supprimer.")
            return

        write_log(f"🗑️ {len(files)} fichiers à supprimer…")

        # ============================
        # 🗑️ SUPPRESSION
        # ============================
        for i, f in enumerate(files, start=1):
            service.files().delete(
                fileId=f["id"],
                supportsAllDrives=True
            ).execute()
            write_log(f"🗑️ [{i}/{len(files)}] Supprimé : {f['name']}")

        write_log("✅ Suppression demandée pour tous les fichiers.")

        # ============================
        # ⏳ ATTENTE VIDAGE RÉEL
        # ============================
        if wait_until_empty:
            write_log("⏳ Attente vidage réel du dossier Drive…")
            start = time.time()

            while time.time() - start < timeout:
                remaining = service.files().list(
                    q=f"'{folder_id}' in parents and trashed=false",
                    corpora="drive",
                    driveId=BA380_SHARED_DRIVE_ID,
                    supportsAllDrives=True,
                    includeItemsFromAllDrives=True,
                    fields="files(id)",
                    pageSize=10
                ).execute().get("files", [])

                if not remaining:
                    write_log("✅ Dossier Drive confirmé vide.")
                    return

                time.sleep(1)

            write_log("⚠️ Timeout attente vidage Drive (poursuite quand même)")

    except Exception as e:
        write_log(f"❌ Erreur delete_drive_folder_contents : {e}")

# ===============================
# 📂 Traitement fichier participation
# ===============================
@traitements_bp.route("/traitement_participation", methods=["GET", "POST"])
@login_required
def traitement_participation():
    """
    - Lit les .txt dans le dossier Drive défini par DOSSIER_PARTICIPATION (.env)
    - Supprime les lignes ven/sam/dim, recalcule les totaux
    - Crée/choisit un sous-dossier TrimN_YYYY sous DOSSIER_PARTICIPATION
      * S'il existe déjà : on le garde, on SUPPRIME TOUT SON CONTENU
      * On supprime aussi d'éventuels DOUBLONS de dossiers homonymes
    - Dépose 3 fichiers dedans (corrigé, lignes_supprimées, analyse),
      suffixés par _TrimN_YYYY
    """
    import io, os, re
    from datetime import datetime
    from collections import OrderedDict

    DOSSIER_PARTICIPATION = os.getenv("DOSSIER_PARTICIPATION")

    if not DOSSIER_PARTICIPATION:
        flash("❌ Variable d’environnement DOSSIER_PARTICIPATION manquante.", "danger")
        return redirect(url_for("traitements.utilitaires"))
    
    client, service, creds = get_google_services()
    if service is None:
        flash("❌ Connexion Google Drive impossible", "danger")
        return redirect(url_for("traitements.utilitaires"))

    DOSSIER_PARTICIPATION = os.getenv("DOSSIER_PARTICIPATION")
    if not DOSSIER_PARTICIPATION:
        flash("❌ Variable d’environnement DOSSIER_PARTICIPATION manquante.", "danger")
        return redirect(url_for("traitements.utilitaires"))

    # ✅ AJOUT ICI
    client, service, creds = get_google_services()
    if service is None:
        flash("❌ Connexion Google Drive impossible", "danger")
        return redirect(url_for("traitements.utilitaires"))

    # -------- Helpers Drive --------
    def ensure_clean_trim_folder(parent_id: str, folder_name: str) -> str:
        """
        - Cherche tous les dossiers nommés `folder_name` sous `parent_id`
        - S'il y en a plusieurs: conserve le plus récent, supprime les autres
        - Vide le contenu du dossier conservé (supprime tous les fichiers)
        - S'il n'existe pas: le crée
        - Retourne l'id du dossier propre prêt à l'emploi
        """
        # Lister les dossiers homonymes
        query = (
            f"'{parent_id}' in parents and "
            f"name='{folder_name}' and "
            f"mimeType='application/vnd.google-apps.folder' and trashed=false"
        )
        res = service.files().list(
            q=query,
            fields="files(id, name, createdTime)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
        folders = res.get("files", [])

        folder_id = None
        if folders:
            # Garde le plus récent
            folders.sort(key=lambda x: x.get("createdTime", ""), reverse=True)
            folder_id = folders[0]["id"]
            # Supprime les doublons homonymes plus anciens
            for dup in folders[1:]:
                try:
                    service.files().delete(
                        fileId=dup["id"], supportsAllDrives=True
                    ).execute()
                    write_log(f"🗑️ Dossier dupliqué supprimé: {dup['id']}")
                except Exception as e:
                    write_log(f"⚠️ Impossible de supprimer un doublon: {e}")
        else:
            # Créer le dossier s'il n'existe pas
            meta = {
                "name": folder_name,
                "mimeType": "application/vnd.google-apps.folder",
                "parents": [parent_id],
            }
            folder = service.files().create(
                body=meta, fields="id", supportsAllDrives=True
            ).execute()
            folder_id = folder["id"]
            write_log(f"📂 Dossier créé: {folder_name} ({folder_id})")

        # Purger le contenu du dossier retenu (pas le dossier lui-même)
        try:
            res_children = service.files().list(
                q=f"'{folder_id}' in parents and trashed=false",
                fields="files(id,name)",
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
            ).execute()
            for f in res_children.get("files", []):
                try:
                    service.files().delete(
                        fileId=f["id"], supportsAllDrives=True
                    ).execute()
                    write_log(f"🧹 Supprimé du dossier {folder_name}: {f['name']}")
                except Exception as e:
                    write_log(f"⚠️ Impossible de supprimer {f['name']}: {e}")
        except Exception as e:
            write_log(f"⚠️ Purge du dossier échouée: {e}")

        return folder_id

    # -------- 1) Lister les .txt dans le dossier d’origine --------
    results = service.files().list(
        q=f"'{DOSSIER_PARTICIPATION}' in parents and trashed=false and name contains '.txt'",
        fields="files(id, name, mimeType)",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()

    fichiers = results.get("files", [])

    if request.method == "POST":
        file_id = request.form.get("file_id")
        if not file_id:
            flash("❌ Aucun fichier sélectionné", "danger")
            return redirect(url_for("traitements.traitement_participation"))

        # -------- 2) Télécharger le fichier source --------
        request_dl = service.files().get_media(fileId=file_id, supportsAllDrives=True)
        fh = io.BytesIO()
        from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
        downloader = MediaIoBaseDownload(fh, request_dl)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        fh.seek(0)

        # -------- 3) Lecture contenu (UTF-8, fallback CP1252) --------
        try:
            contenu = fh.read().decode("utf-8")
        except UnicodeDecodeError:
            fh.seek(0)
            contenu = fh.read().decode("cp1252")

        lignes = contenu.splitlines(keepends=True)

        # Utilitaires
        def jour_semaine(date_str: str):
            try:
                d = datetime.strptime(date_str, "%d/%m/%Y")
                return d.weekday()  # 0=lun … 6=dim
            except Exception:
                return None

        # -------- 4) Suffixe & dossier cible TrimN_YYYY --------
        premiere_date = None
        for l in lignes:
            m = re.match(r"^\s*(\d{2}/\d{2}/\d{4})", l)
            if m:
                try:
                    premiere_date = datetime.strptime(m.group(1), "%d/%m/%Y")
                    break
                except Exception:
                    pass

        suffixe = ""
        folder_id_cible = DOSSIER_PARTICIPATION
        if premiere_date:
            trimestre = (premiere_date.month - 1) // 3 + 1
            suffixe = f"_Trim{trimestre}_{premiere_date.year}"
            folder_name = f"Trim{trimestre}_{premiere_date.year}"
            # 👉 Ici on n’efface plus/ recrée pas le dossier : on le nettoie et supprime les doublons
            folder_id_cible = ensure_clean_trim_folder(DOSSIER_PARTICIPATION, folder_name)
        else:
            folder_name = "(SansDate)"  # info pour le flash

        # -------- 5) Découper en factures --------
        factures, facture = [], []
        for ligne in lignes:
            if ligne.strip().startswith("BA. de l'Isère"):
                if facture:
                    factures.append(facture)
                    facture = []
            facture.append(ligne)
        if facture:
            factures.append(facture)

        pat_detail = re.compile(
            r"^\s*(\d{2}/\d{2}/\d{4})\s+(\d+)\s+([\d\s.,]+)\s+([\d\s.,]+)\s*$"
        )

        # -------- 6) Traiter & cumuler les totaux --------
        factures_corrigees = []
        suppr_par_assoc = OrderedDict()
        total_general_suppr = 0.0
        total_general_corrige = 0.0

        for facture in factures:
            nouvelle_facture = []
            assoc = ""
            garder_facture = False
            total_assoc_suppr = 0.0

            for l in facture:
                ls = l.strip()

                if ls.startswith("Association"):
                    assoc = ls
                    if assoc not in suppr_par_assoc:
                        suppr_par_assoc[assoc] = []

                m = pat_detail.match(ls)
                if m:
                    date_str, nb_ben, participation, total = m.groups()
                    try:
                        total_val = float(total.replace(" ", "").replace(",", "."))
                    except Exception:
                        total_val = 0.0

                    wd = jour_semaine(date_str)
                    if wd in (4, 5, 6):  # ven/sam/dim => suppression
                        suppr_par_assoc.setdefault(assoc, []).append(ls)
                        total_assoc_suppr += total_val
                        total_general_suppr += total_val
                        continue
                    else:
                        garder_facture = True
                        total_general_corrige += total_val
                        nouvelle_facture.append(l)
                else:
                    nouvelle_facture.append(l)

            if garder_facture:
                factures_corrigees.append(nouvelle_facture)

            if assoc and total_assoc_suppr > 0:
                suppr_par_assoc[assoc].append(
                    f"TOTAL supprimé {assoc} : {total_assoc_suppr:.2f} €"
                )

        # -------- 7) Construire les sorties --------
        txt_corrige = "".join("".join(f) for f in factures_corrigees)
        txt_corrige += f"\n=== TOTAL GÉNÉRAL (corrigé) : {total_general_corrige:.2f} € ===\n"

        blocs = []
        for a, lignes_s in suppr_par_assoc.items():
            if not lignes_s:
                continue
            blocs.append(a + "\n" + "\n".join("  " + s for s in lignes_s) + "\n")
        blocs.append(f"\n=== TOTAL GÉNÉRAL SUPPRIMÉ : {total_general_suppr:.2f} € ===\n")
        txt_suppr = "".join(blocs)

        txt_analyse = (
            f"Total général corrigé : {total_general_corrige:.2f} €\n"
            f"Total supprimé : {total_general_suppr:.2f} €\n"
        )

        # -------- 8) Upload dans le dossier cible (UTF-8) --------
        def upload_txt(nom: str, contenu_txt: str, folder_id: str):
            chemin_tmp = f"/tmp/{nom}"
            with open(chemin_tmp, "w", encoding="utf-8", newline="") as f:
                f.write(contenu_txt)
            media = MediaFileUpload(chemin_tmp, mimetype="text/plain", resumable=False)
            meta = {"name": nom, "parents": [folder_id]}
            service.files().create(
                body=meta, media_body=media, fields="id", supportsAllDrives=True
            ).execute()

        fichier_nom = next((f["name"] for f in fichiers if f["id"] == file_id), "parsol2l.txt")
        base = fichier_nom[:-4] if fichier_nom.lower().endswith(".txt") else fichier_nom

        upload_txt(f"{base}_corrigé{suffixe}.txt", txt_corrige, folder_id_cible)
        upload_txt(f"{base}_lignes_supprimees{suffixe}.txt", txt_suppr, folder_id_cible)
        upload_txt(f"{base}_analyse{suffixe}.txt", txt_analyse, folder_id_cible)

        flash(
            f"✅ Traitement terminé — {total_general_suppr:.2f} € supprimés. "
            f"Fichiers déposés dans « {folder_name} ».",
            "success"
        )
        return redirect(url_for("traitements.traitement_participation"))

    return render_template("traitement_participation.html", fichiers=fichiers)

# ===============================
# 🗑️ Ancienne fonction simple (conservée pour tests)
# ===============================
def traiter_parsol(contenu):
    lignes = contenu.splitlines(keepends=True)
    factures, facture = [], []
    for ligne in lignes:
        if ligne.strip().startswith("BA. de l'Isère"):
            if facture:
                factures.append(facture)
                facture = []
        facture.append(ligne)
    if facture:
        factures.append(facture)

    factures_corrigees, lignes_supprimees = [], []
    total_general = 0.0

    for facture in factures:
        nouvelle_facture = []
        assoc = ""
        garder_facture = False

        for l in facture:
            if l.strip().startswith("Association"):
                assoc = l.strip()

            match = re.match(r"(\d{2}/\d{2}/\d{4})\s+(\d+)\s+([\d,]+)\s+([\d,]+)", l.strip())
            if match:
                date_str, nb, prix, total = match.groups()
                total = float(total.replace(",", "."))
                total_general += total
                wd = jour_semaine(date_str)
                if wd in (4, 5, 6):
                    lignes_supprimees.append(f"{assoc} → {l.strip()}\n")
                    continue
                else:
                    garder_facture = True
            nouvelle_facture.append(l)

        if garder_facture:
            factures_corrigees.append(nouvelle_facture)

    txt_corrige = "".join("".join(f) for f in factures_corrigees)
    txt_suppr = "".join(lignes_supprimees)
    txt_analyse = f"Total général du fichier original : {total_general:.2f} €\n"
    return txt_corrige, txt_suppr, txt_analyse



@traitements_bp.route("/aide/ba38_traitements")
def aide_traitements():    
    """
    Sert le fichier Markdown d'aide pour le module ba38_traitements.
    - En DEV : /home/ndprz/dev/docstech/ba38_traitements.md
    - En PROD : /home/ndprz/ba380/docstech/ba38_traitements.md
    """
    env = os.getenv("ENVIRONMENT", "prod").lower()
    base_path = "/home/ndprz/ba380/docstech" if env == "prod" else "/home/ndprz/dev/docstech"
    file_path = os.path.join(base_path, "ba38_traitements.md")

    if not os.path.exists(file_path):
        abort(404, f"Fichier d'aide introuvable : {file_path}")

    return send_file(file_path, mimetype="text/markdown")

@traitements_bp.route("/aide/<page>")
def aide_page(page):
    mapping = {
        "traitement_participation": "ba38_traitements.md",
        # on pourra ajouter d'autres pages ici
    }
    if page not in mapping:
        abort(404)
    env = os.getenv("ENVIRONMENT", "prod").lower()
    base_path = "/home/ndprz/ba380/docstech" if env == "prod" else "/home/ndprz/dev/docstech"
    file_path = os.path.join(base_path, mapping[page])
    return send_file(file_path, mimetype="text/markdown")





def calculer_cotisations_par_annee(db_path, benef_par_vif):
    """
    benef_par_vif : dict { code_vif : total_beneficiaires }

    Retourne un dict :
    {
        "facturables": [ ... ],
        "orphelines": { code_vif: nb_benef }
    }
    """

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # ============================
    # Chargement des associations
    # ============================
    cursor.execute("""
        SELECT
            Id AS id_association,
            code_VIF,
            vif_regroup,
            nom_association,
            compte_comptable,
            adresse_association_1,
            adresse_association_2,
            CP,
            COMMUNE,
            courriel_association
        FROM associations
    """)
    assos = cursor.fetchall()

    # Index par code VIF
    asso_par_vif = {a["code_VIF"]: dict(a) for a in assos}

    # ============================
    # Cumul avec regroupements
    # ============================
    cumuls = defaultdict(int)
    rattachements = defaultdict(list)
    benef_sans_asso = defaultdict(int)

    for code_vif, nb in benef_par_vif.items():
        asso = asso_par_vif.get(code_vif)

        if not asso:
            benef_sans_asso[code_vif] += nb
            continue

        code_facture = asso["vif_regroup"] or code_vif
        cumuls[code_facture] += nb
        rattachements[code_facture].append(code_vif)

    # ============================
    # Construction du résultat
    # ============================
    resultats = []

    for code_facture, total_benef in cumuls.items():
        asso = asso_par_vif.get(code_facture)
        if not asso:
            continue

        # Adresse multi-lignes propre
        adresse = "\n".join(filter(None, [
            asso.get("adresse_association_1"),
            asso.get("adresse_association_2"),
            " ".join(filter(None, [
                asso.get("CP"),
                asso.get("COMMUNE")
            ]))
        ]))

        # Calcul cotisation
        if total_benef <= 1000:
            cotis = 50
        elif total_benef <= 10000:
            cotis = 80
        else:
            cotis = 110

        # ============================
        # Construction nom affiché
        # ============================

        noms_regroupes = []
        for code in rattachements[code_facture]:
            if code != code_facture:
                asso_r = asso_par_vif.get(code)
                if asso_r:
                    noms_regroupes.append(asso_r["nom_association"])

        nom_affiche = asso["nom_association"]
        if noms_regroupes:
            nom_affiche = nom_affiche + " / " + " / ".join(noms_regroupes)

        # ============================
        # Noms des associations regroupées
        # ============================

        noms_regroupes = []
        for code in rattachements[code_facture]:
            if code != code_facture:
                asso_r = asso_par_vif.get(code)
                if asso_r:
                    noms_regroupes.append(asso_r["nom_association"])

        # Nom STRICTEMENT de l'association maître (bloc adresse)
        nom_association_maitre = asso["nom_association"]

        # Nom affiché (tableau / Excel)
        nom_affichage = nom_association_maitre
        if noms_regroupes:
            nom_affichage = nom_association_maitre + " / " + " / ".join(noms_regroupes)

        # Texte concaténé (commentaire uniquement)
        commentaire_regroupement = None
        if noms_regroupes:
            commentaire_regroupement = (
                "Cette cotisation regroupe aussi les bénéficiaires des partenaires suivants :\n"
                + " / ".join(noms_regroupes)
            )


        resultats.append({
            "id_association": asso["id_association"],
            "code_vif_facture": code_facture,
            "compte_comptable": asso["compte_comptable"],
            "nom_association": nom_association_maitre,
            "nom_association_affichage": nom_affichage,
            "adresse": adresse,
            "email": asso.get("courriel_association"),
            "beneficiaires": total_benef,
            "cotisation": cotis,
            "codes_vif_inclus": rattachements[code_facture],
            "commentaire_regroupement": commentaire_regroupement

        })

    conn.close()

    return {
        "facturables": resultats,
        "orphelines": dict(benef_sans_asso)
    }



@traitements_bp.route("/cotisations", methods=["GET", "POST"])
@login_required
def cotisations():
    """
    Module principal de facturation des cotisations.

    Fonctionnement :
    - GET  : affiche les cotisations déjà calculées pour une année
    - POST : calcule les cotisations à partir du fichier PARSOL,
             insère en base (si année non verrouillée),
             puis affiche le résultat.

    Règle métier :
    - Si des factures ont déjà été envoyées en PROD pour une année,
      le recalcul est bloqué.
    - Le mode TEST ne bloque jamais.
    """

    from datetime import datetime

    mail_mode = session.get(
        "MAIL_MODE",
        os.getenv("MAIL_MODE", "PROD").upper()
    )

    mail_test_to = os.getenv(
        "MAIL_TEST_TO",
        "ba380.informatique2@banquealimentaire.org"
    )

    resultats = None
    orphelines = None
    annee = request.args.get("annee")

    conn = sqlite3.connect(get_db_path())
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # ==========================================================
    # POST → CALCUL NOUVELLE ANNÉE
    # ==========================================================
    if request.method == "POST":
        write_log("DEBUG POST COTISATIONS")

        try:
            annee = int(request.form.get("annee"))

            # --------------------------------------------------
            # 🔒 BLOCAGE SI ANNÉE VALIDÉE EN PROD
            # --------------------------------------------------
            cursor.execute("""
                SELECT COUNT(*)
                FROM cotisations
                WHERE annee = ?
                AND date_envoi_mail IS NOT NULL
                AND mode_test = 0
            """, (annee,))

            if cursor.fetchone()[0] > 0:
                flash(
                    "⚠️ Cette année a déjà été validée "
                    "(factures envoyées en PROD). "
                    "Recalcul interdit.",
                    "danger"
                )
                conn.close()
                return redirect(url_for("traitements.cotisations"))

            # --------------------------------------------------
            # SUPPRESSION DES CALCULS PRÉCÉDENTS
            # (uniquement si non verrouillée)
            # --------------------------------------------------
            cursor.execute(
                "DELETE FROM cotisations WHERE annee = ?",
                (annee,)
            )

            # --------------------------------------------------
            # PARSE FICHIER PARSOL
            # --------------------------------------------------
            fichier = request.files.get("parsol_file")
            if not fichier:
                raise ValueError("Fichier PARSOL manquant")

            with tempfile.NamedTemporaryFile(delete=False) as tmp:
                fichier.save(tmp.name)
                parsol_path = tmp.name

            benefs = parse_parsol2l_annuel(parsol_path)

            data = calculer_cotisations_par_annee(
                get_db_path(),
                benefs
            )

            resultats = data["facturables"]
            orphelines = data["orphelines"]

            facture_start = int(request.form.get("facture_start"))

            # Tri par compte comptable
            resultats = sorted(
                resultats,
                key=lambda x: (x.get("compte_comptable") or "")
            )

            # --------------------------------------------------
            # INSERTION EN BASE
            # --------------------------------------------------
            for i, r in enumerate(resultats):
                numero_facture = facture_start + i

                cursor.execute("""
                    INSERT INTO cotisations (
                        annee,
                        id_association,
                        numero_facture,
                        code_vif,
                        beneficiaires,
                        montant,
                        date_calcul,
                        statut,
                        commentaire_regroupement
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    annee,
                    r["id_association"],
                    numero_facture,
                    r["code_vif_facture"],
                    r["beneficiaires"],
                    r["cotisation"],
                    datetime.now().isoformat(),
                    "calcule",
                    r.get("commentaire_regroupement")
                ))

            conn.commit()

            flash(
                f"✅ Cotisations {annee} calculées et enregistrées.",
                "success"
            )

            return redirect(
                url_for("traitements.cotisations", annee=annee)
            )

        except Exception:
            current_app.logger.exception(
                "Erreur calcul cotisations"
            )
            flash(
                "Erreur lors du calcul des cotisations",
                "danger"
            )
            conn.close()
            return redirect(url_for("traitements.cotisations"))

    # ==========================================================
    # GET → AFFICHAGE ANNÉE
    # ==========================================================
    if annee:
        cursor.execute("""
            SELECT *
            FROM cotisations
            WHERE annee = ?
            ORDER BY numero_facture
        """, (annee,))

        lignes = cursor.fetchall()

        if lignes:
            resultats = []
            for l in lignes:
                resultats.append({
                    "id_association": l["id_association"],
                    "code_vif_facture": l["code_vif"],
                    "numero_facture": l["numero_facture"],
                    "beneficiaires": l["beneficiaires"],
                    "cotisation": l["montant"],
                    "statut": l["statut"]
                })

    conn.close()

    # ==========================================================
    # RENDU TEMPLATE
    # ==========================================================
    return render_template(
        "cotisations.html",
        resultats=resultats,
        orphelines=orphelines,
        annee=annee,
        mail_mode=mail_mode,
        mail_test_to=mail_test_to,
        job_done=False,
        mode_relance=False
    )





@traitements_bp.route("/cotisations/toggle_test_mode", methods=["POST"])
@login_required
def cotisations_toggle_test_mode():

    current = session.get(
        "MAIL_MODE",
        os.getenv("MAIL_MODE", "PROD").upper()
    )

    if current == "PROD":
        session["MAIL_MODE"] = "TEST"
        flash("🧪 Mode TEST activé (mails redirigés)", "warning")
    else:
        session["MAIL_MODE"] = "PROD"
        flash("✅ Mode PROD réactivé", "success")

    return redirect(url_for("traitements.cotisations"))



@traitements_bp.route("/cotisations/generer_pdfs", methods=["GET"])
@login_required
def cotisations_generer_pdfs():
    """
    Génère les PDF des cotisations pour une année donnée.

    - Lecture depuis cotisations
    - JOIN avec associations
    - Batch pour éviter timeout
    - Mise à jour date_generation_pdf + statut
    """

    from datetime import datetime
    import shutil

    BATCH_SIZE = 20

    annee = request.args.get("annee")
    offset = int(request.args.get("offset", 0))

    if not annee:
        flash("Année manquante", "danger")
        return redirect(url_for("traitements.cotisations"))

    annee = int(annee)

    conn = sqlite3.connect(get_db_path())
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # ==========================================================
    # Lecture cotisations + associations
    # ==========================================================
    cursor.execute("""
        SELECT
            c.*,
            a.nom_association,
            a.adresse_association_1,
            a.adresse_association_2,
            a.CP,
            a.COMMUNE,
            a.courriel_association
        FROM cotisations c
        JOIN associations a
            ON a.Id = c.id_association
        WHERE c.annee = ?
        ORDER BY c.numero_facture
    """, (annee,))

    lignes = cursor.fetchall()

    total = len(lignes)
    end = min(offset + BATCH_SIZE, total)

    if total == 0:
        flash("Aucune cotisation trouvée", "warning")
        conn.close()
        return redirect(url_for("traitements.cotisations", annee=annee))

    write_log(f"📄 PDF {offset+1} à {end} / {total}")

    base_drive = f"COTISATIONS/Cotisations {annee}/Factures PDF"

    tmp_dir = Path(tempfile.mkdtemp(prefix="cotisations_pdf_"))

    try:
        for ligne in lignes[offset:end]:

            # --------------------------------------------------
            # Construction adresse multi-lignes
            # --------------------------------------------------
            adresse = "\n".join(filter(None, [
                ligne["adresse_association_1"],
                ligne["adresse_association_2"],
                " ".join(filter(None, [
                    ligne["CP"],
                    ligne["COMMUNE"]
                ]))
            ]))

            nom_asso_slug = slugify_filename(ligne["nom_association"])
            code_vif = ligne["code_vif"]

            nom_pdf = (
                f"FACTURE_{ligne['numero_facture']}_"
                f"{code_vif}_{nom_asso_slug}.pdf"
            )

            pdf_path = tmp_dir / nom_pdf

            # --------------------------------------------------
            # Génération PDF
            # --------------------------------------------------
            generer_facture_pdf(
                {
                    "nom_association": ligne["nom_association"],
                    "adresse": adresse,
                    "cotisation": ligne["montant"],
                    "annee": annee,
                    "code_vif_facture": code_vif,
                    "numero_facture": ligne["numero_facture"],
                    "commentaire_regroupement": ligne["commentaire_regroupement"]
                },
                pdf_path
            )

            # --------------------------------------------------
            # Upload Drive
            # --------------------------------------------------
            upload_file_to_drive_path(
                local_path=str(pdf_path),
                drive_path=base_drive,
                filename=nom_pdf,
                shared_drive_id=BA380_SHARED_DRIVE_ID
            )

            # --------------------------------------------------
            # Mise à jour base
            # --------------------------------------------------
            cursor.execute("""
                UPDATE cotisations
                SET date_generation_pdf = ?,
                    statut = 'pdf_genere'
                WHERE id = ?
            """, (
                datetime.now().isoformat(),
                ligne["id"]
            ))

        conn.commit()

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

    # ==========================================================
    # Batch suivant ?
    # ==========================================================
    if end < total:
        conn.close()
        return redirect(
            url_for(
                "traitements.cotisations_generer_pdfs",
                annee=annee,
                offset=end
            )
        )

    conn.close()

    flash(f"✅ {total} factures PDF générées.", "success")

    return redirect(
        url_for("traitements.cotisations", annee=annee)
    )




@traitements_bp.route("/cotisations/envoyer_mails", methods=["POST"])
@login_required
def cotisations_envoyer_mails():
    """
    Envoi des mails de cotisation pour une année.

    - Lecture base + JOIN associations
    - Respect mode TEST / PROD
    - Mise à jour date_envoi_mail + statut
    """

    from datetime import datetime

    annee = request.form.get("annee")
    mail_sender = request.form.get("mail_sender")
    confirmation_prod = request.form.get("confirmation_prod")

    if not annee:
        flash("Année manquante", "danger")
        return redirect(url_for("traitements.cotisations"))

    if not mail_sender:
        flash("Expéditeur manquant", "danger")
        return redirect(url_for("traitements.cotisations", annee=annee))

    annee = int(annee)

    mail_mode = session.get(
        "MAIL_MODE",
        os.getenv("MAIL_MODE", "PROD").upper()
    )

    mail_test_to = os.getenv(
        "MAIL_TEST_TO",
        "ba380.informatique2@banquealimentaire.org"
    )

    # ==========================================================
    # Sécurité PROD
    # ==========================================================
    if mail_mode == "PROD" and confirmation_prod != "1":
        flash(
            "⚠️ Envoi bloqué : confirmation PROD requise.",
            "danger"
        )
        return redirect(url_for("traitements.cotisations", annee=annee))

    conn = sqlite3.connect(get_db_path())
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # ==========================================================
    # Lecture cotisations non encore envoyées
    # ==========================================================
    cursor.execute("""
        SELECT
            c.*,
            a.nom_association,
            a.adresse_association_1,
            a.adresse_association_2,
            a.CP,
            a.COMMUNE,
            a.courriel_association
        FROM cotisations c
        JOIN associations a
            ON a.Id = c.id_association
        WHERE c.annee = ?
        AND c.statut = 'pdf_genere'
        ORDER BY c.numero_facture
    """, (annee,))

    lignes = cursor.fetchall()

    if not lignes:
        flash("Aucune facture prête à être envoyée.", "warning")
        conn.close()
        return redirect(url_for("traitements.cotisations", annee=annee))

    nb_mails = 0

    for ligne in lignes:

        if not ligne["courriel_association"]:
            continue

        # --------------------------------------------------
        # Construction adresse
        # --------------------------------------------------
        adresse = "\n".join(filter(None, [
            ligne["adresse_association_1"],
            ligne["adresse_association_2"],
            " ".join(filter(None, [
                ligne["CP"],
                ligne["COMMUNE"]
            ]))
        ]))

        # --------------------------------------------------
        # Corps du mail
        # --------------------------------------------------
        texte_mail = f"""Association : {ligne['nom_association']}

        Madame, Monsieur,

        Conformément à la convention signée avec la Banque Alimentaire de l'Isère,
        votre cotisation pour l'année {annee} s'élève à {ligne['montant']} €.

        Nous vous remercions d'avance pour votre règlement avant le 28 février {annee}.
        """

        if ligne.get("commentaire_regroupement"):
            texte_mail += "\n" + ligne["commentaire_regroupement"] + "\n"

        texte_mail += """

        Bien cordialement,

        Christian GRAFF
        Comptable BAI
        """

        # --------------------------------------------------
        # Destinataires
        # --------------------------------------------------
        if mail_mode == "TEST":
            destinataires = [mail_test_to]
            sujet = (
                f"[TEST] Appel de cotisation {annee} – "
                f"{ligne['nom_association']}"
            )
        else:
            destinataires = [ligne["courriel_association"]]
            sujet = (
                f"Appel de cotisation {annee} – "
                f"{ligne['nom_association']}"
            )

        # --------------------------------------------------
        # Génération PDF temporaire
        # --------------------------------------------------
        with tempfile.TemporaryDirectory() as tmpdir:

            nom_asso_slug = slugify_filename(
                ligne["nom_association"]
            )

            nom_pdf = (
                f"FACTURE_{ligne['numero_facture']}_"
                f"{ligne['code_vif']}_{nom_asso_slug}.pdf"
            )

            pdf_path = Path(tmpdir) / nom_pdf

            generer_facture_pdf(
                {
                    "nom_association": ligne["nom_association"],
                    "adresse": adresse,
                    "cotisation": ligne["montant"],
                    "annee": annee,
                    "code_vif_facture": ligne["code_vif"],
                    "numero_facture": ligne["numero_facture"],
                    "commentaire_regroupement": ligne["commentaire_regroupement"]
                },
                pdf_path
            )

            envoyer_mail(
                sujet=sujet,
                destinataires=destinataires,
                texte=texte_mail,
                sender_override=mail_sender,
                attachment_path=str(pdf_path)
            )

        # --------------------------------------------------
        # Mise à jour base
        # --------------------------------------------------
        mode_test_flag = 1 if mail_mode == "TEST" else 0

        cursor.execute("""
            UPDATE cotisations
            SET date_envoi_mail = ?,
                statut = 'envoye',
                mode_test = ?
            WHERE id = ?
        """, (
            datetime.now().isoformat(),
            mode_test_flag,
            ligne["id"]
        ))

        nb_mails += 1

    conn.commit()
    conn.close()

    if mail_mode == "TEST":
        flash(
            f"🧪 {nb_mails} mails envoyés en MODE TEST.",
            "warning"
        )
    else:
        flash(
            f"✅ {nb_mails} mails envoyés en PROD.",
            "success"
        )

    return redirect(
        url_for("traitements.cotisations", annee=annee)
    )



DATE_REGEX = re.compile(r"\d{2}/\d{2}/\d{4}")
ASSO_REGEX = re.compile(r"Association\s*:\s*(\d{8})")

def parse_parsol2l_annuel(file_path):
    """
    Parse un fichier PARSOL2L annuel (texte)
    Retourne un dict :
        { code_vif (str) : total_beneficiaires (int) }
    """

    totaux = defaultdict(int)
    code_vif_courant = None

    with open(file_path, "r", encoding="latin-1", errors="ignore") as f:
        for line in f:
            line = line.strip()

            if not line:
                continue

            # ==========================
            # Détection nouvelle association
            # ==========================
            m_asso = ASSO_REGEX.search(line)
            if m_asso:
                code_vif_courant = m_asso.group(1)
                continue

            if not code_vif_courant:
                continue

            # ==========================
            # Ligne de passage (date en début)
            # ==========================
            if DATE_REGEX.match(line):
                parts = re.split(r"\s+", line)
                if len(parts) < 2:
                    continue

                try:
                    nb_benef = int(parts[1])
                except ValueError:
                    continue

                totaux[code_vif_courant] += nb_benef

    return dict(totaux)




# ============================
# PARAMÈTRES FIXES BAI
# ============================

BAI_NOM = "BANQUE ALIMENTAIRE DE L'ISÈRE"
BAI_ADRESSE = "11, allée de la Pinéa\n38600 FONTAINE"
BAI_TEL = "04 76 85 92 50"
BAI_MAIL = "ba380@banquealimentaire.org"
BAI_IBAN = "FR76 1027 8089 2200 0598 3594 087"
BAI_BIC = "CMCIFR2A"
BAI_SIREN = "388 092 132 00033"
BAI_NAF = "8899B"


def generer_facture_pdf(data, output_path):
    """
    Génère une facture ou relance PDF individuelle de cotisation BA38.

    data doit contenir :
      - nom_association
      - adresse (multi-lignes)
      - cotisation
      - annee
      - code_vif_facture
    """

    c = canvas.Canvas(str(output_path), pagesize=A4)
    largeur, hauteur = A4

    # ============================
    # LOGO BA (centré page)
    # ============================

    logo_path = (
        Path(current_app.root_path)
        / "static"
        / "images"
        / "logo ba complet.png"
    )

    if logo_path.exists():
        logo = ImageReader(str(logo_path))

        # Taille du logo (+30 % supplémentaires)
        logo_width = 120 * mm
        logo_height = 46 * mm   # ratio conservé

        # --- CENTRAGE OPTIQUE ---
        # Le logo est visuellement décalé dans son PNG,
        # on applique une correction horizontale manuelle
        correction_x = -20 * mm   # ajuste si besoin ±1 mm

        x_logo = (largeur - logo_width) / 2 - correction_x

        # --- POSITION VERTICALE ---
        # REMONTER de 3 cm par rapport à l’état actuel
        y_logo = hauteur - logo_height - 5 * mm

        c.drawImage(
            logo,
            x_logo,
            y_logo,
            width=logo_width,
            height=logo_height,
            preserveAspectRatio=True,
            mask="auto"
        )    
    else:
        # optionnel mais TRÈS utile
        print(f"LOGO INTROUVABLE : {logo_path}")
        
    # ============================
    # Décalage global sous le logo
    # ============================

    y = hauteur - 20 * mm - 40 * mm

    # ============================
    # BLOC BAI (gauche)
    # ============================

    c.setFont("Helvetica-Bold", 11)
    c.drawString(20 * mm, y, BAI_NOM)

    c.setFont("Helvetica", 9)
    y -= 12
    for line in BAI_ADRESSE.split("\n"):
        c.drawString(20 * mm, y, line)
        y -= 10

    c.drawString(20 * mm, y, f"Tél : {BAI_TEL}")
    y -= 10
    c.drawString(20 * mm, y, f"Mail : {BAI_MAIL}")

    # ============================
    # INFOS FACTURE (droite)
    # ============================

    y_fact = hauteur - 20 * mm - 40 * mm
    c.setFont("Helvetica", 9)
    c.drawRightString(
        largeur - 20 * mm,
        y_fact,
        f"Date : {date.today().strftime('%d/%m/%Y')}"
    )
    y_fact -= 15
    c.drawRightString(
        largeur - 20 * mm,
        y_fact,
        f"Échéance : 28/02/{data['annee']}"
    )
    # Numéro de facture uniquement en facturation normale
    if not data.get("mode_relance"):
        y_fact -= 15
        c.drawRightString(
            largeur - 20 * mm,
            y_fact,
            f"Facture n° {data.get('numero_facture')}"
        )
    if data.get("mode_relance"):
        y_fact -= 15
        c.setFont("Helvetica-Bold", 10)
        c.drawRightString(
            largeur - 20 * mm,
            y_fact,
            f"RELANCE n°{data.get('numero_relance')}"
        )
        c.setFont("Helvetica", 9)

        c.setFillColorRGB(0, 0, 0)


    # ============================
    # ADRESSE ASSOCIATION
    # ============================

    y -= 40  # ⬅️ descente volontaire pour éviter chevauchement
    c.setFont("Helvetica-Bold", 10)
    c.drawString(20 * mm, y, data["nom_association"])

    c.setFont("Helvetica", 9)
    y -= 12
    for line in data["adresse"].split("\n"):
        c.drawString(20 * mm, y, line)
        y -= 10

    # ============================
    # TABLE FACTURE
    # ============================

    y -= 20
    c.setFont("Helvetica-Bold", 9)
    c.drawString(20 * mm, y, "Désignation")
    c.drawRightString(largeur - 20 * mm, y, "Montant")

    y -= 10
    c.line(20 * mm, y, largeur - 20 * mm, y)

    y -= 15
    c.setFont("Helvetica", 9)
    c.drawString(20 * mm, y, f"COTISATION {data['annee']}")
    c.drawRightString(
        largeur - 20 * mm,
        y,
        f"{data['cotisation']:.2f} €"
    )

    # ============================
    # TOTAL
    # ============================

    y -= 30
    c.line(120 * mm, y, largeur - 20 * mm, y)
    y -= 15
    c.setFont("Helvetica-Bold", 10)
    c.drawString(120 * mm, y, "Net à payer")
    c.drawRightString(
        largeur - 20 * mm,
        y,
        f"{data['cotisation']:.2f} €"
    )

    # ============================
    # MENTIONS LÉGALES
    # ============================

    y -= 30
    c.setFont("Helvetica", 8)
    c.drawString(20 * mm, y, "TVA non applicable, art. 293B du CGI")

    y -= 15
    c.drawString(20 * mm, y, f"IBAN : {BAI_IBAN}")
    y -= 10
    c.drawString(20 * mm, y, f"BIC : {BAI_BIC}")

    y -= 20
    c.drawString(20 * mm, y, f"SIREN : {BAI_SIREN}")
    y -= 10
    c.drawString(20 * mm, y, f"NAF : {BAI_NAF}")

    # ============================
    # COMMENTAIRE REGROUPEMENT (PIED DE PAGE)
    # ============================

    commentaire = data.get("commentaire_regroupement")
    if commentaire:
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph
        from reportlab.lib.enums import TA_LEFT

        styles = getSampleStyleSheet()
        style = styles["Normal"]
        style.fontName = "Helvetica-Oblique"
        style.fontSize = 8
        style.leading = 10
        style.alignment = TA_LEFT

        # Zone pied de page (largeur utile)
        largeur_bloc = largeur - 40 * mm
        x_bloc = 20 * mm
        y_bloc = 35 * mm   # pied de page sûr

        p = Paragraph(
            commentaire.replace("\n", "<br/>"),
            style
        )

        w, h = p.wrap(largeur_bloc, 60 * mm)
        p.drawOn(c, x_bloc, y_bloc)    
    # ============================
    # FINALISATION
    # ============================

    c.showPage()
    c.save()


@traitements_bp.route("/cotisations/export_excel", methods=["POST"])
@login_required
def cotisations_export_excel():
    import json
    from io import BytesIO
    from openpyxl import Workbook
    from flask import send_file

    annee = request.form.get("annee")
    raw = request.form.get("data")

    if not raw:
        flash("Aucune donnée à exporter", "danger")
        return redirect(url_for("traitements.cotisations"))

    lignes = json.loads(raw)

    # ============================
    # Création Excel en mémoire
    # ============================
    wb = Workbook()
    ws = wb.active
    ws.title = f"Cotisations {annee}"

    # En-têtes
    ws.append([
        "Code VIF",
        "Compte Comptable",
        "Association",
        "Bénéficiaires",
        "Cotisation (€)"
    ])

    # Lignes
    for l in lignes:
        ws.append([
            l.get("code_vif_facture"),
            l.get("compte_comptable"),
            l.get("nom_association_affichage"),  # 👈 ici
            l.get("beneficiaires"),
            l.get("cotisation"),
            l.get("commentaire_regroupement") or ""
        ])

    # Ajustement largeur colonnes
    largeurs = [15, 22, 45, 18, 18]

    for i, largeur in enumerate(largeurs, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = largeur

    # ============================
    # Export HTTP
    # ============================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Cotisations_{annee}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


def wait_until_drive_folder_empty(service, folder_id, drive_id, timeout=30):
    """
    Attend que le dossier Drive soit réellement vide
    (cohérence Drive), avec timeout en secondes.
    """
    import time

    start = time.time()

    while True:
        results = service.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            corpora="drive",
            driveId=drive_id,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            fields="files(id)"
        ).execute()

        if not results.get("files"):
            return True

        if time.time() - start > timeout:
            return False

        time.sleep(1)

@traitements_bp.route("/cotisations/start")
@login_required
def cotisations_start():
    # Reset systématique
    job_id = session.get("COTISATIONS_JOB_ID")
    if job_id:
        Path(tempfile.gettempdir(), f"cotisations_{job_id}.json").unlink(missing_ok=True)

    session.pop("COTISATIONS_JOB_ID", None)
    session.pop("COTISATIONS_ANNEE", None)

    return redirect(url_for("traitements.cotisations"))



@traitements_bp.route("/cotisations/quit")
@login_required
def cotisations_quit():
    job_id = session.get("COTISATIONS_JOB_ID")
    if job_id:
        Path(tempfile.gettempdir(), f"cotisations_{job_id}.json").unlink(missing_ok=True)

    session.pop("COTISATIONS_JOB_ID", None)
    session.pop("COTISATIONS_ANNEE", None)

    flash("Calcul des cotisations fermé", "info")
    return redirect(url_for("traitements.utilitaires"))


# ============================
# RELANCES
# ============================

@traitements_bp.route("/cotisations/relance", methods=["GET"])
@login_required
def cotisations_relance_start():

    mail_mode = session.get(
        "MAIL_MODE",
        os.getenv("MAIL_MODE", "PROD").upper()
    )

    mail_test_to = os.getenv(
        "MAIL_TEST_TO",
        "ba380.informatique2@banquealimentaire.org"
    )

    return render_template(
        "cotisations_relance.html",
        mail_mode=mail_mode,
        mail_test_to=mail_test_to
    )



@traitements_bp.route("/cotisations/relance", methods=["POST"])
@login_required
def cotisations_relance():

    mode_relance=True

    mail_mode = session.get(
        "MAIL_MODE",
        os.getenv("MAIL_MODE", "PROD").upper()
    )

    mail_test_to = os.getenv(
        "MAIL_TEST_TO",
        "ba380.informatique2@banquealimentaire.org"
    )


    try:
        annee = int(request.form.get("annee"))
        numero_relance = int(request.form.get("numero_relance"))

        fichier = request.files.get("parsol_file")
        if not fichier:
            raise ValueError("Fichier PARSOL manquant")

        with tempfile.NamedTemporaryFile(delete=False) as tmp:
            fichier.save(tmp.name)
            parsol_path = tmp.name

        benefs = parse_parsol2l_annuel(parsol_path)

        data = calculer_cotisations_par_annee(get_db_path(), benefs)
        resultats = data["facturables"]

        current_app.logger.info(
            f"NB FACTURABLES = {len(resultats)}"
        )

        # 🔎 Charger relance depuis DB
        conn = sqlite3.connect(get_db_path())
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute("SELECT Id, relance FROM associations")
        relances = {row["Id"]: row["relance"] for row in cursor.fetchall()}
        conn.close()

        write_log(f"EXEMPLE RESULTAT : {resultats[0]}")

        # 🔹 Filtrer uniquement celles à relancer
        resultats = [
            r for r in resultats
            if int(relances.get(r["id_association"], 0)) == numero_relance
        ]
        write_log(f"NB RELANCES TROUVÉES : {len(resultats)}")

        for r in resultats:
            r["mode_relance"] = True
            r["numero_relance"] = numero_relance

        session["RELANCE_DATA"] = resultats
        session["RELANCE_ANNEE"] = annee
        session["RELANCE_NUMERO"] = numero_relance

        return render_template(
            "cotisations.html",
            resultats=resultats,
            orphelines=None,
            annee=annee,
            mail_mode=mail_mode,
            mail_test_to=mail_test_to,
            job_done=True,
            mode_relance=True,         # ← ICI
            numero_relance=numero_relance
        )

    except Exception:
        current_app.logger.exception("Erreur relance cotisations")
        flash("Erreur lors du traitement des relances", "danger")
        return redirect(url_for("traitements.cotisations"))


@traitements_bp.route("/cotisations/saisie-paiements", methods=["GET", "POST"])
@login_required
def saisie_paiements_cotisations():
    """
    Écran de saisie des paiements des cotisations.

    - GET  : affiche les cotisations d’une année
    - POST : enregistre une date de paiement
    """

    from datetime import datetime

    annee = request.args.get("annee") or request.form.get("annee")
    impayes_only = request.args.get("impayes") == "1"

    if not annee:
        flash("Sélectionner une année.", "warning")
        return render_template(
            "saisie_paiements_cotisations.html",
            resultats=None
        )

    annee = int(annee)

    conn = sqlite3.connect(get_db_path())
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # ==========================================================
    # POST → ENREGISTRER PAIEMENT
    # ==========================================================
    if request.method == "POST":

        cotisation_id = request.form.get("cotisation_id")
        date_paiement = request.form.get("date_paiement")

        if cotisation_id and date_paiement:
            cursor.execute("""
                UPDATE cotisations
                SET date_paiement = ?,
                    statut = 'paye'
                WHERE id = ?
            """, (date_paiement, cotisation_id))

            conn.commit()
            flash("Paiement enregistré.", "success")

        return redirect(
            url_for(
                "traitements.saisie_paiements_cotisations",
                annee=annee
            )
        )

    # ==========================================================
    # GET → AFFICHAGE
    # ==========================================================
    sql = """
        SELECT
            c.*,
            a.nom_association,
            a.compte_comptable
        FROM cotisations c
        JOIN associations a
            ON a.Id = c.id_association
        WHERE c.annee = ?
    """

    params = [annee]

    if impayes_only:
        sql += " AND c.statut != 'paye'"

    sql += " ORDER BY c.numero_facture"

    cursor.execute(sql, params)

    # 🔹 ICI on récupère les lignes
    lignes = cursor.fetchall()

    total_facture = 0
    total_paye = 0
    resultats = []

    for l in lignes:
        montant = l["montant"] or 0

        total_facture += montant

        if l["statut"] == "paye":
            total_paye += montant

        resultats.append(dict(l))

    total_restant = total_facture - total_paye

    taux_recouvrement = 0
    if total_facture > 0:
        taux_recouvrement = round((total_paye / total_facture) * 100, 2)

    conn.close()

    return render_template(
        "saisie_paiements_cotisations.html",
        resultats=resultats,
        annee=annee,
        total_facture=total_facture,
        total_paye=total_paye,
        total_restant=total_restant,
        taux_recouvrement=taux_recouvrement
    )

@traitements_bp.route("/cotisations/export/<int:annee>")
@login_required
def export_cotisations_excel(annee):

    import pandas as pd

    conn = sqlite3.connect(get_db_path())
    df = pd.read_sql_query("""
        SELECT
            c.numero_facture,
            a.nom_association,
            a.compte_comptable,
            c.code_vif,
            c.beneficiaires,
            c.montant,
            c.statut,
            c.date_paiement
        FROM cotisations c
        JOIN associations a
            ON a.Id = c.id_association
        WHERE c.annee = ?
        ORDER BY c.numero_facture
    """, conn, params=(annee,))
    conn.close()

    output = io.BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"cotisations_{annee}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
