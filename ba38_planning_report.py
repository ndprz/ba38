# ============================================================
#   BA38 – Module : Rapport activité bénévoles
#   Fichier : ba38_planning_report.py
#   Auteur : ChatGPT + Nicolas
#   Version : 1.1 (filtre mois + correction bug NameError)
# ============================================================


from flask import send_file
import os
from flask import Blueprint, render_template, request, flash, session
from flask_login import login_required
from datetime import datetime, timedelta, date
from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from utils import get_db_connection, write_log

import sqlite3

import io
from datetime import datetime, timedelta
from flask import send_file

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None  # On gérera le cas où la lib n'est pas installée


# 🔧 Paramètre global : activer / désactiver l'export Excel de contrôle
GENERER_EXCEL_CONTROLE = True

planning_report_bp = Blueprint("planning_report", __name__, url_prefix="/planning")

PLANNING_MAP = {
    "plannings_ramasse": "ramasse",
    "plannings_distribution": "distribution",
    "plannings_pal": "pal",
    "plannings_pesee": "pesee",
    "plannings_vif": "vif"
}

DEBUG_PLANNING = ""   # ou "ramasse" ou "distribution" ou "all"

# ------------------------------------------------------------
#  TOOLS : Date helpers
# ------------------------------------------------------------

def parse_date(d):
    return datetime.strptime(d, "%Y-%m-%d").date()


def parse_num_semaine(iso_str):
    """Convertit '2025-W46' → (2025, 46)"""
    try:
        y, w = iso_str.split("-W")
        return int(y), int(w)
    except Exception:
        return None, None


def dates_for_week(year, week):
    """Retourne les dates du lundi → vendredi pour une semaine donnée."""
    lundi = datetime.strptime(f"{year}-W{week}-1", "%Y-W%W-%w").date()
    jours = ["lundi", "mardi", "mercredi", "jeudi", "vendredi"]
    return {j: lundi + timedelta(days=i) for i, j in enumerate(jours)}


# ------------------------------------------------------------
#  LOADERS : Durées + planning réel
# ------------------------------------------------------------

def load_durations(cursor, table_name):
    """Charge {jour: durée_en_minutes} depuis un modèle standard."""
    try:
        rows = cursor.execute(f"SELECT jour, duree FROM {table_name}").fetchall()
        return {r["jour"].strip().lower(): int(r["duree"]) for r in rows}
    except Exception:
        return {}


def load_planning(cursor, table_name, postes, start_date, end_date):
    """
    Charge les lignes d'un planning entre start_date et end_date.
    Ajoute un champ date_jour (date réelle) et, pour la ramasse,
    harmonise les noms des champs de remplaçants.

    NOTE IMPORTANTE :
    - On reconstruit la date à partir de (année de start_date, semaine, jour),
      puis on applique un filtre :
        1) sur la période [start_date, end_date]
        2) si la période est un mois complet (ex. 01/10 → 31/10),
           on ne garde que les dates dont le mois est exactement celui de start_date.
           => ceci évite de compter le 01/11 quand on demande “mois d’octobre”.
    """

    # 🔍 Logs pour VIF
    debug = (DEBUG_PLANNING in ("all", table_name.replace("plannings_", "")))
    prefix = f"[{table_name.replace('plannings_', '').upper()}]"

    if debug:
        write_log(f"{prefix} ----- Chargement planning {table_name} -----")

    rows = cursor.execute(f"SELECT * FROM {table_name}").fetchall()
    resultat = []

    jours_iso = ["lundi", "mardi", "mercredi", "jeudi", "vendredi"]

    # 🧠 Détection d’une période “mois complet”
    # ex : 01/10/2025 → 31/10/2025
    is_full_month = (
        start_date.day == 1
        and end_date.day >= 28            # fin de mois “classique”
        and start_date.month == end_date.month
        and start_date.year == end_date.year
    )

    for r in rows:
        semaine = int(r["semaine"])
        jour_txt = str(r["jour"]).strip().lower()

        if jour_txt not in jours_iso:
            if debug:
                write_log(f"{prefix} Jour ignoré = {jour_txt}")
            continue

        weekday = jours_iso.index(jour_txt) + 1
        year = r["annee"] if "annee" in r.keys() else start_date.year

        # mapping jours → ISO weekday
        JOUR_VERS_ISO = {
            "lundi": 1,
            "mardi": 2,
            "mercredi": 3,
            "jeudi": 4,
            "vendredi": 5,
            "samedi": 6,
            "dimanche": 7,
        }

        try:
            weekday_iso = JOUR_VERS_ISO[jour_txt]
            date_reelle = date.fromisocalendar(year, semaine, weekday_iso)
        except Exception as e:
            if debug:
                write_log(
                    f"{prefix} ❌ Erreur calcul date (annee={year}, semaine={semaine}, jour={jour_txt}) : {e}"
                )
            continue

        # 🎯 Si la période couvre un mois complet, on exclut les jours
        # qui ne sont pas dans le mois concerné (cas des semaines “coupées”).
        if is_full_month and date_reelle.month != start_date.month:
            if debug:
                write_log(
                    f"{prefix}   -> rejeté (date {date_reelle} hors du mois demandé)"
                )
            continue

        if debug:
            write_log(
                f"{prefix} Ligne : semaine={semaine} | jour={jour_txt} "
                f"| date_calculee={date_reelle}"
            )

        # Filtre classique sur la période
        if start_date <= date_reelle <= end_date:
            if debug:
                write_log(f"{prefix}   -> RETENU")
            rr = dict(r)
            rr["date_jour"] = date_reelle

            # 🧩 Harmonisation pour plannings_ramasse
            if table_name == "plannings_ramasse":
                rr["chauffeur_remplacant"] = rr.get("remplacant_chauffeur_id")
                rr["responsable_remplacant"] = rr.get("remplacant_responsable_id")
                rr["equipier_remplacant"] = rr.get("remplacant_equipier_id")

                rr["ramasse_tri1_remplacant"] = rr.get("remplacant_ramasse_tri1_id")
                rr["ramasse_tri2_remplacant"] = rr.get("remplacant_ramasse_tri2_id")
                rr["ramasse_tri3_remplacant"] = rr.get("remplacant_ramasse_tri3_id")

            resultat.append(rr)
        else:
            if debug:
                write_log(f"{prefix}   -> rejeté (hors période)")

    if debug:
        write_log(f"{prefix} TOTAL RETENUES = {len(resultat)}")

    return resultat


# ------------------------------------------------------------
#  CALCUL HEURES – GÉNÉRIQUE
# ------------------------------------------------------------

def accumulate_hours(cursor, table_name, postes, lignes):
    """
    Additionne les durées par bénévole.

    RÈGLE UNIVERSELLE :
    Pour chaque poste 'xxx' dans POSTES[], on lit :
      - titulaire  : xxx_id
      - absent     : xxx_absent
      - remplaçant : xxx_remplacant

    Ces trois champs existent dans TOUTES les tables :
    - plannings_distribution
    - plannings_pal
    - plannings_pesee
    - plannings_vif
    - plannings_ramasse (avec compatibilité fallback)

    LOGIQUE :
      - Si absent = oui ET remplaçant ≠ None → créditer remplaçant
      - Si absent = oui ET remplaçant = None → ignorer
      - Sinon → créditer titulaire
    """

    # Debug contrôlé par DEBUG_PLANNING
    # ex : DEBUG_PLANNING = "vif", ou "ramasse", ou "all", ou ""
    debug = DEBUG_PLANNING in ("all", table_name.replace("plannings_", ""))

    # Préfixe automatique basé sur le nom de la table
    prefix = f"[{table_name.replace('plannings_', '').upper()}]"

    heures = {}

    for r in lignes:

        # ----------------------------------------------
        # DURÉE
        # ----------------------------------------------
        duree = 0

        if table_name == "plannings_ramasse":
            duree = None
            tournee_id = r.get("tournee_id")

            if tournee_id:
                row = cursor.execute(
                    "SELECT MAX(duree) FROM tournees_fournisseurs WHERE tournee_id = ?",
                    (tournee_id,),
                ).fetchone()

                if row and row[0]:
                    duree = int(row[0])

            # aucune durée trouvée → fallback mini (120)
            if not duree or duree <= 0:
                duree = 120

        else:
            val = r.get("duree")
            try:
                duree = int(val) if val not in (None, "") else 0
            except:
                duree = 0

            if duree == 0:
                duree = 120

        if debug:
            write_log(
                f"{prefix} Ligne : jour={r.get('jour')} | duree={duree} | "
                f"semaine={r.get('semaine')} | date={r.get('date_jour')}"
            )

        # ----------------------------------------------
        # ATTRIBUTION DES HEURES
        # ----------------------------------------------
        for poste in postes:

            # Champs standard pour toutes les tables
            titulaire_key = f"{poste}_id"
            absent_key = f"{poste}_absent"
            rempla_key = f"{poste}_remplacant"

            titulaire = r.get(titulaire_key)
            absent = str(r.get(absent_key, "non")).strip().lower() == "oui"
            remplacant = r.get(rempla_key)

            # ----- FALLBACK RAMASSE -----
            if table_name == "plannings_ramasse":
                if not remplacant:
                    # chauffeur/responsable/equipier
                    if poste in ("chauffeur", "responsable", "equipier"):
                        remplacant = r.get(f"remplacant_{poste}_id")
                    # ramasse_tri1 / 2 / 3
                    if poste.startswith("ramasse_tri"):
                        num = poste[-1]
                        remplacant = r.get(f"remplacant_ramasse_tri{num}_id")

            if debug:
                write_log(
                    f"{prefix} Poste={poste} | titulaire={titulaire} | "
                    f"absent={'oui' if absent else 'non'} | remplaçant={remplacant}"
                )

            # AUCUN BÉNÉVOLE
            if not titulaire and not remplacant:
                if debug:
                    write_log(f"{prefix}   -> ignoré (aucun bénévole)")
                continue

            # ABSENT + REMPLAÇANT → crédit remplaçant
            if absent and remplacant:
                bene_id = remplacant

            # ABSENT SANS REMPLAÇANT → ignorer
            elif absent and not remplacant:
                if debug:
                    write_log(f"{prefix}   -> ignoré (absent sans remplaçant)")
                continue

            # TITULAIRE PRÉSENT
            else:
                bene_id = titulaire

            if not bene_id:
                if debug:
                    write_log(f"{prefix}   -> ignoré (bene_id vide)")
                continue

            # Ajout des minutes
            heures[bene_id] = heures.get(bene_id, 0) + duree

            if debug:
                write_log(f"{prefix}   +{duree} min (bénévole {bene_id})")

    if debug:
        total = sum(heures.values())
        write_log(f"{prefix} TOTAL HEURES CALCULEES = {total / 60.0}")
        write_log("==== RÉCAP LIGNES COMPTÉES ====")
        write_log(f"Total lignes acceptées : {len(heures)} bénévoles comptabilisés")
        write_log("================================")


    return heures


# ------------------------------------------------------------
#  POSTES PAR PLANNING
# ------------------------------------------------------------

POSTES = {
    "ramasse": ["chauffeur", "responsable", "equipier",
                "ramasse_tri1", "ramasse_tri2", "ramasse_tri3"],

    "distribution": ["froid1", "froid2", "froid3", "froid4",
                     "frais_sec1", "frais_sec2", "frais_sec3", "frais_sec4"],

    "palettes": [f"pal{str(i).zfill(2)}" for i in range(1, 11)],

    "pesee": [f"pesee{str(i).zfill(2)}" for i in range(1, 9)],

    "vif": [f"vif{str(i).zfill(2)}" for i in range(1, 4)],
}

# ------------------------------------------------------------
#  TABLES PAR PLANNING
# ------------------------------------------------------------

TABLES = {
    "ramasse": ("planning_standard_ramasse_ids", "plannings_ramasse"),
    "distribution": ("planning_standard_distribution_ids", "plannings_distribution"),
    "palettes": ("planning_standard_pal_ids", "plannings_pal"),
    "pesee": ("planning_standard_pesee_ids", "plannings_pesee"),
    "vif": ("planning_standard_vif_ids", "plannings_vif"),
}


# ------------------------------------------------------------
#  ROUTE : Page de sélection + calcul
# ------------------------------------------------------------

@planning_report_bp.route("/rapport_benevoles", methods=["GET", "POST"])
@login_required
def planning_report_run():

    periode_label = ""
    total_general = 0.0

    resultats = {
        "ramasse": 0.0,
        "distribution": 0.0,
        "palettes": 0.0,
        "pesee": 0.0,
        "vif": 0.0,
    }

    # 🔎 Listes pour l'Excel de contrôle
    lignes_ok = []
    lignes_rejetees = []

    # ---------- GET : afficher page avec choix période --------------
    if request.method == "GET":
        return render_template("planning_report.html")

    # ---------- POST : calculer le rapport --------------------------
    mode = request.form.get("periode")
    today = datetime.today().date()

    if mode == "mois_precedent":
        first = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
        last = today.replace(day=1) - timedelta(days=1)
        periode_label = f"Mois précédent ({first.strftime('%d/%m/%Y')} - {last.strftime('%d/%m/%Y')})"

    elif mode == "annee_en_cours":
        first = today.replace(month=1, day=1)
        last = today
        periode_label = f"Année en cours ({first.strftime('%d/%m/%Y')} - {last.strftime('%d/%m/%Y')})"

    elif mode == "annee_precedente":
        year = today.year - 1
        first = datetime(year, 1, 1).date()
        last = datetime(year, 12, 31).date()
        periode_label = f"Année {year} ({first.strftime('%d/%m/%Y')} - {last.strftime('%d/%m/%Y')})"

    elif mode == "personnalisee":
        d1 = request.form.get("debut")
        d2 = request.form.get("fin")

        if not d1 or not d2:
            flash("❌ Merci de sélectionner une période valide", "danger")
            return render_template("planning_report.html")

        first = parse_date(d1)
        last = parse_date(d2)

    else:
        flash("❌ Période invalide", "danger")
        return render_template("planning_report.html")

    # Label générique utilisé pour l’Excel
    periode_label = f"{first.strftime('%d/%m/%Y')} → {last.strftime('%d/%m/%Y')}"

    # ----------- CALCUL GLOBAL --------------
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    for nom in resultats.keys():
        modele, table = TABLES[nom]

        # Planning filtré sur la période (et sur le mois si "mois complet")
        records = load_planning(cur, table, POSTES[nom], first, last)

        # ---------- DIAGNOSTIC : comptage des lignes ------------
        nb_lignes = len(records)
        nb_postes = nb_lignes * len(POSTES[nom])

        write_log(
            f"[REPORT] {nom.upper()} : {nb_lignes} lignes lues, "
            f"{len(POSTES[nom])} postes par ligne → {nb_postes} postes analysés"
        )


        # ⚠️ IMPORTANT : plus de dédoublonnage ici, on garde toutes les lignes
        heures = accumulate_hours(cur, table, POSTES[nom], records)

        # =====================================================
        # CAPTURE DES LIGNES POUR EXCEL (OK + REJETÉES)
        # -> même logique que accumulate_hours, mais détaillée
        # =====================================================
        for r in records:

            # -------- DURÉE brute minutes (même logique qu'accumulate_hours) -----
            duree_brute = 0

            if table == "plannings_ramasse":
                tournee_id = r.get("tournee_id")
                if tournee_id:
                    row = cur.execute(
                        "SELECT MAX(duree) FROM tournees_fournisseurs WHERE tournee_id = ?",
                        (tournee_id,),
                    ).fetchone()
                    if row and row[0]:
                        try:
                            duree_brute = int(row[0])
                        except Exception:
                            duree_brute = 0

                # Sécurité : si jamais rien trouvé on garde 180 comme dans accumulate_hours
                if duree_brute == 0:
                    duree_brute = 180

            else:
                val = r.get("duree")
                try:
                    duree_brute = int(val) if val not in (None, "") else 0
                except Exception:
                    duree_brute = 0

                # Sécurité : si 0, on force 120 min (règle actuelle)
                if duree_brute == 0:
                    duree_brute = 120

            # -------- PARCOURIR TOUS LES POSTES ----------------
            for poste in POSTES[nom]:

                titulaire_key = f"{poste}_id"
                absent_key = f"{poste}_absent"
                rempla_key = f"{poste}_remplacant"

                titulaire = r.get(titulaire_key)
                absent = str(r.get(absent_key, "non")).strip().lower() == "oui"
                remplacant = r.get(rempla_key)

                # ----- FALLBACK RAMASSE pour les champs remplaçants -----
                if table == "plannings_ramasse" and not remplacant:
                    if poste in ("chauffeur", "responsable", "equipier"):
                        remplacant = r.get(f"remplacant_{poste}_id")
                    if poste.startswith("ramasse_tri"):
                        num = poste[-1]
                        remplacant = r.get(f"remplacant_ramasse_tri{num}_id")

                # --------- CAS REJETÉS -----------------
                if not titulaire and not remplacant:
                    # Aucun bénévole sur ce poste
                    lignes_rejetees.append({
                        "type_planning": nom,
                        "semaine": r.get("semaine"),
                        "jour": r.get("jour"),
                        "benevole_id": None,
                        "remplacant_id": None,
                        "absent": absent,
                        "duree_brute_minutes": duree_brute,
                        "duree_retendue_minutes": 0,
                        "duree_retendue_heures": 0.0,
                        "raison_rejet": "Aucun bénévole",
                    })
                    continue

                if absent and not remplacant:
                    # Titulaire absent et pas de remplaçant
                    lignes_rejetees.append({
                        "type_planning": nom,
                        "semaine": r.get("semaine"),
                        "jour": r.get("jour"),
                        "benevole_id": titulaire,
                        "remplacant_id": None,
                        "absent": True,
                        "duree_brute_minutes": duree_brute,
                        "duree_retendue_minutes": 0,
                        "duree_retendue_heures": 0.0,
                        "raison_rejet": "Absent sans remplaçant",
                    })
                    continue

                # --------- CAS ACCEPTÉS -----------------
                if absent and remplacant:
                    bene_id = remplacant
                else:
                    bene_id = titulaire

                if not bene_id:
                    # Sécurité : cas théorique (devrait déjà être couvert ci-dessus)
                    lignes_rejetees.append({
                        "type_planning": nom,
                        "semaine": r.get("semaine"),
                        "jour": r.get("jour"),
                        "benevole_id": None,
                        "remplacant_id": remplacant,
                        "absent": absent,
                        "duree_brute_minutes": duree_brute,
                        "duree_retendue_minutes": 0,
                        "duree_retendue_heures": 0.0,
                        "raison_rejet": "ID bénévole vide",
                    })
                    continue

                # Ligne prise en compte dans le calcul
                lignes_ok.append({
                    "type_planning": nom,
                    "semaine": r.get("semaine"),
                    "jour": r.get("jour"),
                    "benevole_id": bene_id,
                    "remplacant_id": remplacant,
                    "absent": absent,
                    "duree_brute_minutes": duree_brute,
                    "duree_retendue_minutes": duree_brute,
                    "duree_retendue_heures": round(duree_brute / 60, 2),
                })

        # 🔥 total en heures (float) pour ce planning
        total_h = int(sum(heures.values()) / 60)

        # 🔥 stocker directement la valeur totale
        resultats[nom] = total_h

        # 🔥 accumulation du total général
        total_general += total_h

    conn.close()

    # -------------------------------------------------
    # Stockage en session pour la route /rapport_benevoles_excel
    # -------------------------------------------------
    if GENERER_EXCEL_CONTROLE:
        session["lignes_ok"] = lignes_ok
        session["lignes_rejetees"] = lignes_rejetees
        session["periode_excel"] = periode_label

    # ----------- AFFICHAGE RESULTAT ----------
    return render_template(
        "planning_report_result.html",
        first=first,
        last=last,
        total_ram=resultats["ramasse"],
        total_dist=resultats["distribution"],
        total_pal=resultats["palettes"],
        total_pes=resultats["pesee"],
        total_vif=resultats["vif"],
        total_general=round(total_general, 2),
    )


# ------------------------------------------------------------
#  EXPORT CSV
# ------------------------------------------------------------
@planning_report_bp.route("/rapport_benevoles_csv")
@login_required
def planning_report_csv():
    from flask import make_response

    # paramètres envoyés depuis planning_report_result.html
    periode = request.args.get("periode", "periode")

    # Conversion des heures en entiers
    ram = int(float(request.args.get("ram", 0)))
    dist = int(float(request.args.get("dist", 0)))
    pal = int(float(request.args.get("pal", 0)))
    pes = int(float(request.args.get("pes", 0)))
    vif = int(float(request.args.get("vif", 0)))
    total = int(float(request.args.get("total", 0)))

    # Nom fichier propre
    periode_safe = periode.replace("/", "-").replace(" ", "_")
    filename = f"rapport_benevoles_{periode_safe}.csv"

    # Construction CSV
    lignes = [
        "Type;Heures",
        f"Ramasse;{ram}",
        f"Distribution;{dist}",
        f"Palettes;{pal}",
        f"Pesée;{pes}",
        f"VIF;{vif}",
        f"TOTAL;{total}",
    ]
    csv_data = "\n".join(lignes)

    response = make_response(csv_data)
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    return response

@planning_report_bp.route("/rapport_benevoles_excel")
@login_required
def planning_report_excel():
    from flask import session, abort

    lignes_ok = session.get("lignes_ok")
    lignes_rejetees = session.get("lignes_rejetees")
    periode_label = session.get("periode_excel", "")

    if lignes_ok is None or lignes_rejetees is None:
        abort(400)  # rien en session → mauvaise utilisation

    return _export_debug_excel(lignes_ok, lignes_rejetees, periode_label)


@planning_report_bp.route("/rapport_benevoles_pdf")
@login_required
def planning_report_pdf():
    from flask import send_file
    import io
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    )
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    import os

    # Conversion des heures en int
    ram = int(float(request.args.get("ram", 0)))
    dist = int(float(request.args.get("dist", 0)))
    pal = int(float(request.args.get("pal", 0)))
    pes = int(float(request.args.get("pes", 0)))
    vif = int(float(request.args.get("vif", 0)))
    total = int(float(request.args.get("total", 0)))

    periode = request.args.get("periode", "")
    periode_safe = periode.replace("/", "-").replace(" ", "_")
    filename = f"rapport_benevoles_{periode_safe}.pdf"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30)
    styles = getSampleStyleSheet()
    story = []

    # Logo
    logo_path = os.path.join("static", "logo.png")
    if os.path.exists(logo_path):
        img = Image(logo_path, width=80, height=80)
        img.hAlign = 'CENTER'
        story.append(img)
        story.append(Spacer(1, 20))

    # Titre
    story.append(Paragraph(
        f"<para align='center'><b>Rapport d’activité des bénévoles</b></para>",
        styles["Title"]
    ))
    story.append(Paragraph(
        f"<para align='center'>Période : {periode}</para>",
        styles["Normal"]
    ))
    story.append(Spacer(1, 20))

    # Tableau des heures
    data = [
        ["Catégorie", "Heures"],
        ["Ramasse", ram],
        ["Distribution", dist],
        ["Palettes", pal],
        ["Pesée", pes],
        ["VIF", vif],
        ["TOTAL", total],
    ]

    table = Table(data, colWidths=[200, 150])

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(1, 0.55, 0)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 12),

        ("BACKGROUND", (0, 1), (-1, 1), colors.whitesmoke),
        ("BACKGROUND", (0, 3), (-1, 3), colors.whitesmoke),
        ("BACKGROUND", (0, 5), (-1, 5), colors.whitesmoke),

        ("ALIGN", (0, 1), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 11),

        ("BOX", (0, 0), (-1, -1), 1.2, colors.black),
        ("LINEBEFORE", (1, 0), (1, -1), 1.0, colors.black),
        ("LINEABOVE", (0, 0), (-1, 0), 1.2, colors.black),
        ("LINEBELOW", (0, -1), (-1, -1), 1.2, colors.black),
    ]))

    story.append(table)

    doc.build(story)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )


def _export_debug_excel(lignes_ok, lignes_rejetees, periode_label: str):
    """
    Génère un fichier Excel en mémoire avec deux onglets :
    - POSTES_OK
    - POSTES_REJETES

    Chaque élément de lignes_ok / lignes_rejetees est un dict avec au moins :
      - type_planning
      - semaine
      - jour
      - benevole_id
      - remplacant_id
      - absent
      - duree_brute_minutes
      - duree_retendue_minutes
      - duree_retendue_heures
    + pour lignes_rejetees : 'raison_rejet'
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl n'est pas installé (pip install openpyxl)")

    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws_ok = wb.active
    ws_ok.title = "POSTES_OK"
    ws_rej = wb.create_sheet("POSTES_REJETES")

    # Colonnes communes
    colonnes = [
        "type_planning",
        "semaine",
        "jour",
        "benevole_id",
        "remplacant_id",
        "absent",
        "duree_brute_minutes",
        "duree_retendue_minutes",
        "duree_retendue_heures",
    ]

    # -------------------------------
    # FEUILLE POSTES OK
    # -------------------------------
    ws_ok.append(colonnes)
    for ligne in lignes_ok:
        ws_ok.append([
            ligne.get("type_planning"),
            ligne.get("semaine"),
            ligne.get("jour"),
            ligne.get("benevole_id"),
            ligne.get("remplacant_id"),
            ligne.get("absent"),
            ligne.get("duree_brute_minutes"),
            ligne.get("duree_retendue_minutes"),
            ligne.get("duree_retendue_heures"),
        ])

    # -------------------------------
    # FEUILLE POSTES REJETÉS
    # -------------------------------
    colonnes_rej = colonnes + ["raison_rejet"]
    ws_rej.append(colonnes_rej)

    for ligne in lignes_rejetees:
        ws_rej.append([
            ligne.get("type_planning"),
            ligne.get("semaine"),
            ligne.get("jour"),
            ligne.get("benevole_id"),
            ligne.get("remplacant_id"),
            ligne.get("absent"),
            ligne.get("duree_brute_minutes"),
            ligne.get("duree_retendue_minutes"),
            ligne.get("duree_retendue_heures"),
            ligne.get("raison_rejet"),
        ])

    # -------------------------------
    # Ajustement automatique de largeur
    # -------------------------------
    for ws in (ws_ok, ws_rej):
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col:
                v = cell.value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    # -------------------------------
    # Export en mémoire
    # -------------------------------
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Nom du fichier
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"controle_reporting_{ts}.xlsx"

    if periode_label:
        safe_label = "".join(c for c in periode_label if c.isalnum() or c in ("_", "-"))
        filename = f"controle_reporting_{safe_label}_{ts}.xlsx"

    # Envoi en téléchargement
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
