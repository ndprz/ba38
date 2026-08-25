import io
import re
import sqlite3
import tempfile
from collections import defaultdict

from flask import request, render_template, flash, redirect, url_for, send_file, current_app, session
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ba38_utilitaires.core import get_db_path, require_access

from ba38_tresorerie import tresorerie_bp


DATE_REGEX = re.compile(r"\d{2}/\d{2}/\d{4}")
ASSO_REGEX = re.compile(r"Association\s*:\s*(\d{8})")


def parse_parsol2l_annuel_BL(file_path):
    """
    Parse un fichier PARSOL2L annuel.

    Retourne :
        { code_vif : total_beneficiaires }

    Le nombre de bénéficiaires est pris dans la colonne
    'Nbre de Bénéficiaires' (3ème colonne du tableau PARSOL),
    et NON dans le numéro de BL.
    """

    totaux = defaultdict(int)
    code_vif_courant = None

    with open(file_path, "r", encoding="latin-1", errors="ignore") as f:
        for line in f:
            line = line.strip()

            if not line:
                continue

            # ----------------------------------------
            # Changement d'association
            # ----------------------------------------
            m_asso = ASSO_REGEX.search(line)
            if m_asso:
                code_vif_courant = m_asso.group(1)
                continue

            if not code_vif_courant:
                continue

            # ----------------------------------------
            # Ligne de livraison
            # ----------------------------------------
            if DATE_REGEX.match(line):

                parts = re.split(r"\s+", line)

                if len(parts) < 3:
                    continue

                try:
                    nb_benef = int(parts[2])
                except ValueError:
                    continue

                totaux[code_vif_courant] += nb_benef

    return dict(totaux)


def calculer_beneficiaires_par_vif(db_path, benef_par_vif):
    """
    Retourne une ligne par code VIF du fichier PARSOL2L
    sans aucun regroupement.

    benef_par_vif :
        { code_vif : nb_beneficiaires }

    Retour :
        {
            "resultats": [...],
            "orphelines": [...]
        }
    """

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            Id,
            code_VIF,
            nom_association,
            compte_comptable
        FROM associations
    """)

    assos = cursor.fetchall()

    asso_par_vif = {
        str(a["code_VIF"]).zfill(8): dict(a)
        for a in assos
        if a["code_VIF"]
    }

    resultats = []
    orphelines = []

    for code_vif, nb_benef in sorted(benef_par_vif.items()):

        code_vif = str(code_vif).zfill(8)

        asso = asso_par_vif.get(code_vif)

        if asso:

            resultats.append({
                "id_association": asso["Id"],
                "code_vif": code_vif,
                "compte_comptable": asso["compte_comptable"] or "",
                "nom_association": asso["nom_association"] or "",
                "beneficiaires": nb_benef,
                "trouve": "Oui"
            })

        else:

            resultats.append({
                "id_association": None,
                "code_vif": code_vif,
                "compte_comptable": "",
                "nom_association": "",
                "beneficiaires": nb_benef,
                "trouve": "Non"
            })

            orphelines.append(code_vif)

    conn.close()

    return {
        "resultats": resultats,
        "orphelines": orphelines
    }


@tresorerie_bp.route("/parsol2l", methods=["GET", "POST"])
@login_required
@require_access("tresorerie", "ecriture")
def parsol2l():

    resultats = None
    orphelines = None

    if request.method == "POST":

        try:

            fichier = request.files.get("parsol_file")

            if not fichier:
                flash("❌ Aucun fichier sélectionné", "danger")
                return redirect(url_for("tresorerie.parsol2l"))

            with tempfile.NamedTemporaryFile(delete=False) as tmp:
                fichier.save(tmp.name)
                parsol_path = tmp.name

            benefs = parse_parsol2l_annuel_BL(parsol_path)

            data = calculer_beneficiaires_par_vif(
                get_db_path(),
                benefs
            )

            resultats = data["resultats"]
            orphelines = data["orphelines"]

            session["parsol2l_resultats"] = resultats

        except Exception:
            current_app.logger.exception("Erreur analyse PARSOL2L")
            flash("❌ Erreur lors du traitement du fichier", "danger")

    return render_template(
        "tresorerie/parsol2l.html",
        resultats=resultats,
        orphelines=orphelines
    )


@tresorerie_bp.route("/parsol2l/export_excel", methods=["POST"])
@login_required
@require_access("tresorerie", "ecriture")
def parsol2l_export_excel():

    data = session.get("parsol2l_resultats")

    if not data:
        flash("Aucune donnée à exporter", "danger")
        return redirect(url_for("tresorerie.parsol2l"))

    wb = Workbook()
    ws = wb.active

    ws.title = "Beneficiaires"

    ws.append([
        "Code VIF",
        "Compte comptable",
        "Association",
        "Bénéficiaires",
        "Trouvé"
    ])

    for ligne in data:

        ws.append([
            ligne["code_vif"],
            ligne["compte_comptable"],
            ligne["nom_association"],
            ligne["beneficiaires"],
            ligne["trouve"]
        ])

    largeurs = [15, 20, 50, 15, 10]

    for i, largeur in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largeur

    output = io.BytesIO()

    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="Beneficiaires_Parsol2l.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
