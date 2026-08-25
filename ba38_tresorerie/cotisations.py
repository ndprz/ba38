import io
import re
import os
import sqlite3
import tempfile
from datetime import datetime, date
from collections import defaultdict
from pathlib import Path

from flask import (
    request, render_template, flash, redirect, url_for, send_file,
    current_app, session, abort,
)
from flask_login import login_required
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from utils import (
    get_db_path, write_log, envoyer_mail, upload_file_to_drive_path,
    slugify_filename, split_emails, require_access,
)

from ba38_tresorerie import tresorerie_bp
from ba38_tresorerie.constants import (
    BA380_SHARED_DRIVE_ID, BAI_NOM, BAI_ADRESSE, BAI_TEL, BAI_MAIL,
    BAI_IBAN, BAI_BIC, BAI_SIREN, BAI_NAF,
)


def calculer_cotisations_par_annee(db_path, benef_par_vif):
    """
    benef_par_vif : dict { code_vif : total_beneficiaires }

    Retourne un dict :
    {
        "facturables": [ ... ],
        "orphelines": { code_vif: nb_benef }
    }
    """

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # ============================
    # Chargement des associations
    # ============================
    cursor.execute("""
        SELECT
            Id AS id_association,
            code_VIF,
            vif_regroup_cotisation,
            nom_association,
            compte_comptable,
            adresse_association_1,
            adresse_association_2,
            CP,
            COMMUNE,
            courriel_association
        FROM associations
    """)
    assos = cursor.fetchall()

    # Index par code VIF
    asso_par_vif = {a["code_VIF"]: dict(a) for a in assos}

    # ============================
    # Cumul avec regroupements
    # ============================
    cumuls = defaultdict(int)
    rattachements = defaultdict(list)
    benef_sans_asso = defaultdict(int)

    for code_vif, nb in benef_par_vif.items():
        asso = asso_par_vif.get(code_vif)

        if not asso:
            benef_sans_asso[code_vif] += nb
            continue

        code_facture = asso["vif_regroup_cotisation"] or code_vif
        cumuls[code_facture] += nb
        rattachements[code_facture].append(code_vif)

    # ============================
    # Construction du résultat
    # ============================
    resultats = []

    for code_facture, total_benef in cumuls.items():
        asso = asso_par_vif.get(code_facture)
        if not asso:
            continue

        # Adresse multi-lignes propre
        adresse = "\n".join(filter(None, [
            asso.get("adresse_association_1"),
            asso.get("adresse_association_2"),
            " ".join(filter(None, [
                asso.get("CP"),
                asso.get("COMMUNE")
            ]))
        ]))

        # Calcul cotisation
        if total_benef <= 1000:
            cotis = 50
        elif total_benef <= 10000:
            cotis = 80
        else:
            cotis = 110

        # ============================
        # Construction nom affiché
        # ============================

        noms_regroupes = []
        for code in rattachements[code_facture]:
            if code != code_facture:
                asso_r = asso_par_vif.get(code)
                if asso_r:
                    noms_regroupes.append(asso_r["nom_association"])

        nom_affiche = asso["nom_association"]
        if noms_regroupes:
            nom_affiche = nom_affiche + " / " + " / ".join(noms_regroupes)

        # ============================
        # Noms des associations regroupées
        # ============================

        noms_regroupes = []
        for code in rattachements[code_facture]:
            if code != code_facture:
                asso_r = asso_par_vif.get(code)
                if asso_r:
                    noms_regroupes.append(asso_r["nom_association"])

        # Nom STRICTEMENT de l'association maître (bloc adresse)
        nom_association_maitre = asso["nom_association"]

        # Nom affiché (tableau / Excel)
        nom_affichage = nom_association_maitre
        if noms_regroupes:
            nom_affichage = nom_association_maitre + " / " + " / ".join(noms_regroupes)

        # Texte concaténé (commentaire uniquement)
        commentaire_regroupement = None
        if noms_regroupes:
            commentaire_regroupement = (
                "Cette cotisation regroupe aussi les bénéficiaires des partenaires suivants :\n"
                + " / ".join(noms_regroupes)
            )


        resultats.append({
            "id_association": asso["id_association"],
            "code_vif_facture": code_facture,
            "compte_comptable": asso["compte_comptable"],
            "nom_association": nom_association_maitre,
            "nom_association_affichage": nom_affichage,
            "adresse": adresse,
            "email": asso.get("courriel_association"),
            "beneficiaires": total_benef,
            "cotisation": cotis,
            "codes_vif_inclus": rattachements[code_facture],
            "commentaire_regroupement": commentaire_regroupement

        })

    conn.close()

    return {
        "facturables": resultats,
        "orphelines": dict(benef_sans_asso)
    }


@tresorerie_bp.route("/cotisations", methods=["GET", "POST"])
@login_required
@require_access("tresorerie", "ecriture")
def cotisations():
    """
    Module principal de facturation des cotisations.

    Fonctionnement :
    - GET  : affiche les cotisations déjà calculées pour une année
    - POST : calcule les cotisations à partir du fichier PARSOL,
             insère en base (si année non verrouillée),
             puis affiche le résultat.

    Règle métier :
    - Si des factures ont déjà été envoyées en PROD pour une année,
      le recalcul est bloqué.
    - Le mode TEST ne bloque jamais.
    """

    from datetime import datetime

    mail_mode = session.get(
        "MAIL_MODE",
        os.getenv("MAIL_MODE", "PROD").upper()
    )

    mail_test_to = os.getenv(
        "MAIL_TEST_TO",
        "ba380.informatique2@banquealimentaire.org"
    )

    resultats = None
    orphelines = None
    from datetime import datetime

    annee = request.args.get("annee")

    if not annee:
        annee = datetime.now().year
    else:
        annee = int(annee)

    conn = sqlite3.connect(get_db_path())
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # ==========================================================
    # POST → CALCUL NOUVELLE ANNÉE
    # ==========================================================
    if request.method == "POST":

        try:
            annee = int(request.form.get("annee"))

            # --------------------------------------------------
            # 🔒 BLOCAGE SI ANNÉE VALIDÉE EN PROD
            # --------------------------------------------------
            cursor.execute("""
                SELECT COUNT(*)
                FROM cotisations
                WHERE annee = ?
                AND date_envoi_facture IS NOT NULL
                AND mode_test_envoi_facture = 0
            """, (annee,))

            if cursor.fetchone()[0] > 0:
                flash(
                    "⚠️ Cette année a déjà été validée "
                    "(factures envoyées en PROD). "
                    "Recalcul interdit.",
                    "danger"
                )
                conn.close()
                return redirect(url_for("tresorerie.cotisations"))

            # --------------------------------------------------
            # SUPPRESSION DES CALCULS PRÉCÉDENTS
            # (uniquement si non verrouillée)
            # --------------------------------------------------
            cursor.execute(
                "DELETE FROM cotisations WHERE annee = ?",
                (annee,)
            )

            # --------------------------------------------------
            # PARSE FICHIER PARSOL
            # --------------------------------------------------
            fichier = request.files.get("parsol_file")
            if not fichier:
                raise ValueError("Fichier PARSOL manquant")

            with tempfile.NamedTemporaryFile(delete=False) as tmp:
                fichier.save(tmp.name)
                parsol_path = tmp.name

            try:
                benefs = parse_parsol2l_annuel(parsol_path)
            finally:
                os.remove(parsol_path)

            data = calculer_cotisations_par_annee(
                get_db_path(),
                benefs
            )

            resultats = data["facturables"]
            orphelines = data["orphelines"]

            facture_start = int(request.form.get("facture_start"))

            # Tri par compte comptable
            resultats = sorted(
                resultats,
                key=lambda x: (x.get("compte_comptable") or "")
            )

            # --------------------------------------------------
            # INSERTION EN BASE
            # --------------------------------------------------
            for i, r in enumerate(resultats):
                numero_facture = facture_start + i

                cursor.execute("""
                    INSERT INTO cotisations (
                        annee,
                        id_association,
                        numero_facture,
                        code_vif,
                        beneficiaires,
                        montant,
                        date_calcul,
                        statut,
                        commentaire_regroupement
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    annee,
                    r["id_association"],
                    numero_facture,
                    r["code_vif_facture"],
                    r["beneficiaires"],
                    r["cotisation"],
                    datetime.now().isoformat(),
                    "calcule",
                    r.get("commentaire_regroupement")
                ))

            conn.commit()

            flash(
                f"✅ Cotisations {annee} calculées et enregistrées.",
                "success"
            )

            return redirect(
                url_for("tresorerie.cotisations", annee=annee)
            )

        except Exception:
            current_app.logger.exception(
                "Erreur calcul cotisations"
            )
            flash(
                "Erreur lors du calcul des cotisations",
                "danger"
            )
            conn.close()
            return redirect(url_for("tresorerie.cotisations", annee=annee))

    # ==========================================================
    # GET → AFFICHAGE ANNÉE
    # ==========================================================
    if annee:
        cursor.execute("""
            SELECT
                c.*,
                a.nom_association,
                a.compte_comptable,
                c.code_vif AS code_vif_facture,
                c.montant AS cotisation
            FROM cotisations c
            JOIN associations a ON a.Id = c.id_association
            WHERE c.annee = ?
            ORDER BY c.numero_facture
        """, (annee,))

        lignes = cursor.fetchall()

        if lignes:
            resultats = []

            for l in lignes:
                ligne_dict = dict(l)   # ← ICI EXACTEMENT
                # Si vous avez des regroupements, recalculez nom_association_affichage ici
                # Exemple simplifié (à adapter selon votre logique métier) :
                ligne_dict["nom_association_affichage"] = ligne_dict["nom_association"]
                # Si vous avez des regroupements, ajoutez-les ici
                resultats.append(ligne_dict)

    conn.close()

    # ==========================================================
    # RENDU TEMPLATE
    # ==========================================================
    return render_template(
        "tresorerie/cotisations.html",
        resultats=resultats,
        orphelines=orphelines,
        annee=annee,
        mail_mode=mail_mode,
        mail_test_to=mail_test_to,
        job_done=False,
        mode_relance=False,
        manque=(request.args.get("manque") == "1")
    )


@tresorerie_bp.route("/cotisations/toggle_test_mode", methods=["POST"])
@login_required
@require_access("tresorerie", "ecriture")
def cotisations_toggle_test_mode():

    current = session.get(
        "MAIL_MODE",
        os.getenv("MAIL_MODE", "PROD").upper()
    )

    if current == "PROD":
        session["MAIL_MODE"] = "TEST"
        flash("🧪 Mode TEST activé (mails redirigés)", "warning")
    else:
        session["MAIL_MODE"] = "PROD"
        flash("✅ Mode PROD réactivé", "success")

    return redirect(request.referrer or url_for("tresorerie.cotisations"))


@tresorerie_bp.route("/cotisations/generer_pdfs", methods=["GET"])
@login_required
@require_access("tresorerie", "ecriture")
def cotisations_generer_pdfs():
    """
    Génère les PDF des cotisations pour une année donnée.

    - Lecture depuis cotisations
    - JOIN avec associations
    - Batch pour éviter timeout
    - Mise à jour date_generation_pdf + statut
    """

    from datetime import datetime
    import shutil

    BATCH_SIZE = 20

    annee = request.args.get("annee")
    offset = int(request.args.get("offset", 0))

    if not annee:
        flash("Année manquante", "danger")
        return redirect(url_for("tresorerie.cotisations"))

    annee = int(annee)

    conn = sqlite3.connect(get_db_path())
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # ==========================================================
    # Lecture cotisations + associations
    # ==========================================================
    cursor.execute("""
        SELECT
            c.*,
            a.nom_association,
            a.adresse_association_1,
            a.adresse_association_2,
            a.CP,
            a.COMMUNE,
            a.courriel_association
        FROM cotisations c
        JOIN associations a
            ON a.Id = c.id_association
        WHERE c.annee = ?
        ORDER BY c.numero_facture
    """, (annee,))

    lignes = cursor.fetchall()

    total = len(lignes)
    end = min(offset + BATCH_SIZE, total)

    if total == 0:
        flash("Aucune cotisation trouvée", "warning")
        conn.close()
        return redirect(url_for("tresorerie.cotisations", annee=annee))

    # write_log(f"📄 PDF {offset+1} à {end} / {total}")

    base_drive = f"COTISATIONS/Cotisations {annee}/Factures PDF"

    tmp_dir = Path(tempfile.mkdtemp(prefix="cotisations_pdf_"))

    try:
        for ligne in lignes[offset:end]:

            # --------------------------------------------------
            # Construction adresse multi-lignes
            # --------------------------------------------------
            adresse = "\n".join(filter(None, [
                ligne["adresse_association_1"],
                ligne["adresse_association_2"],
                " ".join(filter(None, [
                    ligne["CP"],
                    ligne["COMMUNE"]
                ]))
            ]))

            nom_asso_slug = slugify_filename(ligne["nom_association"])
            code_vif = ligne["code_vif"]

            nom_pdf = (
                f"FACTURE_{ligne['numero_facture']}_"
                f"{code_vif}_{nom_asso_slug}.pdf"
            )

            pdf_path = tmp_dir / nom_pdf

            # --------------------------------------------------
            # Génération PDF
            # --------------------------------------------------
            generer_facture_pdf(
                {
                    "nom_association": ligne["nom_association"],
                    "adresse": adresse,
                    "cotisation": ligne["montant"],
                    "annee": annee,
                    "code_vif_facture": code_vif,
                    "numero_facture": ligne["numero_facture"],
                    "commentaire_regroupement": ligne["commentaire_regroupement"]
                },
                pdf_path
            )

            # --------------------------------------------------
            # Upload Drive
            # --------------------------------------------------
            upload_file_to_drive_path(
                local_path=str(pdf_path),
                drive_path=base_drive,
                filename=nom_pdf,
                shared_drive_id=BA380_SHARED_DRIVE_ID
            )

            # --------------------------------------------------
            # Mise à jour base
            # --------------------------------------------------
            cursor.execute("""
                UPDATE cotisations
                SET date_generation_pdf = ?,
                    statut = 'pdf_genere'
                WHERE id = ?
            """, (
                datetime.now().isoformat(),
                ligne["id"]
            ))

        conn.commit()

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

    # ==========================================================
    # Batch suivant ?
    # ==========================================================
    if end < total:
        conn.close()
        return redirect(
            url_for(
                "tresorerie.cotisations_generer_pdfs",
                annee=annee,
                offset=end
            )
        )

    conn.close()

    flash(f"✅ {total} factures PDF générées.", "success")

    return redirect(
        url_for("tresorerie.cotisations", annee=annee)
    )


@tresorerie_bp.route("/cotisations/envoyer_mails", methods=["POST"])
@login_required
@require_access("tresorerie", "ecriture")
def cotisations_envoyer_mails():
    """
    Envoi des mails de cotisation pour une année.

    - Lecture base + JOIN associations
    - Respect mode TEST / PROD
    - Mise à jour date_envoi_mail + statut
    """

    from datetime import datetime

    annee = request.form.get("annee")
    mail_sender = request.form.get("mail_sender")
    confirmation_prod = request.form.get("confirmation_prod")

    if not annee:
        flash("Année manquante", "danger")
        return redirect(url_for("tresorerie.cotisations"))

    if not mail_sender:
        flash("Expéditeur manquant", "danger")
        return redirect(url_for("tresorerie.cotisations", annee=annee))

    annee = int(annee)

    mail_mode = session.get(
        "MAIL_MODE",
        os.getenv("MAIL_MODE", "PROD").upper()
    )

    mail_test_to = os.getenv(
        "MAIL_TEST_TO",
        "ba380.informatique2@banquealimentaire.org"
    )

    # ==========================================================
    # Sécurité PROD
    # ==========================================================
    if mail_mode == "PROD" and confirmation_prod != "1":
        flash(
            "⚠️ Envoi bloqué : confirmation PROD requise.",
            "danger"
        )
        return redirect(url_for("tresorerie.cotisations", annee=annee))

    conn = sqlite3.connect(get_db_path())
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # ==========================================================
    # Lecture cotisations non encore envoyées
    # ==========================================================
    cursor.execute("""
        SELECT
            c.*,
            a.nom_association,
            a.adresse_association_1,
            a.adresse_association_2,
            a.CP,
            a.COMMUNE,
            a.courriel_association,
            a.courriel_resp_tresorerie
        FROM cotisations c
        JOIN associations a
            ON a.Id = c.id_association
        WHERE c.annee = ?
        AND c.statut = 'pdf_genere'
        ORDER BY c.numero_facture
    """, (annee,))

    lignes = cursor.fetchall()

    if not lignes:
        flash("Aucune facture prête à être envoyée.", "warning")
        conn.close()
        return redirect(url_for("tresorerie.cotisations", annee=annee))

    nb_mails = 0

    for ligne in lignes:

        email_asso = ligne["courriel_resp_tresorerie"] or ligne["courriel_association"]

        if not email_asso:
            continue

        # --------------------------------------------------
        # Construction adresse
        # --------------------------------------------------
        adresse = "\n".join(filter(None, [
            ligne["adresse_association_1"],
            ligne["adresse_association_2"],
            " ".join(filter(None, [
                ligne["CP"],
                ligne["COMMUNE"]
            ]))
        ]))

        # --------------------------------------------------
        # Corps du mail
        # --------------------------------------------------
        texte_mail = f"""Association : {ligne['nom_association']}

        Madame, Monsieur,

        Conformément à la convention signée avec la Banque Alimentaire de l'Isère,
        votre cotisation pour l'année {annee} s'élève à {ligne['montant']} €.

        Nous vous remercions d'avance pour votre règlement avant le 28 février {annee}.
        """

        if ligne.get("commentaire_regroupement"):
            texte_mail += "\n" + ligne["commentaire_regroupement"] + "\n"

        texte_mail += """

        Bien cordialement,

        Christian GRAFF
        Comptable BAI
        """

        # --------------------------------------------------
        # Destinataires
        # --------------------------------------------------
        if mail_mode == "TEST":
            destinataires = [mail_test_to]
            sujet = (
                f"[TEST] Appel de cotisation {annee} – "
                f"{ligne['nom_association']}"
            )
        else:
            destinataires = split_emails(email_asso)
            if not destinataires:
                continue
            sujet = (
                f"Appel de cotisation {annee} – "
                f"{ligne['nom_association']}"
            )

        # --------------------------------------------------
        # Génération PDF temporaire
        # --------------------------------------------------
        with tempfile.TemporaryDirectory() as tmpdir:

            nom_asso_slug = slugify_filename(
                ligne["nom_association"]
            )

            nom_pdf = (
                f"FACTURE_{ligne['numero_facture']}_"
                f"{ligne['code_vif']}_{nom_asso_slug}.pdf"
            )

            pdf_path = Path(tmpdir) / nom_pdf

            generer_facture_pdf(
                {
                    "nom_association": ligne["nom_association"],
                    "adresse": adresse,
                    "cotisation": ligne["montant"],
                    "annee": annee,
                    "code_vif_facture": ligne["code_vif"],
                    "numero_facture": ligne["numero_facture"],
                    "commentaire_regroupement": ligne["commentaire_regroupement"]
                },
                pdf_path
            )

            envoyer_mail(
                sujet=sujet,
                destinataires=destinataires,
                texte=texte_mail,
                sender_override=mail_sender,
                attachment_path=str(pdf_path),
                bcc=mail_sender
            )

        # --------------------------------------------------
        # Mise à jour base
        # --------------------------------------------------
        mode_test_flag = 1 if mail_mode == "TEST" else 0

        cursor.execute("""
            UPDATE cotisations
            SET date_envoi_facture = ?,
                mode_test_envoi_facture = ?
            WHERE id = ?
        """, (
            datetime.now().isoformat(),
            mode_test_flag,
            ligne["id"]
        ))
        nb_mails += 1

    conn.commit()
    conn.close()

    if mail_mode == "TEST":
        flash(
            f"🧪 {nb_mails} mails envoyés en MODE TEST.",
            "warning"
        )
    else:
        flash(
            f"✅ {nb_mails} mails envoyés en PROD.",
            "success"
        )

    return redirect(
        url_for("tresorerie.cotisations", annee=annee)
    )


DATE_REGEX = re.compile(r"\d{2}/\d{2}/\d{4}")
ASSO_REGEX = re.compile(r"Association\s*:\s*(\d{8})")


def parse_parsol2l_annuel(file_path):
    """
    Parse un fichier PARSOL2L annuel (texte)
    Retourne un dict :
        { code_vif (str) : total_beneficiaires (int) }
    """

    totaux = defaultdict(int)
    code_vif_courant = None

    with open(file_path, "r", encoding="latin-1", errors="ignore") as f:
        for line in f:
            line = line.strip()

            if not line:
                continue

            # ==========================
            # Détection nouvelle association
            # ==========================
            m_asso = ASSO_REGEX.search(line)
            if m_asso:
                code_vif_courant = m_asso.group(1)
                continue

            if not code_vif_courant:
                continue

            # ==========================
            # Ligne de passage (date en début)
            # ==========================
            if DATE_REGEX.match(line):
                parts = re.split(r"\s+", line)
                if len(parts) < 2:
                    continue

                try:
                    nb_benef = int(parts[1])
                except ValueError:
                    continue

                totaux[code_vif_courant] += nb_benef

    return dict(totaux)


def generer_facture_pdf(data, output_path):
    """
    Génère une facture ou relance PDF individuelle de cotisation BA38.

    data doit contenir :
      - nom_association
      - adresse (multi-lignes)
      - cotisation
      - annee
      - code_vif_facture
    """

    c = canvas.Canvas(str(output_path), pagesize=A4)
    largeur, hauteur = A4

    # ============================
    # LOGO BA (centré page)
    # ============================

    logo_path = (
        Path(current_app.root_path)
        / "static"
        / "images"
        / "logo_ba_complet.png"
    )

    if logo_path.exists():
        logo = ImageReader(str(logo_path))

        # Taille du logo (+30 % supplémentaires)
        logo_width = 120 * mm
        logo_height = 46 * mm   # ratio conservé

        # --- CENTRAGE OPTIQUE ---
        # Le logo est visuellement décalé dans son PNG,
        # on applique une correction horizontale manuelle
        correction_x = -20 * mm   # ajuste si besoin ±1 mm

        x_logo = (largeur - logo_width) / 2 - correction_x

        # --- POSITION VERTICALE ---
        # REMONTER de 3 cm par rapport à l’état actuel
        y_logo = hauteur - logo_height - 5 * mm

        c.drawImage(
            logo,
            x_logo,
            y_logo,
            width=logo_width,
            height=logo_height,
            preserveAspectRatio=True,
            mask="auto"
        )
    else:
        # optionnel mais TRÈS utile
        print(f"LOGO INTROUVABLE : {logo_path}")

    # ============================
    # Décalage global sous le logo
    # ============================

    y = hauteur - 20 * mm - 40 * mm

    # ============================
    # BLOC BAI (gauche)
    # ============================

    c.setFont("Helvetica-Bold", 11)
    c.drawString(20 * mm, y, BAI_NOM)

    c.setFont("Helvetica", 9)
    y -= 12
    for line in BAI_ADRESSE.split("\n"):
        c.drawString(20 * mm, y, line)
        y -= 10

    c.drawString(20 * mm, y, f"Tél : {BAI_TEL}")
    y -= 10
    c.drawString(20 * mm, y, f"Mail : {BAI_MAIL}")

    # ============================
    # INFOS FACTURE (droite)
    # ============================

    y_fact = hauteur - 20 * mm - 40 * mm
    c.setFont("Helvetica", 9)
    c.drawRightString(
        largeur - 20 * mm,
        y_fact,
        f"Date : {date.today().strftime('%d/%m/%Y')}"
    )
    y_fact -= 15
    c.drawRightString(
        largeur - 20 * mm,
        y_fact,
        f"Échéance : 28/02/{data['annee']}"
    )
    # Numéro de facture uniquement en facturation normale
    if not data.get("mode_relance"):
        y_fact -= 15
        c.drawRightString(
            largeur - 20 * mm,
            y_fact,
            f"Facture n° {data.get('numero_facture')}"
        )
    if data.get("mode_relance"):
        y_fact -= 15
        c.setFont("Helvetica-Bold", 10)
        c.drawRightString(
            largeur - 20 * mm,
            y_fact,
            f"RAPPEL n°{data.get('numero_relance')}"
        )
        c.setFont("Helvetica", 9)

        c.setFillColorRGB(0, 0, 0)


    # ============================
    # ADRESSE ASSOCIATION
    # ============================

    y -= 40  # ⬅️ descente volontaire pour éviter chevauchement
    c.setFont("Helvetica-Bold", 10)
    c.drawString(20 * mm, y, data["nom_association"])

    c.setFont("Helvetica", 9)
    y -= 12
    for line in data["adresse"].split("\n"):
        c.drawString(20 * mm, y, line)
        y -= 10

    # ============================
    # TABLE FACTURE
    # ============================

    y -= 20
    c.setFont("Helvetica-Bold", 9)
    c.drawString(20 * mm, y, "Désignation")
    c.drawRightString(largeur - 20 * mm, y, "Montant")

    y -= 10
    c.line(20 * mm, y, largeur - 20 * mm, y)

    y -= 15
    c.setFont("Helvetica", 9)
    c.drawString(20 * mm, y, f"COTISATION {data['annee']}")
    c.drawRightString(
        largeur - 20 * mm,
        y,
        f"{data['cotisation']:.2f} €"
    )

    # ============================
    # TOTAL
    # ============================

    y -= 30
    c.line(120 * mm, y, largeur - 20 * mm, y)
    y -= 15
    c.setFont("Helvetica-Bold", 10)
    c.drawString(120 * mm, y, "Net à payer")
    c.drawRightString(
        largeur - 20 * mm,
        y,
        f"{data['cotisation']:.2f} €"
    )

    # ============================
    # MENTIONS LÉGALES
    # ============================

    y -= 30
    c.setFont("Helvetica", 8)
    c.drawString(20 * mm, y, "TVA non applicable, art. 293B du CGI")

    y -= 15
    c.drawString(20 * mm, y, f"IBAN : {BAI_IBAN}")
    y -= 10
    c.drawString(20 * mm, y, f"BIC : {BAI_BIC}")

    y -= 20
    c.drawString(20 * mm, y, f"SIREN : {BAI_SIREN}")
    y -= 10
    c.drawString(20 * mm, y, f"NAF : {BAI_NAF}")

    # ============================
    # COMMENTAIRE REGROUPEMENT (PIED DE PAGE)
    # ============================

    commentaire = data.get("commentaire_regroupement")
    if commentaire:
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph
        from reportlab.lib.enums import TA_LEFT

        styles = getSampleStyleSheet()
        style = styles["Normal"]
        style.fontName = "Helvetica-Oblique"
        style.fontSize = 8
        style.leading = 10
        style.alignment = TA_LEFT

        # Zone pied de page (largeur utile)
        largeur_bloc = largeur - 40 * mm
        x_bloc = 20 * mm
        y_bloc = 35 * mm   # pied de page sûr

        p = Paragraph(
            commentaire.replace("\n", "<br/>"),
            style
        )

        w, h = p.wrap(largeur_bloc, 60 * mm)
        p.drawOn(c, x_bloc, y_bloc)
    # ============================
    # FINALISATION
    # ============================

    c.showPage()
    c.save()


@tresorerie_bp.route("/cotisations/export_excel", methods=["POST"])
@login_required
@require_access("tresorerie", "ecriture")
def cotisations_export_excel():
    import json
    from io import BytesIO
    from openpyxl import Workbook
    from flask import send_file

    annee = request.form.get("annee")
    raw = request.form.get("data")

    if not raw:
        flash("Aucune donnée à exporter", "danger")
        return redirect(url_for("tresorerie.cotisations"))

    lignes = json.loads(raw)

    # ============================
    # Création Excel en mémoire
    # ============================
    wb = Workbook()
    ws = wb.active
    ws.title = f"Cotisations {annee}"

    # En-têtes
    ws.append([
        "Code VIF",
        "Compte Comptable",
        "Association",
        "Bénéficiaires",
        "Cotisation (€)"
    ])

    # Lignes
    for l in lignes:
        ws.append([
            l.get("code_vif_facture"),
            l.get("compte_comptable"),
            l.get("nom_association_affichage"),  # 👈 ici
            l.get("beneficiaires"),
            l.get("cotisation"),
            l.get("commentaire_regroupement") or ""
        ])

    # Ajustement largeur colonnes
    largeurs = [15, 22, 45, 18, 18]

    for i, largeur in enumerate(largeurs, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = largeur

    # ============================
    # Export HTTP
    # ============================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Cotisations_{annee}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


@tresorerie_bp.route("/cotisations/export/<int:annee>")
@login_required
@require_access("tresorerie", "ecriture")
def export_cotisations_excel(annee):

    import pandas as pd

    conn = sqlite3.connect(get_db_path())
    df = pd.read_sql_query("""
        SELECT
            c.numero_facture,
            a.nom_association,
            a.compte_comptable,
            c.code_vif,
            c.beneficiaires,
            c.montant,
            c.statut,
            c.date_paiement
        FROM cotisations c
        JOIN associations a
            ON a.Id = c.id_association
        WHERE c.annee = ?
        ORDER BY c.numero_facture
    """, conn, params=(annee,))
    conn.close()

    output = io.BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"cotisations_{annee}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


COTISATIONS_MENU_ACTIONS = {
    "facturation": ("💰", "Facturation des cotisations"),
    "relance": ("🔔", "Relance des cotisations"),
    "paiements": ("💳", "Saisie des paiements de cotisations"),
}


@tresorerie_bp.route("/cotisations/menu/<action>")
@login_required
@require_access("tresorerie", "ecriture")
def cotisations_menu(action):
    from datetime import datetime

    if action not in COTISATIONS_MENU_ACTIONS:
        abort(404)

    icone, titre = COTISATIONS_MENU_ACTIONS[action]

    return render_template(
        "tresorerie/cotisations_menu.html",
        action=action,
        icone=icone,
        titre=titre,
        annee_defaut=datetime.now().year
    )


@tresorerie_bp.route("/cotisations/start")
@login_required
@require_access("tresorerie", "ecriture")
def cotisations_start():
    # Reset systématique
    job_id = session.get("COTISATIONS_JOB_ID")
    if job_id:
        Path(tempfile.gettempdir(), f"cotisations_{job_id}.json").unlink(missing_ok=True)

    session.pop("COTISATIONS_JOB_ID", None)
    session.pop("COTISATIONS_ANNEE", None)

    annee = request.args.get("annee")
    if annee:
        return redirect(url_for("tresorerie.cotisations", annee=annee))

    return redirect(url_for("tresorerie.cotisations"))


@tresorerie_bp.route("/cotisations/quit")
@login_required
@require_access("tresorerie", "ecriture")
def cotisations_quit():
    job_id = session.get("COTISATIONS_JOB_ID")
    if job_id:
        Path(tempfile.gettempdir(), f"cotisations_{job_id}.json").unlink(missing_ok=True)

    session.pop("COTISATIONS_JOB_ID", None)
    session.pop("COTISATIONS_ANNEE", None)

    flash("Calcul des cotisations fermé", "info")
    return redirect(url_for("tresorerie.tresorerie"))
