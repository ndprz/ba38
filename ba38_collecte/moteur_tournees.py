#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Génération des tournées de collecte BAI 38 - 2026

Copie quasi-verbatim de dev/uploads/collecte_fichiers_source/generer_tournees_bai_v2.py
(outil historique lancé via lancer_tournees_bai_v2.bat), importée par ba38_collecte.py.
Seule différence volontaire avec l'original : lire_pages_pdf() lève une exception
au lieu d'un sys.exit(1) — un sys.exit() ferait mourir le worker Flask entier au
lieu de simplement faire échouer la requête. main()/parse_args() sont conservés
tels quels mais ne sont plus appelés (le CLI d'origine reste utilisable tel quel
pour comparaison, hors de l'appli).
"""
import argparse, sys, os, re, subprocess, glob
from math import radians, cos, sin, asin, sqrt
from collections import defaultdict
from datetime import datetime

# ─── CONSTANTES ──────────────────────────────────────────────────────────────
BAI_LAT, BAI_LON = 45.18867, 5.68456  # 11 allée de la Pinea, 38600 Fontaine
VEHICULES_FIGES  = sorted(['V007','V008','V009','V013','V023','V026','V027','V028','V037'], key=lambda v: int(v[1:]))

# Adresse du siège affichée sur la carte des tournées : tirée de la config
# organisation quand ce module tourne dans l'appli Flask (import optionnel —
# le CLI d'origine reste utilisable hors appli, cf. docstring en tête de
# fichier, d'où le repli sur l'adresse BA38 si l'import/la lecture DB échoue).
try:
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from ba38_utilitaires.organisation import get_organisation as _get_organisation
except Exception:
    _get_organisation = None


def _adresse_siege_html():
    if _get_organisation is not None:
        try:
            return _get_organisation()["adresse"].replace("\n", "<br>")
        except Exception:
            pass
    return "11 All&eacute;e de la Pin&eacute;a<br>38600 Fontaine"

# Contraintes de gel par véhicule ET demi-journée :
# (veh, dj) → ne pas modifier cette tournée (ni ajouter ni retirer de magasins)
TOURNEES_GELEES = set([
    # V027 : gel Vendredi et Samedi (toutes DJ)
    ('V027', 'Vendredi Matin'), ('V027', 'Vendredi Apres Midi'),
    ('V027', 'Samedi Matin'),   ('V027', 'Samedi Apres Midi'),
    # V009 : gel toutes DJ (n'existe que le Samedi)
    ('V009', 'Samedi Matin'),   ('V009', 'Samedi Apres Midi'),
    ('V009', 'Vendredi Matin'), ('V009', 'Vendredi Apres Midi'),
    ('V009', 'Jeudi Matin'),    ('V009', 'Jeudi Apres Midi'),
    ('V009', 'Dimanche Matin'),
    # V028 : gel Vendredi, Samedi et Dimanche
    ('V028', 'Vendredi Matin'), ('V028', 'Vendredi Apres Midi'),
    ('V028', 'Samedi Matin'),   ('V028', 'Samedi Apres Midi'),
    ('V028', 'Dimanche Matin'),
    # V007 : gel Vendredi et Samedi
    ('V007', 'Vendredi Matin'), ('V007', 'Vendredi Apres Midi'),
    ('V007', 'Samedi Matin'),   ('V007', 'Samedi Apres Midi'),
    # V008 : gel Vendredi et Samedi
    ('V008', 'Vendredi Matin'), ('V008', 'Vendredi Apres Midi'),
    ('V008', 'Samedi Matin'),   ('V008', 'Samedi Apres Midi'),
    # V013 : gel Vendredi et Samedi
    ('V013', 'Vendredi Matin'), ('V013', 'Vendredi Apres Midi'),
    ('V013', 'Samedi Matin'),   ('V013', 'Samedi Apres Midi'),
    # V023 : gel toutes DJ (Super U Voreppe ramené par association)
    ('V023', 'Vendredi Matin'), ('V023', 'Vendredi Apres Midi'),
    ('V023', 'Samedi Matin'),   ('V023', 'Samedi Apres Midi'),
    ('V023', 'Dimanche Matin'), ('V023', 'Jeudi Matin'), ('V023', 'Jeudi Apres Midi'),
    # V003, V006 : gel Jeudi uniquement
    ('V003', 'Jeudi Matin'), ('V003', 'Jeudi Apres Midi'),
    ('V006', 'Jeudi Matin'), ('V006', 'Jeudi Apres Midi'),
    # V037 : gel toutes DJ (n'existe que le Jeudi)
    ('V037', 'Jeudi Matin'), ('V037', 'Jeudi Apres Midi'),
    ('V037', 'Vendredi Matin'), ('V037', 'Vendredi Apres Midi'),
    ('V037', 'Samedi Matin'),   ('V037', 'Samedi Apres Midi'),
    ('V037', 'Dimanche Matin'),
    # V026 : gel toutes DJ
    ('V026', 'Vendredi Matin'), ('V026', 'Vendredi Apres Midi'),
    ('V026', 'Samedi Matin'),   ('V026', 'Samedi Apres Midi'),
    ('V026', 'Dimanche Matin'), ('V026', 'Jeudi Matin'), ('V026', 'Jeudi Apres Midi'),
])

DEMI_JOURNEES = [
    'Jeudi Matin','Jeudi Apres Midi',
    'Vendredi Matin','Vendredi Apres Midi',
    'Samedi Matin','Samedi Apres Midi',
    'Dimanche Matin',
]
DJ_DIMANCHE = {'Dimanche Matin'}


def parse_creneaux(creneaux_raw):
    """
    Convertit le contenu de la colonne 'Créneaux' du référentiel magasins
    (une ligne par plage horaire, ex. 'vendredi 09h30-11h30\\nsamedi 14h30-17h30\\n...')
    en un ensemble de demi-journées (valeurs de DEMI_JOURNEES) où le magasin est
    disponible pour la collecte.

    Règle Matin/Après-midi : même convention que le reste du projet (heure de
    début du créneau < 13h → Matin, sinon Après-midi). Le dimanche n'a qu'une
    seule demi-journée possible dans DEMI_JOURNEES ('Dimanche Matin'), donc tout
    créneau du dimanche y est rattaché quelle que soit son heure de début.

    Retourne un set (éventuellement vide si aucune ligne n'a pu être interprétée),
    ou None si creneaux_raw est vide — pour distinguer "aucun créneau déclaré"
    (pas de filtrage) de "créneaux déclarés mais tous hors DEMI_JOURNEES".
    """
    if not creneaux_raw or str(creneaux_raw).strip().lower() == 'nan':
        return None

    jours_fr = {'jeudi': 'Jeudi', 'vendredi': 'Vendredi', 'samedi': 'Samedi', 'dimanche': 'Dimanche'}
    djs = set()
    for ligne in re.split(r'[\n;]+', str(creneaux_raw)):
        ligne = ligne.strip().lower()
        if not ligne:
            continue
        m = re.match(r'(\w+)\s+(\d{1,2})h(\d{2})', ligne)
        if not m:
            continue
        jour_brut, heure_str, _ = m.groups()
        jour = jours_fr.get(jour_brut)
        if not jour:
            continue
        if jour == 'Dimanche':
            dj = 'Dimanche Matin'
        else:
            periode = 'Matin' if int(heure_str) < 13 else 'Apres Midi'
            dj = f'{jour} {periode}'
        if dj in DEMI_JOURNEES:
            djs.add(dj)
    return djs

# Mapping texte PDF → nom interne (gère accents et variantes)
DJ_MAP = {
    'Jeudi Matin':'Jeudi Matin',
    'Jeudi Après Midi':'Jeudi Apres Midi', 'Jeudi Apres Midi':'Jeudi Apres Midi',
    'Jeudi AprÃ¨s Midi':'Jeudi Apres Midi',   # Latin-1 mal décodé
    'Vendredi Matin':'Vendredi Matin',
    'Vendredi Après Midi':'Vendredi Apres Midi', 'Vendredi Apres Midi':'Vendredi Apres Midi',
    'Vendredi AprÃ¨s Midi':'Vendredi Apres Midi',  # Latin-1 mal décodé
    'Samedi Matin':'Samedi Matin',
    'Samedi Après Midi':'Samedi Apres Midi', 'Samedi Apres Midi':'Samedi Apres Midi',
    'Samedi AprÃ¨s Midi':'Samedi Apres Midi',   # Latin-1 mal décodé
    'Dimanche Matin':'Dimanche Matin',
}

# ─── ARGUMENTS ───────────────────────────────────────────────────────────────
def parse_args():
    p = argparse.ArgumentParser(description='Génère les tournées BAI 38 - 2026')
    p.add_argument('--camions-supp',  type=int, default=0,   metavar='N',  help='Camions supplémentaires (défaut: 0)')
    p.add_argument('--poids-nouveaux',type=int, default=200, metavar='KG', help='Poids nouveaux magasins kg (défaut: 200)')
    p.add_argument('--max-magasins',  type=int, default=4,   metavar='N',  help='Max magasins/tournée hors dimanche (défaut: 4, dimanche: +1)')
    p.add_argument('--corriger-mal-places', action='store_true', default=False,
                   help='Corriger automatiquement les magasins mal placés géographiquement')
    p.add_argument('--fusionner-legeres', action='store_true', default=False,
                   help='Fusionner automatiquement les tournées légères (≤ 2 mag) faisables')
    p.add_argument('--optimiser-anciens', action='store_true', default=False,
                   help='Autoriser le déplacement des anciens magasins (2025) pour optimiser les tournées')
    p.add_argument('--output',   type=str, default=None, metavar='F')
    p.add_argument('--pdf',      type=str, default='fiches jour-véhicule-magasin.pdf', metavar='F')
    p.add_argument('--magasins', type=str, default='liste-magasins2026.xlsx', metavar='F')
    p.add_argument('--nouveaux', type=str, default=None, metavar='F',
                   help='(obsolète) Ignoré — les nouveaux magasins sont détectés automatiquement')
    p.add_argument('--debug-pdf',   action='store_true', help='Affiche les 3 premières pages brutes')
    p.add_argument('--debug-pages', action='store_true', help='Résumé de toutes les pages')
    p.add_argument('--debug-vides', action='store_true', help='Affiche le contenu brut des pages non détectées')
    return p.parse_args()

# ─── GÉO ─────────────────────────────────────────────────────────────────────
def haversine(lat1, lon1, lat2, lon2):
    R = 6371
    dlat, dlon = radians(lat2-lat1), radians(lon2-lon1)
    a = sin(dlat/2)**2 + cos(radians(lat1))*cos(radians(lat2))*sin(dlon/2)**2
    return 2*R*asin(sqrt(a))

def dist_bai(lat, lon):
    return haversine(BAI_LAT, BAI_LON, lat, lon)

# ─── EXTRACTION PDF ───────────────────────────────────────────────────────────
def lire_pages_pdf(pdf_path):
    """Lit le PDF et retourne la liste des pages (texte)."""
    # Tentative pdftotext - PDF généré par Access 2007 = encodage CP1252 (Windows-1252)
    # On lit en bytes et décode en CP1252 pour éviter la corruption des accents
    try:
        r = subprocess.run(['pdftotext', '-enc', 'Latin1', pdf_path, '-'],
                           capture_output=True, timeout=30)
        if r.returncode == 0 and r.stdout:
            # Décoder en CP1252 (Windows-1252) = encodage natif Access 2007
            try:
                texte = r.stdout.decode('cp1252')
            except Exception:
                texte = r.stdout.decode('latin-1')
            pages = texte.split('\f')
            print(f"  → pdftotext CP1252 : {len(pages)} pages")
            return pages
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass
    # Tentative pypdf (déjà installé dans l'environnement de l'appli — pas
    # d'installation à la volée ici, contrairement au script d'origine)
    try:
        from pypdf import PdfReader
        reader = PdfReader(pdf_path)
        pages = [p.extract_text() or '' for p in reader.pages]
        print(f"  → pypdf : {len(pages)} pages")
        return pages
    except Exception as e:
        print(f"  ERREUR pypdf: {e}")
    raise RuntimeError("Impossible de lire le PDF (pdftotext et pypdf ont échoué).")

def normaliser(s):
    """Supprime les diacritiques pour comparaison robuste."""
    import unicodedata
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

DJ_MAP_NORM = {normaliser(k): v for k, v in DJ_MAP.items()}

def parse_page(lines):
    """
    Parse une page du PDF : 1 page = 1 demi-journée = 1 véhicule.
    Retourne (demi_journee, vehicule, vif_codes).
    """
    demi_j         = None
    vehicule       = None
    nom_vehicule   = ''
    vif_codes      = []
    attente_veh    = False
    attente_nom_veh = False

    for line in lines:
        ls = line.strip()
        if not ls:
            continue

        # Demi-journée
        if ls in DJ_MAP:
            demi_j = DJ_MAP[ls]
            continue
        ls_norm = normaliser(ls)
        if ls_norm in DJ_MAP_NORM:
            demi_j = DJ_MAP_NORM[ls_norm]
            continue

        # Véhicule
        if re.search(r'vehicule\s*:', ls, re.IGNORECASE):
            attente_veh = True
            continue
        if attente_veh:
            m = re.match(r'^(V[X]?\d{3})\b', ls, re.IGNORECASE)
            if m:
                vehicule = m.group(1).upper()
                reste = ls[m.end():].strip()
                if reste:
                    nom_vehicule = reste
                else:
                    attente_nom_veh = True
            attente_veh = False
            continue
        if re.match(r'^V[X]?\d{3}$', ls, re.IGNORECASE):
            vehicule = ls.upper()
            attente_nom_veh = True
            continue

        if attente_nom_veh and vehicule:
            MOTS_VEHICULE = ('BAI','CAMION','RENAULT','IVECO','PEUGEOT','HERTZ',
                             'CROIX','DLM','ENTREPRISE','GNV','BD ')
            if re.match(r'\b0\d{7}\b', ls):
                attente_nom_veh = False
            elif (ls.startswith('Quai') or ls.startswith('Enregistr') or
                  ls.startswith('Tournée') or ls.startswith('prendre') or
                  ls.startswith('cagettes') or ls.startswith('Collecté') or
                  ls.startswith('Ramener') or re.match(r'^\d+$', ls) or
                  ls in ('V', 'S', 'VIF')):
                pass  # ignorer, continuer à chercher
            elif ls and len(ls) > 2 and any(m in ls.upper() for m in MOTS_VEHICULE):
                nom_vehicule = ls
                attente_nom_veh = False
            # sinon continuer (adresse, nom magasin...)

        # Codes VIF (8 chiffres commençant par 0)
        for code in re.findall(r'\b(0\d{7})\b', ls):
            vif_norm = code.lstrip('0')
            if vif_norm and vif_norm not in vif_codes:
                vif_codes.append(vif_norm)

    return [(demi_j, vehicule, vif_codes, nom_vehicule)] if demi_j and vehicule else []

def extraire_pdf(pdf_path):
    """Extrait toutes les fiches du PDF. Retourne liste de dicts."""
    print(f"  Lecture PDF : {pdf_path}")
    pages = lire_pages_pdf(pdf_path)
    fiches = []

    for i, page in enumerate(pages):
        lines = [l.strip() for l in page.split('\n') if l.strip()]
        if len(lines) < 3:
            continue
        for res in parse_page(lines):
            demi_j, vehicule, vif_codes = res[0], res[1], res[2]
            nom_veh = res[3] if len(res) > 3 else ''
            fiches.append({'demi_journee': demi_j, 'vehicule': vehicule, 'vif_codes': vif_codes, 'nom_vehicule': nom_veh})

    print(f"  → {len(fiches)} fiches extraites")
    from collections import Counter
    for dj, n in sorted(Counter(f['demi_journee'] for f in fiches).items()):
        print(f"      {dj}: {n} tournées")
    if not any('Jeudi' in f['demi_journee'] for f in fiches):
        print("  AVERTISSEMENT: aucune fiche Jeudi détectée !")
    return fiches

# ─── MAGASINS ─────────────────────────────────────────────────────────────────
def lire_magasins(path_mag, vifs_pdf_2025, poids_nouveaux):
    import pandas as pd

    print(f"  Lecture magasins : {path_mag}")
    df = pd.read_excel(path_mag)
    df.columns = [c.strip() for c in df.columns]
    df = df.loc[:, ~df.columns.duplicated()]

    # Renommage colonnes utiles
    rmap = {}
    for col in df.columns:
        cl = col.lower()
        if 'vif' in cl or cl == 'code':        rmap[col] = 'Code VIF'
        elif ('nom' in cl or 'magasin' in cl) and 'fiche' not in cl: rmap[col] = 'Nom'
        elif 'ville' in cl:                     rmap[col] = 'Ville'
        elif 'lat' in cl:                       rmap[col] = 'Latitude'
        elif 'lon' in cl:                       rmap[col] = 'Longitude'
        elif 'tonnage' in cl and '2025' in cl:  rmap[col] = 'Tonnage 2025'
        elif 'secteur' in cl:                   rmap[col] = 'Secteur'
    df = df.rename(columns=rmap)
    df = df.loc[:, ~df.columns.duplicated()]

    for col in ['Code VIF','Nom','Ville','Latitude','Longitude','Tonnage 2025','Secteur','État','Stockage']:
        if col not in df.columns:
            df[col] = ''

    # Filtrer les magasins actifs : État = 'Collecté par la BAI', PLUS les magasins
    # en 'Collecte gardée' (collecte assurée par une autre association) dont le
    # Stockage reste néanmoins partagé avec la BAI (ex. 'BAI + EQUILIBRE',
    # 'BAI+3ABI') — la BAI doit alors quand même y passer un camion pour le
    # stockage même si elle n'assure pas la collecte elle-même.
    # Exclut toujours 'Non collecté' et 'Collecte gardée' sans stockage BAI+.
    nb_avant = len(df)
    if 'État' in df.columns:
        etat = df['État'].astype(str).str.strip()
        stockage = df['Stockage'].astype(str).str.strip()
        mask_collecte_bai = etat == 'Collecté par la BAI'
        mask_gardee_bai_plus = (etat == 'Collecte gardée') & stockage.str.contains(r'BAI\s*\+', case=False, regex=True, na=False)
        nb_gardee_bai_plus = int(mask_gardee_bai_plus.sum())
        df = df[mask_collecte_bai | mask_gardee_bai_plus].reset_index(drop=True)
        print(f"  → Filtre État='Collecté par la BAI' + 'Collecte gardée' avec Stockage BAI+ : {nb_avant} → {len(df)} magasins retenus"
              f" (dont {nb_gardee_bai_plus} en collecte gardée/stockage BAI+)")
    else:
        print(f"  INFO: colonne 'État' absente, tous les {nb_avant} magasins conservés")

    df['Tonnage 2025'] = pd.to_numeric(df['Tonnage 2025'], errors='coerce').fillna(0)
    df['Nouveau'] = False

    # Construire les secteurs géographiques
    # Normaliser les noms de villes (casse, accents)
    import unicodedata as _ud
    def _norm_ville(v):
        v = str(v).strip()
        # Normaliser casse
        v = v.title()
        # Regroupements géographiques
        equivalences = {
            "Grenoble": ["Grenoble", "GRENOBLE", "grenoble"],
            "Echirolles": ["Echirolles", "Echirolles", "ECHIROLLES"],
            "Saint Martin D'Heres": ["Saint Martin D'Heres", "Saint Martin D'Heres"],
            "Saint Egreve": ["Saint Egreve", "Saint Egreve", "ST EGREVE"],
        }
        for ref, variantes in equivalences.items():
            if v in variantes or v.lower() in [x.lower() for x in variantes]:
                return ref
        return v

    if 'Secteur' not in df.columns or df['Secteur'].astype(str).str.strip().eq('').all():
        df['Ville_norm'] = df['Ville'].apply(_norm_ville)
    else:
        df['Ville_norm'] = df['Ville'].apply(_norm_ville)

    # Compter magasins par ville normalisée
    nb_par_ville = df['Ville_norm'].value_counts()

    # Secteur = ville, mais les grosses villes (>8 magasins) sont découpées
    # par secteur géographique (Nord/Sud/Centre/Est/Ouest)
    SEUIL_GROS = 6

    def _secteur_grenoble(lat, lon):
        """7 secteurs Grenoble basés sur les coordonnées GPS réelles des 30 magasins."""
        if lat > 45.193:                    return 'Grenoble Nord'        # Ile Verte, Esplanade
        if lat < 45.173:                    return 'Grenoble Sud'         # Reynoard, Libérations, Lidl
        if lon < 5.714:                     return 'Grenoble Ouest'       # Fusillés, Rhin Danube, Vallier
        if lon > 5.729:                     return 'Grenoble Est'         # Albert 1er, Lafayette, Spars
        if lat > 45.187:                    return 'Grenoble Centre Nord' # Jaurès, Berriat, Viallet
        if lon < 5.721:                     return 'Grenoble Centre Ouest'# Monoprix Lory, U Express, Itm Vallier
        return 'Grenoble Centre Est'                                      # Foch, Stalingrad, Aldi, Super U

    GRESIVAUDAN_VILLES = {'Biviers', 'Crolles', 'Froges', 'Saint Ismier', 'Domene', 'Le Versoud'}

    def _attribuer_secteur(row):
        ville = row['Ville_norm']
        # Regroupement Grésivaudan en priorité (avant le test SEUIL_GROS)
        if ville in GRESIVAUDAN_VILLES:
            return 'Gresivaudan'
        if ville in {'Voiron', 'Saint Jean De Moirans'}:                return 'Voiron'
        if ville in {'Rives', 'Renage'}:                                return 'Rives'
        if ville in {'Seyssinet Pariset', 'Seyssins', 'Fontaine'}:     return 'Seyssinet'
        if ville in {'Apprieu', 'Colombe', 'Le Grand Lemps'}:          return 'Plateau Nord'
        if ville in {'Varces Allieres Et Risset', 'Vif', 'Vizille', 'Claix'}: return 'Sud Vercors'
        if ville in {'La Terrasse', 'Le Touvet', 'Tencin'}:            return 'Chartreuse Est'
        if ville in {'Peage De Roussillon', 'Salaise Sur Sanne'}:      return 'Roussillon'
        if ville in {'Poisat', 'Saint Martin Dheres', 'Saint Martin D Heres', 'Gieres', 'Gières'}: return 'Saint Martin Dheres'
        nb = nb_par_ville.get(ville, 0)
        if nb <= SEUIL_GROS:
            return ville
        lat = float(row['Latitude']) if row['Latitude'] else 0
        lon = float(row['Longitude']) if row['Longitude'] else 0
        if 'grenoble' in ville.lower():
            return _secteur_grenoble(lat, lon)
        # Autres grosses villes (ex: Saint Martin d'Hères) : Nord/Sud
        lat_moy = df[df['Ville_norm']==ville]['Latitude'].astype(float).mean()
        sub = 'Nord' if lat > lat_moy else 'Sud'
        return f'{ville} {sub}'

    df['Secteur'] = df.apply(_attribuer_secteur, axis=1)
    df = df.reset_index(drop=True)

    # Calculer le nombre de passages par VIF (= nb de fiches dans le PDF qui le contiennent)
    # -> sera utilisé plus tard pour diviser le tonnage annuel
    df['NbPassages'] = 0  # sera mis à jour après extraction PDF si disponible
    df['Tonnage 2025'] = pd.to_numeric(df['Tonnage 2025'], errors='coerce').fillna(0)

    # Détecter les nouveaux magasins 2026 :
    # Nouveau = présent dans liste-magasins2026.xlsx (État='Collecté par la BAI')
    #           ET absent du PDF des tournées 2025 (VIF non reconnu)
    # vifs_pdf_2025 = ensemble des codes VIF extraits du PDF (sans zéro initial)
    vifs_df_str = df['Code VIF'].astype(str).str.strip().str.lstrip('0')
    nb_nouveaux = 0
    for idx_r in df.index:
        vif_norm = vifs_df_str[idx_r]
        if vif_norm not in vifs_pdf_2025:
            df.loc[idx_r, 'Nouveau'] = True
            df.loc[idx_r, 'Tonnage 2025'] = poids_nouveaux
            nb_nouveaux += 1
    print(f"  → {nb_nouveaux} nouveaux magasins détectés (absents du PDF 2025, poids: {poids_nouveaux} kg)")

    # Détecter les magasins sans coordonnées GPS avant de mettre les valeurs par défaut
    df['Latitude_raw']  = pd.to_numeric(df['Latitude'],  errors='coerce')
    df['Longitude_raw'] = pd.to_numeric(df['Longitude'], errors='coerce')
    df['Sans_GPS'] = (df['Latitude_raw'].isna() | df['Longitude_raw'].isna() |
                      (df['Latitude_raw'] == 0)  | (df['Longitude_raw'] == 0))
    sans_gps = df[df['Sans_GPS']]
    if len(sans_gps) > 0:
        print(f"  AVERTISSEMENT: {len(sans_gps)} magasin(s) sans coordonnées GPS :")
        for _, r in sans_gps.iterrows():
            print(f"    - {r['Nom']} ({r.get('Ville','')})")
    df['Latitude']  = df['Latitude_raw'].fillna(BAI_LAT)
    df['Longitude'] = df['Longitude_raw'].fillna(BAI_LON)
    df = df.drop(columns=['Latitude_raw','Longitude_raw'])
    df['Dist_BAI']  = df.apply(lambda r: dist_bai(r['Latitude'], r['Longitude']), axis=1)
    df = df.reset_index(drop=True)
    print(f"  → {len(df)} magasins ({int(df['Nouveau'].sum())} nouveaux)")
    return df

# ─── OPTIMISATION TOURNÉES ────────────────────────────────────────────────────
def optimiser_tournees(fiches, df_mag, args):
    import pandas as pd
    from collections import Counter

    max_norm = args.max_magasins
    max_dim  = max_norm + 1

    # Calculer le nombre de passages par VIF (nb de demi-journées distinctes où il apparaît)
    from collections import Counter as _Counter
    vif_passages = _Counter()
    for f in fiches:
        for v in f['vif_codes']:
            vif_passages[v] += 1

    # Index véhicule → nom depuis les fiches PDF
    veh2nom = {}
    for f in fiches:
        veh = f['vehicule']
        nom = f.get('nom_vehicule', '').strip()
        if nom and veh not in veh2nom:
            veh2nom[veh] = nom
    sans_nom = sorted([v for v in set(f['vehicule'] for f in fiches) if v not in veh2nom])
    if sans_nom:
        print(f"  INFO: {len(sans_nom)} camions sans nom dans le PDF: {sans_nom}")
    else:
        print(f"  → {len(veh2nom)} noms de camions extraits du PDF")

    # Index VIF → ligne magasin depuis le référentiel COMPLET (toutes lignes, sans filtre État)
    # Nécessaire pour ne pas perdre les magasins du PDF 2025 devenus "Non collecté"
    import pandas as _pd
    df_complet_opt = _pd.read_excel(args.magasins)
    df_complet_opt.columns = [c.strip() for c in df_complet_opt.columns]
    df_complet_opt = df_complet_opt.loc[:, ~df_complet_opt.columns.duplicated()]
    # Renommer colonnes
    _rmap = {}
    for _col in df_complet_opt.columns:
        _cl = _col.lower()
        if 'vif' in _cl or _cl == 'code':       _rmap[_col] = 'Code VIF'
        elif ('nom' in _cl or 'magasin' in _cl) and 'fiche' not in _cl: _rmap[_col] = 'Nom'
        elif 'ville' in _cl:                     _rmap[_col] = 'Ville'
        elif 'lat' in _cl:                       _rmap[_col] = 'Latitude'
        elif 'lon' in _cl:                       _rmap[_col] = 'Longitude'
        elif 'tonnage' in _cl and '2025' in _cl: _rmap[_col] = 'Tonnage 2025'
        elif 'secteur' in _cl:                   _rmap[_col] = 'Secteur'
    df_complet_opt = df_complet_opt.rename(columns=_rmap)
    df_complet_opt = df_complet_opt.loc[:, ~df_complet_opt.columns.duplicated()]
    for _col in ['Code VIF','Nom','Ville','Latitude','Longitude','Tonnage 2025','Secteur']:
        if _col not in df_complet_opt.columns:
            df_complet_opt[_col] = ''
    df_complet_opt['Latitude']  = _pd.to_numeric(df_complet_opt['Latitude'],  errors='coerce').fillna(BAI_LAT)
    df_complet_opt['Longitude'] = _pd.to_numeric(df_complet_opt['Longitude'], errors='coerce').fillna(BAI_LON)
    df_complet_opt['Tonnage 2025'] = _pd.to_numeric(df_complet_opt['Tonnage 2025'], errors='coerce').fillna(0)
    df_complet_opt['Nouveau'] = False
    df_complet_opt = df_complet_opt.reset_index(drop=True)

    # Les codes VIF issus du PDF 2025 n'ont pas de zéro initial (ex: "2380059"),
    # contrairement à ceux des fichiers magasins Excel qui en ont parfois un
    # (ex: "02380249") : on normalise systématiquement via lstrip('0') pour que
    # les deux référentiels se rejoignent (même convention que lire_magasins()).
    vif2row = {str(r['Code VIF']).strip().lstrip('0'): r
               for _, r in df_complet_opt.iterrows()}
    # Enrichir vif2row avec les secteurs calculés dans df_mag (absents de df_complet_opt)
    import unicodedata as _ud_sec
    import re as _re_sec
    def _norm_sec(s):
        s = _ud_sec.normalize('NFD', str(s).strip()).encode('ascii','ignore').decode('ascii')
        s = _re_sec.sub(r"[-'’]", ' ', s)
        return _re_sec.sub(r'\s+', ' ', s).strip().title()
    secteurs_mag = {str(r['Code VIF']).strip().lstrip('0'): _norm_sec(r.get('Secteur',''))
                    for _, r in df_mag.iterrows() if r.get('Secteur','')}
    for vif, row in vif2row.items():
        if vif in secteurs_mag:
            # Modifier le secteur dans la Series (créer une copie modifiable)
            row_copy = row.copy()
            row_copy['Secteur'] = secteurs_mag[vif]
            # Stocker les DJ déduites des créneaux
            creneaux_raw = str(row_copy.get('Créneaux', '')).strip()
            row_copy['djs_creneaux'] = parse_creneaux(creneaux_raw) if creneaux_raw and creneaux_raw != 'nan' else None
            vif2row[vif] = row_copy
    new_vifs = set(str(r['Code VIF']).strip().lstrip('0') for _, r in df_mag.iterrows() if r['Nouveau'])
    nouveaux_vehs = set(f'VX{300+i:03d}' for i in range(args.camions_supp))

    # Structure (demi_journee, vehicule) → [vif, ...]
    # On ne garde que les VIFs encore actifs en 2026 (présents dans df_mag = État='Collecté par la BAI')
    vifs_actifs = set(str(r['Code VIF']).strip().lstrip('0') for _, r in df_mag.iterrows())
    dj_veh = defaultdict(list)
    nb_supprimes_pdf = 0
    for f in fiches:
        key = (f['demi_journee'], f['vehicule'])
        for v in f['vif_codes']:
            if v not in vifs_actifs:
                nb_supprimes_pdf += 1
                continue  # magasin supprimé ou inactif en 2026
            if v not in dj_veh[key]:
                dj_veh[key].append(v)
    if nb_supprimes_pdf:
        print(f"  → {nb_supprimes_pdf} magasin(s) du PDF 2025 non reconduit(s) en 2026 (retirés des tournées)")

    # ── INJECTION DES NOUVEAUX MAGASINS PAR SECTEUR STRICT ─────────────────
    # Un nouveau magasin va UNIQUEMENT dans une tournée de son secteur exact.
    # Si aucune place → nouvelle tournée avec camion supplémentaire disponible.
    # Si pas de camion → signalé comme non affecté dans la console.
    DJ_NOUVEAUX = ['Jeudi Matin', 'Jeudi Apres Midi', 'Vendredi Matin', 'Vendredi Apres Midi',
              'Samedi Matin', 'Samedi Apres Midi', 'Dimanche Matin']  # Toutes DJ — filtrage par créneaux

    import unicodedata as _ud
    import re as _re_nv
    def _norm_ville(s):
        s = _ud.normalize('NFD', str(s).strip()).encode('ascii', 'ignore').decode('ascii')
        s = _re_nv.sub(r"[-'’]", ' ', s)
        return _re_nv.sub(r'\s+', ' ', s).strip().title()

    import unicodedata as _ud_loc
    import re as _re_loc
    def _norm_ville_loc(s):
        s = _ud_loc.normalize('NFD', str(s).strip()).encode('ascii','ignore').decode('ascii')
        s = _re_loc.sub(r"[-'’]", ' ', s)
        return _re_loc.sub(r'\s+', ' ', s).strip().title()

    def _calc_secteur_local(lat, lon, ville):
        """Calcule le secteur géographique depuis lat/lon/ville.
        Normalise les accents pour éviter les faux non-matchs (ex: Egreve/Egrève).
        """
        v = _norm_ville_loc(ville)
        if 'grenoble' in v.lower():
            if lat > 45.193: return 'Grenoble Nord'
            if lat < 45.173: return 'Grenoble Sud'
            if lon < 5.714:  return 'Grenoble Ouest'
            if lon > 5.729:  return 'Grenoble Est'
            if lat > 45.187: return 'Grenoble Centre Nord'
            if lon < 5.721:  return 'Grenoble Centre Ouest'
            return 'Grenoble Centre Est'
        # ── Regroupements géographiques ───────────────────────────────────────
        GRESIVAUDAN   = {'Biviers', 'Crolles', 'Froges', 'Saint Ismier', 'Domene', 'Le Versoud'}
        VOIRON_GRP    = {'Voiron', 'Saint Jean De Moirans'}
        RIVES_GRP     = {'Rives', 'Renage'}
        SEYSSINET_GRP = {'Seyssinet Pariset', 'Seyssins', 'Fontaine'}
        PLATEAU_NORD  = {'Apprieu', 'Colombe', 'Le Grand Lemps'}
        SUD_VERCORS   = {'Varces Allieres Et Risset', 'Vif', 'Vizille', 'Claix'}
        CHARTREUSE_E  = {'La Terrasse', 'Le Touvet', 'Tencin'}
        ROUSSILLON    = {'Peage De Roussillon', 'Salaise Sur Sanne'}
        SMH_GRP       = {'Poisat', 'Saint Martin Dheres', 'Saint Martin D Heres', 'Gieres', 'Gières'}
        if v in GRESIVAUDAN:   return 'Gresivaudan'
        if v in VOIRON_GRP:    return 'Voiron'
        if v in RIVES_GRP:     return 'Rives'
        if v in SEYSSINET_GRP: return 'Seyssinet'
        if v in PLATEAU_NORD:  return 'Plateau Nord'
        if v in SUD_VERCORS:   return 'Sud Vercors'
        if v in CHARTREUSE_E:  return 'Chartreuse Est'
        if v in ROUSSILLON:    return 'Roussillon'
        if v in SMH_GRP:       return 'Saint Martin Dheres'
        return v

    def _secteur_vif_calc2(vif):
        r = vif2row.get(vif)
        if r is None: return ''
        try:
            return _calc_secteur_local(float(r.get('Latitude', BAI_LAT)),
                                       float(r.get('Longitude', BAI_LON)),
                                       str(r.get('Ville', '')))
        except Exception:
            return ''

    def _centre_vifs(vifs):
        pts = [(float(vif2row.get(v, {}).get('Latitude', BAI_LAT)),
                float(vif2row.get(v, {}).get('Longitude', BAI_LON)))
               for v in vifs if v in vif2row]
        if not pts: return BAI_LAT, BAI_LON
        return sum(p[0] for p in pts)/len(pts), sum(p[1] for p in pts)/len(pts)

    def _sec_dominant(vifs):
        from collections import Counter as _C
        secs = [_secteur_vif_calc2(v) for v in vifs
                if v not in new_vifs and _secteur_vif_calc2(v)]
        return _C(secs).most_common(1)[0][0] if secs else ''

    camions_existants = set(veh for (dj, veh) in dj_veh.keys())
    camions_supp_dispo = [f'VX{300+i:03d}' for i in range(args.camions_supp)]
    non_affectes = []

    # Nouveaux VIFs = magasins actifs 2026 absents du PDF 2025
    nouveaux_vifs = [str(r['Code VIF']).strip().lstrip('0') for _, r in df_mag.iterrows() if r['Nouveau']]
    print(f"  → {len(nouveaux_vifs)} nouveaux magasins à injecter dans les tournées")

    # Statistiques créneaux
    nb_avec_creneaux = sum(1 for v in nouveaux_vifs if vif2row.get(v, {}).get('djs_creneaux') is not None)
    print(f"  Injection des nouveaux magasins par secteur strict (créneaux : {nb_avec_creneaux}/{len(nouveaux_vifs)} renseignés)...")
    for vif_new in nouveaux_vifs:
        row_new = vif2row.get(vif_new)
        if row_new is None:
            non_affectes.append((vif_new, '?', 'VIF absent du référentiel'))
            continue
        nom_new = str(row_new.get('Nom', vif_new))
        lat_new = float(row_new.get('Latitude', BAI_LAT))
        lon_new = float(row_new.get('Longitude', BAI_LON))
        sec_new = _secteur_vif_calc2(vif_new)
        if sec_new.lower() in ('nan', 'none', ''):
            # Ville manquante → calculer depuis coordonnées uniquement
            try:
                r_tmp = vif2row.get(vif_new)
                if r_tmp is not None:
                    lat_t = float(r_tmp.get('Latitude', BAI_LAT))
                    lon_t = float(r_tmp.get('Longitude', BAI_LON))
                    # Trouver la ville la plus proche parmi les tournées existantes
                    sec_new = ''  # restera vide → fallback géo
            except Exception:
                sec_new = ''


        for dj in DJ_NOUVEAUX:
            # Filtrage par créneaux : ne pas injecter si le magasin n'est pas disponible ce jour
            djs_cr_new = row_new.get('djs_creneaux')
            if djs_cr_new is not None and dj not in djs_cr_new:
                continue  # Créneau incompatible → magasin non disponible cette DJ
            max_c_dj = max_dim if dj in DJ_DIMANCHE else max_norm

            # 1. Chercher la tournée du même secteur la plus proche avec de la place
            best_key = None
            best_dist = float('inf')
            for (d, veh), vifs in dj_veh.items():
                if d != dj: continue
                if (veh, dj) in TOURNEES_GELEES: continue
                if vif_new in vifs: continue
                sec_dom = _sec_dominant(vifs)
                if not sec_dom: continue  # secteur indéterminé → ignorer
                if sec_dom != sec_new: continue  # STRICT : même secteur uniquement
                # Tolérance +1 si tournée mono-secteur et non figée
                secs_existants = {_secteur_vif_calc2(v) for v in vifs if _secteur_vif_calc2(v)}
                mono = (len(secs_existants) == 1 and sec_new in secs_existants
                        and veh not in VEHICULES_FIGES and (veh, dj) not in TOURNEES_GELEES)
                limite = max_c_dj + 2 if mono else max_c_dj
                if len(vifs) >= limite: continue
                # Vérifier ≤ 3 secteurs après ajout
                secs_apres = secs_existants | {sec_new}
                if len(secs_apres) > 3: continue
                clat, clon = _centre_vifs(vifs)
                d_km = haversine(lat_new, lon_new, clat, clon)

                if d_km < best_dist:
                    best_dist = d_km
                    best_key = (dj, veh)

            if best_key:

                if vif_new not in dj_veh[best_key]: dj_veh[best_key].append(vif_new)
                continue

            # 2. Aucune tournée du même secteur exact
            # Chercher une tournée à moins de 10 km (secteur voisin acceptable)
            best_proche = None
            best_dist_proche = float('inf')
            for (d, veh), vifs in dj_veh.items():
                if d != dj: continue
                if (veh, dj) in TOURNEES_GELEES: continue
                if vif_new in vifs: continue
                if len(vifs) >= max_c_dj: continue  # strict
                clat, clon = _centre_vifs(vifs)
                d_km = haversine(lat_new, lon_new, clat, clon)
                if d_km < 7.0 and d_km < best_dist_proche:
                    best_dist_proche = d_km
                    best_proche = (dj, veh)

            if best_proche:
                # Vérifier cohérence sectorielle : max 2 secteurs distincts après ajout
                vifs_proche = dj_veh[best_proche]
                secs_proche = set(s for v in vifs_proche if (s := _secteur_vif_calc2(v)))
                secs_proche.add(sec_new)
                if len(secs_proche) <= 3:
                    # Tournée voisine trouvée → affecter sans signaler
                    if vif_new not in dj_veh[best_proche]: dj_veh[best_proche].append(vif_new)
                    continue

            # 3. Aucune tournée voisine → camion supplémentaire libre
            veh_new = None
            for vx in camions_supp_dispo:
                if (dj, vx) not in dj_veh and (vx, dj) not in TOURNEES_GELEES:
                    veh_new = vx
                    break
            if veh_new is None:
                for veh_c in sorted(camions_existants):
                    if (dj, veh_c) not in dj_veh and (veh_c, dj) not in TOURNEES_GELEES:
                        veh_new = veh_c
                        break
            if veh_new:
                dj_veh[(dj, veh_new)] = [vif_new]
            else:
                # 3bis. Retry avec MAX+1 si même secteur (hors dimanche)
                if dj not in DJ_DIMANCHE:
                    best_p1 = None; best_dist_p1 = float('inf')
                    for (d2, veh2), vifs2 in dj_veh.items():
                        if d2 != dj: continue
                        if veh2 in VEHICULES_FIGES: continue
                        if (veh2, dj) in TOURNEES_GELEES: continue
                        if len(vifs2) >= max_c_dj + 2: continue  # accepter jusqu'à max+2
                        sec2 = _sec_dominant(vifs2)
                        if not sec2 or sec2 != sec_new: continue  # même secteur dominant
                        secs2 = {_secteur_vif_calc2(v) for v in vifs2 if _secteur_vif_calc2(v)} | {sec_new}
                        if len(secs2) > 3: continue
                        clat2, clon2 = _centre_vifs(list(vifs2))
                        d_km2 = haversine(lat_new, lon_new, clat2, clon2)
                        if d_km2 < best_dist_p1:
                            best_dist_p1 = d_km2; best_p1 = (dj, veh2)
                    if best_p1:
                        if vif_new not in dj_veh[best_p1]: dj_veh[best_p1].append(vif_new)
                        print(f"  Injection MAX+1 ({max_c_dj+1} mag): '{nom_new}' → {best_p1[1]} ({dj})")
                        continue  # magasin placé → passer au suivant

                # 4. Fallback : tournée la plus proche tous secteurs, signaler
                best_key3 = None
                best_dist3 = float('inf')
                for (d, veh), vifs in dj_veh.items():
                    if d != dj: continue
                    if (veh, dj) in TOURNEES_GELEES: continue
                    if vif_new in vifs: continue
                    if len(vifs) >= max_c_dj: continue  # strict
                    clat, clon = _centre_vifs(vifs)
                    d_km = haversine(lat_new, lon_new, clat, clon)
                    if d_km < best_dist3:
                        best_dist3 = d_km
                        best_key3 = (dj, veh)
                if best_key3:
                    if vif_new not in dj_veh[best_key3]: dj_veh[best_key3].append(vif_new)
                    non_affectes.append((vif_new, nom_new,
                        f'{dj}/{best_key3[1]}: secteur {sec_new} → placé hors secteur ({best_dist3:.1f}km)'))
                else:
                    non_affectes.append((vif_new, nom_new,
                        f'{dj}: aucun camion disponible'))

    if non_affectes:
        print(f"  ATTENTION: {len(non_affectes)} affectation(s) impossible(s):")
        for _, nom, msg in non_affectes:
            print(f"    - {nom}: {msg}")
    else:
        print("  Tous les nouveaux magasins affectés")

    def _dist(vif):
        r = vif2row.get(vif)
        if r is None: return 0
        return dist_bai(float(r['Latitude']), float(r['Longitude']))

    def _coords(vif):
        r = vif2row.get(vif)
        if r is None: return BAI_LAT, BAI_LON
        return float(r['Latitude']), float(r['Longitude'])

    # ── CAMIONS SUPPLÉMENTAIRES ──────────────────────────────────────────────
    # Stratégie : chaque camion VX prend max_c magasins depuis les tournées les plus
    # chargées de la demi-journée la plus surchargée (hors dimanche).
    # On vise à remplir le VX à max_c magasins en prenant des anciens magasins.
    if args.camions_supp > 0:
        print(f"  Affectation de {args.camions_supp} camions supplémentaires...")

        def _delester_dj_secteur(veh_new, dj, max_c_vx, sec_cible):
            """Déleste vers veh_new les magasins du secteur sec_cible
            qui sont isolés dans des tournées d'un autre secteur dominant."""
            candidats = []
            for (d, veh_src), vifs in dj_veh.items():
                if d != dj: continue
                if veh_src == veh_new: continue
                if (veh_src, dj) in TOURNEES_GELEES: continue
                if veh_src in VEHICULES_FIGES: continue
                sec_dom = _sec_dominant([v for v in vifs if v not in new_vifs]) or _sec_dominant(vifs)
                if sec_dom == sec_cible: continue  # tournée déjà du bon secteur
                # Chercher les VIFs du secteur cible dans cette tournée
                vifs_cible = [v for v in vifs if _secteur_vif_calc2(v) == sec_cible]
                if not vifs_cible: continue
                if len(vifs) - len(vifs_cible) < 1: continue  # garder au moins 1
                candidats.extend([(v, veh_src) for v in vifs_cible])

            if not candidats or len(candidats) < 2:
                return 0  # pas assez pour créer une tournée VX

            # Prendre jusqu'à max_c_vx magasins
            pris = candidats[:max_c_vx]
            sources = {}
            for vif, veh_src in pris:
                sources.setdefault(veh_src, []).append(vif)

            for veh_src, vifs_pris in sources.items():
                dj_veh[(dj, veh_src)] = [v for v in dj_veh[(dj, veh_src)] if v not in vifs_pris]
            dj_veh[(dj, veh_new)] = [v for v, _ in pris]
            print(f"    {veh_new} → {dj}: {len(pris)} mag secteur '{sec_cible}' extraits de {list(sources.keys())}")
            return len(pris)

        def _delester_dj(veh_new, dj, max_c_vx):
            """Déleste la demi-journée dj vers veh_new.

            Étape 1 : essaie chaque camion source chargé, du plus chargé au moins
            chargé (au lieu de s'arrêter au premier candidat comme auparavant), et
            prend le premier qui peut donner au moins 2 magasins d'un secteur clair
            tout en en gardant au moins 2 — pour garantir la cohérence géographique
            du VX sans dépendre d'un unique camion source qui n'aurait pas assez
            d'excédent ce jour-là.

            Étape 2 (repli) : si aucun camion seul n'a assez d'excédent, combine
            les magasins de plusieurs camions partageant le même secteur dominant
            majoritaire sur cette demi-journée (même logique que
            _delester_dj_secteur, généralisée à un secteur déterminé automatiquement
            plutôt qu'imposé). Ceci évite qu'un camion supplémentaire reste inutilisé
            sur une demi-journée simplement parce qu'aucune tournée isolée n'était
            assez chargée pour être délestée à elle seule.
            """
            groupes_dj = sorted(
                [(veh, list(vifs)) for (d, veh), vifs in dj_veh.items()
                 if d == dj and veh != veh_new and (veh, dj) not in TOURNEES_GELEES
                 and veh not in VEHICULES_FIGES],
                key=lambda x: len(x[1]), reverse=True
            )
            if not groupes_dj:
                return 0

            # Étape 1 : un seul camion source, en essayant tous les candidats
            for veh_src, vifs_src in groupes_dj:
                anciens = [v for v in vifs_src if v not in new_vifs]
                if len(anciens) < 2:
                    continue
                sec = _sec_dominant(anciens)
                if not sec:
                    continue
                anciens_tries = sorted(anciens, key=_dist, reverse=True)
                nb_garder = max(2, len(anciens_tries) // 2)
                nb_donner = min(len(anciens_tries) - nb_garder, max_c_vx)
                if nb_donner < 2:
                    continue  # cette tournée n'a pas assez d'excédent, essayer la suivante
                pris = anciens_tries[:nb_donner]
                dj_veh[(dj, veh_src)] = [v for v in vifs_src if v not in pris]
                dj_veh[(dj, veh_new)] = list(pris)
                print(f"    {veh_new} → {dj}: {len(pris)} mag de {veh_src} (secteur {sec})")
                return len(pris)

            # Étape 2 : repli — combiner plusieurs camions du même secteur dominant
            from collections import Counter as _C_deles
            secteurs_count = _C_deles()
            for veh_src, vifs_src in groupes_dj:
                for v in vifs_src:
                    if v in new_vifs:
                        continue
                    s = _secteur_vif_calc2(v)
                    if s:
                        secteurs_count[s] += 1
            if not secteurs_count:
                return 0
            sec_cible = secteurs_count.most_common(1)[0][0]

            candidats = []
            for veh_src, vifs_src in groupes_dj:
                anciens_cible = [v for v in vifs_src
                                  if v not in new_vifs and _secteur_vif_calc2(v) == sec_cible]
                if not anciens_cible:
                    continue
                # Plafonner la contribution de cette tournée pour garder au moins 1
                # magasin en son sein, plutôt que d'écarter toute la tournée si elle
                # est entièrement du secteur cible (cas d'une tournée homogène).
                cap = len(vifs_src) - 1
                if cap <= 0:
                    continue
                candidats.extend([(v, veh_src) for v in anciens_cible[:cap]])

            if len(candidats) < 2:
                return 0

            pris = candidats[:max_c_vx]
            sources = {}
            for vif, veh_src in pris:
                sources.setdefault(veh_src, []).append(vif)
            for veh_src, vifs_pris in sources.items():
                dj_veh[(dj, veh_src)] = [v for v in dj_veh[(dj, veh_src)] if v not in vifs_pris]
            dj_veh[(dj, veh_new)] = [v for v, _ in pris]
            print(f"    {veh_new} → {dj}: {len(pris)} mag secteur '{sec_cible}' combinés depuis {list(sources.keys())}")
            return len(pris)

        # Chaque VX couvre les 4 demi-journées Vendredi+Samedi
        # Les VX sont affectés en parallèle : VX300 et VX301 sur les mêmes DJ
        DJ_VX = ['Vendredi Matin', 'Vendredi Apres Midi', 'Samedi Matin', 'Samedi Apres Midi']

        # Secteurs prioritaires par numéro de VX :
        # VX300 → délestage général (secteur dominant)
        # VX301 → priorité aux magasins Uriage isolés dans des tournées mixtes
        SECTEURS_PRIORITAIRES_VX = {'VX301': 'Saint Martin D Uriage'}

        for i in range(args.camions_supp):
            veh_new = f'VX{300+i:03d}'
            sec_prio = SECTEURS_PRIORITAIRES_VX.get(veh_new)
            total = 0
            for dj in DJ_VX:
                if sec_prio:
                    nb = _delester_dj_secteur(veh_new, dj, max_norm, sec_prio)
                    if nb == 0:
                        nb = _delester_dj(veh_new, dj, max_norm)
                else:
                    nb = _delester_dj(veh_new, dj, max_norm)
                total += nb
            print(f"    {veh_new}: {total} magasins au total sur {len(DJ_VX)} demi-journées")

    # ── OPTIMISATION 1 : rééquilibrage des charges par demi-journée ──────────
    print("  Rééquilibrage des charges...")
    nb_reequil = 0
    for dj in DEMI_JOURNEES:
        max_c = max_dim if dj in DJ_DIMANCHE else max_norm
        groupes_dj = {veh: list(vifs) for (d, veh), vifs in dj_veh.items()
                      if d == dj and (veh, dj) not in TOURNEES_GELEES
                      and veh not in VEHICULES_FIGES}
        if len(groupes_dj) < 2:
            continue
        nb_total = sum(len(v) for v in groupes_dj.values())
        moy = nb_total / len(groupes_dj)

        for _ in range(20):
            modifie = False
            surcharges  = sorted([v for v in groupes_dj if len(groupes_dj[v]) > max_c],
                                   key=lambda v: len(groupes_dj[v]), reverse=True)
            sous_charges = sorted([v for v in groupes_dj if len(groupes_dj[v]) < max_c],
                                    key=lambda v: len(groupes_dj[v]))
            for veh_src in surcharges:
                for veh_dst in sous_charges:
                    if len(groupes_dj[veh_src]) <= len(groupes_dj[veh_dst]):
                        continue
                    if len(groupes_dj[veh_dst]) >= max_c or not groupes_dj[veh_dst]:
                        continue
                    lat_dst = sum(_coords(v)[0] for v in groupes_dj[veh_dst]) / len(groupes_dj[veh_dst])
                    lon_dst = sum(_coords(v)[1] for v in groupes_dj[veh_dst]) / len(groupes_dj[veh_dst])
                    # Rééquilibrage : déplacer uniquement les nouveaux magasins
                    # ET uniquement si leur secteur correspond à la tournée destination
                    sec_dst = _sec_dominant(groupes_dj[veh_dst])
                    if args.optimiser_anciens:
                        # Déplacer anciens ET nouveaux si secteur correspond
                        candidats = sorted(
                            [v for v in groupes_dj[veh_src]
                             if _secteur_vif_calc2(v) == sec_dst],
                            key=lambda v: haversine(*_coords(v), lat_dst, lon_dst))
                    else:
                        candidats = sorted(
                            [v for v in groupes_dj[veh_src] if v in new_vifs
                             and _secteur_vif_calc2(v) == sec_dst],
                            key=lambda v: haversine(*_coords(v), lat_dst, lon_dst))
                    if not candidats:
                        continue
                    vif_mv = candidats[0]
                    # Vérifier que la destination ne dépassera pas 2 secteurs distincts
                    secs_dst_apres = {_secteur_vif_calc2(v) for v in groupes_dj[veh_dst] if _secteur_vif_calc2(v)}
                    secs_dst_apres.add(_secteur_vif_calc2(vif_mv))
                    if len(secs_dst_apres) > 3:
                        continue
                    groupes_dj[veh_src].remove(vif_mv)
                    groupes_dj[veh_dst].append(vif_mv)
                    nb_reequil += 1
                    modifie = True
            if not modifie:
                break

        for veh, vifs in groupes_dj.items():
            dj_veh[(dj, veh)] = vifs
    print(f"    -> {nb_reequil} transferts de magasins effectués")

    # ── OPTIMISATION 2 : échanges géographiques entre tournées ───────────────
    print("  Regroupement géographique...")
    nb_swaps = 0

    def _km_groupe(vifs_list):
        if not vifs_list: return 0
        pts = sorted([_coords(v) for v in vifs_list],
                     key=lambda p: haversine(BAI_LAT, BAI_LON, p[0], p[1]))
        km = haversine(BAI_LAT, BAI_LON, pts[0][0], pts[0][1])
        for j in range(len(pts)-1):
            km += haversine(pts[j][0], pts[j][1], pts[j+1][0], pts[j+1][1])
        km += haversine(pts[-1][0], pts[-1][1], BAI_LAT, BAI_LON)
        return km

    def _secteur_vif(vif):
        r = vif2row.get(vif)
        if r is None: return ''
        sec = str(r.get('Secteur', '')).strip()
        if sec: return sec
        try:
            lat = float(r.get('Latitude', BAI_LAT))
            lon = float(r.get('Longitude', BAI_LON))
            ville = str(r.get('Ville', ''))
            return _calc_secteur(lat, lon, ville)
        except Exception:
            return ''

    def _score_groupe(vifs_list):
        """Score = km + pénalité pour mélange de secteurs géographiques.
        Pénalise les groupes qui mélangent des secteurs différents.
        """
        km = _km_groupe(vifs_list)
        if len(vifs_list) < 2: return km
        secteurs = [_secteur_vif(v) for v in vifs_list if _secteur_vif(v)]
        if not secteurs: return km
        # Compter les secteurs distincts - pénalité de 3 km par secteur supplémentaire
        nb_secteurs = len(set(secteurs))
        penalite = (nb_secteurs - 1) * 3.0
        return km + penalite

    for dj in DEMI_JOURNEES:
        groupes_dj = {veh: list(vifs) for (d, veh), vifs in dj_veh.items()
                      if d == dj and (veh, dj) not in TOURNEES_GELEES
                      and veh not in VEHICULES_FIGES}
        vehs = list(groupes_dj.keys())
        ameliore = True
        nb_passes = 0
        while ameliore and nb_passes < 20:
            ameliore = False
            nb_passes += 1
            for i in range(len(vehs)):
                for j in range(i+1, len(vehs)):
                    va, vb = vehs[i], vehs[j]
                    la, lb = groupes_dj[va], groupes_dj[vb]
                    if not la or not lb:
                        continue
                    km_av = _score_groupe(la) + _score_groupe(lb)
                    for vif_a in list(la):
                        if not args.optimiser_anciens and vif_a not in new_vifs: continue
                        for vif_b in list(lb):
                            if not args.optimiser_anciens and vif_b not in new_vifs: continue
                            new_a = [v for v in la if v != vif_a] + [vif_b]
                            new_b = [v for v in lb if v != vif_b] + [vif_a]
                            # Vérifier que les tailles ne dépassent pas max_c
                            max_c_echange = max_dim if dj in DJ_DIMANCHE else max_norm
                            if len(new_a) > max_c_echange or len(new_b) > max_c_echange:
                                continue
                            # Vérifier cohérence sectorielle : max 2 secteurs distincts
                            secs_a = set(s for v in new_a if (s := _secteur_vif_calc2(v)))
                            secs_b = set(s for v in new_b if (s := _secteur_vif_calc2(v)))
                            if len(secs_a) > 3 or len(secs_b) > 3:
                                continue
                            if _score_groupe(new_a) + _score_groupe(new_b) < km_av - 0.1:
                                groupes_dj[va] = new_a
                                groupes_dj[vb] = new_b
                                la, lb = new_a, new_b
                                km_av = _score_groupe(la) + _score_groupe(lb)
                                nb_swaps += 1
                                ameliore = True
        for veh, vifs in groupes_dj.items():
            dj_veh[(dj, veh)] = vifs
    print(f"    -> {nb_swaps} échanges géographiques effectués")

    # ── Remplissage des tournées légères (2-3 mag) ──────────────────────────
    print("  Remplissage des tournées légères...")
    nb_remplissages = 0
    for dj in [d for d in DEMI_JOURNEES if d not in DJ_DIMANCHE and 'Jeudi' not in d]:
        max_c_dj = max_norm
        deja_traites = set()  # éviter les boucles A→B→A→B
        modifie = True
        while modifie:
            modifie = False
            # Identifier tournées légères (2-3 mag, non figées)
            legeres = sorted(
                [(veh, list(vifs)) for (d, veh), vifs in dj_veh.items()
                 if d == dj and len(vifs) in (2, 3)
                 and veh not in VEHICULES_FIGES
                 and (veh, dj) not in TOURNEES_GELEES],
                key=lambda x: len(x[1]))  # les plus légères en premier
            if not legeres: break
            for veh_dst, vifs_dst in legeres:
                if len(dj_veh.get((dj, veh_dst), [])) >= max_c_dj: continue
                sec_dst = _sec_dominant(dj_veh.get((dj, veh_dst), []))
                if not sec_dst: continue
                lat_dst, lon_dst = _centre_vifs(dj_veh.get((dj, veh_dst), []))
                best_vif = None; best_src = None; best_dist = float('inf')
                for (d2, veh_src), vifs_src in dj_veh.items():
                    if d2 != dj or veh_src == veh_dst: continue
                    if veh_src in VEHICULES_FIGES: continue
                    if (veh_src, dj) in TOURNEES_GELEES: continue
                    if len(vifs_src) <= 3: continue  # prendre uniquement dans tournées chargées (>3)
                    if len(vifs_src) - 1 < 2: continue  # garder au moins 2 dans la source
                    for vif in vifs_src:
                        if (vif, veh_dst) in deja_traites: continue  # déjà tenté
                        sec_v = _secteur_vif_calc2(vif)
                        if sec_v != sec_dst: continue
                        secs_src_apres = {_secteur_vif_calc2(v) for v in vifs_src if v != vif and _secteur_vif_calc2(v)}
                        if len(secs_src_apres) > 3: continue
                        secs_dst_apres = {_secteur_vif_calc2(v) for v in dj_veh.get((dj,veh_dst),[]) if _secteur_vif_calc2(v)} | {sec_v}
                        if len(secs_dst_apres) > 3: continue
                        r_v = vif2row.get(vif)
                        if r_v is None: continue
                        lat_v = float(r_v.get('Latitude', BAI_LAT))
                        lon_v = float(r_v.get('Longitude', BAI_LON))
                        d_km = haversine(lat_v, lon_v, lat_dst, lon_dst)
                        if d_km < best_dist:
                            best_dist = d_km; best_vif = vif; best_src = veh_src
                if best_vif and best_src:
                    r_v = vif2row.get(best_vif, {})
                    nom_v = str(r_v.get('Nom', best_vif))
                    dej_traites = deja_traites  # référence locale
                    deja_traites.add((best_vif, veh_dst))
                    deja_traites.add((best_vif, best_src))  # bloquer retour
                    dj_veh[(dj, best_src)] = [v for v in dj_veh[(dj, best_src)] if v != best_vif]
                    if best_vif not in dj_veh[(dj, veh_dst)]: dj_veh[(dj, veh_dst)].append(best_vif)
                    print(f"  Remplissage: '{nom_v}' {best_src}→{veh_dst} ({dj}, {len(dj_veh[(dj,veh_dst)])} mag)")
                    nb_remplissages += 1
                    modifie = True
                    break
    if nb_remplissages:
        print(f"  → {nb_remplissages} magasin(s) transféré(s) vers tournées légères")
    else:
        print("  → Aucun transfert possible vers tournées légères")

    # ── Réduction des secteurs : extraire les magasins isolés des tournées à 3+ secteurs ──
    print("  Réduction des secteurs...")
    nb_reductions = 0
    for dj in [d for d in DEMI_JOURNEES if d not in DJ_DIMANCHE]:
        max_c_dj = max_norm
        modifie = True
        while modifie:
            modifie = False
            for (d, veh_src), vifs in list(dj_veh.items()):
                if d != dj: continue
                if veh_src in VEHICULES_FIGES: continue
                if (veh_src, dj) in TOURNEES_GELEES: continue
                if len(vifs) < 2: continue
                # Calculer les secteurs de la tournée
                secs = [_secteur_vif_calc2(v) for v in vifs if _secteur_vif_calc2(v)]
                from collections import Counter as _C
                cnt = _C(secs)
                if len(cnt) < 3: continue  # déjà ≤ 2 secteurs → OK
                # Identifier les secteurs minoritaires (1 seul magasin)
                secs_min = {s for s,n in cnt.items() if n == 1}
                if not secs_min: continue
                # Tenter de déplacer chaque magasin isolé
                for vif in list(vifs):
                    sec_v = _secteur_vif_calc2(vif)
                    if sec_v not in secs_min: continue
                    r_v = vif2row.get(vif)
                    if r_v is None: continue
                    lat_v = float(r_v.get('Latitude', BAI_LAT))
                    lon_v = float(r_v.get('Longitude', BAI_LON))
                    # Chercher une tournée du même secteur avec de la place
                    best_key = None; best_score = float('inf')
                    for (d2, veh2), vifs2 in dj_veh.items():
                        if d2 != dj or veh2 == veh_src: continue
                        if (veh2, dj) in TOURNEES_GELEES: continue
                        if veh2 in VEHICULES_FIGES: continue
                        if len(vifs2) >= max_c_dj: continue
                        sec2 = _sec_dominant(vifs2)
                        if not sec2 or sec2 != sec_v: continue
                        # Vérifier que le déplacement ne crée pas de 3ème secteur dans la source
                        secs_src_apres = set(_secteur_vif_calc2(v) for v in vifs if v != vif and _secteur_vif_calc2(v))
                        if len(secs_src_apres) > 2: continue
                        clat2, clon2 = _centre_vifs(vifs2)
                        d_km = haversine(lat_v, lon_v, clat2, clon2)
                        if d_km < best_score:
                            best_score = d_km; best_key = (d2, veh2)
                    if best_key:
                        nom_v = str(r_v.get('Nom', vif))
                        dj_veh[(dj, veh_src)] = [v for v in dj_veh[(dj, veh_src)] if v != vif]
                        if vif not in dj_veh[best_key]: dj_veh[best_key].append(vif)
                        print(f"  Réduction secteur: '{nom_v}' {veh_src}→{best_key[1]} ({dj})")
                        nb_reductions += 1
                        modifie = True
                        break  # recommencer la boucle
    if nb_reductions:
        print(f"  → {nb_reductions} magasin(s) déplacé(s) pour réduire les secteurs")

    # ── Corriger les magasins mal placés si option activée ──────────────────
    if args.corriger_mal_places:
        nb_corrections_mp = 0
        for dj in [d for d in DEMI_JOURNEES if d not in DJ_DIMANCHE]:
            max_c_dj = max_norm
            for (d, veh_src), vifs in list(dj_veh.items()):
                if d != dj: continue
                if veh_src in VEHICULES_FIGES: continue
                if (veh_src, dj) in TOURNEES_GELEES: continue
                if len(vifs) < 2: continue
                # Calculer le centre de la tournée
                coords = [(float(vif2row.get(v, {}).get('Latitude', BAI_LAT)),
                           float(vif2row.get(v, {}).get('Longitude', BAI_LON)))
                          for v in vifs if v in vif2row]
                if not coords: continue
                clat = sum(c[0] for c in coords) / len(coords)
                clon = sum(c[1] for c in coords) / len(coords)
                for vif in list(vifs):
                    if vif not in vif2row: continue
                    if vif not in new_vifs and not args.optimiser_anciens: continue
                    r_v = vif2row.get(vif)
                    lat_v = float(r_v.get('Latitude', BAI_LAT))
                    lon_v = float(r_v.get('Longitude', BAI_LON))
                    dist = haversine(lat_v, lon_v, clat, clon)
                    if dist <= 8: continue  # < 8km → pas mal placé
                    sec_v = _secteur_vif_calc2(vif)
                    # Chercher meilleure tournée (même secteur, plus proche)
                    best_key = None; best_score = float('inf')
                    for (d2, veh2), vifs2 in dj_veh.items():
                        if d2 != dj or veh2 == veh_src: continue
                        if (veh2, dj) in TOURNEES_GELEES: continue
                        if veh2 in VEHICULES_FIGES: continue
                        if vif in vifs2: continue
                        if len(vifs2) >= max_c_dj: continue
                        sec2 = _sec_dominant([v for v in vifs2 if v not in new_vifs])
                        if sec2 and sec_v and sec2 != sec_v: continue
                        clat2, clon2 = _centre_vifs(vifs2)
                        d_km = haversine(lat_v, lon_v, clat2, clon2)
                        if d_km < dist and d_km < best_score:
                            best_score = d_km; best_key = (d2, veh2)
                    if best_key and len(dj_veh[(dj, veh_src)]) > 1:
                        dj_veh[(dj, veh_src)] = [v for v in dj_veh[(dj, veh_src)] if v != vif]
                        if vif not in dj_veh[best_key]: dj_veh[best_key].append(vif)
                        nom_v = str(r_v.get('Nom', vif))
                        print(f"  Correction mal placé: '{nom_v}' {veh_src}→{best_key[1]} ({dj})")
                        nb_corrections_mp += 1
        if nb_corrections_mp:
            print(f"  → {nb_corrections_mp} magasin(s) mal placé(s) corrigé(s)")

    # ── Fusionner les tournées légères si option activée ────────────────────
    if args.fusionner_legeres:
        nb_fusions = 0
        for dj in [d for d in DEMI_JOURNEES if d not in DJ_DIMANCHE]:
            max_c_dj = max_norm
            # Trouver les tournées légères (≤ 2 mag, non figées, non gelées)
            legeres = [(veh, list(vifs)) for (d, veh), vifs in dj_veh.items()
                       if d == dj and len(vifs) <= 2
                       and veh not in VEHICULES_FIGES
                       and (veh, dj) not in TOURNEES_GELEES]
            # Trier par taille croissante
            legeres.sort(key=lambda x: len(x[1]))
            fusionne = set()
            for i in range(len(legeres)):
                veh1, vifs1 = legeres[i]
                if veh1 in fusionne: continue
                for j in range(i+1, len(legeres)):
                    veh2, vifs2 = legeres[j]
                    if veh2 in fusionne: continue
                    total = len(vifs1) + len(vifs2)
                    if total > max_c_dj: continue
                    # Vérifier cohérence sectorielle (max 2 secteurs)
                    secs = set()
                    for v in vifs1 + vifs2:
                        s = _secteur_vif_calc2(v)
                        if s: secs.add(s)
                    if len(secs) > 3: continue
                    # Fusionner veh2 dans veh1
                    dj_veh[(dj, veh1)] = vifs1 + vifs2
                    del dj_veh[(dj, veh2)]
                    fusionne.add(veh2)
                    vifs1 = dj_veh[(dj, veh1)]
                    print(f"  Fusion: {veh1}+{veh2} → {veh1} ({dj}, {total} mag)")
                    nb_fusions += 1
                    break
        if nb_fusions:
            print(f"  → {nb_fusions} fusion(s) effectuée(s)")


    # ── Passe finale : injection des manquants sur certaines DJ ──────────────
    # Détecte les magasins présents en 2025 sur une DJ mais absents en 2026,
    # et les nouveaux présents sur d'autres DJ mais absents sur la DJ courante.
    print("  Injection des magasins manquants par DJ...")
    nb_manquants = 0

    # Construire l'index 2025 : vif → set de DJ où il était présent
    vif_djs_2025 = defaultdict(set)
    for fiche in fiches:
        dj_f = fiche.get('demi_journee','')
        if dj_f in [x for x in DEMI_JOURNEES if x not in DJ_DIMANCHE and 'Jeudi' not in x]:
            for vif in fiche.get('vif_codes',[]):
                if vif in vifs_actifs:  # seulement les actifs en 2026
                    vif_djs_2025[vif].add(dj_f)

    for dj in [d for d in DEMI_JOURNEES if d not in DJ_DIMANCHE and 'Jeudi' not in d]:
        max_c_dj = max_norm
        vifs_dj = set(v for (d,_), vifs in dj_veh.items() if d==dj for v in vifs)
        # Manquants = anciens présents en 2025 sur cette DJ mais absents en 2026
        #           + nouveaux présents sur d'autres DJ mais absents sur celle-ci
        manquants_anciens = {v for v, djs in vif_djs_2025.items()
                             if dj in djs and v not in new_vifs and v not in vifs_dj}
        # Nouveaux absents de cette DJ mais présents sur d'autres
        manquants_nouveaux = {v for (d,_), vifs in dj_veh.items()
                              if d != dj and d not in DJ_DIMANCHE and 'Jeudi' not in d
                              for v in vifs if v in new_vifs and v not in vifs_dj}
        # Nouveaux absents de TOUTES les DJ V/S sur cette DJ
        manquants_nouveaux |= {v for v in new_vifs
                                if v not in vifs_dj
                                and v not in {x for (d,_), vl in dj_veh.items()
                                              if d not in DJ_DIMANCHE and 'Jeudi' not in d
                                              for x in vl}}
        vifs_autres_dj = manquants_anciens | manquants_nouveaux
        manquants = vifs_autres_dj - vifs_dj
        if not manquants: continue
        for vif in sorted(manquants):
            r_v = vif2row.get(vif)
            if r_v is None: continue
            # Filtrage créneaux : ne pas injecter si magasin non disponible cette DJ
            djs_cr_v = r_v.get('djs_creneaux')
            if djs_cr_v is not None and dj not in djs_cr_v:
                continue
            sec_v = _secteur_vif_calc2(vif)
            lat_v = float(r_v.get('Latitude', BAI_LAT))
            lon_v = float(r_v.get('Longitude', BAI_LON))
            nom_v = str(r_v.get('Nom', vif))
            # Chercher la meilleure tournée disponible
            est_absent_partout = vif not in {v for (d,_), vl in dj_veh.items()
                                             if d not in DJ_DIMANCHE and 'Jeudi' not in d
                                             for v in vl}
            best_key = None; best_dist = float('inf')
            for (d, veh), vifs in dj_veh.items():
                if d != dj: continue
                if veh in VEHICULES_FIGES: continue
                if (veh, dj) in TOURNEES_GELEES: continue
                nb = len(vifs)
                secs = {_secteur_vif_calc2(v) for v in vifs if _secteur_vif_calc2(v)}
                mono = len(secs)==1 and sec_v in secs
                # Tolérance +1 si mono-secteur OU si nouveau absent de toutes les DJ
                limite = max_c_dj + 2 if mono else (max_c_dj + 2 if (vif in new_vifs and est_absent_partout) else max_c_dj)
                if nb >= limite: continue
                # Vérifier contrainte ≤ 3 secteurs
                secs_apres = secs | {sec_v}
                if len(secs_apres) > 3: continue
                clat, clon = _centre_vifs(list(vifs))
                d_km = haversine(lat_v, lon_v, clat, clon)
                if d_km < best_dist:
                    best_dist = d_km; best_key = (d, veh)
            if best_key:
                if vif not in dj_veh[best_key]:  # éviter les doublons
                    if vif not in dj_veh[best_key]: dj_veh[best_key].append(vif)
                    print(f"  Manquant injecté: '{nom_v}' → {best_key[1]} ({dj}, dist={best_dist:.1f}km)")
                    nb_manquants += 1
    if nb_manquants:
        print(f"  → {nb_manquants} magasin(s) manquant(s) injecté(s)")
    else:
        print("  → Aucun manquant détecté")

    # ── Déduplification : supprimer les doublons dans chaque tournée ─────────
    nb_doublons = 0
    for key in list(dj_veh.keys()):
        vifs = dj_veh[key]
        vifs_uniq = list(dict.fromkeys(vifs))  # préserve l'ordre, supprime doublons
        if len(vifs_uniq) < len(vifs):
            nb_doublons += len(vifs) - len(vifs_uniq)
            dj_veh[key] = vifs_uniq
    if nb_doublons:
        print(f"  [WARN] {nb_doublons} doublon(s) intra-tournée supprimé(s)")

    # ── Vérification : un magasin ne doit pas être dans plusieurs tournées d'une DJ ─
    nb_inter = 0
    for dj in DEMI_JOURNEES:
        # Construire l'index vif → liste de camions pour cette DJ
        vif2cams = {}
        for (d, veh), vifs in dj_veh.items():
            if d != dj: continue
            for vif in vifs:
                vif2cams.setdefault(vif, []).append(veh)
        # Détecter les doublons inter-tournées
        for vif, cams in vif2cams.items():
            if len(cams) > 1:
                r = vif2row.get(vif, {})
                nom = str(r.get('Nom', vif)) if r else vif
                print(f"  [WARN] '{nom}' présent dans {len(cams)} tournées le {dj} : {cams}")
                # Garder uniquement le premier camion (celui avec le plus de magasins)
                cam_garder = max(cams, key=lambda v: len(dj_veh.get((dj, v), [])))
                for cam_retirer in cams:
                    if cam_retirer != cam_garder:
                        dj_veh[(dj, cam_retirer)] = [v for v in dj_veh[(dj, cam_retirer)] if v != vif]
                        print(f"    → Retiré de {cam_retirer}, conservé dans {cam_garder}")
                        nb_inter += 1
    if nb_inter:
        print(f"  [WARN] {nb_inter} magasin(s) inter-tournées corrigé(s)")
    elif nb_doublons == 0:
        print("  ✓ Aucun doublon détecté")

    # ── Tournées à 1 seul magasin (sauf camions figés) ────────────────────────
    # Un camion supplémentaire (VX) réduit à 1 magasin est éliminé : mieux vaut
    # rattacher ce magasin à une tournée voisine que de mobiliser un véhicule
    # de plus. En revanche, un véhicule EXISTANT non figé sortira de toute façon
    # ce jour-là — le vider complètement (comme on faisait avant) le laisse
    # inutilisé alors qu'il pourrait emporter 1-2 magasins proches en plus.
    # On tente donc de le compléter, et on ne le vide que si rien de proche
    # n'a pu lui être ajouté.
    nb_elimines = 0
    nb_completes_1mag = 0
    for dj in [d for d in DEMI_JOURNEES if d not in DJ_DIMANCHE]:
        max_c_dj = max_norm
        for (d, veh), vifs in list(dj_veh.items()):
            if d != dj: continue
            if veh in VEHICULES_FIGES: continue  # préserver les figés
            if (veh, dj) in TOURNEES_GELEES: continue
            if len(vifs) != 1: continue
            vif_seul = vifs[0]
            r_seul = vif2row.get(vif_seul)
            if r_seul is None: continue
            lat_s = float(r_seul.get('Latitude', BAI_LAT))
            lon_s = float(r_seul.get('Longitude', BAI_LON))
            sec_s = _secteur_vif_calc2(vif_seul)
            est_vx = veh in nouveaux_vehs  # camion supplémentaire (VX3xx)

            if not est_vx:
                # Véhicule existant : chercher un magasin proche du même secteur
                # dans une tournée qui a de la marge, pour compléter plutôt que
                # de vider ce camion.
                best_vif = None; best_src = None; best_dist = float('inf')
                for (d2, veh2), vifs2 in dj_veh.items():
                    if d2 != dj or veh2 == veh: continue
                    if veh2 in VEHICULES_FIGES: continue
                    if (veh2, dj) in TOURNEES_GELEES: continue
                    if len(vifs2) <= 2: continue  # ne pas dégarnir une tournée déjà légère
                    for v in vifs2:
                        if _secteur_vif_calc2(v) != sec_s: continue
                        r_v = vif2row.get(v)
                        if r_v is None: continue
                        d_km = haversine(float(r_v.get('Latitude', BAI_LAT)),
                                          float(r_v.get('Longitude', BAI_LON)), lat_s, lon_s)
                        if d_km < best_dist:
                            best_dist = d_km; best_vif = v; best_src = veh2
                if best_vif:
                    nom_a = str(vif2row.get(best_vif, {}).get('Nom', best_vif))
                    dj_veh[(dj, best_src)] = [v for v in dj_veh[(dj, best_src)] if v != best_vif]
                    dj_veh[(dj, veh)].append(best_vif)
                    print(f"  Complément tournée 1 mag: '{nom_a}' {best_src}→{veh} ({dj}, évite un véhicule existant inutilisé)")
                    nb_completes_1mag += 1
                    continue  # ne pas éliminer ce camion, il n'est plus à 1 magasin
                # sinon : rien de proche à ajouter, on laisse la tournée à 1
                # magasin telle quelle (mieux qu'un véhicule totalement vide).
                continue

            # Chercher la tournée la plus proche avec de la place
            best_key = None; best_score = float('inf')
            for (d2, veh2), vifs2 in dj_veh.items():
                if d2 != dj or veh2 == veh: continue
                if (veh2, dj) in TOURNEES_GELEES: continue
                if veh2 in VEHICULES_FIGES: continue
                sec2 = _sec_dominant([v for v in vifs2 if v not in new_vifs])
                limite2 = max_c_dj + 2 if (sec2 and sec_s and sec2 == sec_s) else max_c_dj
                if len(vifs2) >= limite2: continue
                clat, clon = _centre_vifs(vifs2)
                d_km = haversine(lat_s, lon_s, clat, clon)
                pen = 0.0 if (not sec2 or not sec_s or sec2 == sec_s) else max(10.0, d_km)
                score = d_km + pen
                if score < best_score:
                    best_score = score; best_key = (d2, veh2)
            if best_key:
                nom_s = str(r_seul.get('Nom', vif_seul))
                if vif_seul not in dj_veh[best_key]: dj_veh[best_key].append(vif_seul)
                del dj_veh[(dj, veh)]
                print(f"  Tournée 1 mag éliminée: '{nom_s}' ({veh}/{dj}) → {best_key[1]}")
                nb_elimines += 1
    if nb_elimines:
        print(f"  → {nb_elimines} tournée(s) à 1 magasin éliminée(s) (camions supplémentaires)")
    if nb_completes_1mag:
        print(f"  → {nb_completes_1mag} tournée(s) à 1 magasin complétée(s) (véhicules existants, pour éviter un camion inutilisé)")


    # Coefficients trafic TomTom 2025 Grenoble
    SECTEURS_METRO_XL = {
        'Grenoble Nord','Grenoble Sud','Grenoble Est','Grenoble Ouest',
        'Grenoble Centre Nord','Grenoble Centre Est','Grenoble Centre Ouest',
        'Echirolles','Seyssinet','Saint Martin Dheres'
    }
    TRAFIC_XL = {
        'Vendredi Matin':      {'m': 1.30, 'h': 1.10},
        'Vendredi Apres Midi': {'m': 1.40, 'h': 1.15},
        'Samedi Matin':        {'m': 1.10, 'h': 1.05},
        'Samedi Apres Midi':   {'m': 1.15, 'h': 1.05},
    }
    def _fourchette_xl(dj, secteur, nb_mag, km_estimes):
        """Calcule la fourchette de durée estimée (conduite + collecte)."""
        import math as _m
        # Vitesse moyenne hors trafic : ~40 km/h en zone urbaine, 60 hors
        sec_dom = (secteur or '').split('|')[0].strip()
        metro = sec_dom in SECTEURS_METRO_XL
        vitesse_base = 35 if metro else 55  # km/h
        duree_base = (km_estimes / vitesse_base) * 60  # minutes
        coef = TRAFIC_XL.get(dj, {'m': 1.20, 'h': 1.10})
        coef_val = coef['m'] if metro else coef['h']
        cond_min = int(duree_base * coef_val)
        cond_max = int(duree_base * (coef_val + 0.10))
        total_min = cond_min + nb_mag * 30
        total_max = cond_max + nb_mag * 35
        def fmt(v):
            h, m = divmod(v, 60)
            return f'{h}h{m:02d}' if h else f'{m}min'
        zone = 'Métropole' if metro else 'Hors métropole'
        return f'{fmt(total_min)} – {fmt(total_max)} ({zone})'

    # Construire les lignes de tournées
    rows = []
    dj_order = {dj: i for i, dj in enumerate(DEMI_JOURNEES)}

    for (dj, veh), vifs_all in dj_veh.items():
        max_c = max_dim if dj in DJ_DIMANCHE else max_norm

        # Enrichir avec coordonnées et trier par distance BAI
        infos = []
        for vif in vifs_all:
            row = vif2row.get(vif)
            if row is not None:
                # Tonnage par demi-journée = tonnage annuel / nb de passages du magasin
                nb_passages = vif_passages.get(vif, 1)
                tonnage_annuel = float(row['Tonnage 2025']) if row['Tonnage 2025'] else 0
                tonnage_par_dj = round(tonnage_annuel / nb_passages) if nb_passages > 0 else 0
                infos.append({
                    'vif': vif, 'nom': str(row['Nom']),
                    'lat': float(row['Latitude']), 'lon': float(row['Longitude']),
                    'tonnage': tonnage_par_dj,
                    'nouveau': vif in new_vifs,
                    'secteur': _secteur_vif_calc2(vif) or str(row.get('Secteur',''))
                })
        infos.sort(key=lambda x: dist_bai(x['lat'], x['lon']))

        # Une entrée (dj, veh) = une seule tournée (pas de découpage en sous-groupes)
        groupe = infos
        if groupe:
            pts = [(m['lat'], m['lon']) for m in groupe]
            km = dist_bai(pts[0][0], pts[0][1])
            for j in range(len(pts)-1):
                km += haversine(pts[j][0], pts[j][1], pts[j+1][0], pts[j+1][1])
            km += dist_bai(pts[-1][0], pts[-1][1])

            tonnage = sum(m['tonnage'] for m in groupe)
            has_new = any(m['nouveau'] for m in groupe)
            from collections import Counter as C2
            secs = [m['secteur'] for m in groupe if m['secteur']]
            # Afficher tous les secteurs distincts (ordre fréquence décroissante)
            secs_uniq = [s for s, _ in C2(secs).most_common()] if secs else []
            secteur = ' | '.join(secs_uniq) if secs_uniq else ''

            parts = []
            if has_new: parts.append('Nouveau magasin 2026')
            if veh in nouveaux_vehs: parts.append('Camion supplementaire')
            # Optimisation 3 : signaler les tournées légères et surchargées
            if len(groupe) <= 2:
                capacite_restante = max_c - len(groupe)
                parts.append(f'Tournee legere ({len(groupe)} mag) - peut absorber {capacite_restante} mag supp')
            elif len(groupe) >= max_c:
                parts.append(f'Tournee pleine ({len(groupe)}/{max_c})')
            else:
                parts.append(f'{len(groupe)}/{max_c} magasins')

            nb_mag_t = len(groupe)
            duree_est = _fourchette_xl(dj, secteur, nb_mag_t, round(km, 1))
            row_t = {
                'Demi-journee': dj, 'Camion': veh,
                'Nom camion': veh2nom.get(veh, ''),
                'Tonnage': round(tonnage), 'Km estimes': round(km, 1),
                'Duree estimee': duree_est,
                'Secteur': secteur, 'Commentaire optimisation': ' | '.join(parts),
            }
            for k, m in enumerate(groupe):
                row_t[f'Magasin {k+1}'] = m['nom']
            for k in range(len(groupe), max_dim):
                row_t[f'Magasin {k+1}'] = ''
            rows.append(row_t)

    df_t = pd.DataFrame(rows)
    df_t['_ord'] = df_t['Demi-journee'].map(dj_order).fillna(99)
    df_t = df_t.sort_values(['_ord','Camion']).drop(columns=['_ord']).reset_index(drop=True)
    # Stocker les magasins hors secteur pour l'onglet Tournées
    global non_affectes_global
    non_affectes_global = [(nom, msg) for _, nom, msg in non_affectes if 'hors secteur' in msg]
    return df_t

# ─── GÉNÉRATION EXCEL ─────────────────────────────────────────────────────────
def _ordonner_magasins_nn(mags, nom2coords, bai_lat, bai_lon):
    """Ordonne les magasins par plus proche voisin depuis la BAI (approximation TSP)."""
    from math import radians, cos, sin, asin, sqrt
    def hav(la1,lo1,la2,lo2):
        R=6371; dlat=radians(la2-la1); dlon=radians(lo2-lo1)
        a=sin(dlat/2)**2+cos(radians(la1))*cos(radians(la2))*sin(dlon/2)**2
        return 2*R*asin(sqrt(a))
    if len(mags) <= 1:
        return mags
    restants = list(mags)
    ordre = []
    lat_cur, lon_cur = bai_lat, bai_lon
    while restants:
        best = min(restants, key=lambda m: hav(lat_cur, lon_cur,
                   nom2coords.get(m,(bai_lat,bai_lon))[0],
                   nom2coords.get(m,(bai_lat,bai_lon))[1]))
        ordre.append(best)
        restants.remove(best)
        lat_cur, lon_cur = nom2coords.get(best,(bai_lat,bai_lon))
    return ordre


def generer_excel(df_t, df_mag, args, output_path, fiches_2025=None, hors_secteur=None, log_bat=None):
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    print(f"  Génération Excel : {output_path}")

    C_HDR  = PatternFill("solid", fgColor="1F4E79")
    C_ORG  = PatternFill("solid", fgColor="FFE699")
    C_VRT  = PatternFill("solid", fgColor="E2EFDA")
    C_RGE  = PatternFill("solid", fgColor="FFD7D7")
    C_BLC  = PatternFill("solid", fgColor="DDEEFF")
    C_GRS  = PatternFill("solid", fgColor="F2F2F2")
    C_WHT  = PatternFill("solid", fgColor="FFFFFF")

    F_HDR  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    F_NRM  = Font(name='Calibri', size=10)
    thin   = Side(style='thin', color='AAAAAA')
    BRD    = Border(left=thin, right=thin, top=thin, bottom=thin)
    A_C    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    A_L    = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    nouveaux_vehs = set(f'VX{300+i:03d}' for i in range(args.camions_supp))
    noms_nouveaux  = set(str(r['Nom']) for _, r in df_mag[df_mag['Nouveau']=='OUI'].iterrows())

    wb = openpyxl.Workbook()

    # ── Paramètres ──────────────────────────────────────────────────────────
    ws_p = wb.active; ws_p.title = 'Parametres'
    params = [
        ('Paramètre','Valeur','Description'),
        ('--camions-supp',   args.camions_supp,    'Camions supplémentaires'),
        ('--poids-nouveaux', args.poids_nouveaux,  'Poids kg nouveaux magasins'),
        ('--max-magasins',   args.max_magasins,    'Max magasins/tournée hors dimanche'),
        ('Max dimanche',     args.max_magasins+1,  'Max dimanche (auto)'),
        ('Optimiser anciens', 'OUI' if args.optimiser_anciens else 'NON', 'Déplacer anciens magasins 2025'),
        ('Fusionner légères',  'OUI' if args.fusionner_legeres else 'NON', 'Fusionner tournées légères (≤ 2 mag)'),
        ('Corriger mal placés','OUI' if args.corriger_mal_places else 'NON', 'Réaffecter magasins éloignés du centre'),
        ('Génération',       datetime.now().strftime('%d/%m/%Y %H:%M'), ''),
        ('PDF source',       args.pdf, ''),
        ('Magasins',         args.magasins, ''),
        ('Nouveaux',         ', '.join(str(v) for v in args.nouveaux) if args.nouveaux else '(aucun)', ''),
    ]
    for r, row in enumerate(params, 1):
        for c, v in enumerate(row, 1):
            cell = ws_p.cell(row=r, column=c, value=v)
            cell.font = F_HDR if r==1 else F_NRM
            cell.fill = C_HDR if r==1 else (C_GRS if r%2==0 else C_WHT)
            cell.border = BRD; cell.alignment = A_L
    for w, col in zip([22,30,45], ['A','B','C']):
        ws_p.column_dimensions[col].width = w

    # Compte rendu d'exécution au pied de l'onglet Parametres
    if log_bat:
        row_log = len(params) + 3
        # Titre
        cell_titre = ws_p.cell(row=row_log, column=1, value='COMPTE RENDU D EXECUTION')
        cell_titre.font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        cell_titre.fill = C_HDR
        cell_titre.border = BRD
        row_log += 1
        # Lignes du log - colonnes A (contenu) B et C vides
        fill_log = PatternFill("solid", fgColor="F8F8F8")
        font_log = Font(name='Courier New', size=9)
        for ligne in log_bat.split('\n'):
            ligne = ligne.rstrip()
            if not ligne:
                continue
            # Tronquer les lignes trop longues
            if len(ligne) > 200:
                ligne = ligne[:200] + '...'
            # Préfixer avec espace si la ligne commence par =,+,-,@ pour éviter
            # qu'Excel interprète la valeur comme une formule
            if ligne and ligne[0] in ('=', '+', '-', '@', '*'):
                ligne = ' ' + ligne
            cell = ws_p.cell(row=row_log, column=1, value=str(ligne))
            cell.data_type = 's'  # forcer type string
            cell.font = font_log
            cell.fill = fill_log
            cell.border = BRD
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            row_log += 1

    # ── Tournées ─────────────────────────────────────────────────────────────
    ws_t = wb.create_sheet('Tournees')

    # ── Lignes de paramètres en en-tête (avant la ligne de filtres) ───────────
    params_tournees = [
        ('Camions supplémentaires', args.camions_supp),
        ('Max magasins/tournée',    args.max_magasins),
        ('Max dimanche',            args.max_magasins + 1),
        ('Poids nouveaux magasins', f'{args.poids_nouveaux} kg'),
        ('Optimiser anciens',       'OUI' if args.optimiser_anciens else 'NON'),
        ('Fusionner légères',       'OUI' if args.fusionner_legeres else 'NON'),
        ('Corriger mal placés',     'OUI' if args.corriger_mal_places else 'NON'),
    ]
    nb_cols_t = 13  # nombre de colonnes de l'onglet Tournées
    fill_param = PatternFill("solid", fgColor="D9E1F2")  # bleu très clair
    font_param_lbl = Font(name='Calibri', bold=True, size=9, color='1F4E79')
    font_param_val = Font(name='Calibri', size=9, color='1F4E79')
    for i, (lbl, val) in enumerate(params_tournees, 1):
        cell_l = ws_t.cell(row=i, column=1, value=lbl)
        cell_l.font = font_param_lbl; cell_l.fill = fill_param
        cell_l.alignment = Alignment(horizontal='right', vertical='center')
        cell_v = ws_t.cell(row=i, column=2, value=val)
        cell_v.font = font_param_val; cell_v.fill = fill_param
        cell_v.alignment = Alignment(horizontal='left', vertical='center')
        # Remplir le reste de la ligne en bleu clair
        for c in range(3, nb_cols_t + 1):
            ws_t.cell(row=i, column=c).fill = fill_param
        ws_t.row_dimensions[i].height = 14

    row_hdr = len(params_tournees) + 1  # ligne d'en-tête des colonnes

    cols_t = ['Demi-journee','Camion','Nom camion','Tonnage','Km estimes','Duree estimee','Secteur',
              'Magasin 1','Magasin 2','Magasin 3','Magasin 4','Magasin 5','Magasin 6',
              'Commentaire optimisation']
    for c, col in enumerate(cols_t, 1):
        cell = ws_t.cell(row=row_hdr, column=c, value=col)
        cell.font = F_HDR; cell.fill = C_HDR; cell.border = BRD; cell.alignment = A_C
    ws_t.row_dimensions[row_hdr].height = 22

    # Construire l'index 2025 : nom_magasin → (demi_journee, camion)
    # Pour détecter les changements de tournée entre 2025 et 2026
    index_2025 = {}  # nom → (dj_2025, veh_2025)
    if fiches_2025:
        vif2nom_t = {}
        df_ref_t = pd.read_excel(args.magasins)
        df_ref_t.columns = [c.strip() for c in df_ref_t.columns]
        for col in df_ref_t.columns:
            if 'vif' in col.lower() or col.lower() == 'code':
                df_ref_t = df_ref_t.rename(columns={col: 'Code VIF'}); break
        for _, r in df_ref_t.iterrows():
            v = str(r.get('Code VIF','')).strip().lstrip('0')
            n = str(r.get('Nom','')).strip()
            if v and n:
                vif2nom_t[v] = n
        # index_2025 : nom → set de (dj, veh) — toutes les combinaisons 2025
        index_2025_multi = {}
        for fiche in (fiches_2025 or []):
            dj_25  = fiche['demi_journee']
            veh_25 = fiche['vehicule']
            for vif in fiche['vif_codes']:
                nom = vif2nom_t.get(vif.lstrip('0'), '')
                if nom:
                    if nom not in index_2025_multi:
                        index_2025_multi[nom] = set()
                    index_2025_multi[nom].add((dj_25, veh_25))
                    # Pour compatibilité : garder index_2025 avec la première occurrence
                    if nom not in index_2025:
                        index_2025[nom] = (dj_25, veh_25)

    C_CHG_CAM = PatternFill("solid", fgColor="FF6B6B")   # Rouge  = changement de camion
    C_CHG_DJ  = PatternFill("solid", fgColor="FFB347")   # Orange = changement de demi-journée

    dj_cur = None; alt = True
    for r, (_, row) in enumerate(df_t.iterrows(), row_hdr + 1):
        if row['Demi-journee'] != dj_cur:
            dj_cur = row['Demi-journee']; alt = not alt
        veh = str(row.get('Camion',''))
        dj  = str(row.get('Demi-journee',''))
        if veh in nouveaux_vehs: fill_ligne = C_VRT
        elif veh in VEHICULES_FIGES: fill_ligne = C_RGE
        elif alt: fill_ligne = C_BLC
        else: fill_ligne = C_WHT
        for c, col in enumerate(cols_t, 1):
            val = row.get(col,'')
            if col in ('Tonnage','Km estimes') and val != '':
                try: val = round(float(val), 1)
                except: pass
            cell = ws_t.cell(row=r, column=c, value=val)
            cell.border = BRD
            cell.alignment = A_L if col in ('Commentaire optimisation','Secteur') else A_C
            # Colorier les cellules magasin :
            # Orange si nouveau magasin 2026, Rouge/Orange si changement vs 2025
            if col.startswith('Magasin ') and str(val).strip() not in ('','nan'):
                nom_mag = str(val).strip()
                if nom_mag in noms_nouveaux:
                    # Nouveau magasin 2026 → orange sur la cellule uniquement
                    cell.fill = C_ORG
                    cell.font = F_NRM
                    # Violet si hors secteur
                    if hors_secteur and any(nom_mag == n for n, _ in hors_secteur):
                        cell.fill = PatternFill("solid", fgColor="CC99FF")  # violet
                        cell.font = Font(name='Calibri', size=10, bold=True, color='4B0082')
                    continue
            if col.startswith('Magasin ') and str(val).strip() not in ('','nan') and index_2025:
                nom_mag = str(val).strip()
                if nom_mag in index_2025_multi:
                    combinaisons_25 = index_2025_multi[nom_mag]
                    # Vérifier si la combinaison (dj, veh) 2026 existait en 2025
                    meme_combo = (dj, veh) in combinaisons_25
                    meme_veh   = any(v == veh for _, v in combinaisons_25)
                    meme_dj    = any(d == dj  for d, _ in combinaisons_25)
                    if meme_combo:
                        # Même camion ET même demi-journée → pas de changement
                        cell.fill = fill_ligne
                        cell.font = F_NRM
                    elif not meme_veh:
                        # Changement de camion → rouge
                        cell.fill = C_CHG_CAM
                        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                    elif not meme_dj:
                        # Même camion mais demi-journée différente → orange
                        cell.fill = C_CHG_DJ
                        cell.font = Font(name='Calibri', size=10, bold=True)
                    else:
                        cell.fill = fill_ligne
                        cell.font = F_NRM
                else:
                    # Nouveau magasin 2026 (absent en 2025) → garder couleur ligne
                    cell.fill = fill_ligne
                    cell.font = F_NRM
            else:
                cell.fill = fill_ligne
                cell.font = F_NRM
            # En dernier : colorer Secteur selon nombre de secteurs
            if col == 'Secteur' and str(val).strip() not in ('','nan'):
                nb_sec = len([s for s in str(val).split('|') if s.strip()])
                if nb_sec > 3:
                    cell.font = Font(name='Calibri', size=10, bold=True, color='C00000')  # rouge
                elif nb_sec == 3:
                    cell.font = Font(name='Calibri', size=10, bold=True, color='C65911')  # orange
    # Largeurs colonnes : DJ, Camion, Nom camion, Tonnage, Km, Secteur, Mag1-6, Commentaire
    for i, w in enumerate([22,8,28,9,9,22,22,25,25,25,25,25,25,35], 1):
        ws_t.column_dimensions[get_column_letter(i)].width = w
    # Hauteur de ligne automatique (wrap_text)
    for row_idx in range(row_hdr + 1, row_hdr + len(df_t) + 1):
        ws_t.row_dimensions[row_idx].height = None
    ws_t.freeze_panes = f'A{row_hdr + 1}'
    ws_t.auto_filter.ref = f'A{row_hdr}:{get_column_letter(len(cols_t))}{row_hdr + len(df_t)}'

    # ── Configuration impression ────────────────────────────────────────────
    from openpyxl.worksheet.page import PageMargins
    ws_t.page_setup.orientation = ws_t.ORIENTATION_LANDSCAPE
    ws_t.page_setup.paperSize   = ws_t.PAPERSIZE_A4
    ws_t.page_setup.fitToWidth  = 1
    ws_t.page_setup.fitToHeight = 4
    ws_t.page_setup.fitToPage   = True
    ws_t.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5,
                                    header=0.2, footer=0.2)
    ws_t.oddHeader.center.text = "BAI 38 — Tournées de Collecte 2026"
    ws_t.oddHeader.center.size = 10
    ws_t.oddFooter.left.text   = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws_t.oddFooter.right.text  = "Page &P / &N"
    ws_t.oddFooter.left.size   = 8
    ws_t.oddFooter.right.size  = 8
    ws_t.print_title_rows      = f'{row_hdr}:{row_hdr}'
    # Bouton impression en bas du tableau
    row_btn = row_hdr + len(df_t) + 1
    ws_t.merge_cells(f'A{row_btn}:{get_column_letter(len(cols_t))}{row_btn}')
    btn = ws_t.cell(row=row_btn, column=1,
        value='🖨  Imprimer cet onglet : Ctrl+P  —  Paysage A4, 1 page en largeur, ~4 pages, en-têtes répétés')
    btn.font      = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    btn.fill      = PatternFill("solid", fgColor="1F4E79")
    btn.alignment = Alignment(horizontal='center', vertical='center')
    ws_t.row_dimensions[row_btn].height = 20

    # ── Onglet VIF : une ligne par tournée avec codes VIF et noms ────────────
    ws_v = wb.create_sheet('Tournees VIF')

    # Index nom_magasin → code VIF depuis df_mag
    nom2vif = {}
    for _, rmag in df_mag.iterrows():
        nom = str(rmag.get('Nom', '')).strip()
        vif = str(rmag.get('Code VIF', '')).strip()
        if nom and vif and vif != 'nan':
            nom2vif[nom] = vif.zfill(8) if vif.isdigit() else vif

    # Colonnes dynamiques : DJ, Camion, Nom camion, puis VIF1/Nom1, VIF2/Nom2...
    nb_max_mag = max((sum(1 for k in range(1,7) if str(row.get(f'Magasin {k}','')).strip() not in ('','nan'))
                      for _,row in df_t.iterrows()), default=0)
    cols_vif = ['Demi-journee', 'Camion', 'Nom camion']
    for k in range(1, nb_max_mag + 1):
        cols_vif += [f'Code VIF {k}', f'Magasin {k}']

    # En-tête
    for c, col in enumerate(cols_vif, 1):
        cell = ws_v.cell(row=1, column=c, value=col)
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.fill = C_HDR
        cell.border = BRD
        cell.alignment = A_C

    # Index nom → coords pour le tri NNF
    nom2coords_vif = {}
    for _, rmag2 in df_mag.iterrows():
        nom2 = str(rmag2.get('Nom', '')).strip()
        try:
            nom2coords_vif[nom2] = (float(rmag2.get('Latitude', BAI_LAT)),
                                    float(rmag2.get('Longitude', BAI_LON)))
        except (ValueError, TypeError):
            nom2coords_vif[nom2] = (BAI_LAT, BAI_LON)

    # Données : trier par Camion puis Demi-journée
    DJ_ORDER = {dj: i for i, dj in enumerate(DEMI_JOURNEES)}
    df_sorted = df_t.copy()
    df_sorted['_dj_ord'] = df_sorted['Demi-journee'].map(DJ_ORDER).fillna(99)
    df_sorted = df_sorted.sort_values(['Camion', '_dj_ord']).drop(columns=['_dj_ord'])

    alt_row_v = 0
    for row_v, (_, row) in enumerate(df_sorted.iterrows(), 2):
        fill_v = C_BLC if alt_row_v % 2 == 0 else C_WHT
        alt_row_v += 1
        mags_raw = [str(row.get(f'Magasin {k}', '')).strip() for k in range(1, 7)
                if str(row.get(f'Magasin {k}', '')).strip() not in ('', 'nan')]
        # Ordonner par plus proche voisin depuis la BAI
        mags = _ordonner_magasins_nn(mags_raw, nom2coords_vif, BAI_LAT, BAI_LON)

        row_data = [
            row.get('Demi-journee', ''),
            row.get('Camion', ''),
            row.get('Nom camion', ''),
        ]
        for k, mag in enumerate(mags):
            vif = nom2vif.get(mag, '')
            # Ajouter le zéro initial si VIF numérique (format 8 chiffres)
            if vif and vif.isdigit():
                vif = vif.zfill(8)
            row_data += [vif, mag]
        # Compléter jusqu'à nb_max_mag
        while len(row_data) < len(cols_vif):
            row_data.append('')

        for c, val in enumerate(row_data, 1):
            cell = ws_v.cell(row=row_v, column=c, value=val)
            cell.font = F_NRM
            cell.fill = fill_v
            cell.border = BRD
            cell.alignment = A_C if c <= 3 else A_L

    # Largeurs colonnes
    ws_v.column_dimensions['A'].width = 22  # DJ
    ws_v.column_dimensions['B'].width = 9   # Camion
    ws_v.column_dimensions['C'].width = 28  # Nom camion
    for k in range(nb_max_mag):
        col_vif = get_column_letter(4 + k * 2)
        col_nom = get_column_letter(5 + k * 2)
        ws_v.column_dimensions[col_vif].width = 12
        ws_v.column_dimensions[col_nom].width = 32
    ws_v.freeze_panes = 'A2'
    ws_v.auto_filter.ref = f"A1:{get_column_letter(len(cols_vif))}1"

    # ── Magasins ─────────────────────────────────────────────────────────────
    ws_m = wb.create_sheet('Magasins')
    cols_fix = ['Code VIF','Nom','Ville','Secteur','Tonnage 2025','Latitude','Longitude','Dist_BAI','Nouveau']
    cols_m   = cols_fix + DEMI_JOURNEES
    for c, col in enumerate(cols_m, 1):
        cell = ws_m.cell(row=1, column=c, value=col)
        cell.font = F_HDR; cell.fill = C_HDR; cell.border = BRD; cell.alignment = A_C

    # Index VIF → planning
    vif_plan = defaultdict(dict)
    for _, tr in df_t.iterrows():
        dj = tr['Demi-journee']; cam = tr['Camion']
        for k in range(1,7):
            nom = str(tr.get(f'Magasin {k}',''))
            if nom and nom != 'nan':
                matches = df_mag[df_mag['Nom'].astype(str).values == nom]
                if len(matches) > 0:
                    vif = str(matches.iloc[0]['Code VIF'])
                    vif_plan[vif][dj] = cam

    for r, (_, mag) in enumerate(df_mag.iterrows(), 2):
        is_new = bool(mag.get('Nouveau', False))
        fill_r = C_ORG if is_new else (C_GRS if r%2==0 else C_WHT)
        vif = str(mag['Code VIF'])
        for c, col in enumerate(cols_m, 1):
            if col == 'Nouveau':     val = 'OUI' if is_new else ''
            elif col == 'Dist_BAI': val = round(float(mag['Dist_BAI']), 1)
            elif col == 'Tonnage 2025': val = int(float(mag.get('Tonnage 2025', 0)))
            elif col in DEMI_JOURNEES:  val = vif_plan.get(vif, {}).get(col, '')
            else:
                val = mag.get(col, '')
                if hasattr(val, 'iloc'): val = val.iloc[0] if len(val)>0 else ''
                val = str(val) if val else ''
            cell = ws_m.cell(row=r, column=c, value=val)
            cell.font = F_NRM; cell.fill = fill_r; cell.border = BRD; cell.alignment = A_C
    for i, w in enumerate([12,30,20,20,12,12,12,12,8]+[16]*len(DEMI_JOURNEES), 1):
        ws_m.column_dimensions[get_column_letter(i)].width = w
    ws_m.freeze_panes = 'A2'
    ws_m.auto_filter.ref = ws_m.dimensions

    # ── Secteurs ─────────────────────────────────────────────────────────────
    ws_s = wb.create_sheet('Secteurs')
    cols_s = ['Secteur','Nb Magasins','Tonnage Total','Magasins']
    for c, col in enumerate(cols_s, 1):
        cell = ws_s.cell(row=1, column=c, value=col)
        cell.font = F_HDR; cell.fill = C_HDR; cell.border = BRD; cell.alignment = A_C

    # Calculer les stats par secteur
    from collections import defaultdict as _dd
    secteur_stats = _dd(lambda: {'noms': [], 'tonnage': 0})
    for _, mag in df_mag.iterrows():
        sec = str(mag.get('Secteur',''))
        nom = str(mag.get('Nom',''))
        ton = float(mag.get('Tonnage 2025', 0) or 0)
        if sec and nom:
            secteur_stats[sec]['noms'].append(nom)
            secteur_stats[sec]['tonnage'] += ton

    for r, (secteur, stats) in enumerate(sorted(secteur_stats.items()), 2):
        fill_r = C_GRS if r%2==0 else C_WHT
        vals = [secteur, len(stats['noms']), round(stats['tonnage']), ', '.join(sorted(stats['noms']))]
        for c, val in enumerate(vals, 1):
            cell = ws_s.cell(row=r, column=c, value=val)
            cell.font = F_NRM; cell.fill = fill_r; cell.border = BRD
            cell.alignment = A_L if c in (1,4) else A_C
    ws_s.column_dimensions['A'].width = 25
    ws_s.column_dimensions['B'].width = 15
    ws_s.column_dimensions['C'].width = 15
    ws_s.column_dimensions['D'].width = 80
    ws_s.freeze_panes = 'A2'
    ws_s.auto_filter.ref = ws_s.dimensions

    # ── Analyse ──────────────────────────────────────────────────────────────
    ws_a = wb.create_sheet('Analyse')

    # Titre
    ws_a.merge_cells('A1:H1')
    cell = ws_a.cell(row=1, column=1, value='ANALYSE ET OPTIMISATION DES TOURNÉES BAI 38 - 2026')
    cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    cell.fill = C_HDR
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_a.row_dimensions[1].height = 28

    row_a = 3  # ligne courante dans l'onglet Analyse

    # ── POINT 1 : Camions sous-utilisés ──────────────────────────────────────
    # ── SECTION 0 : Magasins sans coordonnées GPS ───────────────────────────
    sans_gps_df = df_mag[df_mag.get('Sans_GPS', False)] if 'Sans_GPS' in df_mag.columns else df_mag[df_mag['Latitude'] == BAI_LAT]
    # Recalculer proprement depuis la colonne Sans_GPS
    # Magasins sans GPS : ceux du référentiel (Sans_GPS) + ceux dans les tournées sans coords
    sans_gps_noms = set()
    if 'Sans_GPS' in df_mag.columns:
        sans_gps_noms.update(df_mag[df_mag['Sans_GPS'] == True]['Nom'].astype(str).tolist())
    sans_gps_noms.update(noms_sans_gps if 'noms_sans_gps' in dir() else set())

    if sans_gps_noms:
        ws_a.cell(row=row_a, column=1, value='0. MAGASINS SANS COORDONNÉES GPS').font = Font(name='Calibri', bold=True, size=11, color='C00000')
        row_a += 1
        p_note = ws_a.cell(row=row_a, column=1,
                            value='Ces magasins utilisent les coordonnées de la BAI par défaut — leurs tournées et liens Google Maps sont approximatifs.')
        p_note.font = Font(name='Calibri', size=10, italic=True, color='C00000')
        row_a += 1
        hdrs0 = ['Code VIF','Nom','Ville','Adresse','Nouveau']
        for c, h in enumerate(hdrs0, 1):
            cell = ws_a.cell(row=row_a, column=c, value=h)
            cell.font = F_HDR; cell.fill = C_HDR; cell.border = BRD; cell.alignment = A_C
        row_a += 1
        fill_sgps = PatternFill("solid", fgColor="FFD7D7")
        for nom_sgps in sorted(sans_gps_noms):
            # Récupérer les infos depuis df_mag
            matches = df_mag[df_mag['Nom'].astype(str) == nom_sgps]
            if len(matches) > 0:
                r = matches.iloc[0]
                vals = [str(r.get('Code VIF','')), str(r.get('Nom','')),
                        str(r.get('Ville','')),    str(r.get('Adresse','')),
                        'OUI' if r.get('Nouveau', False) else '']
            else:
                vals = ['', nom_sgps, '', '', '']
            for c, v in enumerate(vals, 1):
                cell = ws_a.cell(row=row_a, column=c, value=v)
                cell.font = F_NRM; cell.fill = fill_sgps; cell.border = BRD; cell.alignment = A_L
            row_a += 1
        row_a += 1

    ws_a.cell(row=row_a, column=1, value='1. CAMIONS SOUS-UTILISÉS (≤ 2 magasins)').font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
    row_a += 1
    hdrs = ['Demi-journée','Camion','Nb Magasins','Capacité restante','Magasins']
    for c, h in enumerate(hdrs, 1):
        cell = ws_a.cell(row=row_a, column=c, value=h)
        cell.font = F_HDR; cell.fill = C_HDR; cell.border = BRD; cell.alignment = A_C
    row_a += 1

    DJ_ANALYSE = {'Vendredi Matin','Vendredi Apres Midi','Samedi Matin','Samedi Apres Midi'}

    # Regrouper df_t par (Demi-journee, Camion) pour compter le total réel de magasins
    # (un camion peut avoir plusieurs lignes dans df_t si rééquilibrage)
    def _mags_camion(dj, veh):
        """Retourne la liste de tous les magasins d'un camion sur une DJ (toutes lignes confondues)."""
        rows = df_t[(df_t['Demi-journee']==dj) & (df_t['Camion']==veh)]
        mags = []
        for _, r in rows.iterrows():
            for k in range(1,7):
                m = str(r.get(f'Magasin {k}','')).strip()
                if m and m not in ('','nan'):
                    mags.append(m)
        return mags

    # Construire la liste unique des (DJ, camion) analysés
    paires_analyse = df_t[df_t['Demi-journee'].isin(DJ_ANALYSE)][['Demi-journee','Camion']].drop_duplicates()

    sous_util_filtre = []
    for _, pair in paires_analyse.iterrows():
        dj, veh = pair['Demi-journee'], pair['Camion']
        if veh in VEHICULES_FIGES: continue  # ignorer les camions figés
        mags = _mags_camion(dj, veh)
        if len(mags) <= 2:
            sous_util_filtre.append((dj, veh, len(mags), mags))
    sous_util_filtre.sort(key=lambda x: (x[0], x[2]))

    for dj, veh, nb, mags in sous_util_filtre:
        max_c = args.max_magasins + 1 if dj == 'Dimanche Matin' else args.max_magasins
        vals = [dj, veh, nb, max_c - nb, ', '.join(mags)]
        fill_r = PatternFill("solid", fgColor="FFD7D7")
        for c, v in enumerate(vals, 1):
            cell = ws_a.cell(row=row_a, column=c, value=v)
            cell.font = F_NRM; cell.fill = fill_r; cell.border = BRD; cell.alignment = A_L
        row_a += 1

    if not sous_util_filtre:
        ws_a.cell(row=row_a, column=1, value='Aucun camion sous-utilisé détecté').font = F_NRM
        row_a += 1
    row_a += 1

    # ── POINT 2 : Fusions possibles de tournées légères ──────────────────────
    ws_a.cell(row=row_a, column=1, value='2. FUSIONS POSSIBLES DE TOURNÉES LÉGÈRES').font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
    row_a += 1
    hdrs2 = ['Demi-journée','Camion 1','Nb mag 1','Camion 2','Nb mag 2','Total','Capacité max','Faisable']
    for c, h in enumerate(hdrs2, 1):
        cell = ws_a.cell(row=row_a, column=c, value=h)
        cell.font = F_HDR; cell.fill = C_HDR; cell.border = BRD; cell.alignment = A_C
    row_a += 1

    fusions = []
    for dj in df_t['Demi-journee'].unique():
        if dj not in DJ_ANALYSE:
            continue
        max_c = args.max_magasins + 1 if dj == 'Dimanche Matin' else args.max_magasins
        # Regrouper par camion pour avoir le vrai total de magasins
        vehs_dj = df_t[df_t['Demi-journee']==dj]['Camion'].unique()
        legeres = [(veh, len(_mags_camion(dj, veh)))
                   for veh in vehs_dj
                   if len(_mags_camion(dj, veh)) <= 3
                   and veh not in VEHICULES_FIGES
                   and (veh, dj) not in TOURNEES_GELEES]
        for i in range(len(legeres)):
            for j in range(i+1, len(legeres)):
                v1, n1 = legeres[i]; v2, n2 = legeres[j]
                total = n1 + n2
                faisable = 'OUI ✓' if total <= max_c else f'NON ({total} > {max_c})'
                fusions.append((dj, v1, n1, v2, n2, total, max_c, faisable))

    fusions.sort(key=lambda x: (x[0], x[5]))
    for f in fusions:
        fill_r = PatternFill("solid", fgColor="E2EFDA") if 'OUI' in f[7] else PatternFill("solid", fgColor="FFF2CC")
        for c, v in enumerate(f, 1):
            cell = ws_a.cell(row=row_a, column=c, value=v)
            cell.font = F_NRM; cell.fill = fill_r; cell.border = BRD; cell.alignment = A_C
        row_a += 1

    if not fusions:
        ws_a.cell(row=row_a, column=1, value='Aucune fusion possible détectée').font = F_NRM
        row_a += 1
    row_a += 1

    # ── POINT 3 : Magasins mal placés géographiquement ───────────────────────
    ws_a.cell(row=row_a, column=1, value='3. MAGASINS MAL PLACÉS (éloignés du centre géo de leur tournée)').font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
    row_a += 1
    hdrs3 = ['Demi-journée','Camion','Magasin','Distance au centre (km)','Suggestion']
    for c, h in enumerate(hdrs3, 1):
        cell = ws_a.cell(row=row_a, column=c, value=h)
        cell.font = F_HDR; cell.fill = C_HDR; cell.border = BRD; cell.alignment = A_C
    row_a += 1

    # Index nom → coordonnées
    # Construire index nom → coordonnées GPS
    # Pour les magasins sans coordonnées, géocoder via Nominatim (OpenStreetMap) à partir de l'adresse
    import urllib.request, json, time

    def geocoder_adresse(adresse, ville):
        """Géocode une adresse via Nominatim (OpenStreetMap, gratuit, pas de clé API)."""
        try:
            q = f'{adresse}, {ville}, France'.replace(' ', '+')
            url = f'https://nominatim.openstreetmap.org/search?q={q}&format=json&limit=1'
            req = urllib.request.Request(url, headers={'User-Agent': 'BAI38-tournees/1.0'})
            with urllib.request.urlopen(req, timeout=5) as resp:
                data = json.loads(resp.read())
            if data:
                return float(data[0]['lat']), float(data[0]['lon'])
        except Exception:
            pass
        return None

    # Construire nom → coordonnées GPS depuis le référentiel COMPLET (toutes lignes, sans filtre État)
    # Nécessaire pour les magasins 'Collecte gardée' ou 'Non collecté' présents dans les tournées 2025
    df_ref_gps = pd.read_excel(args.magasins)
    df_ref_gps.columns = [c.strip() for c in df_ref_gps.columns]
    nom2coords = {}
    for _, r in df_ref_gps.iterrows():
        nom = str(r.get('Nom', '')).strip()
        try:
            lat_f = float(r.get('Latitude', 0))
            lon_f = float(r.get('Longitude', 0))
            if lat_f != 0 and lon_f != 0 and abs(lat_f - BAI_LAT) > 0.0001:
                nom2coords[nom] = (lat_f, lon_f)
        except (TypeError, ValueError):
            pass
    print(f"  → {len(nom2coords)}/{len(df_ref_gps)} magasins avec coordonnées GPS")
    # Vérifier correspondance avec les noms dans df_t
    noms_t = set()
    for k in range(1,7):
        col = f'Magasin {k}'
        if col in df_t.columns:
            noms_t.update(df_t[col].dropna().astype(str).str.strip())
    noms_t.discard(''); noms_t.discard('nan')
    noms_trouves = noms_t & set(nom2coords.keys())
    noms_sans_gps = noms_t - set(nom2coords.keys())
    if noms_sans_gps:
        print(f"  ATTENTION: {len(noms_sans_gps)} magasins sans coordonnées GPS dans les tournées :")
        for m in sorted(noms_sans_gps):
            print(f"    - {m}")

    mal_places = []
    for dj in df_t['Demi-journee'].unique():
        if dj not in DJ_ANALYSE:
            continue
        tournees_dj = df_t[df_t['Demi-journee']==dj]
        vehs_dj2 = tournees_dj['Camion'].unique()
        for veh in vehs_dj2:
            if veh in VEHICULES_FIGES: continue
            if (veh, dj) in TOURNEES_GELEES: continue
            mags = _mags_camion(dj, veh)
            if len(mags) < 2:
                continue
            coords = [nom2coords[m] for m in mags if m in nom2coords]
            if len(coords) < 2:
                continue
            lat_c = sum(c[0] for c in coords) / len(coords)
            lon_c = sum(c[1] for c in coords) / len(coords)
            for m, (lat, lon) in zip(mags, coords):
                dist_c = haversine(lat, lon, lat_c, lon_c)
                if dist_c > 8:  # plus de 8 km du centre de la tournée
                    # Trouver la tournée la plus proche pour ce magasin
                    best_veh = veh; best_d = dist_c
                    for _, row2 in tournees_dj[tournees_dj['Camion']!=veh].iterrows():
                        mags2 = [str(row2.get(f'Magasin {k}','')) for k in range(1,7)
                                  if str(row2.get(f'Magasin {k}','')).strip() not in ('','nan')]
                        coords2 = [nom2coords[m2] for m2 in mags2 if m2 in nom2coords]
                        if not coords2: continue
                        lat_c2 = sum(c[0] for c in coords2)/len(coords2)
                        lon_c2 = sum(c[1] for c in coords2)/len(coords2)
                        d2 = haversine(lat, lon, lat_c2, lon_c2)
                        if d2 < best_d:
                            best_d = d2; best_veh = row2['Camion']
                    suggestion = f'Plutôt avec {best_veh}' if best_veh != veh else 'Isolé géographiquement'
                    mal_places.append((dj, veh, m, round(dist_c,1), suggestion))

    mal_places.sort(key=lambda x: -x[3])
    for mp in mal_places[:20]:  # Top 20
        fill_r = PatternFill("solid", fgColor="FFE699")
        for c, v in enumerate(mp, 1):
            cell = ws_a.cell(row=row_a, column=c, value=v)
            cell.font = F_NRM; cell.fill = fill_r; cell.border = BRD; cell.alignment = A_L
        row_a += 1

    if not mal_places:
        ws_a.cell(row=row_a, column=1, value='Aucun magasin mal placé détecté (seuil 8 km)').font = F_NRM
        row_a += 1
    row_a += 1

    # Largeurs colonnes Analyse
    for col, w in zip(['A','B','C','D','E','F','G','H'], [22,10,12,10,12,8,12,60]):
        ws_a.column_dimensions[col].width = w
    ws_a.freeze_panes = 'A2'
    ws_a.auto_filter.ref = ws_a.dimensions

    # ── Tournées 2025 (source PDF) ───────────────────────────────────────────
    if fiches_2025:
        ws_25 = wb.create_sheet('Tournees 2025')

        # Index VIF → nom magasin depuis le référentiel COMPLET (toutes lignes, sans filtre État)
        # Permet de résoudre les magasins qui ne sont plus collectés, changés de nom ou supprimés
        df_complet = pd.read_excel(args.magasins)
        df_complet.columns = [c.strip() for c in df_complet.columns]
        df_complet = df_complet.loc[:, ~df_complet.columns.duplicated()]
        # Renommer Code VIF et Nom si nécessaire
        for col in df_complet.columns:
            cl = col.lower()
            if 'vif' in cl or cl == 'code':
                df_complet = df_complet.rename(columns={col: 'Code VIF'})
                break
        vif2nom = {}
        vif2secteur_25 = {}
        for _, r in df_complet.iterrows():
            vif = str(r.get('Code VIF', '')).strip().lstrip('0')
            nom = str(r.get('Nom', '')).strip()
            if vif and nom and vif not in vif2nom:
                vif2nom[vif] = nom
            # Calculer le secteur depuis lat/lon/ville
            try:
                import unicodedata as _ud2, re as _re2
                lat_c = float(r.get('Latitude', 0))
                lon_c = float(r.get('Longitude', 0))
                _v2 = _ud2.normalize('NFD', str(r.get('Ville',''))).encode('ascii','ignore').decode('ascii')
                _v2 = _re2.sub(r"[-'\u2019]", ' ', _v2)
                vil_c = _re2.sub(r'\s+', ' ', _v2).strip().title()
                if 'Grenoble' in vil_c:
                    if lat_c > 45.193: sec_c = 'Grenoble Nord'
                    elif lat_c < 45.173: sec_c = 'Grenoble Sud'
                    elif lon_c < 5.714: sec_c = 'Grenoble Ouest'
                    elif lon_c > 5.729: sec_c = 'Grenoble Est'
                    elif lat_c > 45.187: sec_c = 'Grenoble Centre Nord'
                    elif lon_c < 5.721: sec_c = 'Grenoble Centre Ouest'
                    else: sec_c = 'Grenoble Centre Est'
                elif vil_c in {'Biviers', 'Crolles', 'Froges', 'Saint Ismier', 'Domene', 'Le Versoud'}: sec_c = 'Gresivaudan'
                elif vil_c in {'Voiron', 'Saint Jean De Moirans'}:                sec_c = 'Voiron'
                elif vil_c in {'Rives', 'Renage'}:                                sec_c = 'Rives'
                elif vil_c in {'Seyssinet Pariset', 'Seyssins', 'Fontaine'}:     sec_c = 'Seyssinet'
                elif vil_c in {'Apprieu', 'Colombe', 'Le Grand Lemps'}:          sec_c = 'Plateau Nord'
                elif vil_c in {'Varces Allieres Et Risset', 'Vif', 'Vizille', 'Claix'}: sec_c = 'Sud Vercors'
                elif vil_c in {'La Terrasse', 'Le Touvet', 'Tencin'}:            sec_c = 'Chartreuse Est'
                elif vil_c in {'Peage De Roussillon', 'Salaise Sur Sanne'}:      sec_c = 'Roussillon'
                elif vil_c in {'Poisat', 'Saint Martin Dheres', 'Saint Martin D Heres', 'Gieres', 'Gières'}: sec_c = 'Saint Martin Dheres'
                else:
                    sec_c = vil_c
                if vif: vif2secteur_25[vif] = sec_c
            except Exception:
                pass

        # Colonnes Code VIF k / Magasin k intercalées, sur le même modèle que
        # l'onglet 'Tournees VIF' (les codes VIF sont déjà résolus ci-dessus
        # via fiche['vif_codes'], il ne restait qu'à les écrire dans le fichier).
        cols_25 = ['Demi-journée', 'Camion', 'Nom camion', 'Nb Magasins', 'Secteur dominant']
        for k in range(1, 7):
            cols_25 += [f'Code VIF {k}', f'Magasin {k}']
        for c, col in enumerate(cols_25, 1):
            cell = ws_25.cell(row=1, column=c, value=col)
            cell.font = F_HDR; cell.fill = C_HDR; cell.border = BRD; cell.alignment = A_C
        ws_25.row_dimensions[1].height = 22

        # Trier les fiches par demi-journée puis véhicule
        dj_order_25 = {dj: i for i, dj in enumerate(DEMI_JOURNEES)}
        fiches_triees = sorted(fiches_2025,
                               key=lambda f: (dj_order_25.get(f['demi_journee'], 99), f['vehicule']))

        dj_cur_25 = None; alt_25 = True
        for r_25, fiche in enumerate(fiches_triees, 2):
            dj  = fiche['demi_journee']
            veh = fiche['vehicule']
            vifs = fiche['vif_codes']

            if dj != dj_cur_25:
                dj_cur_25 = dj; alt_25 = not alt_25
            fill_25 = C_BLC if alt_25 else C_WHT

            # Résoudre les noms et secteurs depuis les VIF
            noms = [vif2nom.get(v.lstrip('0'), v) for v in vifs]
            secs_25 = [vif2secteur_25.get(v.lstrip('0'), '') for v in vifs if vif2secteur_25.get(v.lstrip('0'), '')]
            from collections import Counter as _C25
            secs_25_uniq = [s for s, _ in _C25(secs_25).most_common()] if secs_25 else []
            sec_dom_25 = ' | '.join(secs_25_uniq) if secs_25_uniq else ''

            # Codes VIF formatés (8 chiffres avec zéro initial), même convention que
            # l'onglet 'Tournees VIF', intercalés avec le nom du magasin correspondant.
            vifs_fmt = [v.zfill(8) if v.isdigit() else v for v in vifs]
            vif_mag_interleaved = []
            for k in range(6):
                if k < len(noms):
                    vif_mag_interleaved += [vifs_fmt[k], noms[k]]
                else:
                    vif_mag_interleaved += ['', '']

            nom_veh_25 = fiche.get('nom_vehicule', '')
            row_vals = [dj, veh, nom_veh_25, len(noms), sec_dom_25] + vif_mag_interleaved
            nb_cols_25 = len(row_vals)
            for c, val in enumerate(row_vals[:nb_cols_25], 1):
                cell = ws_25.cell(row=r_25, column=c, value=val)
                cell.font = F_NRM; cell.fill = fill_25; cell.border = BRD; cell.alignment = A_C
                # Coloration secteur : orange si 3 secteurs, rouge si > 3
                col_name_25 = cols_25[c-1] if c-1 < len(cols_25) else ''
                if col_name_25 == 'Secteur dominant' and str(val).strip() not in ('','nan'):
                    nb_sec_25 = len([s for s in str(val).split('|') if s.strip()])
                    if nb_sec_25 > 3:
                        cell.font = Font(name='Calibri', size=10, bold=True, color='C00000')
                    elif nb_sec_25 == 3:
                        cell.font = Font(name='Calibri', size=10, bold=True, color='C65911')

        # Largeurs colonnes
        ws_25.column_dimensions['A'].width = 22
        ws_25.column_dimensions['B'].width = 10
        ws_25.column_dimensions['C'].width = 28
        ws_25.column_dimensions['D'].width = 10
        ws_25.column_dimensions['E'].width = 40
        for k in range(6, 18):
            # Colonnes F,H,J,L,N,P = Code VIF k (étroites) ; G,I,K,M,O,Q = Magasin k (larges)
            largeur = 12 if (k - 6) % 2 == 0 else 28
            ws_25.column_dimensions[get_column_letter(k)].width = largeur
        ws_25.freeze_panes = 'A2'
        ws_25.auto_filter.ref = ws_25.dimensions

    # ── Évolutions : magasins ajoutés et supprimés ────────────────────────────
    ws_ev = wb.create_sheet('Evolutions')
    ws_ev.merge_cells('A1:D1')
    cell = ws_ev.cell(row=1, column=1, value='ÉVOLUTIONS DES MAGASINS 2025 → 2026')
    cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    cell.fill = C_HDR
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_ev.row_dimensions[1].height = 28

    df_ref_ev = pd.read_excel(args.magasins)
    df_ref_ev.columns = [c.strip() for c in df_ref_ev.columns]
    vif2nom_ev = {}
    for _, r in df_ref_ev.iterrows():
        vif = str(r.get('Code VIF', '')).strip().lstrip('0')
        nom = str(r.get('Nom', '')).strip()
        if vif and nom: vif2nom_ev[vif] = nom

    noms_2025 = set()
    if fiches_2025:
        for fiche in (fiches_2025 or []):
            for vif in fiche['vif_codes']:
                nom = vif2nom_ev.get(vif.lstrip('0'), '')
                if nom: noms_2025.add(nom)

    noms_2026  = set(str(r['Nom']) for _, r in df_mag.iterrows())
    ajoutes    = sorted(noms_2026 - noms_2025)
    supprimes  = sorted(noms_2025 - noms_2026)

    row_ev = 3
    for col, label, couleur, nb in [
        (1, f'NOUVEAUX MAGASINS 2026 ({len(ajoutes)})', '2E7D32', len(ajoutes)),
        (3, f'SUPPRIMÉS / NON RECONDUITS ({len(supprimes)})', 'C62828', len(supprimes))]:
        cell = ws_ev.cell(row=row_ev, column=col, value=label)
        cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill("solid", fgColor=couleur)
        cell.border = BRD; cell.alignment = A_C
        ws_ev.merge_cells(f'{get_column_letter(col)}{row_ev}:{get_column_letter(col+1)}{row_ev}')
    row_ev += 1

    fill_aj  = PatternFill("solid", fgColor="E8F5E9")
    fill_sup = PatternFill("solid", fgColor="FFEBEE")
    for i in range(max(len(ajoutes), len(supprimes))):
        if i < len(ajoutes):
            cell = ws_ev.cell(row=row_ev+i, column=1, value=ajoutes[i])
            cell.font = F_NRM; cell.fill = fill_aj; cell.border = BRD; cell.alignment = A_L
            ws_ev.merge_cells(f'A{row_ev+i}:B{row_ev+i}')
        if i < len(supprimes):
            cell = ws_ev.cell(row=row_ev+i, column=3, value=supprimes[i])
            cell.font = F_NRM; cell.fill = fill_sup; cell.border = BRD; cell.alignment = A_L
            ws_ev.merge_cells(f'C{row_ev+i}:D{row_ev+i}')

    for col, w in zip(['A','B','C','D'], [35,5,35,5]):
        ws_ev.column_dimensions[col].width = w
    print(f"  → Évolutions : {len(ajoutes)} nouveaux, {len(supprimes)} supprimés")

    # ── Onglet Configuration : véhicules figés et tournées gelées ───────────
    ws_cfg = wb.create_sheet('Configuration')

    # Titre
    ws_cfg.merge_cells('A1:E1')
    cell = ws_cfg.cell(row=1, column=1, value='CONFIGURATION — VÉHICULES FIGÉS ET TOURNÉES GELÉES')
    cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    cell.fill = C_HDR
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_cfg.row_dimensions[1].height = 28

    # Section véhicules figés
    ws_cfg.cell(row=3, column=1, value='VÉHICULES FIGÉS').font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    ws_cfg.cell(row=3, column=1).fill = PatternFill("solid", fgColor="C65C1E")
    ws_cfg.cell(row=3, column=1).alignment = A_C
    ws_cfg.merge_cells('A3:E3')
    ws_cfg.cell(row=4, column=1, value='Véhicule').font = F_HDR
    ws_cfg.cell(row=4, column=1).fill = C_HDR
    ws_cfg.cell(row=4, column=1).alignment = A_C
    ws_cfg.cell(row=4, column=2, value='Rôle').font = F_HDR
    ws_cfg.cell(row=4, column=2).fill = C_HDR
    ws_cfg.cell(row=4, column=2).alignment = A_C
    ws_cfg.merge_cells('B4:E4')

    figes_info = {
        'V007': 'CAMION 3ABI 1',
        'V008': 'CAMION 3ABI 2',
        'V009': 'CAMION SOLIFAIM',
        'V013': 'CAMION BD 1',
        'V023': 'CAMION RIGODON — Super U Voreppe ramené par association',
        'V026': 'Tournée spéciale',
        'V027': 'CAMION DLM 20 m3 1',
        'V028': 'CAMION DLM 20 m3 2 Roussillon',
    }
    fill_fige = PatternFill("solid", fgColor="FFE0CC")
    for i, veh in enumerate(sorted(VEHICULES_FIGES, key=lambda v: int(v[1:])), 5):
        ws_cfg.cell(row=i, column=1, value=veh).fill = fill_fige
        ws_cfg.cell(row=i, column=1).font = F_NRM
        ws_cfg.cell(row=i, column=1).alignment = A_C
        ws_cfg.cell(row=i, column=1).border = BRD
        # Récupérer le vrai nom depuis veh2nom si disponible
        nom_veh = df_t[df_t['Camion']==veh]['Nom camion'].dropna().iloc[0] if len(df_t[df_t['Camion']==veh]['Nom camion'].dropna()) > 0 else figes_info.get(veh, '')
        cell_nom = ws_cfg.cell(row=i, column=2, value=nom_veh)
        cell_nom.fill = fill_fige; cell_nom.font = F_NRM
        cell_nom.alignment = A_L; cell_nom.border = BRD
        ws_cfg.merge_cells(f'B{i}:E{i}')

    # Section tournées gelées
    row_gel = len(VEHICULES_FIGES) + 7
    ws_cfg.merge_cells(f'A{row_gel}:E{row_gel}')
    ws_cfg.cell(row=row_gel, column=1, value='TOURNEES GELEES (contenu non modifiable par l optimisation)').font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    ws_cfg.cell(row=row_gel, column=1).fill = PatternFill("solid", fgColor="1F4E79")
    ws_cfg.cell(row=row_gel, column=1).alignment = A_C

    row_gel += 1
    for col, label in [(1,'Véhicule'),(2,'Demi-journée gelée')]:
        ws_cfg.cell(row=row_gel, column=col, value=label).font = F_HDR
        ws_cfg.cell(row=row_gel, column=col).fill = C_HDR
        ws_cfg.cell(row=row_gel, column=col).alignment = A_C
    ws_cfg.merge_cells(f'B{row_gel}:E{row_gel}')

    # Regrouper par véhicule
    from collections import defaultdict as _dd
    gel_par_veh = _dd(list)
    for veh_g, dj_g in sorted(TOURNEES_GELEES):
        gel_par_veh[veh_g].append(dj_g)

    fill_gel = PatternFill("solid", fgColor="DCE6F1")
    for veh_g in sorted(gel_par_veh, key=lambda v: int(v[1:]) if v[1:].isdigit() else 999):
        row_gel += 1
        djs = ', '.join(sorted(gel_par_veh[veh_g]))
        ws_cfg.cell(row=row_gel, column=1, value=veh_g).fill = fill_gel
        ws_cfg.cell(row=row_gel, column=1).font = F_NRM
        ws_cfg.cell(row=row_gel, column=1).alignment = A_C
        ws_cfg.cell(row=row_gel, column=1).border = BRD
        cell_dj = ws_cfg.cell(row=row_gel, column=2, value=djs)
        cell_dj.fill = fill_gel; cell_dj.font = F_NRM
        cell_dj.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_dj.border = BRD
        ws_cfg.merge_cells(f'B{row_gel}:E{row_gel}')
        ws_cfg.row_dimensions[row_gel].height = None

    for col, w in zip([1,2,3,4,5],[12,60,5,5,5]):
        ws_cfg.column_dimensions[get_column_letter(col)].width = w

    # ── Légende ───────────────────────────────────────────────────────────────
    ws_l = wb.create_sheet('Legende')
    C_CHG_CAM_L = PatternFill("solid", fgColor="FF6B6B")
    C_CHG_DJ_L  = PatternFill("solid", fgColor="FFB347")
    leg = [
        ('Couleur','Signification',C_HDR,F_HDR),
        ('Orange (#FFE699)','Nouveau magasin 2026 (ligne)',C_ORG,F_NRM),
        ('Vert clair','Camion supplémentaire (ligne)',C_VRT,F_NRM),
        ('Rouge clair','Camion figé (ligne)',C_RGE,F_NRM),
        ('Bleu clair','Alternance demi-journée (ligne)',C_BLC,F_NRM),
        ('Rouge (#FF6B6B)','Magasin : changement de camion vs 2025',C_CHG_CAM_L,Font(name='Calibri',size=10,bold=True,color='FFFFFF')),
        ('Orange (#FFB347)','Magasin : changement de demi-journée vs 2025',C_CHG_DJ_L,Font(name='Calibri',size=10,bold=True)),
        ('Violet (#CC99FF)','Nouveau magasin placé hors secteur (à vérifier)',PatternFill("solid", fgColor="CC99FF"),Font(name='Calibri',size=10,bold=True,color='4B0082')),
        ('Texte orange foncé','Secteur : tournée avec exactement 3 secteurs (attention)',C_WHT,Font(name='Calibri',size=10,bold=True,color='C65911')),
        ('Texte rouge foncé','Secteur : tournée avec > 3 secteurs distincts (incohérence)',C_WHT,Font(name='Calibri',size=10,bold=True,color='C00000')),
    ]
    for r, (c1, c2, fill, font) in enumerate(leg, 1):
        for c, v in enumerate([c1, c2], 1):
            cell = ws_l.cell(row=r, column=c, value=v)
            cell.font = font; cell.fill = fill; cell.border = BRD; cell.alignment = A_L
    ws_l.column_dimensions['A'].width = 18
    ws_l.column_dimensions['B'].width = 50

    wb.save(output_path)
    print(f"  → Sauvegardé : {output_path}")

# ─── MAIN ─────────────────────────────────────────────────────────────────────
non_affectes_global = []

class _TeeOutput:
    """Duplique stdout vers un buffer pour capturer le compte rendu."""
    def __init__(self, orig, buf):
        self._orig = orig; self._buf = buf
    def write(self, s):
        self._orig.write(s); self._buf.write(s)
    def flush(self):
        self._orig.flush(); self._buf.flush()

def generer_carte_tournees(df_t, df_mag, args, dossier_resultat):
    """Génère une carte HTML interactive des tournées avec itinéraire routier OSRM."""
    import json, os
    from collections import defaultdict

    FIGES_CT = set(VEHICULES_FIGES)
    DJ_LIST_CT = ['Vendredi Matin', 'Vendredi Apres Midi', 'Samedi Matin', 'Samedi Apres Midi']
    BAI_LAT_CT, BAI_LON_CT = BAI_LAT, BAI_LON

    # Index nom → coordonnées + adresse
    nom2coords_ct, nom2adresse_ct = {}, {}
    for _, r in df_mag.iterrows():
        nom = str(r['Nom']).strip()
        try:
            lat, lon = float(r.get('Latitude', 0)), float(r.get('Longitude', 0))
            if lat and lon: nom2coords_ct[nom] = [lat, lon]
        except: pass
        adr = str(r.get('Adresse', '')).strip()
        vil = str(r.get('Ville', '')).strip()
        cp  = str(r.get('C.P.', '')).strip()
        if adr and vil: nom2adresse_ct[nom] = f'{adr}, {cp} {vil}'

    # Construire les données par DJ → camion
    data_ct = defaultdict(dict)
    for _, row in df_t[df_t['Demi-journee'].isin(DJ_LIST_CT)].iterrows():
        veh = str(row['Camion']).strip()
        dj  = str(row['Demi-journee']).strip()
        if veh in FIGES_CT: continue
        nom_cam = str(row.get('Nom camion', '')).strip()
        if nom_cam in ('nan', ''): nom_cam = ''
        mags = [str(row.get(f'Magasin {k}', '')).strip()
                for k in range(1, 7)
                if str(row.get(f'Magasin {k}', '')).strip() not in ('', 'nan')]
        sec = str(row.get('Secteur', '')).strip()
        if sec == 'nan': sec = ''
        points = []
        for mag in mags:
            pt = {'nom': mag, 'adresse': nom2adresse_ct.get(mag, ''), 'secteur': sec}
            pt['lat'] = nom2coords_ct[mag][0] if mag in nom2coords_ct else BAI_LAT_CT
            pt['lon'] = nom2coords_ct[mag][1] if mag in nom2coords_ct else BAI_LON_CT
            points.append(pt)
        data_ct[dj][veh] = {'nom_cam': nom_cam, 'secteur': sec, 'magasins': points}

    DATA_JSON_CT = json.dumps(data_ct, ensure_ascii=False)
    DJ_JSON_CT   = json.dumps(DJ_LIST_CT, ensure_ascii=False)

    parts = []
    parts.append("""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>BAI 38 — Carte des tourn&#233;es 2026</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/leaflet/1.9.4/leaflet.min.css"/>
<script src="https://cdnjs.cloudflare.com/ajax/libs/leaflet/1.9.4/leaflet.min.js"></script>
<style>
*{box-sizing:border-box;margin:0;padding:0;font-family:'Segoe UI',Arial,sans-serif}
body{display:flex;flex-direction:column;height:100vh;background:#f0f2f5}
header{background:#1F4E79;color:#fff;padding:10px 18px;display:flex;align-items:center;gap:16px;flex-shrink:0;box-shadow:0 2px 6px rgba(0,0,0,.3)}
header h1{font-size:16px;font-weight:600;white-space:nowrap}
.controls{display:flex;gap:10px;align-items:center;flex-wrap:wrap}
select{padding:6px 10px;border-radius:6px;border:none;font-size:13px;background:#fff;color:#1F4E79;font-weight:600;cursor:pointer;min-width:130px}
#btn-gmaps{padding:6px 14px;background:#E8A020;color:#fff;border:none;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;white-space:nowrap}
#btn-gmaps:hover{background:#d4901a}
#btn-gmaps:disabled{background:#888;cursor:default}
.badge{background:#E8A020;color:#fff;border-radius:10px;padding:2px 8px;font-size:12px;font-weight:700}
#container{display:flex;flex:1;overflow:hidden}
#sidebar{width:300px;flex-shrink:0;background:#fff;display:flex;flex-direction:column;border-right:1px solid #ddd;overflow:hidden}
#sidebar-header{padding:12px;background:#D9E1F2;border-bottom:1px solid #c0cbdf}
#sidebar-header h2{font-size:13px;color:#1F4E79;font-weight:700}
#sidebar-header .sub{font-size:11px;color:#555;margin-top:3px}
#mag-list{overflow-y:auto;flex:1}
.mag-item{display:flex;align-items:flex-start;padding:10px 12px;border-bottom:1px solid #f0f0f0;cursor:pointer;transition:background .15s}
.mag-item:hover{background:#f5f8ff}
.mag-item.active{background:#D9E1F2}
.mag-num{width:22px;height:22px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;color:#fff;flex-shrink:0;margin-right:10px;margin-top:1px}
.mag-info{flex:1;min-width:0}
.mag-name{font-size:12px;font-weight:600;color:#1F4E79;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.mag-addr{font-size:11px;color:#666;margin-top:2px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.mag-sec{font-size:10px;color:#888;margin-top:2px}
.bai-item{background:#1F4E79;color:#fff;padding:8px 12px;font-size:12px;font-weight:600;display:flex;align-items:center;gap:8px}
#map{flex:1}
.popup-title{font-weight:700;font-size:13px;color:#1F4E79;margin-bottom:4px}
.popup-addr{font-size:11px;color:#555;margin-bottom:3px}
.popup-sec{font-size:11px;color:#888}
.no-tournee{padding:20px;text-align:center;color:#999;font-size:13px}
</style>
</head>
<body>
<header>
  <h1>&#128666; BAI 38 &#8212; Tourn&#233;es 2026</h1>
  <div class="controls">
    <select id="sel-dj" onchange="onDjChange()">
      <option value="">&#8212; Demi-journ&#233;e &#8212;</option>
    </select>
    <select id="sel-cam" onchange="onCamChange()" disabled>
      <option value="">&#8212; Camion &#8212;</option>
    </select>
    <button id="btn-gmaps" disabled onclick="openGmaps()">&#128506; Google Maps</button>
    <span id="badge-nb" class="badge" style="display:none"></span>
  </div>
</header>
<div id="container">
  <div id="sidebar">
    <div id="sidebar-header" style="display:none">
      <h2 id="sb-title">&#8212;</h2>
      <div class="sub" id="sb-sec">&#8212;</div>
    </div>
    <div id="mag-list"><div class="no-tournee">S&#233;lectionnez une demi-journ&#233;e et un camion</div></div>
  </div>
  <div id="map"></div>
</div>
<script>
""")
    parts.append(f'const DATA = {DATA_JSON_CT};\n')
    parts.append(f'const DJ_LIST = {DJ_JSON_CT};\n')
    parts.append(f'const BAI = [{BAI_LAT_CT}, {BAI_LON_CT}];\n')
    parts.append("""
const COULEURS = ['#e74c3c','#e67e22','#2980b9','#27ae60','#8e44ad',
  '#16a085','#d35400','#2c3e50','#f39c12','#1abc9c',
  '#c0392b','#7f8c8d','#6c5ce7','#00b894','#fd79a8'];

const map = L.map('map').setView(BAI, 11);
L.tileLayer('https://{s}.basemaps.cartocdn.com/rastertiles/voyager/{z}/{x}/{y}{r}.png',
  {attribution:'&copy; OpenStreetMap &copy; CARTO',maxZoom:19,subdomains:'abcd'}).addTo(map);

L.marker(BAI, {icon: L.divIcon({
  className:'',
  html:'<div style="background:#1F4E79;color:#fff;border-radius:50%;width:32px;height:32px;display:flex;align-items:center;justify-content:center;font-size:16px;border:3px solid #fff;box-shadow:0 2px 6px rgba(0,0,0,.4)">&#127981;</div>',
  iconSize:[32,32],iconAnchor:[16,16]
})}).addTo(map).bindPopup('<b>BAI 38</b><br>""")
    parts.append(_adresse_siege_html())
    parts.append("""');

let layerGroup = L.layerGroup().addTo(map);
let currentTournee = null;

const selDj = document.getElementById('sel-dj');
DJ_LIST.forEach(function(dj) {
  var o = document.createElement('option');
  o.value = dj; o.textContent = dj;
  selDj.appendChild(o);
});

function onDjChange() {
  var dj = selDj.value;
  var selCam = document.getElementById('sel-cam');
  selCam.innerHTML = '<option value="">&#8212; Camion &#8212;</option>';
  selCam.disabled = !dj;
  if (!dj) { clearMap(); return; }
  var cams = Object.keys(DATA[dj] || {}).sort();
  cams.forEach(function(veh) {
    var t = DATA[dj][veh];
    var label = t.nom_cam ? (veh + ' \u2014 ' + t.nom_cam) : veh;
    var o = document.createElement('option');
    o.value = veh; o.textContent = label;
    selCam.appendChild(o);
  });
  clearMap();
}

function onCamChange() {
  var dj = selDj.value;
  var veh = document.getElementById('sel-cam').value;
  if (!dj || !veh) { clearMap(); return; }
  afficherTournee(dj, veh);
}

function clearMap() {
  layerGroup.clearLayers();
  currentTournee = null;
  document.getElementById('mag-list').innerHTML = '<div class="no-tournee">S\u00e9lectionnez une demi-journ\u00e9e et un camion</div>';
  document.getElementById('sidebar-header').style.display = 'none';
  document.getElementById('badge-nb').style.display = 'none';
  document.getElementById('btn-gmaps').disabled = true;
}

function makeIcon(num, color) {
  return L.divIcon({
    className:'',
    html:'<div style="background:'+color+';color:#fff;border-radius:50%;width:28px;height:28px;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;border:2px solid #fff;box-shadow:0 2px 5px rgba(0,0,0,.35)">'+num+'</div>',
    iconSize:[28,28],iconAnchor:[14,14],popupAnchor:[0,-14]
  });
}

function afficherTournee(dj, veh) {
  layerGroup.clearLayers();
  var t = DATA[dj][veh];
  var mags = t.magasins;
  currentTournee = {dj:dj, veh:veh, mags:mags};
  var cams = Object.keys(DATA[dj]).sort();
  var color = COULEURS[cams.indexOf(veh) % COULEURS.length];

  document.getElementById('sb-title').textContent = t.nom_cam ? (veh + ' \u2014 ' + t.nom_cam) : veh;
  document.getElementById('sb-sec').textContent = t.secteur || '';
  document.getElementById('sidebar-header').style.display = 'block';
  document.getElementById('badge-nb').style.display = 'inline';
  document.getElementById('badge-nb').textContent = mags.length + ' mag.';
  document.getElementById('btn-gmaps').disabled = false;

  // Ligne droite provisoire pendant chargement OSRM
  var latlngsProvis = [BAI];
  mags.forEach(function(m) { latlngsProvis.push([m.lat, m.lon]); });
  latlngsProvis.push(BAI);
  var polyProvis = L.polyline(latlngsProvis, {color:color, weight:3, opacity:0.4, dashArray:'8,6'}).addTo(layerGroup);

  // Itinéraire routier via OSRM
  var waypoints = [BAI[1]+','+BAI[0]];
  mags.forEach(function(m) { waypoints.push(m.lon+','+m.lat); });
  waypoints.push(BAI[1]+','+BAI[0]);
  var osrmUrl = 'https://router.project-osrm.org/route/v1/driving/' + waypoints.join(';')
    + '?overview=full&geometries=geojson';

  fetch(osrmUrl)
    .then(function(r) { return r.json(); })
    .then(function(data) {
      if (data.routes && data.routes[0]) {
        layerGroup.removeLayer(polyProvis);
        var coords = data.routes[0].geometry.coordinates.map(function(c) { return [c[1],c[0]]; });
        var dist_km = (data.routes[0].distance/1000).toFixed(1);
        var dur_min = Math.round(data.routes[0].duration/60);
        L.polyline(coords, {color:color, weight:4, opacity:0.85}).addTo(layerGroup);
        document.getElementById('sb-sec').textContent =
          (t.secteur||'') + ' \u2014 ' + dist_km + ' km, ' + dur_min + ' min';
      }
    })
    .catch(function() { polyProvis.setStyle({opacity:0.7, dashArray:''}); });

  // Marqueurs numérotés
  var markers = [];
  mags.forEach(function(mag, i) {
    var marker = L.marker([mag.lat, mag.lon], {icon: makeIcon(i+1, color)})
      .bindPopup(
        '<div class="popup-title">'+(i+1)+'. '+mag.nom+'</div>'+
        '<div class="popup-addr">'+(mag.adresse||'')+'</div>'+
        '<div class="popup-sec">'+(mag.secteur||'')+'</div>'
      ).addTo(layerGroup);
    markers.push(marker);
    marker.on('click', function() { highlightMag(i); });
  });

  // Sidebar
  var list = document.getElementById('mag-list');
  list.innerHTML = '';
  var baiD = document.createElement('div');
  baiD.className = 'bai-item';
  baiD.innerHTML = '&#127981; BAI &mdash; Fontaine (d&eacute;part)';
  list.appendChild(baiD);
  mags.forEach(function(mag, i) {
    var item = document.createElement('div');
    item.className = 'mag-item';
    item.id = 'mag-'+i;
    item.innerHTML =
      '<div class="mag-num" style="background:'+color+'">'+(i+1)+'</div>'+
      '<div class="mag-info">'+
        '<div class="mag-name">'+mag.nom+'</div>'+
        '<div class="mag-addr">'+(mag.adresse||'&mdash;')+'</div>'+
        '<div class="mag-sec">'+(mag.secteur||'')+'</div>'+
      '</div>';
    (function(idx, m) {
      item.onclick = function() { markers[idx].openPopup(); highlightMag(idx); map.panTo([m.lat, m.lon]); };
    })(i, mag);
    list.appendChild(item);
  });
  var baiA = document.createElement('div');
  baiA.className = 'bai-item';
  baiA.innerHTML = '&#127981; BAI &mdash; Fontaine (arriv&eacute;e)';
  list.appendChild(baiA);

  var bounds = L.latLngBounds([BAI]);
  mags.forEach(function(m) { bounds.extend([m.lat, m.lon]); });
  map.fitBounds(bounds.pad(0.15));
}

function highlightMag(idx) {
  document.querySelectorAll('.mag-item').forEach(function(el, i) {
    el.classList.toggle('active', i === idx);
  });
  var el = document.getElementById('mag-'+idx);
  if (el) el.scrollIntoView({block:'nearest',behavior:'smooth'});
}

function openGmaps() {
  if (!currentTournee) return;
  var mags = currentTournee.mags;
  var bai = '45.18867,5.68456';
  var pts = [bai];
  mags.forEach(function(m) { pts.push(m.lat+','+m.lon); });
  pts.push(bai);
  window.open('https://www.google.com/maps/dir/'+pts.join('/')+'/data=!4m2!4m1!3e0', '_blank');
}
</script>
</body>
</html>""")

    html_out = ''.join(parts)
    if dossier_resultat:
        out_dir = dossier_resultat
    else:
        out_dir = os.path.dirname(os.path.abspath(args.output)) if args.output else '.'
    os.makedirs(out_dir, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    html_path = os.path.join(out_dir, f'carte_tournees_bai38_{ts}.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_out)
    print(f"  Carte HTML générée : {html_path}")
    return html_path


def generer_carte_tournees(df_t, df_mag, args, dossier_resultat):
    """Génère une carte HTML interactive des tournées avec itinéraire routier OSRM."""
    import json, os
    from collections import defaultdict

    FIGES_CT = set(VEHICULES_FIGES)
    DJ_LIST_CT = ['Vendredi Matin', 'Vendredi Apres Midi', 'Samedi Matin', 'Samedi Apres Midi']
    BAI_LAT_CT, BAI_LON_CT = BAI_LAT, BAI_LON

    # Index nom → coordonnées + adresse
    nom2coords_ct, nom2adresse_ct = {}, {}
    for _, r in df_mag.iterrows():
        nom = str(r['Nom']).strip()
        try:
            lat, lon = float(r.get('Latitude', 0)), float(r.get('Longitude', 0))
            if lat and lon: nom2coords_ct[nom] = [lat, lon]
        except: pass
        adr = str(r.get('Adresse', '')).strip()
        vil = str(r.get('Ville', '')).strip()
        cp  = str(r.get('C.P.', '')).strip()
        if adr and vil: nom2adresse_ct[nom] = f'{adr}, {cp} {vil}'

    # Construire les données par DJ → camion
    data_ct = defaultdict(dict)
    for _, row in df_t[df_t['Demi-journee'].isin(DJ_LIST_CT)].iterrows():
        veh = str(row['Camion']).strip()
        dj  = str(row['Demi-journee']).strip()
        if veh in FIGES_CT: continue
        nom_cam = str(row.get('Nom camion', '')).strip()
        if nom_cam in ('nan', ''): nom_cam = ''
        mags = [str(row.get(f'Magasin {k}', '')).strip()
                for k in range(1, 7)
                if str(row.get(f'Magasin {k}', '')).strip() not in ('', 'nan')]
        sec = str(row.get('Secteur', '')).strip()
        if sec == 'nan': sec = ''
        points = []
        for mag in mags:
            pt = {'nom': mag, 'adresse': nom2adresse_ct.get(mag, ''), 'secteur': sec}
            pt['lat'] = nom2coords_ct[mag][0] if mag in nom2coords_ct else BAI_LAT_CT
            pt['lon'] = nom2coords_ct[mag][1] if mag in nom2coords_ct else BAI_LON_CT
            points.append(pt)
        data_ct[dj][veh] = {'nom_cam': nom_cam, 'secteur': sec, 'magasins': points}

    DATA_JSON_CT = json.dumps(data_ct, ensure_ascii=False)
    DJ_JSON_CT   = json.dumps(DJ_LIST_CT, ensure_ascii=False)

    parts = []
    parts.append("""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>BAI 38 — Carte des tourn&#233;es 2026</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/leaflet/1.9.4/leaflet.min.css"/>
<script src="https://cdnjs.cloudflare.com/ajax/libs/leaflet/1.9.4/leaflet.min.js"></script>
<style>
*{box-sizing:border-box;margin:0;padding:0;font-family:'Segoe UI',Arial,sans-serif}
body{display:flex;flex-direction:column;height:100vh;background:#f0f2f5}
header{background:#1F4E79;color:#fff;padding:10px 18px;display:flex;align-items:center;gap:16px;flex-shrink:0;box-shadow:0 2px 6px rgba(0,0,0,.3)}
header h1{font-size:16px;font-weight:600;white-space:nowrap}
.controls{display:flex;gap:10px;align-items:center;flex-wrap:wrap}
select{padding:6px 10px;border-radius:6px;border:none;font-size:13px;background:#fff;color:#1F4E79;font-weight:600;cursor:pointer;min-width:130px}
#btn-gmaps{padding:6px 14px;background:#E8A020;color:#fff;border:none;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;white-space:nowrap}
#btn-gmaps:hover{background:#d4901a}
#btn-gmaps:disabled{background:#888;cursor:default}
.badge{background:#E8A020;color:#fff;border-radius:10px;padding:2px 8px;font-size:12px;font-weight:700}
#container{display:flex;flex:1;overflow:hidden}
#sidebar{width:300px;flex-shrink:0;background:#fff;display:flex;flex-direction:column;border-right:1px solid #ddd;overflow:hidden}
#sidebar-header{padding:12px;background:#D9E1F2;border-bottom:1px solid #c0cbdf}
#sidebar-header h2{font-size:13px;color:#1F4E79;font-weight:700}
#sidebar-header .sub{font-size:11px;color:#555;margin-top:3px}
#mag-list{overflow-y:auto;flex:1}
.mag-item{display:flex;align-items:flex-start;padding:10px 12px;border-bottom:1px solid #f0f0f0;cursor:pointer;transition:background .15s}
.mag-item:hover{background:#f5f8ff}
.mag-item.active{background:#D9E1F2}
.mag-num{width:22px;height:22px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;color:#fff;flex-shrink:0;margin-right:10px;margin-top:1px}
.mag-info{flex:1;min-width:0}
.mag-name{font-size:12px;font-weight:600;color:#1F4E79;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.mag-addr{font-size:11px;color:#666;margin-top:2px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.mag-sec{font-size:10px;color:#888;margin-top:2px}
.bai-item{background:#1F4E79;color:#fff;padding:8px 12px;font-size:12px;font-weight:600;display:flex;align-items:center;gap:8px}
#map{flex:1}
.popup-title{font-weight:700;font-size:13px;color:#1F4E79;margin-bottom:4px}
.popup-addr{font-size:11px;color:#555;margin-bottom:3px}
.popup-sec{font-size:11px;color:#888}
.no-tournee{padding:20px;text-align:center;color:#999;font-size:13px}
</style>
</head>
<body>
<header>
  <h1>&#128666; BAI 38 &#8212; Tourn&#233;es 2026</h1>
  <div class="controls">
    <select id="sel-dj" onchange="onDjChange()">
      <option value="">&#8212; Demi-journ&#233;e &#8212;</option>
    </select>
    <select id="sel-cam" onchange="onCamChange()" disabled>
      <option value="">&#8212; Camion &#8212;</option>
    </select>
    <button id="btn-gmaps" disabled onclick="openGmaps()">&#128506; Google Maps</button>
    <span id="badge-nb" class="badge" style="display:none"></span>
  </div>
</header>
<div id="container">
  <div id="sidebar">
    <div id="sidebar-header" style="display:none">
      <h2 id="sb-title">&#8212;</h2>
      <div class="sub" id="sb-sec">&#8212;</div>
    </div>
    <div id="mag-list"><div class="no-tournee">S&#233;lectionnez une demi-journ&#233;e et un camion</div></div>
  </div>
  <div id="map"></div>
</div>
<script>
""")
    parts.append(f'const DATA = {DATA_JSON_CT};\n')
    parts.append(f'const DJ_LIST = {DJ_JSON_CT};\n')
    parts.append(f'const BAI = [{BAI_LAT_CT}, {BAI_LON_CT}];\n')
    parts.append("""
const COULEURS = ['#e74c3c','#e67e22','#2980b9','#27ae60','#8e44ad',
  '#16a085','#d35400','#2c3e50','#f39c12','#1abc9c',
  '#c0392b','#7f8c8d','#6c5ce7','#00b894','#fd79a8'];

const map = L.map('map').setView(BAI, 11);
L.tileLayer('https://{s}.basemaps.cartocdn.com/rastertiles/voyager/{z}/{x}/{y}{r}.png',
  {attribution:'&copy; OpenStreetMap &copy; CARTO',maxZoom:19,subdomains:'abcd'}).addTo(map);

L.marker(BAI, {icon: L.divIcon({
  className:'',
  html:'<div style="background:#1F4E79;color:#fff;border-radius:50%;width:32px;height:32px;display:flex;align-items:center;justify-content:center;font-size:16px;border:3px solid #fff;box-shadow:0 2px 6px rgba(0,0,0,.4)">&#127981;</div>',
  iconSize:[32,32],iconAnchor:[16,16]
})}).addTo(map).bindPopup('<b>BAI 38</b><br>""")
    parts.append(_adresse_siege_html())
    parts.append("""');

let layerGroup = L.layerGroup().addTo(map);
let currentTournee = null;

const selDj = document.getElementById('sel-dj');
DJ_LIST.forEach(function(dj) {
  var o = document.createElement('option');
  o.value = dj; o.textContent = dj;
  selDj.appendChild(o);
});

function onDjChange() {
  var dj = selDj.value;
  var selCam = document.getElementById('sel-cam');
  selCam.innerHTML = '<option value="">&#8212; Camion &#8212;</option>';
  selCam.disabled = !dj;
  if (!dj) { clearMap(); return; }
  var cams = Object.keys(DATA[dj] || {}).sort();
  cams.forEach(function(veh) {
    var t = DATA[dj][veh];
    var label = t.nom_cam ? (veh + ' \u2014 ' + t.nom_cam) : veh;
    var o = document.createElement('option');
    o.value = veh; o.textContent = label;
    selCam.appendChild(o);
  });
  clearMap();
}

function onCamChange() {
  var dj = selDj.value;
  var veh = document.getElementById('sel-cam').value;
  if (!dj || !veh) { clearMap(); return; }
  afficherTournee(dj, veh);
}

function clearMap() {
  layerGroup.clearLayers();
  currentTournee = null;
  document.getElementById('mag-list').innerHTML = '<div class="no-tournee">S\u00e9lectionnez une demi-journ\u00e9e et un camion</div>';
  document.getElementById('sidebar-header').style.display = 'none';
  document.getElementById('badge-nb').style.display = 'none';
  document.getElementById('btn-gmaps').disabled = true;
}

function makeIcon(num, color) {
  return L.divIcon({
    className:'',
    html:'<div style="background:'+color+';color:#fff;border-radius:50%;width:28px;height:28px;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;border:2px solid #fff;box-shadow:0 2px 5px rgba(0,0,0,.35)">'+num+'</div>',
    iconSize:[28,28],iconAnchor:[14,14],popupAnchor:[0,-14]
  });
}

function afficherTournee(dj, veh) {
  layerGroup.clearLayers();
  var t = DATA[dj][veh];
  var mags = t.magasins;
  currentTournee = {dj:dj, veh:veh, mags:mags};
  var cams = Object.keys(DATA[dj]).sort();
  var color = COULEURS[cams.indexOf(veh) % COULEURS.length];

  document.getElementById('sb-title').textContent = t.nom_cam ? (veh + ' \u2014 ' + t.nom_cam) : veh;
  document.getElementById('sb-sec').textContent = t.secteur || '';
  document.getElementById('sidebar-header').style.display = 'block';
  document.getElementById('badge-nb').style.display = 'inline';
  document.getElementById('badge-nb').textContent = mags.length + ' mag.';
  document.getElementById('btn-gmaps').disabled = false;

  // Ligne droite provisoire pendant chargement OSRM
  var latlngsProvis = [BAI];
  mags.forEach(function(m) { latlngsProvis.push([m.lat, m.lon]); });
  latlngsProvis.push(BAI);
  var polyProvis = L.polyline(latlngsProvis, {color:color, weight:3, opacity:0.4, dashArray:'8,6'}).addTo(layerGroup);

  // Itinéraire routier via OSRM
  var waypoints = [BAI[1]+','+BAI[0]];
  mags.forEach(function(m) { waypoints.push(m.lon+','+m.lat); });
  waypoints.push(BAI[1]+','+BAI[0]);
  var osrmUrl = 'https://router.project-osrm.org/route/v1/driving/' + waypoints.join(';')
    + '?overview=full&geometries=geojson';

  fetch(osrmUrl)
    .then(function(r) { return r.json(); })
    .then(function(data) {
      if (data.routes && data.routes[0]) {
        layerGroup.removeLayer(polyProvis);
        var coords = data.routes[0].geometry.coordinates.map(function(c) { return [c[1],c[0]]; });
        var dist_km = (data.routes[0].distance/1000).toFixed(1);
        var dur_min = Math.round(data.routes[0].duration/60);
        L.polyline(coords, {color:color, weight:4, opacity:0.85}).addTo(layerGroup);
        document.getElementById('sb-sec').textContent =
          (t.secteur||'') + ' \u2014 ' + dist_km + ' km, ' + dur_min + ' min';
      }
    })
    .catch(function() { polyProvis.setStyle({opacity:0.7, dashArray:''}); });

  // Marqueurs numérotés
  var markers = [];
  mags.forEach(function(mag, i) {
    var marker = L.marker([mag.lat, mag.lon], {icon: makeIcon(i+1, color)})
      .bindPopup(
        '<div class="popup-title">'+(i+1)+'. '+mag.nom+'</div>'+
        '<div class="popup-addr">'+(mag.adresse||'')+'</div>'+
        '<div class="popup-sec">'+(mag.secteur||'')+'</div>'
      ).addTo(layerGroup);
    markers.push(marker);
    marker.on('click', function() { highlightMag(i); });
  });

  // Sidebar
  var list = document.getElementById('mag-list');
  list.innerHTML = '';
  var baiD = document.createElement('div');
  baiD.className = 'bai-item';
  baiD.innerHTML = '&#127981; BAI &mdash; Fontaine (d&eacute;part)';
  list.appendChild(baiD);
  mags.forEach(function(mag, i) {
    var item = document.createElement('div');
    item.className = 'mag-item';
    item.id = 'mag-'+i;
    item.innerHTML =
      '<div class="mag-num" style="background:'+color+'">'+(i+1)+'</div>'+
      '<div class="mag-info">'+
        '<div class="mag-name">'+mag.nom+'</div>'+
        '<div class="mag-addr">'+(mag.adresse||'&mdash;')+'</div>'+
        '<div class="mag-sec">'+(mag.secteur||'')+'</div>'+
      '</div>';
    (function(idx, m) {
      item.onclick = function() { markers[idx].openPopup(); highlightMag(idx); map.panTo([m.lat, m.lon]); };
    })(i, mag);
    list.appendChild(item);
  });
  var baiA = document.createElement('div');
  baiA.className = 'bai-item';
  baiA.innerHTML = '&#127981; BAI &mdash; Fontaine (arriv&eacute;e)';
  list.appendChild(baiA);

  var bounds = L.latLngBounds([BAI]);
  mags.forEach(function(m) { bounds.extend([m.lat, m.lon]); });
  map.fitBounds(bounds.pad(0.15));
}

function highlightMag(idx) {
  document.querySelectorAll('.mag-item').forEach(function(el, i) {
    el.classList.toggle('active', i === idx);
  });
  var el = document.getElementById('mag-'+idx);
  if (el) el.scrollIntoView({block:'nearest',behavior:'smooth'});
}

function openGmaps() {
  if (!currentTournee) return;
  var mags = currentTournee.mags;
  var bai = '45.18867,5.68456';
  var pts = [bai];
  mags.forEach(function(m) { pts.push(m.lat+','+m.lon); });
  pts.push(bai);
  window.open('https://www.google.com/maps/dir/'+pts.join('/')+'/data=!4m2!4m1!3e0', '_blank');
}
</script>
</body>
</html>""")

    html_out = ''.join(parts)
    if dossier_resultat:
        out_dir = dossier_resultat
    else:
        out_dir = os.path.dirname(os.path.abspath(args.output)) if args.output else '.'
    os.makedirs(out_dir, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    html_path = os.path.join(out_dir, f'carte_tournees_bai38_{ts}.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_out)
    print(f"  Carte HTML générée : {html_path}")
    return html_path


def main():
    import io, sys as _sys
    _log_buffer = io.StringIO()
    _tee = _TeeOutput(_sys.stdout, _log_buffer)
    _sys.stdout = _tee
    args = parse_args()

    if args.output is None:
        # Chaque simulation (= une combinaison de paramètres) va dans son propre
        # sous-dossier de resultat/simulation tournées/, nommé d'après ses
        # paramètres (ex: vx3-mag5) — pour comparer facilement plusieurs essais
        # sans que l'un écrase le résultat d'un autre.
        # Le nom du sous-dossier ne reprend QUE camions-supp et max-magasins (pas
        # les options optanciens/fuslegeres/cormalplaces/poids) : sur un chemin
        # Google Drive déjà long, un nom de dossier trop détaillé peut dépasser
        # la limite de chemin d'Excel (~218 caractères) et donner un faux
        # "fichier introuvable" alors que le fichier existe bel et bien.
        dossier_resultat_base = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                              'resultat', 'simulation tournées')
        sous_dossier = f'vx{args.camions_supp}-mag{args.max_magasins}'
        dossier_resultat = os.path.join(dossier_resultat_base, sous_dossier)
        os.makedirs(dossier_resultat, exist_ok=True)
        args.output = os.path.join(dossier_resultat, (lambda s: f'Tournees_BAI38_2026_{datetime.now().strftime("%Y%m%d_%H%M")}_VX{args.camions_supp}{s}.xlsx')(
            ('_OptAnciens' if args.optimiser_anciens else '') +
            ('_FusLegeres' if args.fusionner_legeres else '') +
            ('_CorMalPlaces' if args.corriger_mal_places else '')))
    else:
        # Sortie explicitement fournie (--output) : la carte HTML (générée plus
        # loin) va dans le même dossier que ce fichier Excel.
        dossier_resultat = os.path.dirname(os.path.abspath(args.output))

    print()
    print('='*65)
    print('  GÉNÉRATION TOURNÉES BAI 38 - 2026')
    print('='*65)
    print(f'  Camions supplémentaires : {args.camions_supp}')
    print(f'  Poids nouveaux magasins : {args.poids_nouveaux} kg')
    print(f'  Max magasins/tournée    : {args.max_magasins} (dimanche: {args.max_magasins+1})')
    print(f'  Fichier PDF source      : {args.pdf}')
    print(f'  Fichier magasins        : {args.magasins}')
    print(f'  Nouveaux magasins       : détectés automatiquement (absents du PDF 2025)')
    print('='*65)
    print()

    # Trouver le PDF
    if not os.path.exists(args.pdf):
        candidats = glob.glob('fiches jour*.pdf') + glob.glob('fiches_jour*.pdf')
        if candidats:
            args.pdf = candidats[0]
            print(f"  INFO: PDF détecté : {args.pdf}")
        else:
            print(f"ERREUR: PDF non trouvé : {args.pdf}")
            print(f"  PDF disponibles : {glob.glob('*.pdf') or ['(aucun)']}")
            sys.exit(1)

    if not os.path.exists(args.magasins):
        print(f"ERREUR: {args.magasins} non trouvé.")
        sys.exit(1)

    try:
        import pandas, openpyxl
    except ImportError as e:
        print(f"ERREUR: {e} — pip install pandas openpyxl")
        sys.exit(1)

    # Modes debug
    if args.debug_pdf:
        pages = lire_pages_pdf(args.pdf)
        for i, p in enumerate(pages[:3]):
            print(f"\n{'='*50}\n  PAGE {i+1}\n{'='*50}")
            print(p[:800])
        sys.exit(0)

    if args.debug_vides:
        pages = lire_pages_pdf(args.pdf)
        nb = 0
        for i, p in enumerate(pages):
            lines = [l.strip() for l in p.split('\n') if l.strip()]
            blocs = parse_page(lines)
            if not blocs and len(lines) >= 3:
                print(f"\n{'='*50}  PAGE {i+1} (vide)  {'='*50}")
                for j, l in enumerate(lines[:20]):
                    print(f"  {j:2d}: {repr(l)}")
                nb += 1
                if nb >= 3:
                    break
        print(f"\n  {nb} pages vides affichées. Arrêt.")
        sys.exit(0)

    if args.debug_pages:
        pages = lire_pages_pdf(args.pdf)
        print(f"  Total pages: {len(pages)}")
        print(f"  {'PAGE':>4}  {'DEMI-JOURNEE':<22}  {'VEHICULE':<8}  {'NB_VIF':>6}")
        print(f"  {'-'*50}")
        for i, p in enumerate(pages):
            lines = [l.strip() for l in p.split('\n') if l.strip()]
            blocs = parse_page(lines)
            if blocs:
                for dj, veh, vifs in blocs:
                    print(f"  {i+1:>4}  {dj:<22}  {veh:<8}  {len(vifs):>6}")
            else:
                print(f"  {i+1:>4}  {'':<22}  {'':<8}  {0:>6}")
        sys.exit(0)

    print("ÉTAPE 1 — Extraction du PDF...")
    fiches = extraire_pdf(args.pdf)

    print("\nÉTAPE 2 — Lecture des magasins...")
    # Construire l'ensemble des VIF connus du PDF 2025 (normalisés sans zéro initial)
    vifs_pdf_2025 = set()
    for f in fiches:
        for v in f['vif_codes']:
            vifs_pdf_2025.add(str(v).lstrip('0'))
    print(f"  → {len(vifs_pdf_2025)} codes VIF distincts dans le PDF 2025")
    df_mag = lire_magasins(args.magasins, vifs_pdf_2025, args.poids_nouveaux)

    print("\nÉTAPE 3 — Optimisation des tournées...")
    df_t = optimiser_tournees(fiches, df_mag, args)
    print(f"  → {len(df_t)} tournées, {df_t['Camion'].nunique()} camions")
    from collections import Counter
    for dj, n in sorted(Counter(df_t['Demi-journee']).items()):
        print(f"      {dj}: {n}")

    print("\nÉTAPE 4 — Génération du fichier Excel...")
    # Restaurer stdout et récupérer le log
    _sys.stdout = _tee._orig
    log_contenu = _log_buffer.getvalue()
    generer_excel(df_t, df_mag, args, args.output, fiches, non_affectes_global, log_contenu)



    # Génération de la carte HTML interactive des tournées
    print()
    print('  Génération de la carte HTML des tournées...')
    generer_carte_tournees(df_t, df_mag, args, dossier_resultat)


    # ── Contrôle final : magasins actifs absents de certaines DJ ─────────────
    print()
    print("  Contrôle couverture des magasins...")
    DJ_VS_CTRL = ['Vendredi Matin', 'Vendredi Apres Midi', 'Samedi Matin', 'Samedi Apres Midi']

    # Reconstruire la présence depuis df_t
    presence = {dj: set() for dj in DJ_VS_CTRL}
    for _, row in df_t[df_t['Demi-journee'].isin(DJ_VS_CTRL)].iterrows():
        dj = row['Demi-journee']
        for k in range(1, 7):
            mag = str(row.get(f'Magasin {k}', '')).strip()
            if mag and mag != 'nan':
                presence[dj].add(mag)

    # Magasins uniquement Jeudi (PDF 2025) — via df_mag et onglet Tournées 2025
    # Un magasin est "Jeudi only" s'il est dans le PDF 2025 mais PAS sur une DJ V/S
    vifs_jeudi = set()
    vifs_vs = set()
    for fiche in fiches:
        dj_f = fiche.get('demi_journee', '')
        for vif in fiche.get('vif_codes', []):
            if 'Jeudi' in dj_f:
                vifs_jeudi.add(vif)
            elif any(d in dj_f for d in ['Vendredi','Samedi']):
                vifs_vs.add(vif)
    vifs_jeudi_only = vifs_jeudi - vifs_vs
    # Convertir en noms via df_mag
    vif2nom = {str(r['Code VIF']).strip().lstrip('0'): str(r['Nom']).strip()
               for _, r in df_mag.iterrows()}
    noms_jeudi_only = {vif2nom[v] for v in vifs_jeudi_only if v in vif2nom}

    # Tous les magasins actifs depuis df_mag
    tous_actifs = set(df_mag['Nom'].astype(str).str.strip().tolist())

    nb_alertes = 0
    for dj in DJ_VS_CTRL:
        absents = []
        for nom in sorted(tous_actifs):
            if not nom or nom == 'nan': continue
            if nom in noms_jeudi_only: continue
            if nom not in presence[dj]:
                autres = [d[:3] for d in DJ_VS_CTRL if nom in presence[d]]
                statut = ', '.join(autres) if autres else 'AUCUNE DJ V/S'
                absents.append((nom, statut))
        if absents:
            nb_alertes += len(absents)
            print(f"  [ALERTE] {dj} — {len(absents)} magasin(s) manquant(s) :")
            for nom, statut in absents:
                print(f"    ✗ {nom} (présent sur : {statut})")

    if nb_alertes == 0:
        print("  ✓ Tous les magasins actifs sont couverts sur les 4 DJ Vendredi/Samedi")
    else:
        print(f"  → {nb_alertes} alerte(s) au total — vérifier l\'onglet Magasins")

    print()
    print('='*65)
    print('  TERMINÉ')
    print(f'  Fichier  : {args.output}')
    print(f'  Tournées : {len(df_t)}')
    nb_new = int(df_mag["Nouveau"].sum())
    print(f'  Magasins : {len(df_mag)} ({nb_new} nouveaux détectés automatiquement)')
    print('='*65)
    print()

if __name__ == '__main__':
    main()