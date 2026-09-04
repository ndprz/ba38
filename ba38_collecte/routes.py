# =========================================
# 🧺 Module Collecte — tournées de collecte annuelle
# =========================================
# Reprend dans Basilic l'outillage jusqu'ici lancé "à la main" via des
# scripts Python + fichiers .bat (generer_tournees_bai_v2.py,
# Generer_documents_bai38.py, Generer_fiches_2025.py,
# generer_carte_secteurs.py — voir dev/uploads/collecte_fichiers_source).
#
# Étape 1 : page principale du module avec l'upload des fichiers nécessaires
# pour une campagne (année) de collecte donnée :
#   - liste des magasins (export go-on-web de l'année en cours)
#   - PDF des tournées de la collecte précédente (dossier camions du drive
#     collecte), utilisé comme point de départ par l'algorithme d'optimisation
#     — les camions (codes + noms) en sont aussi extraits, pas besoin d'un
#     fichier véhicules séparé pour la simulation
#
# Étape 2 : génération des tournées, en réutilisant telle quelle la logique de
# generer_tournees_bai_v2.py (copiée dans collecte_moteur_tournees.py) plutôt
# que de la réécrire — script déjà validé sur plusieurs campagnes réelles.
# Affichage du résultat dans l'appli + export Excel (fichier natif du script,
# identique à celui produit par l'outil historique) et PDF (rendu du même
# tableau via weasyprint).
#
# Carte des secteurs (generer_carte_secteurs.py, copiée en version allégée
# dans collecte_moteur_carte_secteurs.py) générée automatiquement à chaque
# version de tournées, à partir du même référentiel magasins (donc des mêmes
# secteurs) que la génération en question.
#
# Analyse comparative multi-scénarios (8 configurations camions-supp × max-
# magasins) : lancée en arrière-plan (Thread, cf. ba38_participation.py pour
# le même pattern), stockage temporaire des seuls indicateurs chiffrés en
# JSON dans collecte_analyses — pas de fichier Excel/carte généré pour ces
# 8 scénarios (ce serait 8x plus de travail pour un résultat jetable).
#
# Étape suivante (documents/fiches équipiers) : pas encore ajoutée.

import io
import os
import re
import glob
import json
import copy
import sqlite3
import argparse
import subprocess
from datetime import datetime
from threading import Thread

import anthropic
import markdown
import requests
import pandas as pd
from docx import Document
from openpyxl import load_workbook
from flask import (
    render_template, request, redirect, url_for, flash,
    current_app, send_file
)
from flask_login import login_required, current_user
from weasyprint import HTML
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.lib.utils import ImageReader

from ba38_utilitaires.core import (
    get_db_connection, get_db_path, require_access, write_log, date_fr,
    envoyer_mail, render_modele_email,
)
from ba38_utilitaires.organisation import get_organisation
from ba38_collecte import collecte_bp
from ba38_collecte import moteur_tournees as moteur
from ba38_collecte import moteur_carte_secteurs as carte_secteurs

EXTENSIONS_EXCEL = {".xlsx", ".xls"}
EXTENSIONS_PDF = {".pdf"}
MODELES_GARDEE_DIR = "/srv/ba38/uploads/collecte_fichiers_source"

# Un type de fichier = une colonne (chemin/date/auteur) dans collecte_campagnes
# + un nom de stockage fixe (les scripts de génération, ajoutés dans une
# prochaine itération, s'appuieront sur ce nommage plutôt que sur le nom
# d'origine du fichier importé).
FICHIERS = {
    "magasins": {
        "champ_chemin": "fichier_magasins",
        "champ_le": "fichier_magasins_le",
        "champ_par": "fichier_magasins_par",
        "nom_stockage": "liste_magasins.xlsx",
        "extensions": EXTENSIONS_EXCEL,
        "label": "Liste des magasins",
        "aide": "Export go-on-web de l'année en cours",
    },
    "pdf_precedent": {
        "champ_chemin": "fichier_pdf_precedent",
        "champ_le": "fichier_pdf_precedent_le",
        "champ_par": "fichier_pdf_precedent_par",
        "nom_stockage": "tournees_precedentes.pdf",
        "extensions": EXTENSIONS_PDF,
        "label": "Tournées de la collecte précédente",
        "aide": "PDF \"fiches jour-véhicules-magasins\", dossier camions de la collecte précédente sur le drive collecte",
    },
    "groupes": {
        "champ_chemin": "fichier_groupes",
        "champ_le": "fichier_groupes_le",
        "champ_par": "fichier_groupes_par",
        "nom_stockage": "liste_groupes.xlsx",
        "extensions": EXTENSIONS_EXCEL,
        "label": "Liste des groupes (associations go-on-web)",
        "aide": "Export go-on-web Association > Groupes — utilisé pour identifier les associations qui gardent leur collecte",
    },
    "participants": {
        "champ_chemin": "fichier_participants",
        "champ_le": "fichier_participants_le",
        "champ_par": "fichier_participants_par",
        "nom_stockage": "liste_participants.xlsx",
        "extensions": EXTENSIONS_EXCEL,
        "label": "Liste des participants (contacts go-on-web)",
        "aide": "Export go-on-web des participants/contacts par groupe — utilisé pour retrouver le référent (nom, email, téléphone) de chaque association qui garde sa collecte",
    },
}


def _dossier_annee(annee):
    return os.path.join(current_app.root_path, "uploads", "collecte", str(annee))


def _dossier_resultats(annee):
    return os.path.join(_dossier_annee(annee), "resultats")


def _dossier_production(annee):
    return os.path.join(_dossier_annee(annee), "production")


# ============================================================================
# 🚛 PRODUCTION — génération des documents réels (fiches, pointage, équipier,
# index, carte) directement depuis les 3 exports go-on-web du drive collecte,
# en réutilisant tel quel Generer_documents_bai38_depuis_listes.py (copié
# dans ba38_collecte/scripts/generer_documents_production.py). Contrairement
# à la simulation ci-dessus (tournées calculées par optimisation), ici les
# tournées sont déjà décidées dans liste-vehicule.xlsx : ce module se
# contente de mettre en forme les documents à partir de ce qui existe déjà.
#
# Les 3 fichiers sources sont téléchargés à la volée depuis le drive (Google
# Sheets/Excel partagés en "Toute personne disposant du lien"), pas importés
# à la main : la page production doit toujours refléter le dernier état du
# planning go-on-web. Pas d'historique de versions ici (contrairement aux
# générations de simulation) : chaque lancement écrase les documents
# précédents.
#
# Un nouveau dossier Drive (donc 3 nouveaux liens) est créé chaque année par
# le club : les 3 liens sont donc stockés par année dans collecte_campagnes
# (drive_magasins/drive_vehicules/drive_cagettes), pas codés en dur — voir
# la section "🔗 Fichiers Drive" de la page principale du module.
# ============================================================================

DRIVE_CHAMPS = {
    "magasins":  {"champ": "drive_magasins",  "label": "Liste des magasins"},
    "vehicules": {"champ": "drive_vehicules", "label": "Liste des véhicules / planning"},
    "cagettes":  {"champ": "drive_cagettes",  "label": "Historique cagettes"},
    "groupes":   {"champ": "drive_groupes",   "label": "Liste des groupes"},
    "participants": {"champ": "drive_participants", "label": "Liste des participants"},
    "participants_mailing": {"champ": "drive_participants_mailing", "label": "Liste des participants pour mailing"},
}
DRIVE_CHAMPS_PRODUCTION = {cle: DRIVE_CHAMPS[cle] for cle in ("magasins", "vehicules", "cagettes")}


def _id_drive(url):
    """Extrait l'identifiant Drive (segment /d/{ID}/) d'un lien Google Sheets,
    quel que soit le format exact du lien collé (édition, partage...)."""
    if not url:
        return None
    m = re.search(r"/d/([a-zA-Z0-9_-]+)", url)
    return m.group(1) if m else None


def _url_export_drive(url):
    """Construit l'URL d'export xlsx à partir d'un lien Drive collé par
    l'utilisateur. None si le lien est vide ou ne contient pas d'identifiant
    Drive reconnaissable."""
    fid = _id_drive(url)
    return f"https://docs.google.com/spreadsheets/d/{fid}/export?format=xlsx" if fid else None


def _fichier_drive(annee, cle):
    """Télécharge un export Google Sheets et retourne son chemin local."""
    conf = DRIVE_CHAMPS[cle]
    with get_db_connection() as conn:
        campagne = conn.execute(
            "SELECT * FROM collecte_campagnes WHERE annee = ?", (annee,)
        ).fetchone()
    url = _url_export_drive(campagne[conf["champ"]]) if campagne else None
    if not url:
        return None
    dossier = _dossier_annee(annee)
    os.makedirs(dossier, exist_ok=True)
    noms_stockage = {"vehicules": "liste_vehicules.xlsx"}
    nom_stockage = FICHIERS.get(cle, {}).get("nom_stockage") or noms_stockage[cle]
    chemin = os.path.join(dossier, nom_stockage)
    reponse = requests.get(url, timeout=30)
    reponse.raise_for_status()
    if not reponse.content.startswith(b"PK"):
        raise ValueError(f"contenu invalide pour « {conf['label']} »")
    with open(chemin, "wb") as fichier:
        fichier.write(reponse.content)
    return chemin


def _charger_affectations_chauffeurs_equipiers(annee):
    """Construit les affectations par personne depuis le planning véhicules."""
    chemin = _fichier_drive(annee, "vehicules")
    if not chemin:
        chemin = os.path.join(_dossier_annee(annee), "liste_vehicules.xlsx")
    if not os.path.exists(chemin):
        raise FileNotFoundError("Le fichier liste_vehicules.xlsx est introuvable")

    df = pd.read_excel(chemin)
    df.columns = [str(col).strip() for col in df.columns]
    personnes_par_affectation = {}
    affectations_par_personne = {}
    personnes_sans_email = set()
    jours = {"jeudi": "Jeudi", "vendredi": "Vendredi", "samedi": "Samedi", "dimanche": "Dimanche"}

    for _, ligne in df.iterrows():
        code = str(ligne.get("Code", "")).strip()
        personne = str(ligne.get("Équipier", ligne.get("equipier", ""))).strip()
        jour = jours.get(str(ligne.get("Tournée", "")).strip().lower())
        debut = str(ligne.get("Début", "")).strip()
        if not code or code == "nan" or not personne or personne == "nan" or not jour:
            continue
        match = re.match(r"(\d+)", debut)
        periode = "Matin" if not match or int(match.group(1)) < 13 else "Après-midi"
        demi_journee = f"{jour} {periode}"
        email = str(ligne.get("Email", "")).strip().lower()
        if email == "nan":
            email = ""
        role = "chauffeur" if personne.lower().startswith("chauffeur") else "équipier"
        cle_personne = email or f"nom:{personne.lower()}"
        affectation = {
            "demi_journee": demi_journee,
            "camion": code,
            "nom_camion": str(ligne.get("Véhicule", "")).strip(),
            "magasin": str(ligne.get("Magasin", "")).strip(),
            "personne": personne,
            "role": role,
            "email": email,
        }
        affectations_par_personne.setdefault(cle_personne, {"nom": personne, "role": role, "email": email, "affectations": []})["affectations"].append(affectation)
        personnes_par_affectation.setdefault((demi_journee, code), set()).add(personne)
        if not email:
            personnes_sans_email.add(personne)

    destinataires = []
    for personne in affectations_par_personne.values():
        affectations_regroupees = {}
        for affectation in personne["affectations"]:
            cle_affectation = (affectation["demi_journee"], affectation["camion"], affectation["nom_camion"])
            groupe = affectations_regroupees.setdefault(cle_affectation, {**affectation, "magasins": []})
            if affectation["magasin"] and affectation["magasin"] != "nan" and affectation["magasin"] not in groupe["magasins"]:
                groupe["magasins"].append(affectation["magasin"])

        lignes_personne = []
        for affectation in affectations_regroupees.values():
            autres = sorted(personnes_par_affectation[(affectation["demi_journee"], affectation["camion"])] - {personne["nom"]})
            lignes_personne.append({**affectation, "autres": autres})
        lignes_personne.sort(key=lambda item: (item["demi_journee"], item["camion"]))
        personne["affectations"] = lignes_personne
        if personne["email"]:
            destinataires.append(personne)

    destinataires.sort(key=lambda personne: personne["nom"].lower())
    return destinataires, sorted(personnes_sans_email, key=str.lower)


def _corps_mail_affectations(personne):
    lignes = [
        f"Bonjour {personne['nom']},",
        "",
        "Voici la liste de vos tournées camions et les personnes affectées avec vous :",
        "",
    ]
    for affectation in personne["affectations"]:
        camion = affectation["camion"]
        if affectation["nom_camion"] and affectation["nom_camion"] != "nan":
            camion += f" - {affectation['nom_camion']}"
        lignes.append(f"- {affectation['demi_journee']} : {camion}")
        if affectation["magasins"]:
            lignes.append(f"    Magasins : {', '.join(affectation['magasins'])}")
        if affectation["autres"]:
            lignes.append(f"  Avec : {', '.join(affectation['autres'])}")
    lignes += ["", "Merci."]
    return "\n".join(lignes)

PRODUCTION_FICHIERS_SORTIE = {
    "excel":     {"nom": "Tournees_BAI38_{annee}_GOTW.xlsx", "label": "Classeur Excel (tournées + contrôles)"},
    "fiches":    {"nom": "fiches_jour_vehicule_magasin_{annee}_GOTW.pdf", "label": "Fiches de collecte"},
    "pointage":  {"nom": "pointage_vehicules_{annee}_GOTW.pdf", "label": "Pointage véhicules"},
    "equipier":  {"nom": "fiches_jour_vehicule_magasin_equipier_{annee}_GOTW.pdf", "label": "Fiches équipier"},
    "index":     {"nom": "fiches_equipier_jour_vehicule_{annee}_GOTW.pdf", "label": "Index alphabétique équipiers"},
    "consignes": {"nom": "vehicule_consignes.xlsx", "label": "Consignes véhicules (1 ligne/camion)"},
    "carte":     {"nom": "carte_tournees_production.html", "label": "Carte interactive des tournées"},
}


# Défauts repris de lancer_tournees_bai_v2.bat (section PARAMETRES), pas des
# valeurs par défaut de argparse dans le script (différentes, pensées pour un
# usage en ligne de commande sans .bat).
PARAMS_DEFAUT = {
    "camions_supp": 3,
    "poids_nouveaux": 200,
    "max_magasins": 5,
    "optimiser_anciens": True,
    "fusionner_legeres": False,
    "corriger_mal_places": True,
}

COLONNES_MAGASINS = [f"Magasin {i}" for i in range(1, 7)]


def _nom_fichier_genere(annee, params, horodatage):
    """Nom de fichier encodant les paramètres utilisés, pour pouvoir garder
    plusieurs versions par année sans les confondre — même logique de
    suffixes que le script d'origine (VX{camions-supp} + _OptAnciens/
    _FusLegeres/_CorMalPlaces), étendue à max-magasins et poids-nouveaux qui
    sont ici aussi modifiables par l'utilisateur."""
    suffixes = ""
    suffixes += "_OptAnciens" if params["optimiser_anciens"] else ""
    suffixes += "_FusLegeres" if params["fusionner_legeres"] else ""
    suffixes += "_CorMalPlaces" if params["corriger_mal_places"] else ""
    return (
        f"Tournees_BAI38_{annee}_{horodatage}"
        f"_VX{params['camions_supp']}_MAX{params['max_magasins']}_PN{params['poids_nouveaux']}"
        f"{suffixes}.xlsx"
    )


def _generer_tournees(campagne, params):
    """Exécute le pipeline complet (extraction PDF + magasins + optimisation
    + export Excel) via collecte_moteur_tournees, tel qu'orchestré par
    main() dans le script d'origine. Retourne le nom du fichier Excel généré
    et quelques compteurs pour l'affichage/la BDD."""
    annee = campagne["annee"]
    dossier = _dossier_annee(annee)
    pdf_path = os.path.join(dossier, campagne["fichier_pdf_precedent"])
    magasins_path = _fichier_drive(annee, "magasins") or os.path.join(dossier, campagne["fichier_magasins"])

    args_ns = argparse.Namespace(
        camions_supp=params["camions_supp"],
        poids_nouveaux=params["poids_nouveaux"],
        max_magasins=params["max_magasins"],
        corriger_mal_places=params["corriger_mal_places"],
        fusionner_legeres=params["fusionner_legeres"],
        optimiser_anciens=params["optimiser_anciens"],
        output=None,
        pdf=pdf_path,
        magasins=magasins_path,
        nouveaux=None,
    )

    fiches = moteur.extraire_pdf(pdf_path)

    vifs_pdf_2025 = set()
    for f in fiches:
        for v in f["vif_codes"]:
            vifs_pdf_2025.add(str(v).lstrip("0"))

    df_mag = moteur.lire_magasins(magasins_path, vifs_pdf_2025, params["poids_nouveaux"])
    df_t = moteur.optimiser_tournees(fiches, df_mag, args_ns)

    dossier_resultats = _dossier_resultats(annee)
    os.makedirs(dossier_resultats, exist_ok=True)
    horodatage = datetime.now().strftime("%Y%m%d_%H%M")
    nom_fichier = _nom_fichier_genere(annee, params, horodatage)
    args_ns.output = os.path.join(dossier_resultats, nom_fichier)

    moteur.generer_excel(df_t, df_mag, args_ns, args_ns.output, fiches, moteur.non_affectes_global, "")

    nom_carte_secteurs = os.path.splitext(nom_fichier)[0] + "_secteurs.html"
    data_secteurs = df_mag[["Nom", "Latitude", "Longitude", "Secteur"]].astype({
        "Nom": str, "Latitude": float, "Longitude": float, "Secteur": str,
    }).to_dict("records")
    polygones = carte_secteurs.calculer_polygones(data_secteurs)
    carte_secteurs.generer_html(data_secteurs, polygones, os.path.join(dossier_resultats, nom_carte_secteurs), annee)

    # Carte des tournées optimisées (itinéraire OSRM réel par demi-journée/camion,
    # Vendredi/Samedi hors véhicules figés — cf. doc §8) : contrairement à la carte
    # des secteurs, celle-ci tient compte du résultat de l'optimisation (df_t).
    chemin_carte_tournees = moteur.generer_carte_tournees(df_t, df_mag, args_ns, dossier_resultats)
    nom_carte_tournees = os.path.basename(chemin_carte_tournees)

    return {
        "nom_fichier": nom_fichier,
        "nom_carte_secteurs": nom_carte_secteurs,
        "nom_carte_tournees": nom_carte_tournees,
        "nb_tournees": len(df_t),
        "nb_magasins": len(df_mag),
        "nb_nouveaux_magasins": int(df_mag["Nouveau"].sum()),
    }


def _charger_tournees(generation):
    """Relit le fichier Excel généré (onglet Tournees) pour l'affichage web.
    header=7 : même décalage que celui utilisé par Generer_documents_bai38.py
    et Generer_fiches_2025.py pour lire ce même onglet (5 lignes de titre +
    1 ligne de légende + 1 ligne d'en-têtes vides au-dessus des en-têtes
    réels)."""
    annee = generation["annee"]
    chemin = os.path.join(_dossier_resultats(annee), generation["fichier_excel"])

    df = pd.read_excel(chemin, sheet_name="Tournees", header=7)
    df = df[df["Camion"].notna()].reset_index(drop=True)

    ordre_dj = {dj: i for i, dj in enumerate(moteur.DEMI_JOURNEES)}
    df["_ordre_dj"] = df["Demi-journee"].map(ordre_dj).fillna(99)
    df = df.sort_values(["_ordre_dj", "Camion"]).reset_index(drop=True)

    lignes = []
    for _, row in df.iterrows():
        magasins = [
            str(row[c]).strip() for c in COLONNES_MAGASINS
            if c in row and str(row[c]).strip() and str(row[c]).strip().lower() != "nan"
        ]
        lignes.append({
            "demi_journee": str(row["Demi-journee"]),
            "camion": str(row["Camion"]),
            "nom_camion": "" if pd.isna(row.get("Nom camion")) else str(row["Nom camion"]),
            "tonnage": 0.0 if pd.isna(row.get("Tonnage")) else float(row["Tonnage"]),
            "km": "" if pd.isna(row.get("Km estimes")) else float(row["Km estimes"]),
            "duree": "" if pd.isna(row.get("Duree estimee")) else str(row["Duree estimee"]),
            "secteur": "" if pd.isna(row.get("Secteur")) else str(row["Secteur"]),
            "magasins": ", ".join(magasins),
            "commentaire": "" if pd.isna(row.get("Commentaire optimisation")) else str(row["Commentaire optimisation"]),
            "figee": bool(row["Camion"] in moteur.VEHICULES_FIGES),
        })
    return lignes


# ============================================================================
# 🔬 ANALYSE COMPARATIVE 8 SCÉNARIOS (camions-supp 1-4 × max-magasins 4-5)
# ============================================================================
# Indicateurs repris de l'analyse manuelle de référence
# (uploads/collecte_fichiers_source/analyse_simulations_8configs.docx),
# recalculés directement depuis df_t — seuls les tableaux chiffrés sont
# reproduits ici (§1 "Indicateurs de performance" et §2 "Kilométrage par
# demi-journée" du document) ; l'analyse qualitative du document (rôle de
# chaque camion VX, verdicts, recommandation) reste une lecture humaine du
# tableau, pas quelque chose de recalculable de façon fiable.

SCENARIOS_ANALYSE = [(c, m) for m in (4, 5) for c in (1, 2, 3, 4)]
DJ_VS = ['Vendredi Matin', 'Vendredi Apres Midi', 'Samedi Matin', 'Samedi Apres Midi']
DJ_VS_LABELS = {
    'Vendredi Matin': 'Vendredi Matin', 'Vendredi Apres Midi': 'Vendredi Après-Midi',
    'Samedi Matin': 'Samedi Matin', 'Samedi Apres Midi': 'Samedi Après-Midi',
}


def _duree_token_en_minutes(token):
    """'2h17' -> 137, '45min' -> 45."""
    m = re.match(r"(\d+)h(\d{2})", token)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    m = re.match(r"(\d+)min", token)
    if m:
        return int(m.group(1))
    return None


def _parse_duree_minutes(duree_str):
    """'45min – 51min (Hors métropole)' -> 48.0 (moyenne des deux bornes).
    Au-delà d'1h, _fourchette_xl() du moteur passe au format 'XhYY' sans
    suffixe 'min' (ex. '2h17 – 2h38 (Métropole)') — les deux formats
    doivent être reconnus."""
    tokens = re.findall(r"\d+h\d{2}|\d+min", str(duree_str))
    valeurs = [v for v in (_duree_token_en_minutes(t) for t in tokens) if v is not None]
    if len(valeurs) >= 2:
        return (valeurs[0] + valeurs[1]) / 2
    if len(valeurs) == 1:
        return valeurs[0]
    return None


def _compter_secteurs(secteur_str):
    """'Grenoble Centre Est | Grenoble Nord' -> 2."""
    s = str(secteur_str).strip()
    if not s or s.lower() == "nan":
        return 0
    return len([p for p in s.split("|") if p.strip()])


def _calculer_indicateurs_scenario(df_t, camions_supp, max_magasins):
    # Colonnes "Magasin N" dynamiques : le moteur ne crée "Magasin 6" (ou plus)
    # que si une tournée de CE scénario atteint réellement ce nombre de
    # magasins (tolérance de surcharge) — un range(1, 7) fixe plante avec
    # "['Magasin 6'] not in index" dès qu'un scénario n'a aucune surcharge
    # aussi élevée (cas rencontré en PROD, absent des données de test DEV).
    cols_mag = sorted(
        (c for c in df_t.columns if re.fullmatch(r"Magasin \d+", c)),
        key=lambda c: int(c.split(" ")[1])
    )

    df = df_t[df_t["Demi-journee"].isin(DJ_VS)].copy()
    df = df[~df["Camion"].astype(str).isin(moteur.VEHICULES_FIGES)]

    df["_nb_mag"] = df[cols_mag].apply(
        lambda r: sum(1 for v in r if str(v).strip() and str(v).strip().lower() != "nan"), axis=1
    )
    df["_nb_sec"] = df["Secteur"].apply(_compter_secteurs)
    df["_duree_min"] = df["Duree estimee"].apply(_parse_duree_minutes)

    nb_tournees = len(df)
    est_vx = df["Camion"].astype(str).str.startswith("VX")
    surcharges = df[df["_nb_mag"] > max_magasins]

    km_par_dj = {dj: round(float(df[df["Demi-journee"] == dj]["Km estimes"].sum()), 1) for dj in DJ_VS}

    return {
        "camions_supp": camions_supp,
        "max_magasins": max_magasins,
        "nb_tournees": nb_tournees,
        "tournees_2": int((df["_nb_mag"] == 2).sum()),
        "tournees_3": int((df["_nb_mag"] == 3).sum()),
        "tournees_4": int((df["_nb_mag"] == 4).sum()),
        "tournees_5": int((df["_nb_mag"] == 5).sum()),
        "tournees_6plus": int((df["_nb_mag"] >= 6).sum()),
        "surcharges": int(len(surcharges)),
        "surcharges_detail": [
            {"camion": str(r["Camion"]), "demi_journee": str(r["Demi-journee"]), "nb_magasins": int(r["_nb_mag"])}
            for _, r in surcharges.iterrows()
        ],
        "tournees_3_secteurs": int((df["_nb_sec"] == 3).sum()),
        "tournees_plus3_secteurs": int((df["_nb_sec"] > 3).sum()),
        "secteurs_multiples_detail": [
            {"camion": str(r["Camion"]), "demi_journee": str(r["Demi-journee"]),
             "secteurs": str(r["Secteur"]), "nb_magasins": int(r["_nb_mag"])}
            for _, r in df[df["_nb_sec"] >= 3].iterrows()
        ],
        "km_total": round(float(df["Km estimes"].sum()), 1),
        "moy_magasins_tournee": round(float(df["_nb_mag"].mean()), 2) if nb_tournees else 0,
        "camions_vx_crees": int(df.loc[est_vx, "Camion"].nunique()),
        "magasins_dans_vx": int(df.loc[est_vx, "_nb_mag"].sum()),
        "vx_detail": [
            {"camion": str(r["Camion"]), "demi_journee": str(r["Demi-journee"]),
             "secteur": str(r["Secteur"]), "nb_magasins": int(r["_nb_mag"])}
            for _, r in df[est_vx].iterrows()
        ],
        "duree_moy_min": round(float(df["_duree_min"].mean()), 1) if df["_duree_min"].notna().any() else None,
        "km_par_dj": km_par_dj,
    }


def _lancer_analyse_8configs_background(app, analyse_id, pdf_path, magasins_path, params_communs):
    with app.app_context():
        db_path = get_db_path()
        try:
            fiches = moteur.extraire_pdf(pdf_path)

            vifs_pdf_2025 = set()
            for f in fiches:
                for v in f["vif_codes"]:
                    vifs_pdf_2025.add(str(v).lstrip("0"))

            df_mag = moteur.lire_magasins(magasins_path, vifs_pdf_2025, params_communs["poids_nouveaux"])

            resultats = []
            for camions_supp, max_magasins in SCENARIOS_ANALYSE:
                args_ns = argparse.Namespace(
                    camions_supp=camions_supp,
                    poids_nouveaux=params_communs["poids_nouveaux"],
                    max_magasins=max_magasins,
                    corriger_mal_places=params_communs["corriger_mal_places"],
                    fusionner_legeres=params_communs["fusionner_legeres"],
                    optimiser_anciens=params_communs["optimiser_anciens"],
                    output=None, pdf=pdf_path, magasins=magasins_path, nouveaux=None,
                )
                df_t = moteur.optimiser_tournees(copy.deepcopy(fiches), df_mag.copy(), args_ns)
                resultats.append(_calculer_indicateurs_scenario(df_t, camions_supp, max_magasins))
                write_log(f"🔬 Analyse 8 scénarios #{analyse_id} : VX{camions_supp}-MAX{max_magasins} terminé")

            conn = sqlite3.connect(db_path)
            conn.execute("""
                UPDATE collecte_analyses SET statut = 'termine', resultat = ?, termine_le = ?
                WHERE id = ?
            """, (json.dumps(resultats), datetime.now().strftime("%Y-%m-%d %H:%M:%S"), analyse_id))
            conn.commit()
            conn.close()
            write_log(f"✅ Analyse 8 scénarios #{analyse_id} terminée")

        except Exception as e:
            write_log(f"❌ Analyse 8 scénarios #{analyse_id} en échec : {e}")
            conn = sqlite3.connect(db_path)
            conn.execute(
                "UPDATE collecte_analyses SET statut = 'erreur', erreur = ?, termine_le = ? WHERE id = ?",
                (str(e), datetime.now().strftime("%Y-%m-%d %H:%M:%S"), analyse_id)
            )
            conn.commit()
            conn.close()


@collecte_bp.route("/collecte")
@login_required
@require_access("collecte", "lecture")
def collecte_main():
    annee = request.args.get("annee", type=int) or datetime.now().year

    with get_db_connection() as conn:
        campagne = conn.execute(
            "SELECT * FROM collecte_campagnes WHERE annee = ?", (annee,)
        ).fetchone()

        annees_existantes = [
            row["annee"] for row in conn.execute(
                "SELECT annee FROM collecte_campagnes ORDER BY annee DESC"
            ).fetchall()
        ]

        generations = conn.execute(
            "SELECT * FROM collecte_generations WHERE annee = ? ORDER BY id DESC", (annee,)
        ).fetchall()

        derniere_analyse = conn.execute(
            "SELECT * FROM collecte_analyses WHERE annee = ? ORDER BY id DESC LIMIT 1", (annee,)
        ).fetchone()

    annee_now = datetime.now().year
    if annee_now not in annees_existantes and annee_now not in (annee,):
        annees_existantes = sorted(set(annees_existantes) | {annee_now}, reverse=True)
    if annee not in annees_existantes:
        annees_existantes = sorted(set(annees_existantes) | {annee}, reverse=True)

    return render_template(
        "collecte/index.html",
        annee=annee,
        campagne=campagne,
        annees_existantes=annees_existantes,
        fichiers=FICHIERS,
        derniere_generation=generations[0] if generations else None,
        nb_generations=len(generations),
        derniere_analyse=derniere_analyse,
    )


@collecte_bp.route("/collecte/chauffeurs_equipiers", methods=["GET", "POST"])
@login_required
@require_access("collecte", "lecture")
def chauffeurs_equipiers():
    annee = request.args.get("annee", type=int) or request.form.get("annee", type=int) or datetime.now().year
    try:
        destinataires, personnes_sans_email = _charger_affectations_chauffeurs_equipiers(annee)
    except Exception as erreur:
        flash(f"❌ Impossible de charger les affectations : {erreur}", "danger")
        return redirect(url_for("collecte.collecte_main", annee=annee))

    if request.method == "POST":
        action = request.form.get("action")
        if not destinataires:
            flash("❌ Aucun chauffeur ou équipier avec une adresse email.", "danger")
        elif action == "test":
            email_test = request.form.get("test_destinataire", "")
            exemple = next((personne for personne in destinataires if personne["email"] == email_test), destinataires[0])
            adresse_test = getattr(current_user, "email", "") or ""
            if not adresse_test:
                flash("❌ Votre compte n’a pas d’adresse email pour le test.", "danger")
            else:
                envoyer_mail(
                    f"[TEST] Affectations tournées {annee}",
                    [adresse_test],
                    _corps_mail_affectations(exemple),
                    sender_override="ba380.directeur@banquealimentaire.org",
                )
                flash(f"✅ Mail de test envoyé à {adresse_test}.", "success")
        elif action == "envoyer":
            for personne in destinataires:
                envoyer_mail(
                    f"Vos affectations tournées {annee}",
                    [personne["email"]],
                    _corps_mail_affectations(personne),
                    sender_override="ba380.directeur@banquealimentaire.org",
                )
            flash(f"✅ {len(destinataires)} mail(s) préparé(s).", "success")

    return render_template(
        "collecte/chauffeurs_equipiers.html",
        annee=annee,
        destinataires=destinataires,
        personnes_sans_email=personnes_sans_email,
    )


@collecte_bp.route("/collecte/upload/<type_fichier>", methods=["POST"])
@login_required
@require_access("collecte", "ecriture")
def upload_fichier(type_fichier):
    annee = request.form.get("annee", type=int)

    if type_fichier not in FICHIERS:
        flash("❌ Type de fichier inconnu", "danger")
        return redirect(url_for("collecte.collecte_main", annee=annee))

    if not annee:
        flash("❌ Année manquante", "danger")
        return redirect(url_for("collecte.collecte_main"))

    conf = FICHIERS[type_fichier]
    fichier = request.files.get("fichier")

    if not fichier or not fichier.filename:
        flash("❌ Aucun fichier sélectionné", "danger")
        return redirect(url_for("collecte.collecte_main", annee=annee))

    ext = os.path.splitext(fichier.filename)[1].lower()
    if ext not in conf["extensions"]:
        flash(
            f"❌ Extension invalide pour « {conf['label']} » "
            f"(attendu : {', '.join(sorted(conf['extensions']))})",
            "danger"
        )
        return redirect(url_for("collecte.collecte_main", annee=annee))

    dossier = _dossier_annee(annee)
    os.makedirs(dossier, exist_ok=True)
    fichier.save(os.path.join(dossier, conf["nom_stockage"]))

    maintenant = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with get_db_connection() as conn:
        existante = conn.execute(
            "SELECT id FROM collecte_campagnes WHERE annee = ?", (annee,)
        ).fetchone()

        if existante:
            conn.execute(f"""
                UPDATE collecte_campagnes
                SET {conf['champ_chemin']} = ?, {conf['champ_le']} = ?, {conf['champ_par']} = ?
                WHERE annee = ?
            """, (conf["nom_stockage"], maintenant, current_user.email, annee))
        else:
            conn.execute(f"""
                INSERT INTO collecte_campagnes
                    (annee, {conf['champ_chemin']}, {conf['champ_le']}, {conf['champ_par']},
                     date_creation, cree_par)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (annee, conf["nom_stockage"], maintenant, current_user.email,
                  maintenant, current_user.email))

        conn.commit()

    write_log(f"✅ Collecte {annee} : fichier « {conf['label']} » importé par {current_user.email}")
    flash(f"✅ {conf['label']} importé(e) pour la collecte {annee}", "success")

    return redirect(url_for("collecte.collecte_main", annee=annee))


@collecte_bp.route("/collecte/dates", methods=["POST"])
@login_required
@require_access("collecte", "ecriture")
def enregistrer_dates():
    annee = request.form.get("annee", type=int)
    date_debut = request.form.get("date_debut", "").strip() or None
    date_fin = request.form.get("date_fin", "").strip() or None

    if not annee:
        flash("❌ Année manquante", "danger")
        return redirect(url_for("collecte.collecte_main"))

    with get_db_connection() as conn:
        existante = conn.execute(
            "SELECT id FROM collecte_campagnes WHERE annee = ?", (annee,)
        ).fetchone()

        if existante:
            conn.execute(
                "UPDATE collecte_campagnes SET date_debut = ?, date_fin = ? WHERE annee = ?",
                (date_debut, date_fin, annee)
            )
        else:
            maintenant = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn.execute("""
                INSERT INTO collecte_campagnes (annee, date_debut, date_fin, date_creation, cree_par)
                VALUES (?, ?, ?, ?, ?)
            """, (annee, date_debut, date_fin, maintenant, current_user.email))

        conn.commit()

    write_log(f"📅 Collecte {annee} : dates mises à jour par {current_user.email} ({date_debut} → {date_fin})")
    flash(f"✅ Dates de la collecte {annee} enregistrées", "success")

    return redirect(url_for("collecte.collecte_main", annee=annee))


@collecte_bp.route("/collecte/drive_liens", methods=["POST"])
@login_required
@require_access("collecte", "ecriture")
def enregistrer_liens_drive():
    annee = request.form.get("annee", type=int)

    if not annee:
        flash("❌ Année manquante", "danger")
        return redirect(url_for("collecte.collecte_main"))

    valeurs = {}
    invalides = []
    for cle, conf in DRIVE_CHAMPS.items():
        brut = request.form.get(cle, "").strip()
        if brut and not _id_drive(brut):
            invalides.append(conf["label"])
        valeurs[conf["champ"]] = brut or None

    if invalides:
        flash(
            f"❌ Lien(s) Drive non reconnu(s), non enregistré(s) : {', '.join(invalides)} "
            f"— coller le lien de partage complet du fichier Google Sheets",
            "danger"
        )
        return redirect(url_for("collecte.collecte_main", annee=annee))

    with get_db_connection() as conn:
        existante = conn.execute(
            "SELECT id FROM collecte_campagnes WHERE annee = ?", (annee,)
        ).fetchone()

        if existante:
            conn.execute(
                "UPDATE collecte_campagnes SET drive_magasins = ?, drive_vehicules = ?, "
                 "drive_cagettes = ?, drive_groupes = ?, drive_participants = ?, "
                 "drive_participants_mailing = ? WHERE annee = ?",
                (valeurs["drive_magasins"], valeurs["drive_vehicules"], valeurs["drive_cagettes"],
                  valeurs["drive_groupes"], valeurs["drive_participants"],
                  valeurs["drive_participants_mailing"], annee)
            )
        else:
            maintenant = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn.execute("""
                INSERT INTO collecte_campagnes
                    (annee, drive_magasins, drive_vehicules, drive_cagettes, drive_groupes,
                                         drive_participants, drive_participants_mailing, date_creation, cree_par)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (annee, valeurs["drive_magasins"], valeurs["drive_vehicules"], valeurs["drive_cagettes"],
                                    valeurs["drive_groupes"], valeurs["drive_participants"],
                                    valeurs["drive_participants_mailing"], maintenant, current_user.email))

        conn.commit()

    write_log(f"🔗 Collecte {annee} : liens Drive mis à jour par {current_user.email}")
    flash(f"✅ Liens Drive de la collecte {annee} enregistrés", "success")

    return redirect(url_for("collecte.collecte_main", annee=annee))


def _get_campagne_ou_redirect(annee):
    with get_db_connection() as conn:
        campagne = conn.execute(
            "SELECT * FROM collecte_campagnes WHERE annee = ?", (annee,)
        ).fetchone()

    if not campagne or not (campagne["fichier_magasins"] or campagne["drive_magasins"]) or not campagne["fichier_pdf_precedent"]:
        flash(
            "⛔ Liste des magasins et PDF des tournées précédentes requis avant de générer "
            f"les tournées {annee}",
            "danger"
        )
        return None
    return campagne


@collecte_bp.route("/collecte/analyse", methods=["POST"])
@login_required
@require_access("collecte", "ecriture")
def lancer_analyse():
    annee = request.form.get("annee", type=int)

    campagne = _get_campagne_ou_redirect(annee)
    if campagne is None:
        return redirect(url_for("collecte.collecte_main", annee=annee))

    params_communs = {
        "poids_nouveaux": PARAMS_DEFAUT["poids_nouveaux"],
        "optimiser_anciens": True,
        "fusionner_legeres": False,
        "corriger_mal_places": True,
    }

    maintenant = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db_connection() as conn:
        cur = conn.execute("""
            INSERT INTO collecte_analyses (annee, statut, parametres_communs, lance_le, lance_par)
            VALUES (?, 'en_cours', ?, ?, ?)
        """, (annee, json.dumps(params_communs), maintenant, current_user.email))
        analyse_id = cur.lastrowid
        conn.commit()

    dossier = _dossier_annee(annee)
    pdf_path = os.path.join(dossier, campagne["fichier_pdf_precedent"])
    magasins_path = _fichier_drive(annee, "magasins") or os.path.join(dossier, campagne["fichier_magasins"])
    app_reel = current_app._get_current_object()

    Thread(
        target=_lancer_analyse_8configs_background,
        args=(app_reel, analyse_id, pdf_path, magasins_path, params_communs)
    ).start()

    write_log(f"🔬 Collecte {annee} : analyse 8 scénarios #{analyse_id} lancée par {current_user.email}")
    flash("🔬 Analyse des 8 scénarios lancée en arrière-plan (compter 1 à 2 minutes)", "info")

    return redirect(url_for("collecte.analyse", analyse_id=analyse_id))


INDICATEURS_PROMPT = [
    ("Nb tournées V/S (hors figés)", "nb_tournees"),
    ("Tournées à 2 magasins", "tournees_2"),
    ("Tournées à 3 magasins", "tournees_3"),
    ("Tournées à 4 magasins", "tournees_4"),
    ("Tournées à 5 magasins", "tournees_5"),
    ("Tournées à 6+ magasins (MAX+2)", "tournees_6plus"),
    ("Tournées > MAX (surcharges)", "surcharges"),
    ("Tournées à 3 secteurs", "tournees_3_secteurs"),
    ("Tournées à > 3 secteurs", "tournees_plus3_secteurs"),
    ("Km total estimé", "km_total"),
    ("Moy. magasins/tournée", "moy_magasins_tournee"),
    ("Camions VX créés", "camions_vx_crees"),
    ("Magasins dans les VX", "magasins_dans_vx"),
    ("Durée moy. estimée (min)", "duree_moy_min"),
]


def _construire_prompt_analyse(annee, scenarios):
    """Prompt prêt à coller dans Claude.ai / ChatGPT pour obtenir la partie
    rédactionnelle (rôle des camions VX, verdicts, recommandation) — sur le
    modèle de l'analyse manuelle de référence
    (uploads/collecte_fichiers_source/analyse_simulations_8configs.docx).
    Pas d'appel API depuis l'appli (choix du 2026-08-17) : uniquement les
    données chiffrées, à coller dans l'outil IA de son choix."""
    entetes = [f"VX{s['camions_supp']}-MAX{s['max_magasins']}" for s in scenarios]

    lignes_tableau = ["| Indicateur | " + " | ".join(entetes) + " |",
                      "|---|" + "|".join(["---"] * len(entetes)) + "|"]
    for label, champ in INDICATEURS_PROMPT:
        lignes_tableau.append("| " + label + " | " + " | ".join(str(s.get(champ, "")) for s in scenarios) + " |")

    parties_surcharges = []
    for s in scenarios:
        detail = s.get("surcharges_detail") or []
        if detail:
            items = "; ".join(f"{d['camion']} ({d['demi_journee']}, {d['nb_magasins']} mag.)" for d in detail)
            parties_surcharges.append(f"- VX{s['camions_supp']}-MAX{s['max_magasins']} ({s['surcharges']}) : {items}")

    parties_vx = []
    for s in scenarios:
        detail = s.get("vx_detail") or []
        if detail:
            items = "; ".join(f"{d['camion']} {d['demi_journee']} → {d['secteur']} ({d['nb_magasins']} mag.)" for d in detail)
            parties_vx.append(f"- VX{s['camions_supp']}-MAX{s['max_magasins']} : {items}")

    parties_secteurs = []
    for s in scenarios:
        detail = s.get("secteurs_multiples_detail") or []
        if detail:
            items = "; ".join(f"{d['camion']} {d['demi_journee']} ({d['secteurs']}, {d['nb_magasins']} mag.)" for d in detail)
            parties_secteurs.append(f"- VX{s['camions_supp']}-MAX{s['max_magasins']} : {items}")

    return f"""Tu es un analyste logistique pour une banque alimentaire (BAI 38 — Isère). Voici les résultats \
de 8 simulations de génération de tournées de collecte alimentaire pour la collecte {annee}, calculés par un \
algorithme d'optimisation réel (pas une estimation manuelle), en faisant varier :
- le nombre de camions supplémentaires (VX) : 1 à 4
- le nombre maximum de magasins par tournée : 4 ou 5

Toutes les données ci-dessous portent uniquement sur le Vendredi et le Samedi (Matin + Après-midi), hors \
véhicules figés (gérés par des associations partenaires, jamais réoptimisés).

## 1. Indicateurs de performance

{chr(10).join(lignes_tableau)}

## 2. Détail des tournées en surcharge (nb magasins > max autorisé)

{chr(10).join(parties_surcharges) if parties_surcharges else "Aucune surcharge sur aucun scénario."}

## 3. Détail des camions supplémentaires (VX) — secteur desservi, nombre de magasins

{chr(10).join(parties_vx) if parties_vx else "Aucun camion supplémentaire créé sur aucun scénario."}

## 4. Détail des tournées à 3 secteurs géographiques ou plus

{chr(10).join(parties_secteurs) if parties_secteurs else "Aucune tournée à 3 secteurs ou plus."}

---

En te basant UNIQUEMENT sur les données ci-dessus (n'invente aucun chiffre), rédige une analyse comparative \
structurée avec :
1. Une analyse des dépassements de capacité : MAX4 est-il structurellement adapté ou non ? à partir de quel \
nombre de camions VX les surcharges disparaissent-elles avec MAX5 ?
2. Le rôle de chaque camion supplémentaire VX dans chaque scénario où il existe (quel secteur il déleste, \
combien de magasins).
3. Une analyse de la cohérence sectorielle (tournées à 3 secteurs ou plus) : quelles tournées reviennent \
dans plusieurs scénarios, signe d'une contrainte géographique structurelle plutôt que d'un mauvais réglage.
4. Une analyse détaillée par configuration (points positifs/négatifs de chaque VX×MAX).
5. Une recommandation finale argumentée avec un tableau de synthèse (surcharges, km total, tournées à 3 \
secteurs) et un verdict par scénario : ★ recommandé, ✓ bonne alternative, ~ acceptable avec nuances, ✗ à éviter.

Style : synthétique, factuel, orienté décision opérationnelle — pas de généralités, uniquement des \
observations appuyées sur les chiffres fournis."""


def _generer_redaction_background(app, analyse_id, prompt):
    """Envoie le prompt à l'API Claude et enregistre le texte rédigé.
    Tourne en arrière-plan (comme l'analyse elle-même) car un appel avec
    réflexion étendue peut dépasser le timeout par défaut de gunicorn (30s)."""
    with app.app_context():
        try:
            api_key = os.getenv("CLAUDE_API_KEY")
            if not api_key:
                raise RuntimeError("CLAUDE_API_KEY absente du .env")

            client = anthropic.Anthropic(api_key=api_key)
            response = client.messages.create(
                model="claude-sonnet-5",
                max_tokens=16000,
                messages=[{"role": "user", "content": prompt}],
            )
            texte = "".join(bloc.text for bloc in response.content if bloc.type == "text")
            if not texte:
                raise RuntimeError(f"Réponse vide de l'API (stop_reason={response.stop_reason})")

            with get_db_connection() as conn:
                conn.execute("""
                    UPDATE collecte_analyses
                    SET redaction_statut = 'termine', redaction_texte = ?, redaction_genere_le = ?
                    WHERE id = ?
                """, (texte, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), analyse_id))
                conn.commit()

            write_log(f"🤖 Collecte : rédaction IA de l'analyse #{analyse_id} terminée")

        except Exception as e:
            write_log(f"❌ Collecte : erreur rédaction IA analyse #{analyse_id} : {e}")
            with get_db_connection() as conn:
                conn.execute(
                    "UPDATE collecte_analyses SET redaction_statut = 'erreur', redaction_erreur = ? WHERE id = ?",
                    (str(e), analyse_id)
                )
                conn.commit()


@collecte_bp.route("/collecte/analyse/<int:analyse_id>/generer_redaction", methods=["POST"])
@login_required
@require_access("collecte", "ecriture")
def generer_redaction(analyse_id):
    with get_db_connection() as conn:
        analyse_row = conn.execute(
            "SELECT * FROM collecte_analyses WHERE id = ?", (analyse_id,)
        ).fetchone()

    if not analyse_row or not analyse_row["resultat"]:
        flash("⛔ Analyse introuvable ou incomplète", "warning")
        return redirect(url_for("collecte.collecte_main"))

    scenarios = json.loads(analyse_row["resultat"])
    prompt = _construire_prompt_analyse(analyse_row["annee"], scenarios)

    with get_db_connection() as conn:
        conn.execute(
            "UPDATE collecte_analyses SET redaction_statut = 'en_cours', redaction_erreur = NULL WHERE id = ?",
            (analyse_id,)
        )
        conn.commit()

    app_reel = current_app._get_current_object()
    Thread(target=_generer_redaction_background, args=(app_reel, analyse_id, prompt)).start()

    write_log(f"🤖 Collecte : rédaction IA de l'analyse #{analyse_id} lancée par {current_user.email}")
    flash("🤖 Génération de la partie rédactionnelle lancée (compter 30 à 60 secondes)", "info")

    return redirect(url_for("collecte.analyse", analyse_id=analyse_id))


@collecte_bp.route("/collecte/analyse/<int:analyse_id>")
@login_required
@require_access("collecte", "lecture")
def analyse(analyse_id):
    with get_db_connection() as conn:
        analyse_row = conn.execute(
            "SELECT * FROM collecte_analyses WHERE id = ?", (analyse_id,)
        ).fetchone()

    if not analyse_row:
        flash("⛔ Analyse introuvable", "warning")
        return redirect(url_for("collecte.collecte_main"))

    scenarios = json.loads(analyse_row["resultat"]) if analyse_row["resultat"] else []
    prompt_analyse = _construire_prompt_analyse(analyse_row["annee"], scenarios) if scenarios else None
    redaction_html = None
    if analyse_row["redaction_texte"]:
        redaction_html = markdown.markdown(analyse_row["redaction_texte"], extensions=["tables"])
        # Tableaux larges (8 scénarios en colonnes) → scroll horizontal contenu
        # dans la carte plutôt que débordement de toute la page.
        redaction_html = redaction_html.replace("<table>", '<div class="table-responsive"><table>')
        redaction_html = redaction_html.replace("</table>", "</table></div>")

    return render_template(
        "collecte/analyse.html",
        annee=analyse_row["annee"],
        analyse=analyse_row,
        scenarios=scenarios,
        dj_vs=DJ_VS,
        dj_vs_labels=DJ_VS_LABELS,
        prompt_analyse=prompt_analyse,
        redaction_html=redaction_html,
    )


@collecte_bp.route("/collecte/generations/<int:annee>")
@login_required
@require_access("collecte", "lecture")
def generations(annee):
    with get_db_connection() as conn:
        lignes = conn.execute(
            "SELECT * FROM collecte_generations WHERE annee = ? ORDER BY id DESC", (annee,)
        ).fetchall()

    versions = []
    for gen in lignes:
        versions.append({
            "gen": gen,
            "params": json.loads(gen["parametres"]) if gen["parametres"] else {},
        })

    return render_template("collecte/generations.html", annee=annee, versions=versions)


@collecte_bp.route("/collecte/generer")
@login_required
@require_access("collecte", "ecriture")
def generer_form():
    annee = request.args.get("annee", type=int) or datetime.now().year

    campagne = _get_campagne_ou_redirect(annee)
    if campagne is None:
        return redirect(url_for("collecte.collecte_main", annee=annee))

    return render_template(
        "collecte/generer.html",
        annee=annee,
        params=PARAMS_DEFAUT,
    )


@collecte_bp.route("/collecte/generer", methods=["POST"])
@login_required
@require_access("collecte", "ecriture")
def generer():
    annee = request.form.get("annee", type=int)

    campagne = _get_campagne_ou_redirect(annee)
    if campagne is None:
        return redirect(url_for("collecte.collecte_main", annee=annee))

    params = {
        "camions_supp": request.form.get("camions_supp", type=int, default=PARAMS_DEFAUT["camions_supp"]),
        "poids_nouveaux": request.form.get("poids_nouveaux", type=int, default=PARAMS_DEFAUT["poids_nouveaux"]),
        "max_magasins": request.form.get("max_magasins", type=int, default=PARAMS_DEFAUT["max_magasins"]),
        "optimiser_anciens": bool(request.form.get("optimiser_anciens")),
        "fusionner_legeres": bool(request.form.get("fusionner_legeres")),
        "corriger_mal_places": bool(request.form.get("corriger_mal_places")),
    }

    try:
        resultat = _generer_tournees(campagne, params)
    except Exception as e:
        write_log(f"❌ Collecte {annee} : échec génération tournées — {e}")
        flash(f"❌ Échec de la génération des tournées : {e}", "danger")
        return redirect(url_for("collecte.generer_form", annee=annee))

    maintenant = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db_connection() as conn:
        cur = conn.execute("""
            INSERT INTO collecte_generations
                (annee, fichier_excel, fichier_carte_secteurs, fichier_carte_tournees, parametres,
                 genere_le, genere_par, nb_tournees, nb_magasins, nb_nouveaux_magasins)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            annee, resultat["nom_fichier"], resultat["nom_carte_secteurs"], resultat["nom_carte_tournees"],
            json.dumps(params), maintenant, current_user.email,
            resultat["nb_tournees"], resultat["nb_magasins"], resultat["nb_nouveaux_magasins"],
        ))
        generation_id = cur.lastrowid
        conn.commit()

    write_log(
        f"✅ Collecte {annee} : tournées générées par {current_user.email} "
        f"({resultat['nb_tournees']} tournées, {resultat['nb_magasins']} magasins) — {resultat['nom_fichier']}"
    )
    flash(f"✅ Tournées {annee} générées : {resultat['nb_tournees']} tournées", "success")

    return redirect(url_for("collecte.resultats", generation_id=generation_id))


def _get_generation_ou_redirect(generation_id):
    with get_db_connection() as conn:
        generation = conn.execute(
            "SELECT * FROM collecte_generations WHERE id = ?", (generation_id,)
        ).fetchone()

    if not generation:
        flash("⛔ Génération introuvable", "warning")
        return None
    return generation


@collecte_bp.route("/collecte/resultats/<int:generation_id>/supprimer", methods=["POST"])
@login_required
@require_access("collecte", "ecriture")
def supprimer_generation(generation_id):
    generation = _get_generation_ou_redirect(generation_id)
    if generation is None:
        return redirect(url_for("collecte.collecte_main"))

    annee = generation["annee"]
    dossier = _dossier_resultats(annee)

    for champ in ("fichier_excel", "fichier_carte_secteurs", "fichier_carte_tournees"):
        nom = generation[champ]
        if not nom:
            continue
        chemin = os.path.join(dossier, nom)
        if os.path.exists(chemin):
            os.remove(chemin)

    with get_db_connection() as conn:
        conn.execute("DELETE FROM collecte_generations WHERE id = ?", (generation_id,))
        conn.commit()

    write_log(
        f"🗑️ Collecte {annee} : génération #{generation_id} ({generation['fichier_excel']}) "
        f"supprimée par {current_user.email}"
    )
    flash(f"🗑️ Génération du {date_fr(generation['genere_le'])} supprimée", "success")

    return redirect(url_for("collecte.generations", annee=annee))


@collecte_bp.route("/collecte/resultats/<int:generation_id>")
@login_required
@require_access("collecte", "lecture")
def resultats(generation_id):
    generation = _get_generation_ou_redirect(generation_id)
    if generation is None:
        return redirect(url_for("collecte.collecte_main"))

    with get_db_connection() as conn:
        autres_generations = conn.execute(
            "SELECT * FROM collecte_generations WHERE annee = ? ORDER BY id DESC",
            (generation["annee"],)
        ).fetchall()

    lignes = _charger_tournees(generation)
    params_utilises = json.loads(generation["parametres"]) if generation["parametres"] else {}

    return render_template(
        "collecte/resultats.html",
        annee=generation["annee"],
        generation=generation,
        autres_generations=autres_generations,
        lignes=lignes,
        params=params_utilises,
    )


@collecte_bp.route("/collecte/resultats/<int:generation_id>/carte_secteurs")
@login_required
@require_access("collecte", "lecture")
def resultats_carte_secteurs(generation_id):
    generation = _get_generation_ou_redirect(generation_id)
    if generation is None:
        return redirect(url_for("collecte.collecte_main"))

    if not generation["fichier_carte_secteurs"]:
        flash("⛔ Pas de carte des secteurs pour cette génération", "warning")
        return redirect(url_for("collecte.resultats", generation_id=generation_id))

    chemin = os.path.join(_dossier_resultats(generation["annee"]), generation["fichier_carte_secteurs"])
    return send_file(chemin, mimetype="text/html")


@collecte_bp.route("/collecte/resultats/<int:generation_id>/carte_tournees")
@login_required
@require_access("collecte", "lecture")
def resultats_carte_tournees(generation_id):
    generation = _get_generation_ou_redirect(generation_id)
    if generation is None:
        return redirect(url_for("collecte.collecte_main"))

    if not generation["fichier_carte_tournees"]:
        flash("⛔ Pas de carte des tournées pour cette génération", "warning")
        return redirect(url_for("collecte.resultats", generation_id=generation_id))

    chemin = os.path.join(_dossier_resultats(generation["annee"]), generation["fichier_carte_tournees"])
    return send_file(chemin, mimetype="text/html")


@collecte_bp.route("/collecte/resultats/<int:generation_id>/excel")
@login_required
@require_access("collecte", "lecture")
def resultats_excel(generation_id):
    generation = _get_generation_ou_redirect(generation_id)
    if generation is None:
        return redirect(url_for("collecte.collecte_main"))

    chemin = os.path.join(_dossier_resultats(generation["annee"]), generation["fichier_excel"])
    return send_file(
        chemin,
        as_attachment=True,
        download_name=generation["fichier_excel"],
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@collecte_bp.route("/collecte/resultats/<int:generation_id>/pdf")
@login_required
@require_access("collecte", "lecture")
def resultats_pdf(generation_id):
    generation = _get_generation_ou_redirect(generation_id)
    if generation is None:
        return redirect(url_for("collecte.collecte_main"))

    lignes = _charger_tournees(generation)
    html = render_template(
        "collecte/resultats_pdf.html", annee=generation["annee"], lignes=lignes,
        org=get_organisation(),
    )

    pdf_buffer = io.BytesIO()
    HTML(string=html).write_pdf(pdf_buffer)
    pdf_buffer.seek(0)

    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=os.path.splitext(generation["fichier_excel"])[0] + ".pdf",
        mimetype="application/pdf",
    )


def _date_fr_courte(iso):
    """Convertit une date ISO (AAAA-MM-JJ, celle du <input type=date>) en
    JJ/MM/AAAA (format attendu par --date-jeudi du script de production).
    Retourne None si iso est vide ou invalide."""
    if not iso:
        return None
    try:
        return datetime.strptime(iso, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return None


@collecte_bp.route("/collecte/production")
@login_required
@require_access("collecte", "lecture")
def production():
    annee = request.args.get("annee", type=int) or datetime.now().year
    dossier = _dossier_production(annee)

    with get_db_connection() as conn:
        campagne = conn.execute(
            "SELECT * FROM collecte_campagnes WHERE annee = ?", (annee,)
        ).fetchone()

    liens_manquants = [
        conf["label"] for cle, conf in DRIVE_CHAMPS_PRODUCTION.items()
        if not (campagne and _url_export_drive(campagne[conf["champ"]]))
    ]

    fichiers = {}
    for cle, conf in PRODUCTION_FICHIERS_SORTIE.items():
        nom = conf["nom"].format(annee=annee)
        chemin = os.path.join(dossier, nom)
        existe = os.path.exists(chemin)
        fichiers[cle] = {
            "label": conf["label"],
            "nom": nom,
            "existe": existe,
            "genere_le": datetime.fromtimestamp(os.path.getmtime(chemin)) if existe else None,
        }

    derniere_generation = max(
        (f["genere_le"] for f in fichiers.values() if f["genere_le"]), default=None
    )

    journal = None
    chemin_journal = os.path.join(dossier, "dernier_journal.log")
    if os.path.exists(chemin_journal):
        with open(chemin_journal, "r", encoding="utf-8") as f:
            journal = f.read()

    return render_template(
        "collecte/collecte_production.html",
        annee=annee,
        fichiers=fichiers,
        derniere_generation=derniere_generation,
        journal=journal,
        date_debut=_date_fr_courte(campagne["date_debut"]) if campagne else None,
        date_fin=_date_fr_courte(campagne["date_fin"]) if campagne else None,
        liens_manquants=liens_manquants,
    )


@collecte_bp.route("/collecte/production/generer", methods=["POST"])
@login_required
@require_access("collecte", "ecriture")
def production_generer():
    annee = request.form.get("annee", type=int) or datetime.now().year
    camion = request.form.get("camion", "").strip().upper()
    # Sécurité + nommage de fichier : un code camion go-on-web est toujours
    # alphanumérique (ex. V003) — on rejette tout le reste plutôt que de
    # laisser passer un caractère de type '/', '..' etc. dans un nom de fichier.
    camion = re.sub(r"[^A-Z0-9]", "", camion)

    # Date du jeudi + liens Drive : renseignés une fois pour toutes sur la
    # page principale du module (un nouveau dossier/jeu de 3 liens est créé
    # par le club chaque année), plus besoin de les ressaisir ici.
    with get_db_connection() as conn:
        campagne = conn.execute(
            "SELECT * FROM collecte_campagnes WHERE annee = ?", (annee,)
        ).fetchone()
    date_jeudi = _date_fr_courte(campagne["date_debut"]) if campagne else None

    urls_drive = {}
    liens_manquants = []
    for cle, conf in DRIVE_CHAMPS_PRODUCTION.items():
        url = _url_export_drive(campagne[conf["champ"]]) if campagne else None
        if url:
            urls_drive[cle] = url
        else:
            liens_manquants.append(conf["label"])
    if liens_manquants:
        flash(
            f"❌ Lien(s) Drive non configuré(s) pour {annee} : {', '.join(liens_manquants)} — "
            f"à renseigner sur la page principale du module (section « Fichiers Drive »)",
            "danger"
        )
        return redirect(url_for("collecte.production", annee=annee))

    dossier = _dossier_production(annee)
    os.makedirs(dossier, exist_ok=True)

    # Téléchargement des 3 fichiers depuis le drive (partagés "Toute personne
    # disposant du lien") : la page production doit toujours refléter le
    # dernier état du planning go-on-web, pas un import ponctuel.
    try:
        for cle, url in urls_drive.items():
            reponse = requests.get(url, timeout=30)
            reponse.raise_for_status()
            if not reponse.content.startswith(b"PK"):
                raise ValueError(
                    f"contenu invalide pour « {cle} » — vérifier que le fichier est "
                    f"bien partagé en \"Toute personne disposant du lien\""
                )
            with open(os.path.join(dossier, f"{cle}.xlsx"), "wb") as f:
                f.write(reponse.content)
    except Exception as e:
        flash(f"❌ Échec du téléchargement des fichiers depuis le drive : {e}", "danger")
        write_log(f"❌ Génération documents production {annee} : échec téléchargement drive ({e})")
        return redirect(url_for("collecte.production", annee=annee))

    script_path = os.path.join(
        current_app.root_path, "ba38_collecte", "scripts", "generer_documents_production.py"
    )
    venv_python = os.path.join(current_app.root_path, "venv", "bin", "python")

    # Un camion précis : aperçu rapide de sa seule fiche de collecte,
    # ouverte directement dans le navigateur, sans toucher aux documents
    # officiels (classeur Excel, pointage, équipier, index, consignes,
    # carte) — --fiche-seule arrête le script juste après ce document.
    if camion:
        chemin_fiche = os.path.join(dossier, f"fiche_collecte_{annee}_{camion}.pdf")
        cmd = [
            venv_python, script_path,
            "--magasins", os.path.join(dossier, "magasins.xlsx"),
            "--vehicules", os.path.join(dossier, "vehicules.xlsx"),
            "--cagettes", os.path.join(dossier, "cagettes.xlsx"),
            "--annee", str(annee),
            "--camion", camion,
            "--fiche-seule",
            "--output-fiches", chemin_fiche,
            # Jamais réellement écrit en --fiche-seule (le script s'arrête
            # avant), mais nécessaire pour que le script déduise son dossier
            # de sortie par défaut (out_dir) du bon endroit plutôt que du
            # chemin Windows codé en dur.
            "--output-excel", os.path.join(dossier, PRODUCTION_FICHIERS_SORTIE["excel"]["nom"].format(annee=annee)),
        ]
        if date_jeudi:
            cmd += ["--date-jeudi", date_jeudi]

        resultat = subprocess.run(cmd, capture_output=True, text=True)
        sortie = (resultat.stdout or "") + "\n" + (resultat.stderr or "")

        if resultat.returncode != 0 or not os.path.exists(chemin_fiche):
            flash(f"❌ Échec de la génération de la fiche du camion {camion}", "danger")
            write_log(
                f"❌ Fiche camion {camion} ({annee}) en échec par {current_user.email}\n{sortie[-4000:]}"
            )
            return redirect(url_for("collecte.production", annee=annee))

        write_log(f"📄 Fiche camion {camion} ({annee}) générée par {current_user.email}")
        return send_file(chemin_fiche, mimetype="application/pdf", as_attachment=False)

    # Génération complète (tous camions) : régénère les 6 documents officiels.
    # La carte HTML est régénérée à chaque lancement sous un nom horodaté :
    # on supprime les anciennes avant de relancer (pas d'historique ici).
    for ancienne_carte in glob.glob(os.path.join(dossier, "carte_tournees_bai38_*.html")):
        os.remove(ancienne_carte)

    cmd = [
        venv_python, script_path,
        "--magasins", os.path.join(dossier, "magasins.xlsx"),
        "--vehicules", os.path.join(dossier, "vehicules.xlsx"),
        "--cagettes", os.path.join(dossier, "cagettes.xlsx"),
        "--annee", str(annee),
        "--output-excel", os.path.join(dossier, PRODUCTION_FICHIERS_SORTIE["excel"]["nom"].format(annee=annee)),
        "--output-fiches", os.path.join(dossier, PRODUCTION_FICHIERS_SORTIE["fiches"]["nom"].format(annee=annee)),
        "--output-pointage", os.path.join(dossier, PRODUCTION_FICHIERS_SORTIE["pointage"]["nom"].format(annee=annee)),
        "--output-equipier", os.path.join(dossier, PRODUCTION_FICHIERS_SORTIE["equipier"]["nom"].format(annee=annee)),
        "--output-index", os.path.join(dossier, PRODUCTION_FICHIERS_SORTIE["index"]["nom"].format(annee=annee)),
        "--output-vehicule-consignes", os.path.join(dossier, PRODUCTION_FICHIERS_SORTIE["consignes"]["nom"]),
    ]
    if date_jeudi:
        cmd += ["--date-jeudi", date_jeudi]

    resultat = subprocess.run(cmd, capture_output=True, text=True)
    sortie = (resultat.stdout or "") + "\n" + (resultat.stderr or "")

    with open(os.path.join(dossier, "dernier_journal.log"), "w", encoding="utf-8") as f:
        f.write(sortie)

    # Nom fixe pour la carte (générée sous un nom horodaté par le script) afin
    # d'offrir un lien de téléchargement stable.
    cartes = sorted(glob.glob(os.path.join(dossier, "carte_tournees_bai38_*.html")))
    if cartes:
        os.replace(cartes[-1], os.path.join(dossier, PRODUCTION_FICHIERS_SORTIE["carte"]["nom"]))

    if resultat.returncode != 0:
        flash("❌ Échec de la génération — voir le journal ci-dessous", "danger")
        write_log(
            f"❌ Génération documents production {annee} en échec par {current_user.email}\n"
            f"{sortie[-4000:]}"
        )
    else:
        flash("✅ Documents générés avec succès", "success")
        write_log(f"📄 Génération documents production {annee} par {current_user.email}")

    return redirect(url_for("collecte.production", annee=annee))


@collecte_bp.route("/collecte/production/telecharger/<cle>")
@login_required
@require_access("collecte", "lecture")
def production_telecharger(cle):
    annee = request.args.get("annee", type=int) or datetime.now().year

    if cle not in PRODUCTION_FICHIERS_SORTIE:
        flash("❌ Fichier inconnu", "danger")
        return redirect(url_for("collecte.production", annee=annee))

    nom = PRODUCTION_FICHIERS_SORTIE[cle]["nom"].format(annee=annee)
    chemin = os.path.join(_dossier_production(annee), nom)
    if not os.path.exists(chemin):
        flash("❌ Fichier introuvable — lancez une génération", "danger")
        return redirect(url_for("collecte.production", annee=annee))

    mimetype = "text/html" if cle == "carte" else None
    return send_file(chemin, as_attachment=(cle != "carte"), download_name=nom, mimetype=mimetype)


# ============================================================================
# 🏠 ASSOCIATIONS GARDANT — associations qui gardent leur collecte au lieu de
# la remettre à la BAI (magasins État='Collecte gardée' du référentiel).
# Croise deux exports go-on-web déjà utilisés ailleurs dans le module :
#   - liste_magasins.xlsx (FICHIERS['magasins']) : un magasin par ligne,
#     colonne 'Gardée par' = nom de l'association qui en assure la collecte
#     (source de vérité pour "qui garde quoi" — voir _construire_associations)
#   - liste_groupes.xlsx  (FICHIERS['groupes'])  : un groupe/association par
#     ligne, sert uniquement à enrichir chaque association trouvée dans
#     'Gardée par' (type, contacts, fiche groupe) — sa colonne 'Nombre
#     magasins' n'est PAS utilisée pour décider qui apparaît dans la liste,
#     car elle accuse parfois un retard sur 'Gardée par' (association trouvée
#     quand même, juste avec des métadonnées vides — voir 'trouvee').
# Étape 1 du sous-projet : constituer/télécharger la liste croisée. L'envoi
# d'une demande de résultats aux associations et la saisie du tonnage reçu
# sont des étapes suivantes, pas encore implémentées.
# ============================================================================

def _valeur_propre(v):
    return "" if pd.isna(v) else v


def _vif_fmt(v):
    """Formate un Code VIF en texte à 8 chiffres, zéros non significatifs
    conservés (ex. 9990005 → '09990005') — Excel/pandas lisent ce genre de
    code comme un nombre et perdent sinon le zéro de tête."""
    if pd.isna(v):
        return ""
    s = str(v).strip().split(".")[0]
    return s.zfill(8) if s.isdigit() else s


def _normaliser_gardee_par(s):
    """Nettoie la colonne 'Gardée par' pour la faire correspondre au nom du
    groupe : certains magasins à stockage partagé notent l'association sous
    la forme 'BAI+Nom' (ex. 'BAI+Equilibre') plutôt que juste 'Nom' — préfixe
    à retirer avant comparaison."""
    s = str(s).strip()
    return re.sub(r'^BAI\s*\+\s*', '', s, flags=re.IGNORECASE).strip()


def _lire_magasins_gardes(annee):
    """Magasins État='Collecte gardée' du référentiel de l'année (DataFrame
    vide si liste_magasins.xlsx est absent)."""
    try:
        chemin = _fichier_drive(annee, "magasins") or os.path.join(_dossier_annee(annee), FICHIERS["magasins"]["nom_stockage"])
    except Exception as erreur:
        write_log(f"⚠️ Lecture Drive magasins {annee} impossible : {erreur}")
        chemin = os.path.join(_dossier_annee(annee), FICHIERS["magasins"]["nom_stockage"])
    if not os.path.exists(chemin):
        return pd.DataFrame()
    df = pd.read_excel(chemin)
    df.columns = [str(c).strip() for c in df.columns]
    if "État" not in df.columns:
        return pd.DataFrame()
    return df[df["État"].astype(str).str.strip() == "Collecte gardée"].reset_index(drop=True)


def _lire_groupes(annee):
    """Tous les groupes (associations) go-on-web de l'année, colonnes
    normalisées (DataFrame vide si liste_groupes.xlsx est absent). Pas
    filtré sur 'Nombre magasins' : cette colonne, calculée côté go-on-web,
    accuse parfois un retard par rapport à la colonne 'Gardée par' du
    référentiel magasins (ex. une association vient de reprendre un magasin
    mais son compteur n'est pas encore remonté) — s'y fier pour décider
    qu'une association « n'existe pas » ferait disparaître des associations
    bien réelles de la liste."""
    try:
        chemin = _fichier_drive(annee, "groupes") or os.path.join(_dossier_annee(annee), FICHIERS["groupes"]["nom_stockage"])
    except Exception as erreur:
        write_log(f"⚠️ Lecture Drive groupes {annee} impossible : {erreur}")
        chemin = os.path.join(_dossier_annee(annee), FICHIERS["groupes"]["nom_stockage"])
    if not os.path.exists(chemin):
        return pd.DataFrame()
    df = pd.read_excel(chemin)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _lire_participants(annee):
    """Tous les participants/contacts go-on-web de l'année (toutes années de
    collecte confondues dans l'export), colonnes normalisées (DataFrame vide
    si liste_participants.xlsx est absent)."""
    try:
        chemin = _fichier_drive(annee, "participants") or os.path.join(_dossier_annee(annee), FICHIERS["participants"]["nom_stockage"])
    except Exception as erreur:
        write_log(f"⚠️ Lecture Drive participants {annee} impossible : {erreur}")
        chemin = os.path.join(_dossier_annee(annee), FICHIERS["participants"]["nom_stockage"])
    if not os.path.exists(chemin):
        return pd.DataFrame()
    df = pd.read_excel(chemin)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _referents_association(nom_asso, df_participants):
    """Contacts connus pour un groupe (association) donné, à partir de
    liste_participants.xlsx (colonne 'Groupe' — notée 'BAI+Nom' pour les
    magasins à stockage partagé, comme 'Gardée par' dans liste_magasins.xlsx,
    d'où la même normalisation). Dédoublonnés par (Nom, Email) ; ceux dont le
    nom porte la mention '(Ref)' (référent désigné côté go-on-web) sont mis
    en tête."""
    if df_participants.empty or "Groupe" not in df_participants.columns:
        return []
    groupe = df_participants["Groupe"].map(_normaliser_gardee_par)
    sub = df_participants[groupe == nom_asso].fillna("")
    if sub.empty:
        return []

    vus = set()
    contacts = []
    for _, r in sub.iterrows():
        nom = str(r.get("Nom", "")).strip()
        if not nom:
            continue
        email = str(r.get("Email", "")).strip()
        cle = (nom, email)
        if cle in vus:
            continue
        vus.add(cle)
        telephone = str(r.get("Portable", "")).strip() or str(r.get("Téléphone", "")).strip()
        contacts.append({
            "nom": nom,
            "email": email,
            "telephone": telephone,
            "referent": "ref" in nom.lower().replace("é", "e"),
        })
    contacts.sort(key=lambda c: (not c["referent"], c["nom"]))
    return contacts


def _associations_secondaires_stockage(stockage, primaire):
    """Certains magasins sont gardés à tour de rôle par DEUX associations
    (ex. une journée chacune) — la colonne 'Gardée par' n'en retient qu'une,
    la seconde n'apparaît que dans le texte libre 'Stockage' (ex.
    'Beurrepinard+la Roseraie'). Renvoie les noms distincts de la principale
    et de 'BAI' (qui désigne un stockage partagé avec la banque alimentaire,
    pas une association)."""
    stockage = str(stockage)
    if "+" not in stockage:
        return []

    def cle(s):
        # espaces ignorés : 'BAI+3ABI' et 'BAI+3 ABI' doivent être reconnus
        # comme la même association malgré l'incohérence de saisie.
        return re.sub(r"\s+", "", s).lower()

    primaire_cle = cle(primaire)
    tokens = [t.strip() for t in stockage.split("+") if t.strip()]
    return [t for t in tokens if cle(t) not in ("bai", primaire_cle)]


def _modele_gardee(nom):
    chemin = os.path.join(MODELES_GARDEE_DIR, nom)
    if not os.path.exists(chemin):
        chemin = os.path.join(current_app.root_path, "uploads", "collecte_fichiers_source", nom)
    return chemin


def _nom_fichier_association(nom, annee):
    slug = re.sub(r"[^A-Za-z0-9]+", "_", nom).strip("_").lower() or "association"
    return f"association_gardant_{annee}_{slug}.xlsx"


def _date_collecte(annee):
    with get_db_connection() as conn:
        campagne = conn.execute(
            "SELECT date_debut, date_fin FROM collecte_campagnes WHERE annee = ?",
            (annee,),
        ).fetchone()
    if not campagne or not campagne["date_debut"]:
        return str(annee)
    debut = campagne["date_debut"]
    fin = campagne["date_fin"]
    try:
        debut = datetime.strptime(debut, "%Y-%m-%d").strftime("%d/%m/%Y")
        fin = datetime.strptime(fin, "%Y-%m-%d").strftime("%d/%m/%Y") if fin else None
    except ValueError:
        pass
    return f"du {debut} au {fin}" if fin else debut


def _creer_fichier_association(asso, annee, dossier):
    source = _modele_gardee("Modele association.xlsx")
    if not os.path.exists(source):
        raise FileNotFoundError("Modèle Excel association introuvable")

    chemin = os.path.join(dossier, _nom_fichier_association(asso["nom"], annee))
    wb = load_workbook(source)
    code = ""
    date_collecte = _date_collecte(annee)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 2025:
                    cell.value = annee
                if isinstance(cell.value, str):
                    cell.value = cell.value.replace("àassociation", asso["nom"])
                    cell.value = re.sub(r"àcode(?!\d)", code, cell.value)
                    cell.value = cell.value.replace("àdatecollecte", date_collecte)
                    cell.value = cell.value.replace("àannéecollecte", str(annee))
                    cell.value = cell.value.replace("2025", str(annee))

    ws = wb["magasins"]
    for index, magasin in enumerate(asso["magasins"], start=1):
        row = 14 + index
        if row >= 25:
            ws.insert_rows(row)
        ws.cell(row=row, column=1).value = _vif_fmt(magasin.get("Code VIF"))
        ws.cell(row=row, column=2).value = _valeur_propre(magasin.get("Nom"))
        ws.cell(row=row, column=3).value = None
    for row in range(15 + len(asso["magasins"]), 24):
        for column in range(1, 10):
            ws.cell(row=row, column=column).value = None
    ws["C25"] = "=SUM(C15:C23)"
    for feuille, zone in ((wb["produits"], "A1:E44"), (ws, "A1:I27")):
        feuille.page_setup.orientation = "portrait"
        feuille.page_setup.fitToWidth = 1
        feuille.page_setup.fitToHeight = 1
        feuille.sheet_properties.pageSetUpPr.fitToPage = True
        feuille.print_area = zone
    wb.save(chemin)
    return chemin


def _creer_pdf_association(fichier_excel, asso, annee, dossier):
    """Crée directement les deux fiches PDF avec la présentation du modèle Excel."""
    wb = load_workbook(fichier_excel, data_only=False)
    nom = os.path.splitext(os.path.basename(fichier_excel))[0] + ".pdf"
    chemin = os.path.join(dossier, nom)
    page_width, page_height = A4
    pdf = pdf_canvas.Canvas(chemin, pagesize=A4)
    logo = os.path.join(current_app.root_path, "static", "images", "logo_ba_complet.png")

    def texte(valeur):
        return str(valeur or "").replace("\n", " ")

    def entete(ws, titre):
        if os.path.exists(logo):
            pdf.drawImage(ImageReader(logo), 12 * mm, page_height - 25 * mm, width=72 * mm, height=8 * mm, preserveAspectRatio=True, mask="auto")
        pdf.setFont("Helvetica", 8)
        pdf.drawString(24 * mm, page_height - 32 * mm, "11 allée de la Pinéa - 38600 FONTAINE")
        pdf.drawString(24 * mm, page_height - 38 * mm, "Tél. : 04.76.85.92.50")
        pdf.drawString(24 * mm, page_height - 44 * mm, "Mail : ba380.collecte@banquealimentaire.org")
        y = page_height - 53 * mm
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(70 * mm, y, "COLLECTE")
        pdf.drawString(105 * mm, y, str(annee))
        pdf.drawString(24 * mm, y - 8 * mm, _date_collecte(annee))
        pdf.drawString(105 * mm, y - 8 * mm, titre)
        pdf.line(24 * mm, y - 11 * mm, 175 * mm, y - 11 * mm)
        pdf.drawString(24 * mm, y - 19 * mm, asso["nom"])
        return y - 25 * mm

    def grille(lignes, largeurs, x, y, hauteur, gras_premiere=False):
        total = sum(largeurs)
        for index, ligne in enumerate(lignes):
            yy = y - index * hauteur
            pdf.setFont("Helvetica-Bold" if gras_premiere and index == 0 else "Helvetica", 6 if index else 6.5)
            xx = x
            for col, largeur in enumerate(largeurs):
                pdf.rect(xx, yy - hauteur, largeur, hauteur)
                pdf.drawString(xx + 1.2 * mm, yy - hauteur + 2.2 * mm, texte(ligne[col])[:38])
                xx += largeur
        return y - len(lignes) * hauteur

    ws = wb["produits"]
    y = entete(ws, "FICHE PRODUITS")
    lignes = [[ws.cell(14, c).value for c in range(1, 6)]]
    lignes.extend([[ws.cell(r, c).value for c in range(1, 6)] for r in range(16, 41) if any(ws.cell(r, c).value is not None for c in range(1, 6))])
    lignes.append(["", "Total", "", "", ""])
    grille(lignes, [25 * mm, 46 * mm, 22 * mm, 62 * mm, 22 * mm], 10 * mm, y, 6.2 * mm, True)
    pdf.showPage()

    ws = wb["magasins"]
    y = entete(ws, "FICHE MAGASINS")
    lignes = [[ws.cell(13, c).value for c in range(1, 9)]]
    lignes.extend([[ws.cell(r, c).value for c in range(1, 9)] for r in range(15, 24) if ws.cell(r, 1).value or ws.cell(r, 2).value])
    lignes.append(["", "Total", "", "", "", "", "", ""])
    grille(lignes, [18 * mm, 39 * mm, 20 * mm, 20 * mm, 20 * mm, 22 * mm, 20 * mm, 22 * mm], 5 * mm, y, 8 * mm, True)
    pdf.save()
    return chemin


def _texte_modele_gardee(asso, annee):
    source = _modele_gardee("associations-gardant.docx")
    if not os.path.exists(source):
        raise FileNotFoundError("Modèle de mail associations-gardant introuvable")
    document = Document(source)
    magasins = "\n".join(
        f"- {_vif_fmt(m.get('Code VIF'))} — {m.get('Nom', '')} ({m.get('Ville', '')})"
        for m in asso["magasins"]
    )
    lignes = []
    for paragraphe in document.paragraphs:
        texte = paragraphe.text
        texte = texte.replace("àassociation", asso["nom"])
        texte = texte.replace("àcode", "")
        texte = texte.replace("àdatecollecte", _date_collecte(annee))
        texte = texte.replace("àLISTE_MAGASINS", magasins)
        texte = texte.replace("àresponsable collecte", "Responsable collecte de la BA38")
        if texte.strip():
            lignes.append(texte)
    return "\n\n".join(lignes)


def _construire_associations(df_mag, df_groupes):
    """Rattache chaque magasin gardé à son (ou ses) association(s), à partir
    de la colonne 'Gardée par' (normalisée — voir _normaliser_gardee_par) et
    d'une éventuelle seconde association révélée par 'Stockage' (voir
    _associations_secondaires_stockage) : la liste des associations vient
    donc de df_mag, pas de df_groupes. Chaque association trouvée dans
    df_groupes est enrichie de ses métadonnées (type, contacts, fiche
    groupe) ; sinon ('trouvee': False) elle apparaît quand même, avec juste
    son nom et ses magasins — jamais de magasin perdu."""
    df_mag = df_mag.fillna("")
    gardee_par = df_mag.get("Gardée par", pd.Series(dtype=str)).map(_normaliser_gardee_par)

    groupes_par_nom = {}
    groupes_par_nom_lower = {}
    if not df_groupes.empty and "Nom" in df_groupes.columns:
        for _, g in df_groupes.iterrows():
            nom = str(g.get("Nom", "")).strip()
            groupes_par_nom[nom] = g
            groupes_par_nom_lower[nom.lower()] = g

    lignes_par_asso = {}
    for idx, row in df_mag.iterrows():
        primaire = gardee_par.loc[idx]
        if primaire:
            lignes_par_asso.setdefault(primaire, []).append(idx)
        for secondaire in _associations_secondaires_stockage(row.get("Stockage", ""), primaire):
            grp_sec = groupes_par_nom_lower.get(secondaire.lower())
            nom_cle = str(grp_sec.get("Nom")).strip() if grp_sec is not None else secondaire
            lignes_par_asso.setdefault(nom_cle, []).append(idx)

    associations = []
    for nom_asso in sorted(lignes_par_asso.keys(), key=str.lower):
        magasins_grp = df_mag.loc[lignes_par_asso[nom_asso]]
        grp = groupes_par_nom.get(nom_asso)
        associations.append({
            "nom": nom_asso,
            "type": _valeur_propre(grp.get("Type")) if grp is not None else "",
            "membre_bai": _valeur_propre(grp.get("Membre BAI")) if grp is not None else "",
            "nb_contacts": int(grp.get("Nombre contacts")) if grp is not None and pd.notna(grp.get("Nombre contacts")) else 0,
            "nb_leaders": int(grp.get("Nombre leaders")) if grp is not None and pd.notna(grp.get("Nombre leaders")) else 0,
            "fiche_groupe": _valeur_propre(grp.get("Fiche groupe")) if grp is not None else "",
            "trouvee": grp is not None,
            "magasins": magasins_grp[["Code VIF", "Nom", "Ville", "Stockage", "Contact", "Email"]].to_dict("records"),
        })

    sans_association = df_mag[gardee_par == ""][["Code VIF", "Nom", "Ville", "Stockage"]].to_dict("records")
    return associations, sans_association


@collecte_bp.route("/collecte/gardee")
@login_required
@require_access("collecte", "lecture")
def gardee():
    annee = request.args.get("annee", type=int) or datetime.now().year

    df_mag = _lire_magasins_gardes(annee)
    df_groupes = _lire_groupes(annee)
    if "Code VIF" in df_mag.columns:
        df_mag["Code VIF"] = df_mag["Code VIF"].map(_vif_fmt)

    manquants = []
    if df_mag.empty:
        manquants.append(FICHIERS["magasins"]["label"])
    if df_groupes.empty:
        manquants.append(FICHIERS["groupes"]["label"])

    associations, sans_association = ([], [])
    if not manquants:
        associations, sans_association = _construire_associations(df_mag, df_groupes)
        df_participants = _lire_participants(annee)
        for asso in associations:
            asso["referents"] = _referents_association(asso["nom"], df_participants)

    return render_template(
        "collecte/gardee.html",
        annee=annee,
        manquants=manquants,
        associations=associations,
        sans_association=sans_association,
        nb_magasins=len(df_mag),
        fichier_excel=os.path.exists(
            os.path.join(_dossier_annee(annee), f"associations_gardant_{annee}.xlsx")
        ),
    )


@collecte_bp.route("/collecte/gardee/excel")
@login_required
@require_access("collecte", "lecture")
def gardee_excel():
    annee = request.args.get("annee", type=int) or datetime.now().year

    df_mag = _lire_magasins_gardes(annee)
    df_groupes = _lire_groupes(annee)

    if df_mag.empty or df_groupes.empty:
        flash("❌ Fichier(s) manquant(s) (liste des magasins et/ou des groupes)", "danger")
        return redirect(url_for("collecte.gardee", annee=annee))

    if "Code VIF" in df_mag.columns:
        df_mag["Code VIF"] = df_mag["Code VIF"].map(_vif_fmt)

    associations, sans_association = _construire_associations(df_mag, df_groupes)
    df_participants = _lire_participants(annee)
    for asso in associations:
        asso["referents"] = _referents_association(asso["nom"], df_participants)

    def _fmt_referents(referents):
        return " ; ".join(
            f"{r['nom']}" + (f" <{r['email']}>" if r["email"] else "") + (f" ({r['telephone']})" if r["telephone"] else "")
            for r in referents
        )

    df_associations = pd.DataFrame([{
        "Nom": a["nom"],
        "Type": a["type"],
        "Membre BAI": a["membre_bai"],
        "Nombre contacts": a["nb_contacts"],
        "Nombre leaders": a["nb_leaders"],
        "Nombre magasins": len(a["magasins"]),
        "Trouvée dans liste des groupes": "Oui" if a["trouvee"] else "Non — nom absent de liste_groupes.xlsx",
        "Référent - Nom": a["referents"][0]["nom"] if a["referents"] else "",
        "Référent - Email": a["referents"][0]["email"] if a["referents"] else "",
        "Référent - Téléphone": a["referents"][0]["telephone"] if a["referents"] else "",
        "Autres contacts": _fmt_referents(a["referents"][1:]),
        "Fiche groupe": a["fiche_groupe"],
    } for a in associations])

    df_mag_flat = df_mag.copy()
    df_mag_flat["Association"] = df_mag_flat.get("Gardée par", "").map(_normaliser_gardee_par)
    colonnes_mag = ["Association", "Gardée par", "Code VIF", "Nom", "Ville", "Stockage",
                    "Contact", "Email", "Accord", "Commentaire", "Fiche magasin"]
    colonnes_mag = [c for c in colonnes_mag if c in df_mag_flat.columns]
    df_magasins = df_mag_flat[colonnes_mag].sort_values(["Association", "Nom"])

    # Onglet pivot : une association par ligne, ses magasins en colonnes
    # (même présentation que l'onglet 'Tournees' du module Production).
    nb_max_mag = max((len(a["magasins"]) for a in associations), default=0)
    colonnes_mag_pivot = [f"Magasin {i + 1}" for i in range(nb_max_mag)]
    rows_pivot = []
    for a in associations:
        noms_mags = [f"{m['Code VIF']} — {m['Nom']}" for m in a["magasins"]]
        row = {"Association": a["nom"], "Type": a["type"], "Nb magasins": len(noms_mags)}
        for i, col in enumerate(colonnes_mag_pivot):
            row[col] = noms_mags[i] if i < len(noms_mags) else ""
        rows_pivot.append(row)
    df_pivot = pd.DataFrame(rows_pivot, columns=["Association", "Type", "Nb magasins"] + colonnes_mag_pivot)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_associations.to_excel(writer, sheet_name="Associations", index=False)
        df_magasins.to_excel(writer, sheet_name="Magasins gardés", index=False)
        df_pivot.to_excel(writer, sheet_name="Associations - Magasins", index=False)
    buffer.seek(0)

    dossier = _dossier_annee(annee)
    os.makedirs(dossier, exist_ok=True)
    chemin = os.path.join(dossier, f"associations_gardant_{annee}.xlsx")
    try:
        for asso in associations:
            fichier_association = _creer_fichier_association(asso, annee, dossier)
            _creer_pdf_association(fichier_association, asso, annee, dossier)
    except RuntimeError as erreur:
        flash(f"❌ Création PDF impossible : {erreur}", "danger")
        return redirect(url_for("collecte.gardee", annee=annee))
    with open(chemin, "wb") as fichier:
        fichier.write(buffer.getvalue())

    write_log(f"📥 Export Associations gardant {annee} par {current_user.email}")
    flash(
        f"✅ Fichier associations_gardant_{annee}.xlsx créé. Contrôlez-le avant de préparer l'envoi.",
        "success",
    )
    return redirect(url_for("collecte.gardee", annee=annee))


@collecte_bp.route("/collecte/gardee/telecharger")
@login_required
@require_access("collecte", "lecture")
def gardee_telecharger():
    annee = request.args.get("annee", type=int) or datetime.now().year
    nom = f"associations_gardant_{annee}.xlsx"
    chemin = os.path.join(_dossier_annee(annee), nom)
    if not os.path.exists(chemin):
        flash("❌ Le fichier Excel n'a pas encore été créé", "danger")
        return redirect(url_for("collecte.gardee", annee=annee))
    return send_file(
        chemin,
        as_attachment=True,
        download_name=nom,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@collecte_bp.route("/collecte/gardee/envoi", methods=["GET", "POST"])
@login_required
@require_access("collecte", "ecriture")
def gardee_envoi():
    annee = request.args.get("annee", type=int) or request.form.get("annee", type=int) or datetime.now().year
    dossier = _dossier_annee(annee)
    chemin = os.path.join(dossier, f"associations_gardant_{annee}.xlsx")
    if not os.path.exists(chemin):
        flash("❌ Créez et contrôlez d'abord le fichier Excel", "danger")
        return redirect(url_for("collecte.gardee", annee=annee))

    df_mag = _lire_magasins_gardes(annee)
    df_groupes = _lire_groupes(annee)
    df_participants = _lire_participants(annee)
    if df_mag.empty or df_groupes.empty:
        flash("❌ Fichier(s) source(s) manquant(s)", "danger")
        return redirect(url_for("collecte.gardee", annee=annee))
    if "Code VIF" in df_mag.columns:
        df_mag["Code VIF"] = df_mag["Code VIF"].map(_vif_fmt)
    associations, _ = _construire_associations(df_mag, df_groupes)
    for asso in associations:
        asso["referents"] = _referents_association(asso["nom"], df_participants)
        asso["emails"] = sorted({r["email"] for r in asso["referents"] if "@" in r["email"]})

    if request.method == "POST":
        if request.form.get("confirmation") != "oui":
            flash("❌ Confirmez le contrôle du fichier avant l'envoi", "danger")
            return redirect(url_for("collecte.gardee_envoi", annee=annee))
        mode_test = request.form.get("mode_test") == "on"
        test_une_association = request.form.get("test_une_association") == "on"
        nom_test = request.form.get("association_test_nom", "")
        if test_une_association and not mode_test:
            flash("❌ L'option sur une seule association nécessite le mode Test", "danger")
            return redirect(url_for("collecte.gardee_envoi", annee=annee))
        associations_a_traiter = associations
        if test_une_association:
            associations_a_traiter = [a for a in associations if a["nom"] == nom_test]
            if not associations_a_traiter:
                flash("❌ Sélectionnez une association pour le test", "danger")
                return redirect(url_for("collecte.gardee_envoi", annee=annee))
        envoyes = 0
        sans_email = []
        for asso in associations_a_traiter:
            if not asso["emails"] and not mode_test:
                sans_email.append(asso["nom"])
                continue
            fichier_association = os.path.join(dossier, _nom_fichier_association(asso["nom"], annee))
            if not os.path.exists(fichier_association):
                fichier_association = _creer_fichier_association(asso, annee, dossier)
            fichier_pdf = os.path.splitext(fichier_association)[0] + ".pdf"
            if not os.path.exists(fichier_pdf):
                fichier_pdf = _creer_pdf_association(fichier_association, asso, annee, dossier)
            destinataires = [os.getenv("MAIL_TEST_TO") or current_user.email] if mode_test else asso["emails"]
            envoyer_mail(
                sujet=("[TEST] " if mode_test else "") + f"Collecte nationale Banque Alimentaire {annee} — {asso['nom']}",
                destinataires=destinataires,
                texte=_texte_modele_gardee(asso, annee),
                sender_override=os.getenv("MAILJET_SENDER"),
                cc=["ba380.collecte@banquealimentaire.org"],
                attachment_path=fichier_association,
                attachment_paths=[fichier_pdf],
            )
            envoyes += 1

        portee = "pour une association" if test_une_association else "pour toutes les associations"
        message = f"✅ {envoyes} mail(s) {'de test ' if mode_test else ''}envoyé(s) {portee} avec les fichiers Excel personnalisés"
        if sans_email:
            message += f" ; sans adresse : {', '.join(sans_email)}"
        flash(message, "success" if not sans_email else "warning")
        write_log(f"📧 Envoi Associations gardant {annee} : {envoyes} envoyé(s) par {current_user.email}")
        return redirect(url_for("collecte.gardee", annee=annee))

    return render_template(
        "collecte/gardee_envoi.html",
        annee=annee,
        associations=associations,
        fichier_nom=os.path.basename(chemin),
    )
