#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generer_documents_bai38_depuis_listes.py — BAI 38 - 2026

Variante de Generer_documents_bai38.py qui NE PART PAS d'un fichier Excel de
tournées déjà calculé (celui que produit generer_tournees_bai_v2.py à partir
du PDF des tournées de l'année précédente). Ici, les tournées sont
reconstruites directement à partir de deux fichiers déjà à jour :

  --vehicules  liste-vehicule.xlsx     (une ligne par passage : Code camion,
                                        Tournée/jour, Début, Quai, consignes,
                                        Code VIF, Magasin, Équipier, Portable)
  --magasins   liste-magasins.xlsx (référentiel magasins : adresse,
                                        horaires, VIF, coordonnées GPS, État...)

Elle produit les mêmes livrables que Generer_documents_bai38.py :

  A. Un classeur Excel des tournées reconstruit (onglets 'Tournees VIF',
     'Tournees', 'Magasins'), obtenu en regroupant liste-vehicule.xlsx par
     (Camion, Demi-journée) — PAS de calcul d'optimisation géographique ici,
     seulement la mise en forme de ce qui est déjà décidé dans liste-vehicule.

  B. Les 4 mêmes documents PDF que Generer_documents_bai38.py (fiches de
     collecte, pointage, équipier, index) + le rapport manquants.xlsx — en
     réutilisant telles quelles les fonctions de dessin de ce script.

Différences fonctionnelles volontaires par rapport à Generer_documents_bai38.py :
  - Pas de paramètre --excel : il n'y a plus de fichier de tournées externe,
    liste-vehicule.xlsx EST la source de vérité des affectations.
  - Les contrôles « Magasins sans passage » et « Camions différents » du
    rapport manquants.xlsx disparaissent : ils comparaient deux sources
    indépendantes (tournées calculées vs outil de collecte) ; ici il n'y en a
    plus qu'une seule, donc plus rien à comparer.
  - Le contrôle « Magasins non planifiés » (actifs dans le référentiel mais
    absents de liste-vehicule.xlsx) est conservé.

Hypothèse à vérifier sur liste-vehicule.xlsx : colonnes 'Code', 'Quai',
'Tournée', 'Début', 'consigne1', 'Consigne2', 'Code VIF', 'Magasin',
'Équipier', 'Portable' (ou 'Tel') — identiques à celles déjà lues par
Generer_documents_bai38.py. Si une colonne 'Nom camion' (ou similaire) existe
dans ce fichier, elle est détectée et utilisée automatiquement ; sinon le nom
du camion reste vide (sauf pour les véhicules figés connus, voir
NOMS_CAMIONS_CONNUS ci-dessous).

Usage :
    python Generer_documents_bai38_depuis_listes.py
    python Generer_documents_bai38_depuis_listes.py --vehicules liste-vehicule.xlsx --magasins liste-magasins.xlsx
    python Generer_documents_bai38_depuis_listes.py --camion V003   (fiches n°1 filtrée)
"""

import argparse, glob, os, re, datetime
from math import radians, cos, sin, asin, sqrt
from collections import Counter
import pandas as pd
import openpyxl
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.pdfgen import canvas as rl_canvas

# Dossier du script : sert d'ancrage pour le sous-dossier de sortie afin que
# la sortie se retrouve toujours au même endroit, quel que soit le dossier
# courant au moment du lancement (ex. si le script est lancé sans passer par
# le .bat qui fait 'cd' au préalable).
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Demi-journées (commun aux 4 documents) ─────────────────────────────────
DJ_LABELS = {
    'Jeudi Matin':         'Jeudi Matin',
    'Jeudi Apres Midi':    'Jeudi Après Midi',
    'Vendredi Matin':      'Vendredi Matin',
    'Vendredi Apres Midi': 'Vendredi Après Midi',
    'Samedi Matin':        'Samedi Matin',
    'Samedi Apres Midi':   'Samedi Après Midi',
    'Dimanche Matin':      'Dimanche Matin',
}
DJ_ORDER = list(DJ_LABELS.keys())

C_BLACK = colors.black

# Noms de camions de secours (utilisés seulement si liste-vehicule.xlsx ne
# contient aucune colonne 'Nom camion' exploitable) — repris de la config des
# véhicules figés de generer_tournees_bai_v2.py, à ajuster si besoin.
NOMS_CAMIONS_CONNUS = {
    'V007': 'CAMION 3ABI 1',
    'V008': 'CAMION 3ABI 2',
    'V009': 'CAMION SOLIFAIM',
    'V013': 'CAMION BD 1',
    'V023': 'CAMION RIGODON',
    'V026': 'Tournée spéciale',
    'V027': 'CAMION DLM 20 m3 1',
    'V028': 'CAMION DLM 20 m3 2 Roussillon',
}

# Véhicules à tournée figée/spéciale (repris de VEHICULES_FIGES dans
# generer_tournees_bai_v2.py) : leur planning ne suit pas le schéma habituel
# « les deux demi-journées d'un même jour couvertes ensemble » — ex. V026
# ('Tournée spéciale') ou V037 (n'existe que le Jeudi). Un magasin couvert par
# l'un de ces camions sur une seule demi-journée d'un jour n'est donc PAS une
# anomalie de planification et ne doit pas déclencher l'alerte 'Créneaux
# incomplets' (voir plus loin dans main()).
VEHICULES_FIGES = {'V007', 'V008', 'V009', 'V013', 'V023', 'V026', 'V027', 'V028', 'V037'}

# ── Secteurs géographiques : MÊME découpage que generer_tournees_bai_v2.py ──
# (reproduit ici pour que le classeur issu de Go on the Web affiche les mêmes
# secteurs que celui de l'optimisation des tournées).
BAI_LAT, BAI_LON = 45.18867, 5.68456  # 11 allée de la Pinea, 38600 Fontaine


def haversine(lat1, lon1, lat2, lon2):
    R = 6371
    dlat, dlon = radians(lat2 - lat1), radians(lon2 - lon1)
    a = sin(dlat / 2) ** 2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon / 2) ** 2
    return 2 * R * asin(sqrt(a))


def dist_bai(lat, lon):
    return haversine(BAI_LAT, BAI_LON, lat, lon)


def _normaliser_ville_secteur(s):
    import unicodedata
    s = unicodedata.normalize('NFD', str(s).strip()).encode('ascii', 'ignore').decode('ascii')
    s = re.sub(r"[-'’]", ' ', s)
    return re.sub(r'\s+', ' ', s).strip().title()


def calculer_secteur(lat, lon, ville):
    """
    Découpage géographique en secteurs — copie conforme de la fonction
    _calc_secteur_local de generer_tournees_bai_v2.py, pour que les deux
    classeurs de tournées (calcul d'optimisation vs Go on the Web) utilisent
    exactement le même référentiel de secteurs.
    """
    v = _normaliser_ville_secteur(ville)
    if 'grenoble' in v.lower():
        if lat > 45.193: return 'Grenoble Nord'
        if lat < 45.173: return 'Grenoble Sud'
        if lon < 5.714:  return 'Grenoble Ouest'
        if lon > 5.729:  return 'Grenoble Est'
        if lat > 45.187: return 'Grenoble Centre Nord'
        if lon < 5.721:  return 'Grenoble Centre Ouest'
        return 'Grenoble Centre Est'
    GRESIVAUDAN   = {'Biviers', 'Crolles', 'Froges', 'Saint Ismier', 'Domene', 'Le Versoud'}
    VOIRON_GRP    = {'Voiron', 'Saint Jean De Moirans'}
    RIVES_GRP     = {'Rives', 'Renage'}
    SEYSSINET_GRP = {'Seyssinet Pariset', 'Seyssins', 'Fontaine'}
    PLATEAU_NORD  = {'Apprieu', 'Colombe', 'Le Grand Lemps'}
    SUD_VERCORS   = {'Varces Allieres Et Risset', 'Vif', 'Vizille', 'Claix'}
    CHARTREUSE_E  = {'La Terrasse', 'Le Touvet', 'Tencin'}
    ROUSSILLON    = {'Peage De Roussillon', 'Salaise Sur Sanne'}
    SMH_GRP       = {'Poisat', 'Saint Martin Dheres', 'Saint Martin D Heres', 'Gieres', 'Gières'}
    if v in GRESIVAUDAN:   return 'Gresivaudan'
    if v in VOIRON_GRP:    return 'Voiron'
    if v in RIVES_GRP:     return 'Rives'
    if v in SEYSSINET_GRP: return 'Seyssinet'
    if v in PLATEAU_NORD:  return 'Plateau Nord'
    if v in SUD_VERCORS:   return 'Sud Vercors'
    if v in CHARTREUSE_E:  return 'Chartreuse Est'
    if v in ROUSSILLON:    return 'Roussillon'
    if v in SMH_GRP:       return 'Saint Martin Dheres'
    return v


# ══════════════════════════════════════════════════════════════════════════
# DOCUMENT 1 — Fiches de collecte (A4 paysage)
# ══════════════════════════════════════════════════════════════════════════
W, H    = landscape(A4)
ML      = 12 * mm   # marge gauche
MR      = 12 * mm   # marge droite
MT      = 12 * mm   # marge haute
MB      = 20 * mm   # marge basse

C_HDR_TXT   = colors.HexColor('#1F3864')   # bleu foncé titres colonnes
C_GREY      = colors.HexColor('#D9D9D9')   # gris clair lignes alternées
C_RED       = colors.HexColor('#C00000')   # rouge consignes bas de page
C_WHITE     = colors.white

# Largeur totale du tableau
TBL_W = W - ML - MR

# Largeurs des colonnes, calculées comme fractions de TBL_W pour occuper toute
# la largeur disponible (proportions mesurées sur le modèle 2025 :
# nom ≈ 28.1% | chaque tournée ≈ 14.4% | collecté dimanche ≈ 14.5%)
COL_NAME  = TBL_W * 0.2806
_COL_TRN  = TBL_W * 0.1437               # largeur totale d'une colonne "Tournée" (prendre + nb)
COL_PRND  = _COL_TRN * (16/25)           # "prendre" — même ratio prendre/nb qu'avant (16:9)
COL_NB    = _COL_TRN * (9/25)            # nb cagettes
COL_DIM   = TBL_W - COL_NAME - 4*_COL_TRN  # reste, garantit un tableau qui remplit TBL_W pile

# Hauteur fixe de chaque sous-ligne "prendre" / "cagettes ramenées" (mesurée sur le
# modèle 2025 : ~6mm/ligne, indépendante de la hauteur totale de la ligne magasin —
# le reste de la ligne reste blanc, pour l'écriture manuelle du nb de cagettes)
MINI_H = 6 * mm

TEL1 = '04 76 85 92 50'
TEL2 = '07 44 95 17 99'


# ── Utilitaires communs ──────────────────────────────────────────────────────
def vif_fmt(vif):
    s = str(vif).strip().split('.')[0]
    if s in ('', 'nan'): return ''
    return s.zfill(8) if s.isdigit() else s

def parse_creneaux(creneaux_raw):
    """
    Convertit le contenu de la colonne 'Créneaux' du référentiel magasins
    (une ligne par plage horaire, ex. 'vendredi 09h30-11h30\\nsamedi
    14h30-17h30\\n...') en un ensemble de demi-journées (valeurs de DJ_ORDER)
    où le magasin est censé être collecté — copie conforme de la fonction du
    même nom dans generer_tournees_bai_v2.py, pour un calcul identique dans
    les deux outils.

    Règle Matin/Après-midi : heure de début du créneau < 13h → Matin, sinon
    Après-midi. Le dimanche n'a qu'une seule demi-journée possible
    ('Dimanche Matin'), donc tout créneau du dimanche y est rattaché quelle
    que soit son heure de début.

    Retourne un set (éventuellement vide si aucune ligne n'a pu être
    interprétée), ou None si creneaux_raw est vide — pour distinguer "aucun
    créneau déclaré" (rien à contrôler) de "créneaux déclarés mais tous hors
    DJ_ORDER".
    """
    if not creneaux_raw or str(creneaux_raw).strip().lower() == 'nan':
        return None

    jours_fr = {'jeudi': 'Jeudi', 'vendredi': 'Vendredi', 'samedi': 'Samedi', 'dimanche': 'Dimanche'}
    djs = set()
    for ligne in re.split(r'[\n;]+', str(creneaux_raw)):
        ligne = ligne.strip().lower()
        if not ligne:
            continue
        m = re.match(r'(\w+)\s+(\d{1,2})h(\d{2})', ligne)
        if not m:
            continue
        jour_brut, heure_str, _ = m.groups()
        jour = jours_fr.get(jour_brut)
        if not jour:
            continue
        if jour == 'Dimanche':
            dj = 'Dimanche Matin'
        else:
            periode = 'Matin' if int(heure_str) < 13 else 'Apres Midi'
            dj = f'{jour} {periode}'
        if dj in DJ_ORDER:
            djs.add(dj)
    return djs

def adresse_fmt(ref_row):
    adr   = str(ref_row.get('Adresse', '')).strip()
    cp    = str(ref_row.get('C.P.', '')).strip().split('.')[0]
    ville = str(ref_row.get('Ville', '')).strip()
    parts = [p for p in [adr, cp, ville] if p and p != 'nan']
    return ' '.join(parts)

def horaires_fmt(ref_row):
    h = str(ref_row.get('Horaires', '')).strip()
    if h in ('', 'nan'): return ''
    return h.split('\n')[0].strip()

def normaliser_dj(s):
    """Normalise une chaîne 'jour' (avec variations de casse/accents) vers une clé
    DJ_ORDER (ex: 'vendredi Après Midi' / 'Vendredi Apres-midi' → 'Vendredi Apres Midi')."""
    s = str(s).strip()
    if not s or s == 'nan':
        return ''
    s_low = s.lower().replace('è', 'e').replace('é', 'e').replace('à', 'a')
    jours = ['jeudi', 'vendredi', 'samedi', 'dimanche']
    jour = next((j for j in jours if s_low.startswith(j)), '')
    if not jour:
        return ''
    periode = 'Apres Midi' if 'apres' in s_low or 'am' in s_low.split() else 'Matin'
    return f"{jour.capitalize()} {periode}"

def camion_sort_key(code):
    """Trie les codes camion (V001, V002, ..., V037, VX300, ...) numériquement."""
    s = str(code).strip()
    digits = ''.join(c for c in s if c.isdigit())
    prefix = ''.join(c for c in s if not c.isdigit())
    return (prefix, int(digits) if digits else 0)

JOURS_FR = ['lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi', 'samedi', 'dimanche']
MOIS_FR = ['', 'janvier', 'février', 'mars', 'avril', 'mai', 'juin', 'juillet',
           'août', 'septembre', 'octobre', 'novembre', 'décembre']

def date_fr(dt):
    """Formate une date en français ('jeudi 27 novembre 2025 14:56:14')."""
    return f'{JOURS_FR[dt.weekday()]} {dt.day} {MOIS_FR[dt.month]} {dt.year} {dt.strftime("%H:%M:%S")}'


def sauver_pdf_avec_repli(cv, chemin):
    """Sauvegarde un canvas ReportLab ; si le fichier est verrouillé (déjà ouvert
    dans un lecteur, sync Drive/OneDrive en cours...), bascule sur un nom horodaté
    plutôt que de planter. Retourne le chemin réellement utilisé."""
    try:
        cv.save()
        return chemin
    except PermissionError:
        base, ext = os.path.splitext(chemin)
        horodatage = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        chemin_repli = f'{base}_{horodatage}{ext}'
        print(f"AVERTISSEMENT : {chemin} est verrouillé (probablement déjà ouvert) "
              f"→ sauvegarde sous {os.path.basename(chemin_repli)} à la place.")
        cv._filename = chemin_repli
        cv.save()
        return chemin_repli


def sauver_xlsx_avec_repli(wb, chemin):
    """Sauvegarde un classeur openpyxl ; même repli que sauver_pdf_avec_repli."""
    try:
        wb.save(chemin)
        return chemin
    except PermissionError:
        base, ext = os.path.splitext(chemin)
        horodatage = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        chemin_repli = f'{base}_{horodatage}{ext}'
        print(f"AVERTISSEMENT : {chemin} est verrouillé (probablement déjà ouvert) "
              f"→ sauvegarde sous {os.path.basename(chemin_repli)} à la place.")
        wb.save(chemin_repli)
        return chemin_repli


# ── Dessin des lignes de tableau (document 1) ────────────────────────────────
def rect_border(cv, x, y, w, h, fill=None, lw=0.5):
    """Dessine un rectangle avec bordure fine."""
    cv.setLineWidth(lw)
    if fill:
        cv.setFillColor(fill)
        cv.rect(x, y, w, h, fill=1, stroke=0)
        cv.setFillColor(C_BLACK)
    cv.setStrokeColor(C_BLACK)
    cv.rect(x, y, w, h, fill=0, stroke=1)

def txt(cv, x, y, s, font='Helvetica', size=8, color=C_BLACK, align='left', maxw=None, shrink=False, min_size=6):
    """
    Dessine du texte. Si maxw est dépassé :
      - si shrink=True, réduit la taille de police (jusqu'à min_size) avant de tronquer
      - tronque en dernier recours en ajoutant une ellipse '…' pour signaler la coupe
    """
    if maxw and len(s) > 0:
        if shrink:
            while size > min_size and cv.stringWidth(s, font, size) > maxw:
                size -= 1
        if cv.stringWidth(s, font, size) > maxw:
            ell = '…'
            while len(s) > 1 and cv.stringWidth(s + ell, font, size) > maxw:
                s = s[:-1]
            s = s.rstrip() + ell
    cv.setFont(font, size)
    cv.setFillColor(color)
    if align == 'center':
        cv.drawCentredString(x, y, s)
    elif align == 'right':
        cv.drawRightString(x, y, s)
    else:
        cv.drawString(x, y, s)
    cv.setFillColor(C_BLACK)


def draw_fiche(cv, dj, camion, nom_camion, magasins, annee=2026, quai='', consigne1='', consigne2=''):
    """
    Dessine une fiche complète sur la page courante.
    magasins : liste de dicts {nom, vif, adresse, horaires, en_dimanche, cag}
               'cag' est une liste de 4 valeurs (str) pour les nb de cagettes
               des Tournées 1 à 4 (vide si inconnu → case à remplir à la main).
    """
    # ── EN-TÊTE ──────────────────────────────────────────────────────────────
    y = H - MT

    dj_label = DJ_LABELS.get(dj, dj)
    txt(cv, ML, y, str(annee), size=9)
    txt(cv, ML + 20*mm, y, dj_label, size=9)
    txt(cv, ML + 65*mm, y, f'Vehicule:', size=9)
    txt(cv, ML + 80*mm, y, camion, font='Helvetica-Bold', size=9)
    txt(cv, ML + 92*mm, y, nom_camion, size=9)
    txt(cv, W - MR - 30*mm, y, 'Enregistré VIF', font='Helvetica-Bold', size=9)
    rect_border(cv, W - MR - 8*mm, y - 1*mm, 8*mm, 5*mm, lw=0.7)
    y -= 8*mm

    txt(cv, ML, y, 'Quai:', size=9)
    rect_border(cv, ML + 10*mm, y - 1.5*mm, 20*mm, 5*mm, lw=0.5)
    if quai:
        txt(cv, ML + 12*mm, y, quai, font='Helvetica-Bold', size=9)
    y -= 8*mm

    # ── EN-TÊTE TABLEAU ──────────────────────────────────────────────────────
    x0 = ML
    x1 = x0 + COL_NAME
    HDR_H = 7 * mm

    tournee_w = COL_PRND + COL_NB
    for i in range(4):
        xi = x1 + i * tournee_w
        rect_border(cv, xi, y - HDR_H, tournee_w, HDR_H, fill=None, lw=0.7)
        txt(cv, xi + tournee_w/2, y - HDR_H + 2*mm,
            f'Tournée {i+1}', font='Helvetica-Bold', size=8, align='center')

    x_dim = x1 + 4 * tournee_w
    rect_border(cv, x_dim, y - HDR_H, COL_DIM, HDR_H, fill=None, lw=0.7)
    txt(cv, x_dim + COL_DIM/2, y - HDR_H + 2*mm, 'Collecté dimanche',
        font='Helvetica-Bold', size=8, align='center', maxw=COL_DIM - 2*mm, shrink=True, min_size=6)

    y -= HDR_H

    # ── LIGNES MAGASINS ───────────────────────────────────────────────────────
    FOOTER_CLEARANCE = 28 * mm
    ZONE_TABLEAU = H - MT - MB - (22 * mm) - FOOTER_CLEARANCE
    HDR_ROW_H = 7 * mm
    ZONE_MAG = ZONE_TABLEAU - HDR_ROW_H
    nb = len(magasins)
    ROW_H = min(55 * mm, max(20 * mm, ZONE_MAG / nb if nb > 0 else 50 * mm))

    for idx_mag, mag in enumerate(magasins):
        nom      = mag.get('nom', '')
        vif      = mag.get('vif', '')
        adr      = mag.get('adresse', '')
        horaires = mag.get('horaires', '')
        en_dim   = mag.get('en_dimanche', False)
        cag      = mag.get('cag', ['', '', '', ''])

        if y - ROW_H < MB + FOOTER_CLEARANCE:
            _draw_footer(cv, dj, consigne1, consigne2)
            cv.showPage()
            y = H - MT - 5*mm
            txt(cv, ML, y, f'{annee}   {dj_label}   {camion}  (suite)',
                font='Helvetica-Bold', size=9)
            y -= 8*mm
            for i in range(4):
                xi = x1 + i * tournee_w
                rect_border(cv, xi, y - HDR_H, tournee_w, HDR_H, fill=None, lw=0.7)
                txt(cv, xi + tournee_w/2, y - HDR_H + 2*mm,
                    f'Tournée {i+1}', font='Helvetica-Bold', size=8, align='center')
            rect_border(cv, x_dim, y - HDR_H, COL_DIM, HDR_H, fill=None, lw=0.7)
            txt(cv, x_dim + COL_DIM/2, y - HDR_H + 2*mm, 'Collecté dimanche',
                font='Helvetica-Bold', size=8, align='center', maxw=COL_DIM - 2*mm, shrink=True, min_size=6)
            y -= HDR_H

        bg = C_WHITE if idx_mag % 2 == 0 else C_GREY

        rect_border(cv, x0, y - ROW_H, COL_NAME, ROW_H, fill=bg, lw=0.5)

        y_nom = y - ROW_H * 0.18
        y_adr = y - ROW_H * 0.38
        y_hor = y - ROW_H * 0.62
        y_vif = y - ROW_H * 0.82

        txt(cv, x0 + 1.5*mm, y_nom, nom, font='Helvetica-Bold', size=9,
            maxw=COL_NAME - 3*mm, shrink=True, min_size=7)
        txt(cv, x0 + 1.5*mm, y_adr, adr, size=8, maxw=COL_NAME - 3*mm,
            shrink=True, min_size=6)
        if horaires:
            txt(cv, x0 + COL_NAME/2, y_hor, horaires,
                font='Helvetica-Oblique', size=8, align='center')
        txt(cv, x0 + 1.5*mm, y_vif, vif, size=9)

        for i in range(4):
            xi_prd = x1 + i * tournee_w
            xi_nb  = xi_prd + COL_PRND

            rect_border(cv, xi_prd, y - ROW_H, COL_PRND + COL_NB, ROW_H, fill=bg, lw=0.5)

            rect_border(cv, xi_prd, y - MINI_H, COL_PRND, MINI_H, lw=0.5)
            txt(cv, xi_prd + 1.5*mm, y - MINI_H + MINI_H*0.3, 'prendre', size=8)
            rect_border(cv, xi_nb, y - MINI_H, COL_NB, MINI_H, lw=0.7)
            if i < len(cag) and cag[i]:
                txt(cv, xi_nb + COL_NB/2, y - MINI_H + MINI_H*0.3, str(cag[i]),
                    font='Helvetica-Bold', size=8, align='center')

            rect_border(cv, xi_prd, y - 2*MINI_H, COL_PRND + COL_NB, MINI_H, lw=0.5)
            txt(cv, xi_prd + 1.5*mm, y - 2*MINI_H + MINI_H*0.3, 'cagettes ramenées', size=7)

        # « Collecté dimanche » : seule la demi-journée Samedi Après-midi affiche
        # cette information (dernier passage avant la collecte du dimanche), sous
        # forme du mot « Dimanche » écrit en clair — pas une case à cocher — et
        # uniquement pour les magasins concernés. Les autres demi-journées et les
        # magasins non concernés du samedi après-midi laissent la cellule vide,
        # conformément au modèle papier 2025.
        rect_border(cv, x_dim, y - ROW_H, COL_DIM, ROW_H, fill=bg, lw=0.5)
        if normaliser_dj(dj) == 'Samedi Apres Midi' and en_dim:
            txt(cv, x_dim + COL_DIM/2, y - ROW_H/2, 'Dimanche',
                font='Helvetica-Bold', size=9, align='center')

        y -= ROW_H

    cv.setLineWidth(0.5)
    cv.setStrokeColor(C_BLACK)
    cv.line(ML, y, x_dim + COL_DIM, y)

    _draw_footer(cv, dj, consigne1, consigne2)


def _draw_footer(cv, dj, consigne1='', consigne2=''):
    """Pied de page identique au PDF 2025."""
    y_base = MB + 18*mm

    cv.setFont('Helvetica-Bold', 9)
    cv.setFillColor(C_RED)
    cv.drawRightString(W - MR, y_base + 8*mm,
                       'Ramener cette fiche remplie au secrétariat en fin de demi-journée')

    cv.setFont('Helvetica-Bold', 10)
    cv.setFillColor(C_BLACK)
    cv.drawString(W - MR - 55*mm, y_base + 2*mm, f'tel BAI : {TEL1}')
    cv.drawString(W - MR - 55*mm, y_base - 3*mm, f'ou       {TEL2}')

    cv.setFont('Helvetica-Bold', 14)
    cv.setFillColor(C_RED)
    cv.drawRightString(W - MR, y_base - 3*mm, 'MERCI')

    lignes_consignes = [c for c in [consigne1, consigne2] if c]
    cv.setFont('Helvetica-Bold', 10)
    cv.setFillColor(C_RED)
    for idx, ligne in enumerate(lignes_consignes):
        y_ligne = MB + 5*mm + (len(lignes_consignes) - 1 - idx) * 6*mm
        cv.drawString(ML, y_ligne, ligne)

    cv.setFillColor(C_BLACK)


# ══════════════════════════════════════════════════════════════════════════
# DOCUMENTS 2, 3, 4 — Pointage / Équipier / Index (A4 portrait)
# Constantes et fonctions préfixées P* pour ne pas entrer en collision avec
# la géométrie (paysage) du document 1 ci-dessus.
# ══════════════════════════════════════════════════════════════════════════
PW, PH = A4
PML, PMR, PMT, PMB = 20*mm, 20*mm, 20*mm, 15*mm

PC_BLUE = colors.HexColor('#4472C4')   # libellés Départ:/Retour: (modèle 2025)
PC_BORD = colors.HexColor('#8EA9DB')   # bordures de tableau (modèle 2025)

# Document 2 — pointage
P_COL_CODE = 18*mm
P_COL_NOM  = 67*mm
P_COL_LBL  = 14*mm
P_COL_VAL  = 13*mm
P_ROW_H    = 7*mm

# Document 3 — jour/véhicule/magasin/équipier
EQ_COL_CODE = 14*mm
EQ_COL_NOM  = 42*mm
EQ_COL_EQP  = 65*mm
EQ_COL_MAG  = PW - PML - PMR - EQ_COL_CODE - EQ_COL_NOM - EQ_COL_EQP
EQ_LINE_H   = 4.2*mm

# Document 4 — index alphabétique équipier
IDX_COL_NOM  = 65*mm
IDX_COL_CODE = 15*mm
IDX_ROW_H    = 5.2*mm


def fit_text(cv, s, font, size, maxw):
    """Tronque s pour qu'il tienne dans maxw (largeur réelle, pas un nombre de caractères)."""
    if cv.stringWidth(s, font, size) <= maxw:
        return s
    while len(s) > 1 and cv.stringWidth(s + '…', font, size) > maxw:
        s = s[:-1]
    return s.rstrip() + '…'


def draw_page(cv, dj_label, annee, camions):
    """Document 2 : pointage Départ/Retour."""
    y = PH - PMT

    cv.setFont('Helvetica', 11)
    cv.setFillColor(C_BLACK)
    cv.drawString(PML, y, dj_label)
    cv.drawString(PML + 100*mm, y, str(annee))
    y -= 12*mm

    x_code = PML
    x_nom  = x_code + P_COL_CODE
    x_dlbl = x_nom + P_COL_NOM
    x_dval = x_dlbl + P_COL_LBL
    x_rlbl = x_dval + P_COL_VAL
    x_rval = x_rlbl + P_COL_LBL
    x_end  = x_rval + P_COL_VAL

    cv.setLineWidth(0.6)
    for code, nom in camions:
        y_bot = y - P_ROW_H

        cv.setStrokeColor(PC_BORD)
        for xa, xb in [(x_code, x_nom), (x_nom, x_dlbl), (x_dlbl, x_dval),
                       (x_dval, x_rlbl), (x_rlbl, x_rval), (x_rval, x_end)]:
            cv.rect(xa, y_bot, xb - xa, P_ROW_H, fill=0, stroke=1)

        cv.setFillColor(C_BLACK)
        cv.setFont('Helvetica', 9)
        cv.drawString(x_code + 1.5*mm, y_bot + 2.3*mm, code)
        cv.drawString(x_nom + 1.5*mm, y_bot + 2.3*mm, nom)

        cv.setFillColor(PC_BLUE)
        cv.drawString(x_dlbl + 1.5*mm, y_bot + 2.3*mm, 'Départ:')
        cv.drawString(x_rlbl + 1.5*mm, y_bot + 2.3*mm, 'Retour:')

        y -= P_ROW_H
        if y - P_ROW_H < PMB:
            cv.showPage()
            y = PH - PMT
            cv.setFont('Helvetica', 11)
            cv.setFillColor(C_BLACK)
            cv.drawString(PML, y, f'{dj_label} (suite)')
            cv.drawString(PML + 100*mm, y, str(annee))
            y -= 12*mm

    cv.setFillColor(C_BLACK)


def _eq_header(cv, dj_label, date_collecte, date_gen, page_num):
    """En-tête de page du document 3 (sur le modèle 2025)."""
    cv.setFont('Helvetica-Bold', 11)
    cv.setFillColor(C_BLACK)
    cv.drawString(PML, PH - PMT, dj_label)
    cv.setFont('Helvetica', 10)
    x = PML + 45*mm
    if date_collecte:
        cv.drawString(x, PH - PMT, date_collecte)
        x += 30*mm
    cv.drawString(x, PH - PMT, date_gen)
    cv.drawString(PW - PMR - 20*mm, PH - PMT, f'Page {page_num}')
    cv.setFillColor(C_BLACK)


def _draw_box_lines(cv, x, y_top, w, lines, line_h, font='Helvetica', size=8, split_col=None):
    """
    Dessine une bordure rectangulaire dimensionnée exactement au nombre de lignes,
    puis le texte à l'intérieur (une chaîne par ligne, ou un tuple (gauche, droite)
    par ligne si split_col est fourni — utilisé pour 'Nom  téléphone').
    Retourne la hauteur du bloc dessiné.
    """
    nb = max(1, len(lines))
    box_h = nb * line_h
    cv.setStrokeColor(PC_BORD)
    cv.setLineWidth(0.6)
    cv.rect(x, y_top - box_h, w, box_h, fill=0, stroke=1)
    cv.setFillColor(C_BLACK)
    cv.setFont(font, size)
    for i, item in enumerate(lines):
        ly = y_top - (i+1)*line_h + line_h*0.28
        if split_col is not None and isinstance(item, tuple):
            gauche, droite = item
            cv.drawString(x + 1.5*mm, ly, fit_text(cv, gauche, font, size, split_col - 3*mm))
            cv.drawString(x + split_col + 1.5*mm, ly, droite)
        else:
            cv.drawString(x + 1.5*mm, ly, fit_text(cv, item, font, size, w - 3*mm))
    return box_h


def draw_equipier_page(cv, dj_label, camions_rows, date_collecte, date_gen, page_num):
    """
    Document 3 : Code | Nom camion | Équipiers (bloc bordé, dimensionné à son
    contenu) | Magasins (bloc bordé, dimensionné à son contenu — hauteurs
    indépendantes, non partagées).
    camions_rows : liste de (code, nom_camion, [magasins], [(equipier, tel)])
    Retourne le numéro de la dernière page dessinée (pour la numérotation continue).
    """
    _eq_header(cv, dj_label, date_collecte, date_gen, page_num)
    y = PH - PMT - 9*mm

    x_code = PML
    x_nom  = x_code + EQ_COL_CODE
    x_eqp  = x_nom + EQ_COL_NOM
    x_mag  = x_eqp + EQ_COL_EQP
    ROW_GAP = 3*mm

    for code, nom, magasins, equipiers in camions_rows:
        nb_lignes = max(1, len(magasins), len(equipiers))
        row_h_estime = nb_lignes * EQ_LINE_H + ROW_GAP

        if y - row_h_estime < PMB:
            cv.showPage()
            page_num += 1
            _eq_header(cv, dj_label + ' (suite)', date_collecte, date_gen, page_num)
            y = PH - PMT - 9*mm

        y_top = y
        cv.setFillColor(C_BLACK)
        cv.setFont('Helvetica', 9)
        cv.drawString(x_code + 1*mm, y_top - EQ_LINE_H + EQ_LINE_H*0.28, code)
        cv.drawString(x_nom + 1*mm, y_top - EQ_LINE_H + EQ_LINE_H*0.28,
                       fit_text(cv, nom, 'Helvetica', 9, EQ_COL_NOM - 2*mm))

        eq_lines = [(nomp, tel) for nomp, tel in equipiers]
        h_eq  = _draw_box_lines(cv, x_eqp, y_top, EQ_COL_EQP, eq_lines, EQ_LINE_H,
                                 split_col=EQ_COL_EQP * 0.62)
        h_mag = _draw_box_lines(cv, x_mag, y_top, EQ_COL_MAG, magasins, EQ_LINE_H)

        y -= max(h_eq, h_mag) + ROW_GAP

    cv.setFillColor(C_BLACK)
    return page_num


def draw_index_page(cv, dj_label, rows, page_num, page_w, page_h, col_nom=IDX_COL_NOM, ncols=1, boxed=False):
    """
    Document 4 : index alphabétique des équipiers en ncols colonnes équilibrées
    côte à côte (pas un simple débordement — la liste est répartie sur toutes
    les colonnes disponibles dès le départ, pour exploiter toute la largeur
    même quand le contenu tiendrait dans une seule colonne).
    rows : liste de (nom_equipier, code_camion, nom_camion), déjà triée.
    boxed : si True, chaque ligne est encadrée (Nom | Code | Nom camion), avec
    un petit espace blanc entre les lignes, sur le modèle papier fourni.
    Retourne le numéro de la dernière page dessinée.
    """
    GUTTER = 10*mm
    ROW_GAP = 2.5*mm if boxed else 0
    row_pitch = IDX_ROW_H + ROW_GAP
    zone_w = page_w - PML - PMR
    block_w = (zone_w - GUTTER * (ncols - 1)) / ncols
    col_cam = block_w - col_nom - IDX_COL_CODE
    y_top = page_h - PMT - 9*mm
    rows_per_col = max(1, int((y_top - PMB) / row_pitch))
    rows_per_page = rows_per_col * ncols

    def _draw_header():
        cv.setFont('Helvetica-Bold', 11)
        cv.setFillColor(C_BLACK)
        cv.drawString(PML, page_h - PMT, dj_label)

    def _block_x(i):
        return PML + i * (block_w + GUTTER)

    i0 = 0
    while i0 < len(rows):
        _draw_header()
        page_chunk = rows[i0:i0 + rows_per_page]
        rows_per_col_ici = -(-len(page_chunk) // ncols)  # ceil(len/ncols)
        for j, (nom_eq, code, nom_cam) in enumerate(page_chunk):
            col_i = j // rows_per_col_ici
            pos_in_col = j % rows_per_col_ici
            x_nom  = _block_x(col_i)
            x_code = x_nom + col_nom
            x_cam  = x_code + IDX_COL_CODE
            y = y_top - pos_in_col * row_pitch

            if boxed:
                cv.setStrokeColor(PC_BORD)
                cv.setLineWidth(0.6)
                cv.rect(x_nom, y - IDX_ROW_H, col_nom + IDX_COL_CODE + col_cam, IDX_ROW_H, fill=0, stroke=1)
                cv.line(x_code, y - IDX_ROW_H, x_code, y)
                cv.line(x_cam, y - IDX_ROW_H, x_cam, y)
                ty = y - IDX_ROW_H + IDX_ROW_H*0.3
            else:
                ty = y

            cv.setFont('Helvetica', 9)
            cv.setFillColor(C_BLACK)
            cv.drawString(x_nom + (1.5*mm if boxed else 0), ty, fit_text(cv, nom_eq, 'Helvetica', 9, col_nom - (3*mm if boxed else 2*mm)))
            cv.drawString(x_code + (1.5*mm if boxed else 0), ty, code)
            cv.drawString(x_cam + (1.5*mm if boxed else 0), ty, fit_text(cv, nom_cam, 'Helvetica', 9, col_cam - (3*mm if boxed else 2*mm)))

        i0 += rows_per_page
        if i0 < len(rows):
            cv.showPage()
            page_num += 1

    return page_num


# ══════════════════════════════════════════════════════════════════════════
# CONSTRUCTION DES TOURNÉES DEPUIS liste-vehicule.xlsx (au lieu du PDF 2025)
# ══════════════════════════════════════════════════════════════════════════
def _normaliser_colonne(s):
    """Minuscules + accents supprimés, pour comparer les noms de colonnes
    sans se soucier de la casse ni des accents ('Véhicule' -> 'vehicule')."""
    import unicodedata
    s = unicodedata.normalize('NFD', str(s).strip().lower())
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn')


def detecter_colonne(df_cols, contient_tous):
    """Retourne le 1er nom de colonne (parmi df_cols) dont la version
    normalisée (minuscules, sans accents) contient tous les fragments de
    contient_tous, sinon None."""
    for col in df_cols:
        cl = _normaliser_colonne(col)
        if all(frag in cl for frag in contient_tous):
            return col
    return None


def detecter_colonne_exacte(df_cols, noms_possibles):
    """Retourne le 1er nom de colonne (parmi df_cols) dont la version
    normalisée est EXACTEMENT égale à l'un des noms de noms_possibles
    (déjà normalisés), sinon None. Utilisé pour une colonne comme
    'Véhicule' qui contient le nom du camion (à ne pas confondre avec
    'Code' qui contient le code du camion)."""
    for col in df_cols:
        if _normaliser_colonne(col) in noms_possibles:
            return col
    return None


def construire_tournees_depuis_vehicules(args):
    """
    Lit liste-vehicule.xlsx (une ligne par passage magasin/camion/demi-journée)
    et reconstruit la structure normalement issue du fichier Excel de tournées
    (celle que lit Generer_documents_bai38.py) :

      - df           : DataFrame 1 ligne par (Demi-journee, Camion), colonnes
                        'Demi-journee','Camion','Nom camion',
                        'Code VIF 1'/'Magasin 1', 'Code VIF 2'/'Magasin 2', ...
      - mag_cols      : ['Magasin 1', 'Magasin 2', ...]
      - vif_cols      : ['Code VIF 1', 'Code VIF 2', ...]
      - quai_par_camion, consignes_par_camion_dj, equipiers_par_camion_dj,
        codes_veh : mêmes structures que Generer_documents_bai38.py, pour que
        la suite du script (documents 1/2/3/4) reste inchangée.
      - dim_set      : noms des magasins collectés le Dimanche (déduit
                        directement des affectations 'Dimanche Matin').

    Retourne un dict avec toutes ces clés, plus 'df_veh' (le DataFrame brut,
    utile pour le contrôle « magasins non planifiés »).
    """
    df_veh = pd.read_excel(args.vehicules)
    df_veh.columns = [str(c).strip() for c in df_veh.columns]
    cols = list(df_veh.columns)

    # Priorité : une colonne explicite 'Nom camion' / 'Nom véhicule', sinon une
    # colonne 'Véhicule' toute seule (fréquent : 'Code' = code camion type V003,
    # 'Véhicule' = son nom/description type 'CAMION 3ABI 1').
    col_nom_camion = (detecter_colonne(cols, ['nom', 'camion'])
                       or detecter_colonne(cols, ['nom', 'vehicule'])
                       or detecter_colonne_exacte(cols, {'vehicule', 'nom vehicule', 'nom du vehicule'}))

    quai_par_camion = {}
    consignes_par_camion_dj = {}
    equipiers_par_camion_dj = {}   # (code, dj_key) -> [(nom, tel), ...]
    codes_veh = set()
    nom_camion_par_code = {}
    tournee_camion_dj = {}         # (code, dj_key) -> [(vif, nom_mag), ...] ordonné, dédupliqué
    magasins_sans_camion = {}      # vif -> nom_mag : ligne avec un magasin mais Code (camion) vide

    for _, r in df_veh.iterrows():
        code = str(r.get('Code', '')).strip()
        if not code or code == 'nan':
            nom_mag_sc = str(r.get('Magasin', '')).strip()
            if nom_mag_sc and nom_mag_sc != 'nan':
                vif_sc = vif_fmt(r.get('Code VIF', ''))
                magasins_sans_camion[vif_sc or nom_mag_sc] = nom_mag_sc
            continue
        codes_veh.add(code)

        quai = str(r.get('Quai', '')).strip()
        if quai and quai != 'nan' and code not in quai_par_camion:
            quai_par_camion[code] = quai

        if col_nom_camion:
            nom_c = str(r.get(col_nom_camion, '')).strip()
            if nom_c and nom_c != 'nan' and code not in nom_camion_par_code:
                nom_camion_par_code[code] = nom_c

        jour = str(r.get('Tournée', '')).strip()
        debut = str(r.get('Début', '')).strip()
        if not jour or jour == 'nan':
            continue
        m = re.match(r'(\d+)', debut)
        heure = int(m.group(1)) if m else 8
        periode = 'Matin' if heure < 13 else 'Apres Midi'
        dj_key = normaliser_dj(f'{jour} {periode}')
        if not dj_key:
            continue

        entry = consignes_par_camion_dj.setdefault((code, dj_key), {'c1': '', 'c2': ''})
        c1 = str(r.get('consigne1', '')).strip()
        c2 = str(r.get('Consigne2', '')).strip()
        if c1 and c1 != 'nan' and not entry['c1']:
            entry['c1'] = c1
        if c2 and c2 != 'nan' and not entry['c2']:
            entry['c2'] = c2

        eq = str(r.get('Équipier', '')).strip()
        tel = str(r.get('Portable', '')).strip()
        if not tel or tel == 'nan':
            tel = str(r.get('Tel', '')).strip()
        if tel == 'nan':
            tel = ''
        if eq and eq != 'nan':
            lst = equipiers_par_camion_dj.setdefault((code, dj_key), [])
            if eq not in [n for n, _ in lst]:
                lst.append((eq, tel))

        vif = vif_fmt(r.get('Code VIF', ''))
        nom_mag = str(r.get('Magasin', '')).strip()
        if nom_mag and nom_mag != 'nan':
            lst_t = tournee_camion_dj.setdefault((code, dj_key), [])
            if not any(nm == nom_mag for _, nm in lst_t):
                lst_t.append((vif, nom_mag))

    print(f"  → {len(codes_veh)} camions, {len(tournee_camion_dj)} tournées "
          f"(camion x demi-journée) reconstruites depuis {args.vehicules}")
    if not col_nom_camion:
        print("  INFO : aucune colonne 'Nom camion' détectée dans liste-vehicule.xlsx "
              "→ noms de secours utilisés pour les véhicules connus, vide sinon.")

    nb_passages_max = max((len(v) for v in tournee_camion_dj.values()), default=0)
    mag_cols = [f'Magasin {k}' for k in range(1, nb_passages_max + 1)]
    vif_cols = [f'Code VIF {k}' for k in range(1, nb_passages_max + 1)]

    dj_ord = {dj: i for i, dj in enumerate(DJ_ORDER)}
    rows = []
    dim_set = set()
    for (code, dj_key), passages in tournee_camion_dj.items():
        nom_cam = nom_camion_par_code.get(code) or NOMS_CAMIONS_CONNUS.get(code, '')
        row = {'Demi-journee': dj_key, 'Camion': code, 'Nom camion': nom_cam}
        for k in range(nb_passages_max):
            vif_k, nom_k = passages[k] if k < len(passages) else ('', '')
            row[f'Code VIF {k+1}'] = vif_k
            row[f'Magasin {k+1}'] = nom_k
        rows.append(row)
        if dj_key == 'Dimanche Matin':
            for _, nom_k in passages:
                dim_set.add(nom_k)

    df = pd.DataFrame(rows)
    if len(df):
        df['_o'] = df['Demi-journee'].map(dj_ord).fillna(99)
        df = df.sort_values(['_o', 'Camion']).drop(columns=['_o']).reset_index(drop=True)

    return {
        'df': df, 'mag_cols': mag_cols, 'vif_cols': vif_cols, 'use_vif': True,
        'quai_par_camion': quai_par_camion,
        'consignes_par_camion_dj': consignes_par_camion_dj,
        'equipiers_par_camion_dj': equipiers_par_camion_dj,
        'codes_veh': codes_veh, 'dim_set': dim_set, 'df_veh': df_veh,
        'nom_camion_par_code': nom_camion_par_code,
        'magasins_sans_camion': magasins_sans_camion,
    }


def construire_lignes_tournees_secteur(df_t, mag_cols, df_ref):
    """
    Construit, pour chaque ligne de df_t (une tournée = un (Demi-journee,
    Camion)), les mêmes informations que l'onglet 'Tournees' du classeur
    produit par generer_tournees_bai_v2.py : Tonnage, Km estimés et surtout
    Secteur — calculé avec calculer_secteur(), donc identique à celui de
    l'optimisation des tournées — même si ici les tournées viennent de Go on
    the Web et n'ont pas été recalculées géographiquement.

    Retourne une liste de dicts : Demi-journee, Camion, Nom camion, Tonnage,
    Km estimes, Secteur, Magasin 1..N (les magasins sont réordonnés par
    distance croissante à la BAI, comme dans l'onglet d'origine).
    """
    # Colonne tonnage : detection souple (le référentiel peut s'appeler
    # 'Tonnage 2025', 'Tonnage annuel', etc.)
    col_tonnage = None
    for col in df_ref.columns:
        cl = str(col).strip().lower()
        if 'tonnage' in cl:
            col_tonnage = col
            break

    infos_mag = {}
    def _safe_float(val, defaut):
        """float(val), en repliant sur `defaut` si val est vide/NaN/non
        convertible — une cellule Excel vide se lit comme NaN (un float !),
        qui passe le test 'or 0' (NaN est 'truthy' en Python) et fait donc
        planter round()/haversine() plus loin si on ne le filtre pas ici."""
        try:
            f = float(val)
            return defaut if f != f else f  # f != f  <=>  f is NaN
        except (TypeError, ValueError):
            return defaut

    for _, r in df_ref.iterrows():
        nom = str(r.get('Nom', '')).strip()
        if not nom or nom == 'nan':
            continue
        lat = _safe_float(r.get('Latitude'), BAI_LAT)
        lon = _safe_float(r.get('Longitude'), BAI_LON)
        ville = str(r.get('Ville', '')).strip()
        tonnage = _safe_float(r.get(col_tonnage), 0) if col_tonnage else 0
        infos_mag[nom] = {
            'lat': lat, 'lon': lon,
            'secteur': calculer_secteur(lat, lon, ville),
            'tonnage': tonnage,
        }

    # Nb de passages par magasin (= nb de tournées distinctes où il apparaît),
    # pour répartir son tonnage annuel entre ses passages — même logique que
    # vif_passages dans generer_tournees_bai_v2.py.
    nb_passages = Counter()
    for _, row in df_t.iterrows():
        for mc in mag_cols:
            nom = str(row.get(mc, '')).strip()
            if nom and nom != 'nan':
                nb_passages[nom] += 1

    lignes = []
    for _, row in df_t.iterrows():
        noms = [str(row.get(mc, '')).strip() for mc in mag_cols]
        noms = [n for n in noms if n and n != 'nan']
        # Réordonner par distance croissante à la BAI (même convention que
        # l'onglet d'origine)
        noms.sort(key=lambda n: dist_bai(infos_mag.get(n, {}).get('lat', BAI_LAT),
                                          infos_mag.get(n, {}).get('lon', BAI_LON)))

        km = 0.0
        tonnage_total = 0
        secteurs = []
        if noms:
            pts = [(infos_mag.get(n, {}).get('lat', BAI_LAT),
                    infos_mag.get(n, {}).get('lon', BAI_LON)) for n in noms]
            km = dist_bai(pts[0][0], pts[0][1])
            for k in range(len(pts) - 1):
                km += haversine(pts[k][0], pts[k][1], pts[k+1][0], pts[k+1][1])
            km += dist_bai(pts[-1][0], pts[-1][1])
            for n in noms:
                info = infos_mag.get(n)
                if info:
                    secteurs.append(info['secteur'])
                    passages = nb_passages.get(n, 1) or 1
                    tonnage_total += info['tonnage'] / passages

        secteurs_uniq = [s for s, _ in Counter(secteurs).most_common()] if secteurs else []
        ligne = {
            'Demi-journee': row.get('Demi-journee', ''),
            'Camion': row.get('Camion', ''),
            'Nom camion': row.get('Nom camion', ''),
            'Tonnage': round(tonnage_total),
            'Km estimes': round(km, 1),
            'Secteur': ' | '.join(secteurs_uniq),
        }
        for k, n in enumerate(noms):
            ligne[f'Magasin {k+1}'] = n
        lignes.append(ligne)
    return lignes


def ecrire_onglet_tournees_secteur(wb, lignes, nb_max_mag, titre='Tournees'):
    """
    Écrit dans le classeur wb (déjà ouvert) un onglet nommé `titre` avec les
    lignes produites par construire_lignes_tournees_secteur() — même mise en
    forme (en-tête, alternance de couleur par demi-journée, mise en évidence
    des secteurs à risque) que l'onglet 'Tournees' du classeur d'optimisation.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    C_HDR = PatternFill("solid", fgColor="1F4E79")
    C_BLC = PatternFill("solid", fgColor="DDEEFF")
    C_WHT = PatternFill("solid", fgColor="FFFFFF")
    F_HDR = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    F_NRM = Font(name='Calibri', size=10)
    thin = Side(style='thin', color='AAAAAA')
    BRD = Border(left=thin, right=thin, top=thin, bottom=thin)
    A_C = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws = wb.create_sheet(titre)
    cols = ['Demi-journee', 'Camion', 'Nom camion', 'Tonnage', 'Km estimes', 'Secteur'] \
        + [f'Magasin {k}' for k in range(1, nb_max_mag + 1)]
    for c, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = F_HDR; cell.fill = C_HDR; cell.border = BRD; cell.alignment = A_C
    ws.row_dimensions[1].height = 20

    dj_ord = {dj: i for i, dj in enumerate(DJ_ORDER)}
    lignes_triees = sorted(lignes, key=lambda l: (dj_ord.get(l['Demi-journee'], 99), str(l['Camion'])))

    dj_cur, alt = None, True
    for r, ligne in enumerate(lignes_triees, 2):
        if ligne['Demi-journee'] != dj_cur:
            dj_cur = ligne['Demi-journee']; alt = not alt
        fill = C_BLC if alt else C_WHT
        for c, col in enumerate(cols, 1):
            val = ligne.get(col, '')
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = F_NRM; cell.fill = fill; cell.border = BRD
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True) \
                if col in ('Secteur',) + tuple(f'Magasin {k}' for k in range(1, nb_max_mag + 1)) else A_C
            if col == 'Secteur' and str(val).strip():
                nb_sec = len([s for s in str(val).split('|') if s.strip()])
                if nb_sec > 3:
                    cell.font = Font(name='Calibri', size=10, bold=True, color='C00000')
                elif nb_sec == 3:
                    cell.font = Font(name='Calibri', size=10, bold=True, color='C65911')

    for i, w in enumerate([22, 9, 28, 10, 11, 24] + [25] * nb_max_mag, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    if lignes_triees:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(lignes_triees) + 1}"
    return ws


def construire_classeur_tournees(df_t, mag_cols, vif_cols, df_ref):
    """
    Construit (sans le sauvegarder — voir main()) LE classeur unique
    regroupant tous les onglets : 'Tournees VIF' (avec les Code VIF),
    'Tournees' (avec Tonnage / Km estimés / Secteur — mêmes secteurs
    géographiques que le classeur de l'optimisation des tournées) et
    'Magasins' (référentiel + camion affecté par demi-journée). Pas de
    calcul de tournées ici : uniquement la mise en forme de ce qui est
    déjà décidé dans liste-vehicule.xlsx.

    Les onglets de contrôle (Quai manquant, Cagettes manquantes...) sont
    ajoutés séparément par main(), dans ce même classeur.
    """
    import openpyxl as _oxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    C_HDR = PatternFill("solid", fgColor="1F4E79")
    F_HDR = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    F_NRM = Font(name='Calibri', size=10)
    thin = Side(style='thin', color='AAAAAA')
    BRD = Border(left=thin, right=thin, top=thin, bottom=thin)
    A_C = Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb = _oxl.Workbook()
    del wb['Sheet']  # supprime la feuille par défaut vide

    # ── Tournees VIF ──────────────────────────────────────────────────────
    ws1 = wb.create_sheet('Tournees VIF')
    cols_vif = ['Demi-journee', 'Camion', 'Nom camion']
    for vc, mc in zip(vif_cols, mag_cols):
        cols_vif += [vc, mc]
    for c, col in enumerate(cols_vif, 1):
        cell = ws1.cell(row=1, column=c, value=col)
        cell.font = F_HDR; cell.fill = C_HDR; cell.border = BRD; cell.alignment = A_C
    for r, (_, row) in enumerate(df_t.iterrows(), 2):
        for c, col in enumerate(cols_vif, 1):
            cell = ws1.cell(row=r, column=c, value=row.get(col, ''))
            cell.font = F_NRM; cell.border = BRD; cell.alignment = A_C
    ws1.freeze_panes = 'A2'
    if len(df_t):
        ws1.auto_filter.ref = f"A1:{get_column_letter(len(cols_vif))}{len(df_t)+1}"

    # ── Tournees (Tonnage / Km estimés / Secteur — mêmes secteurs que
    # l'optimisation des tournées) ───────────────────────────────────────
    lignes_secteur = construire_lignes_tournees_secteur(df_t, mag_cols, df_ref)
    ecrire_onglet_tournees_secteur(wb, lignes_secteur, nb_max_mag=len(mag_cols))

    # ── Magasins (référentiel + camion affecté par demi-journée) ─────────
    ws3 = wb.create_sheet('Magasins')
    cols_fix = ['Code VIF', 'Nom', 'Ville', 'Adresse', 'État']
    cols_m = cols_fix + DJ_ORDER
    for c, col in enumerate(cols_m, 1):
        cell = ws3.cell(row=1, column=c, value=col)
        cell.font = F_HDR; cell.fill = C_HDR; cell.border = BRD; cell.alignment = A_C

    vif_plan = {}
    for _, tr in df_t.iterrows():
        dj_k, cam = tr['Demi-journee'], tr['Camion']
        for mc in mag_cols:
            nom = str(tr.get(mc, ''))
            if nom and nom != 'nan':
                vif_plan.setdefault(nom, {})[dj_k] = cam

    if 'État' in df_ref.columns:
        df_ref_actifs_m = df_ref[df_ref['État'].astype(str).str.strip() == 'Collecté par la BAI'].reset_index(drop=True)
    else:
        df_ref_actifs_m = df_ref
    for r, (_, mag) in enumerate(df_ref_actifs_m.iterrows(), 2):
        nom_mag = str(mag.get('Nom', '')).strip()
        for c, col in enumerate(cols_m, 1):
            if col in DJ_ORDER:
                val = vif_plan.get(nom_mag, {}).get(col, '')
            else:
                val = mag.get(col, '')
                val = '' if str(val) == 'nan' else val
            cell = ws3.cell(row=r, column=c, value=val)
            cell.font = F_NRM; cell.border = BRD; cell.alignment = A_C
    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = ws3.dimensions
    for i, w in enumerate([12, 30, 20, 35, 20] + [16]*len(DJ_ORDER), 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Secteurs / Secteurs - Magasins (répartition géographique du
    # référentiel, indépendante des tournées actuellement affectées — un
    # magasin actif y figure même s'il est absent de liste-vehicule.xlsx,
    # ex. pour préparer une nouvelle tournée) ────────────────────────────
    _ajouter_onglets_secteurs(wb, df_ref)

    return wb


def generer_carte_tournees(df_t, df_mag, mag_cols, dossier_resultat):
    """Génère une carte HTML interactive des tournées avec itinéraire routier
    OSRM — reprise quasi telle quelle de generer_carte_tournees() dans
    generer_tournees_bai_v2.py (même template Leaflet/HTML/CSS/JS autonome),
    adaptée aux données disponibles ici :

      - df_t : ici une liste de dicts (issue de
        construire_lignes_tournees_secteur(), convertie en DataFrame) avec
        colonnes 'Demi-journee', 'Camion', 'Nom camion', 'Secteur' et
        'Magasin 1'..'Magasin N' — N = len(mag_cols), variable (pas figé à 6
        comme dans l'outil d'optimisation), d'où le paramètre mag_cols
        remplaçant la boucle range(1, 7) de la version d'origine.
      - Les demi-journées couvertes sont celles de DJ_ORDER (Jeudi à
        Dimanche), et pas seulement Vendredi/Samedi (DJ_LIST_CT) comme dans
        generer_tournees_bai_v2.py : ce script gère aussi les tournées du
        jeudi et du dimanche.
      - df_mag : le référentiel magasins (df_ref, mêmes colonnes 'Nom',
        'Latitude', 'Longitude', 'Adresse', 'Ville', 'C.P.' que dans l'outil
        d'optimisation).
    """
    import json
    from collections import defaultdict

    FIGES_CT = set(VEHICULES_FIGES)
    DJ_LIST_CT = DJ_ORDER
    BAI_LAT_CT, BAI_LON_CT = BAI_LAT, BAI_LON

    # Index nom → coordonnées + adresse
    nom2coords_ct, nom2adresse_ct = {}, {}
    for _, r in df_mag.iterrows():
        nom = str(r['Nom']).strip()
        try:
            lat, lon = float(r.get('Latitude', 0)), float(r.get('Longitude', 0))
            if lat and lon: nom2coords_ct[nom] = [lat, lon]
        except: pass
        adr = str(r.get('Adresse', '')).strip()
        vil = str(r.get('Ville', '')).strip()
        cp  = str(r.get('C.P.', '')).strip()
        if adr and vil: nom2adresse_ct[nom] = f'{adr}, {cp} {vil}'

    # Construire les données par DJ → camion
    data_ct = defaultdict(dict)
    for _, row in df_t[df_t['Demi-journee'].isin(DJ_LIST_CT)].iterrows():
        veh = str(row['Camion']).strip()
        dj  = str(row['Demi-journee']).strip()
        if veh in FIGES_CT: continue
        nom_cam = str(row.get('Nom camion', '')).strip()
        if nom_cam in ('nan', ''): nom_cam = ''
        mags = [str(row.get(mc, '')).strip()
                for mc in mag_cols
                if str(row.get(mc, '')).strip() not in ('', 'nan')]
        sec = str(row.get('Secteur', '')).strip()
        if sec == 'nan': sec = ''
        points = []
        for mag in mags:
            pt = {'nom': mag, 'adresse': nom2adresse_ct.get(mag, ''), 'secteur': sec}
            pt['lat'] = nom2coords_ct[mag][0] if mag in nom2coords_ct else BAI_LAT_CT
            pt['lon'] = nom2coords_ct[mag][1] if mag in nom2coords_ct else BAI_LON_CT
            points.append(pt)
        data_ct[dj][veh] = {'nom_cam': nom_cam, 'secteur': sec, 'magasins': points}

    DATA_JSON_CT = json.dumps(data_ct, ensure_ascii=False)
    DJ_JSON_CT   = json.dumps(DJ_LIST_CT, ensure_ascii=False)

    parts = []
    parts.append("""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>BAI 38 — Carte des tourn&#233;es 2026</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/leaflet/1.9.4/leaflet.min.css"/>
<script src="https://cdnjs.cloudflare.com/ajax/libs/leaflet/1.9.4/leaflet.min.js"></script>
<style>
*{box-sizing:border-box;margin:0;padding:0;font-family:'Segoe UI',Arial,sans-serif}
body{display:flex;flex-direction:column;height:100vh;background:#f0f2f5}
header{background:#1F4E79;color:#fff;padding:10px 18px;display:flex;align-items:center;gap:16px;flex-shrink:0;box-shadow:0 2px 6px rgba(0,0,0,.3)}
header h1{font-size:16px;font-weight:600;white-space:nowrap}
.controls{display:flex;gap:10px;align-items:center;flex-wrap:wrap}
select{padding:6px 10px;border-radius:6px;border:none;font-size:13px;background:#fff;color:#1F4E79;font-weight:600;cursor:pointer;min-width:130px}
#btn-gmaps{padding:6px 14px;background:#E8A020;color:#fff;border:none;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;white-space:nowrap}
#btn-gmaps:hover{background:#d4901a}
#btn-gmaps:disabled{background:#888;cursor:default}
.badge{background:#E8A020;color:#fff;border-radius:10px;padding:2px 8px;font-size:12px;font-weight:700}
#container{display:flex;flex:1;overflow:hidden}
#sidebar{width:300px;flex-shrink:0;background:#fff;display:flex;flex-direction:column;border-right:1px solid #ddd;overflow:hidden}
#sidebar-header{padding:12px;background:#D9E1F2;border-bottom:1px solid #c0cbdf}
#sidebar-header h2{font-size:13px;color:#1F4E79;font-weight:700}
#sidebar-header .sub{font-size:11px;color:#555;margin-top:3px}
#mag-list{overflow-y:auto;flex:1}
.mag-item{display:flex;align-items:flex-start;padding:10px 12px;border-bottom:1px solid #f0f0f0;cursor:pointer;transition:background .15s}
.mag-item:hover{background:#f5f8ff}
.mag-item.active{background:#D9E1F2}
.mag-num{width:22px;height:22px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;color:#fff;flex-shrink:0;margin-right:10px;margin-top:1px}
.mag-info{flex:1;min-width:0}
.mag-name{font-size:12px;font-weight:600;color:#1F4E79;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.mag-addr{font-size:11px;color:#666;margin-top:2px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.mag-sec{font-size:10px;color:#888;margin-top:2px}
.bai-item{background:#1F4E79;color:#fff;padding:8px 12px;font-size:12px;font-weight:600;display:flex;align-items:center;gap:8px}
#map{flex:1}
.popup-title{font-weight:700;font-size:13px;color:#1F4E79;margin-bottom:4px}
.popup-addr{font-size:11px;color:#555;margin-bottom:3px}
.popup-sec{font-size:11px;color:#888}
.no-tournee{padding:20px;text-align:center;color:#999;font-size:13px}
</style>
</head>
<body>
<header>
  <h1>&#128666; BAI 38 &#8212; Tourn&#233;es 2026</h1>
  <div class="controls">
    <select id="sel-dj" onchange="onDjChange()">
      <option value="">&#8212; Demi-journ&#233;e &#8212;</option>
    </select>
    <select id="sel-cam" onchange="onCamChange()" disabled>
      <option value="">&#8212; Camion &#8212;</option>
    </select>
    <button id="btn-gmaps" disabled onclick="openGmaps()">&#128506; Google Maps</button>
    <span id="badge-nb" class="badge" style="display:none"></span>
  </div>
</header>
<div id="container">
  <div id="sidebar">
    <div id="sidebar-header" style="display:none">
      <h2 id="sb-title">&#8212;</h2>
      <div class="sub" id="sb-sec">&#8212;</div>
    </div>
    <div id="mag-list"><div class="no-tournee">S&#233;lectionnez une demi-journ&#233;e et un camion</div></div>
  </div>
  <div id="map"></div>
</div>
<script>
""")
    parts.append(f'const DATA = {DATA_JSON_CT};\n')
    parts.append(f'const DJ_LIST = {DJ_JSON_CT};\n')
    parts.append(f'const BAI = [{BAI_LAT_CT}, {BAI_LON_CT}];\n')
    parts.append("""
const COULEURS = ['#e74c3c','#e67e22','#2980b9','#27ae60','#8e44ad',
  '#16a085','#d35400','#2c3e50','#f39c12','#1abc9c',
  '#c0392b','#7f8c8d','#6c5ce7','#00b894','#fd79a8'];

const map = L.map('map').setView(BAI, 11);
L.tileLayer('https://{s}.basemaps.cartocdn.com/rastertiles/voyager/{z}/{x}/{y}{r}.png',
  {attribution:'&copy; OpenStreetMap &copy; CARTO',maxZoom:19,subdomains:'abcd'}).addTo(map);

L.marker(BAI, {icon: L.divIcon({
  className:'',
  html:'<div style="background:#1F4E79;color:#fff;border-radius:50%;width:32px;height:32px;display:flex;align-items:center;justify-content:center;font-size:16px;border:3px solid #fff;box-shadow:0 2px 6px rgba(0,0,0,.4)">&#127981;</div>',
  iconSize:[32,32],iconAnchor:[16,16]
})}).addTo(map).bindPopup('<b>BAI 38</b><br>11 All&eacute;e de la Pin&eacute;a<br>38600 Fontaine');

let layerGroup = L.layerGroup().addTo(map);
let currentTournee = null;

const selDj = document.getElementById('sel-dj');
DJ_LIST.forEach(function(dj) {
  var o = document.createElement('option');
  o.value = dj; o.textContent = dj;
  selDj.appendChild(o);
});

function onDjChange() {
  var dj = selDj.value;
  var selCam = document.getElementById('sel-cam');
  selCam.innerHTML = '<option value="">&#8212; Camion &#8212;</option>';
  selCam.disabled = !dj;
  if (!dj) { clearMap(); return; }
  var cams = Object.keys(DATA[dj] || {}).sort();
  cams.forEach(function(veh) {
    var t = DATA[dj][veh];
    var label = t.nom_cam ? (veh + ' — ' + t.nom_cam) : veh;
    var o = document.createElement('option');
    o.value = veh; o.textContent = label;
    selCam.appendChild(o);
  });
  clearMap();
}

function onCamChange() {
  var dj = selDj.value;
  var veh = document.getElementById('sel-cam').value;
  if (!dj || !veh) { clearMap(); return; }
  afficherTournee(dj, veh);
}

function clearMap() {
  layerGroup.clearLayers();
  currentTournee = null;
  document.getElementById('mag-list').innerHTML = '<div class="no-tournee">Sélectionnez une demi-journée et un camion</div>';
  document.getElementById('sidebar-header').style.display = 'none';
  document.getElementById('badge-nb').style.display = 'none';
  document.getElementById('btn-gmaps').disabled = true;
}

function makeIcon(num, color) {
  return L.divIcon({
    className:'',
    html:'<div style="background:'+color+';color:#fff;border-radius:50%;width:28px;height:28px;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;border:2px solid #fff;box-shadow:0 2px 5px rgba(0,0,0,.35)">'+num+'</div>',
    iconSize:[28,28],iconAnchor:[14,14],popupAnchor:[0,-14]
  });
}

function afficherTournee(dj, veh) {
  layerGroup.clearLayers();
  var t = DATA[dj][veh];
  var mags = t.magasins;
  currentTournee = {dj:dj, veh:veh, mags:mags};
  var cams = Object.keys(DATA[dj]).sort();
  var color = COULEURS[cams.indexOf(veh) % COULEURS.length];

  document.getElementById('sb-title').textContent = t.nom_cam ? (veh + ' — ' + t.nom_cam) : veh;
  document.getElementById('sb-sec').textContent = t.secteur || '';
  document.getElementById('sidebar-header').style.display = 'block';
  document.getElementById('badge-nb').style.display = 'inline';
  document.getElementById('badge-nb').textContent = mags.length + ' mag.';
  document.getElementById('btn-gmaps').disabled = false;

  // Ligne droite provisoire pendant chargement OSRM
  var latlngsProvis = [BAI];
  mags.forEach(function(m) { latlngsProvis.push([m.lat, m.lon]); });
  latlngsProvis.push(BAI);
  var polyProvis = L.polyline(latlngsProvis, {color:color, weight:3, opacity:0.4, dashArray:'8,6'}).addTo(layerGroup);

  // Itinéraire routier via OSRM
  var waypoints = [BAI[1]+','+BAI[0]];
  mags.forEach(function(m) { waypoints.push(m.lon+','+m.lat); });
  waypoints.push(BAI[1]+','+BAI[0]);
  var osrmUrl = 'https://router.project-osrm.org/route/v1/driving/' + waypoints.join(';')
    + '?overview=full&geometries=geojson';

  fetch(osrmUrl)
    .then(function(r) { return r.json(); })
    .then(function(data) {
      if (data.routes && data.routes[0]) {
        layerGroup.removeLayer(polyProvis);
        var coords = data.routes[0].geometry.coordinates.map(function(c) { return [c[1],c[0]]; });
        var dist_km = (data.routes[0].distance/1000).toFixed(1);
        var dur_min = Math.round(data.routes[0].duration/60);
        L.polyline(coords, {color:color, weight:4, opacity:0.85}).addTo(layerGroup);
        document.getElementById('sb-sec').textContent =
          (t.secteur||'') + ' — ' + dist_km + ' km, ' + dur_min + ' min';
      }
    })
    .catch(function() { polyProvis.setStyle({opacity:0.7, dashArray:''}); });

  // Marqueurs numérotés
  var markers = [];
  mags.forEach(function(mag, i) {
    var marker = L.marker([mag.lat, mag.lon], {icon: makeIcon(i+1, color)})
      .bindPopup(
        '<div class="popup-title">'+(i+1)+'. '+mag.nom+'</div>'+
        '<div class="popup-addr">'+(mag.adresse||'')+'</div>'+
        '<div class="popup-sec">'+(mag.secteur||'')+'</div>'
      ).addTo(layerGroup);
    markers.push(marker);
    marker.on('click', function() { highlightMag(i); });
  });

  // Sidebar
  var list = document.getElementById('mag-list');
  list.innerHTML = '';
  var baiD = document.createElement('div');
  baiD.className = 'bai-item';
  baiD.innerHTML = '&#127981; BAI &mdash; Fontaine (d&eacute;part)';
  list.appendChild(baiD);
  mags.forEach(function(mag, i) {
    var item = document.createElement('div');
    item.className = 'mag-item';
    item.id = 'mag-'+i;
    item.innerHTML =
      '<div class="mag-num" style="background:'+color+'">'+(i+1)+'</div>'+
      '<div class="mag-info">'+
        '<div class="mag-name">'+mag.nom+'</div>'+
        '<div class="mag-addr">'+(mag.adresse||'&mdash;')+'</div>'+
        '<div class="mag-sec">'+(mag.secteur||'')+'</div>'+
      '</div>';
    (function(idx, m) {
      item.onclick = function() { markers[idx].openPopup(); highlightMag(idx); map.panTo([m.lat, m.lon]); };
    })(i, mag);
    list.appendChild(item);
  });
  var baiA = document.createElement('div');
  baiA.className = 'bai-item';
  baiA.innerHTML = '&#127981; BAI &mdash; Fontaine (arriv&eacute;e)';
  list.appendChild(baiA);

  var bounds = L.latLngBounds([BAI]);
  mags.forEach(function(m) { bounds.extend([m.lat, m.lon]); });
  map.fitBounds(bounds.pad(0.15));
}

function highlightMag(idx) {
  document.querySelectorAll('.mag-item').forEach(function(el, i) {
    el.classList.toggle('active', i === idx);
  });
  var el = document.getElementById('mag-'+idx);
  if (el) el.scrollIntoView({block:'nearest',behavior:'smooth'});
}

function openGmaps() {
  if (!currentTournee) return;
  var mags = currentTournee.mags;
  var bai = '45.18867,5.68456';
  var pts = [bai];
  mags.forEach(function(m) { pts.push(m.lat+','+m.lon); });
  pts.push(bai);
  window.open('https://www.google.com/maps/dir/'+pts.join('/')+'/data=!4m2!4m1!3e0', '_blank');
}
</script>
</body>
</html>""")

    html_out = ''.join(parts)
    out_dir = dossier_resultat
    os.makedirs(out_dir, exist_ok=True)
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M')
    html_path = os.path.join(out_dir, f'carte_tournees_bai38_{ts}.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_out)
    print(f"  Carte HTML générée : {html_path}")
    return html_path


def _ajouter_onglets_secteurs(wb, df_ref):
    """
    Ajoute deux onglets calculés à partir du référentiel magasins (tous les
    magasins actifs, État = 'Collecté par la BAI'), en réutilisant le même
    calcul de secteur géographique (calculer_secteur) que l'onglet 'Tournees'
    et que l'outil d'optimisation des tournées :

      - 'Secteurs'           : liste simple des secteurs (Nb Magasins,
                                Tonnage Total), sans le détail des magasins.
      - 'Secteurs - Magasins' : même liste, avec en plus la liste des noms
                                de magasins de chaque secteur (même format
                                que l'onglet 'Secteurs' de
                                generer_tournees_bai_v2.py).

    Basé sur le référentiel complet (et non sur liste-vehicule.xlsx) : un
    magasin actif y apparaît même s'il n'a pas encore de tournée affectée
    (voir l'onglet 'Magasins non planifiés' pour ce cas précis).
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    C_HDR = PatternFill("solid", fgColor="1F4E79")
    F_HDR = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    F_NRM = Font(name='Calibri', size=10)
    thin = Side(style='thin', color='AAAAAA')
    BRD = Border(left=thin, right=thin, top=thin, bottom=thin)
    A_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
    A_L = Alignment(horizontal='left', vertical='center', wrap_text=True)

    if 'État' not in df_ref.columns:
        return
    actifs = df_ref[df_ref['État'].astype(str).str.strip() == 'Collecté par la BAI']

    col_tonnage = None
    for col in df_ref.columns:
        if 'tonnage' in str(col).strip().lower():
            col_tonnage = col
            break

    def _safe_float(val, defaut=0):
        try:
            f = float(val)
            return defaut if f != f else f
        except (TypeError, ValueError):
            return defaut

    secteur_stats = {}
    for _, r in actifs.iterrows():
        nom = str(r.get('Nom', '')).strip()
        if not nom or nom == 'nan':
            continue
        lat = _safe_float(r.get('Latitude'), BAI_LAT)
        lon = _safe_float(r.get('Longitude'), BAI_LON)
        ville = str(r.get('Ville', '')).strip()
        secteur = calculer_secteur(lat, lon, ville) or '(secteur inconnu)'
        tonnage = _safe_float(r.get(col_tonnage), 0) if col_tonnage else 0
        stats = secteur_stats.setdefault(secteur, {'noms': [], 'tonnage': 0})
        stats['noms'].append(nom)
        stats['tonnage'] += tonnage

    # ── Secteurs (liste simple) ──────────────────────────────────────────
    ws_s = wb.create_sheet('Secteurs')
    cols_s = ['Secteur', 'Nb Magasins', 'Tonnage Total']
    for c, col in enumerate(cols_s, 1):
        cell = ws_s.cell(row=1, column=c, value=col)
        cell.font = F_HDR; cell.fill = C_HDR; cell.border = BRD; cell.alignment = A_C
    for r, (secteur, stats) in enumerate(sorted(secteur_stats.items()), 2):
        vals = [secteur, len(stats['noms']), round(stats['tonnage'])]
        for c, val in enumerate(vals, 1):
            cell = ws_s.cell(row=r, column=c, value=val)
            cell.font = F_NRM; cell.border = BRD
            cell.alignment = A_L if c == 1 else A_C
    ws_s.column_dimensions['A'].width = 28
    ws_s.column_dimensions['B'].width = 14
    ws_s.column_dimensions['C'].width = 16
    ws_s.freeze_panes = 'A2'
    ws_s.auto_filter.ref = ws_s.dimensions

    # ── Secteurs - Magasins (détail, avec la liste des magasins) ─────────
    ws_sm = wb.create_sheet('Secteurs - Magasins')
    cols_sm = ['Secteur', 'Nb Magasins', 'Tonnage Total', 'Magasins']
    for c, col in enumerate(cols_sm, 1):
        cell = ws_sm.cell(row=1, column=c, value=col)
        cell.font = F_HDR; cell.fill = C_HDR; cell.border = BRD; cell.alignment = A_C
    for r, (secteur, stats) in enumerate(sorted(secteur_stats.items()), 2):
        vals = [secteur, len(stats['noms']), round(stats['tonnage']),
                ', '.join(sorted(stats['noms']))]
        for c, val in enumerate(vals, 1):
            cell = ws_sm.cell(row=r, column=c, value=val)
            cell.font = F_NRM; cell.border = BRD
            cell.alignment = A_L if c in (1, 4) else A_C
    ws_sm.column_dimensions['A'].width = 28
    ws_sm.column_dimensions['B'].width = 14
    ws_sm.column_dimensions['C'].width = 16
    ws_sm.column_dimensions['D'].width = 90
    ws_sm.freeze_panes = 'A2'
    ws_sm.auto_filter.ref = ws_sm.dimensions


def ajouter_onglet_explications(wb):
    """
    Ajoute (et place en premier) un onglet 'Explications' qui décrit le rôle
    de chacun des autres onglets présents dans le classeur — sert de page
    d'accueil / mode d'emploi du fichier. Les onglets de contrôle éventuels
    (Quai manquant, Camions non définis, Cagettes manquantes, Magasins non
    planifiés) ne sont documentés que s'ils existent réellement dans wb.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    C_HDR = PatternFill("solid", fgColor="1F4E79")
    F_HDR = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    F_TITRE = Font(name='Calibri', bold=True, size=14, color='1F4E79')
    F_NOTE = Font(name='Calibri', italic=True, size=10, color='555555')
    F_ONG = Font(name='Calibri', bold=True, size=10)
    F_TXT = Font(name='Calibri', size=10)
    A_TOP = Alignment(vertical='top', wrap_text=True)
    A_TOP_C = Alignment(vertical='top', horizontal='center', wrap_text=True)
    thin = Side(style='thin', color='AAAAAA')
    BRD = Border(left=thin, right=thin, top=thin, bottom=thin)

    EXPLICATIONS = [
        ('Tournees VIF',
         "Détail complet de chaque tournée (un camion pour une demi-journée donnée) avec, "
         "pour chaque magasin collecté, son Code VIF. Onglet technique et exhaustif, destiné "
         "surtout aux traitements automatisés (imports, recoupements avec d'autres outils) : "
         "à privilégier pour retrouver rapidement le Code VIF d'un magasin sur une tournée."),
        ('Tournees',
         "Vue de synthèse par tournée : Tonnage estimé, Km estimés (depuis le dépôt BAI, "
         "11 allée de la Pinea, Fontaine) et Secteur géographique — les MÊMES secteurs que "
         "ceux utilisés par l'outil d'optimisation des tournées, pour rester cohérent d'un "
         "outil à l'autre. La colonne Secteur est mise en couleur quand une tournée traverse "
         "plusieurs secteurs (orange = 3 secteurs distincts, rouge = plus de 3) : ces lignes "
         "sont à vérifier en priorité, une tournée qui zigzague entre plusieurs secteurs est "
         "souvent optimisable."),
        ('Magasins',
         "Référentiel des magasins actifs (Code VIF, nom, ville, adresse, horaires, état) "
         "avec, pour chaque demi-journée de collecte, le camion qui lui est affecté. Permet "
         "de vérifier d'un coup d'œil que tous les magasins du référentiel sont bien couverts."),
        ('Secteurs',
         "Liste des secteurs géographiques (mêmes secteurs que l'onglet 'Tournees' et que "
         "l'optimisation des tournées), avec le nombre de magasins actifs et le tonnage total "
         "de chaque secteur. Basé sur le référentiel complet, indépendamment des tournées "
         "actuellement affectées dans liste-vehicule.xlsx."),
        ('Secteurs - Magasins',
         "Même liste que l'onglet 'Secteurs', avec en plus le détail des magasins de chaque "
         "secteur. Pratique pour retrouver tous les magasins d'un secteur donné, ou pour "
         "préparer une nouvelle tournée dans une zone géographique précise."),
        ('Quai manquant',
         "Camions bien identifiés dans liste-vehicule.xlsx mais pour lesquels aucun numéro de "
         "quai n'est renseigné. À compléter (colonne Quai) avant impression des fiches de "
         "collecte."),
        ('Camions non définis',
         "Camions cités sur des tournées mais totalement absents de liste-vehicule.xlsx (le "
         "code camion ne correspond à aucune ligne du fichier). À vérifier : code erroné, ou "
         "camion réellement manquant dans le fichier source."),
        ('Camions absents',
         "Camions identifiés dans liste-vehicule.xlsx (au moins une ligne, ex. avec un Quai "
         "renseigné) mais auxquels aucun magasin/tournée n'est finalement affecté — planning "
         "vide pour ce camion. À vérifier : camion réellement non utilisé cette année, ou "
         "oubli lors de la saisie des tournées (colonnes 'Tournée'/'Début' non remplies)."),
        ('Camion non défini',
         "Magasins présents dans liste-vehicule.xlsx (une ligne existe, avec un nom de magasin) "
         "mais dont la colonne Code (camion) est vide sur cette ligne : le magasin a été saisi, "
         "mais aucun camion ne lui a été attribué. À corriger dans liste-vehicule.xlsx — colonne "
         "Code à compléter pour cette ligne."),
        ('Cagettes manquantes',
         "Couples magasin (Code VIF) / demi-journée sans historique de nombre de cagettes dans "
         "le fichier Cagettes magasins.xlsx. Les valeurs restent à compléter manuellement sur "
         "les fiches de collecte."),
        ('Magasins non planifiés',
         "Magasins actifs du référentiel (liste-magasins.xlsx) qui n'apparaissent dans AUCUNE "
         "tournée de liste-vehicule.xlsx : personne n'est prévu pour les collecter. À vérifier "
         "en priorité — magasin oublié lors de la planification ? (Les magasins présents dans "
         "liste-vehicule.xlsx mais sans camion attribué sont signalés séparément, voir 'Camion "
         "non défini'.)"),
        ('Créneaux incomplets',
         "Magasins collectés sur une SEULE des deux demi-journées d'un même jour (Vendredi ou "
         "Samedi) alors qu'un camion est prévu sur l'autre — ex. un camion le vendredi après-midi "
         "mais aucun le vendredi matin. Un magasin simplement non collecté un jour donné (ex. "
         "collecté uniquement le samedi), ou couvert par un camion à tournée figée/spéciale "
         "(V026, V037...) dont le planning ne suit pas ce schéma, n'est PAS signalé — ce n'est "
         "pas une anomalie. Signale un oubli de planification très probable — à vérifier et "
         "compléter dans liste-vehicule.xlsx en priorité."),
        ('Magasins-DJ non couverts',
         "Liste précise des couples (magasin, demi-journée) attendus d'après les créneaux "
         "RÉELLEMENT déclarés dans le référentiel (colonne 'Créneaux', ex. 'vendredi "
         "09h30-11h30') mais pour lesquels AUCUN camion n'est prévu dans liste-vehicule.xlsx. "
         "Plus direct et plus complet que 'Créneaux incomplets' (qui ne fait que déduire une "
         "anomalie probable de la symétrie Vendredi/Samedi) : ici on part de ce que le magasin "
         "est censé faire, jour par jour, quel que soit le jour de la semaine (jeudi, dimanche "
         "compris). Un magasin sans colonne 'Créneaux' renseignée n'est pas contrôlé ici. À "
         "compléter dans liste-vehicule.xlsx en priorité."),
    ]

    presents = [(nom, texte) for nom, texte in EXPLICATIONS if nom in wb.sheetnames]

    ws = wb.create_sheet('Explications')

    ws.merge_cells('A1:B1')
    c = ws.cell(row=1, column=1, value="Classeur Tournées BAI 38 — origine : Go on the Web (GOTW)")
    c.font = F_TITRE

    ws.merge_cells('A2:B2')
    c = ws.cell(row=2, column=1,
                value="Ce fichier est reconstruit directement à partir de liste-vehicule.xlsx et "
                      "liste-magasins.xlsx (outil « Go on the Web »), sans passer par l'optimisation "
                      "des tournées. Ci-dessous, le rôle de chaque onglet de ce classeur.")
    c.font = F_NOTE
    c.alignment = A_TOP
    ws.row_dimensions[2].height = 30

    r = 4
    headers = ['Onglet', 'Rôle / contenu']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=col, value=h)
        cell.font = F_HDR; cell.fill = C_HDR; cell.border = BRD; cell.alignment = A_TOP_C
    r += 1

    for nom, texte in presents:
        cell = ws.cell(row=r, column=1, value=nom)
        cell.font = F_ONG; cell.border = BRD; cell.alignment = A_TOP
        cell = ws.cell(row=r, column=2, value=texte)
        cell.font = F_TXT; cell.border = BRD; cell.alignment = A_TOP
        ws.row_dimensions[r].height = 60
        r += 1

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 100
    ws.freeze_panes = 'A5'

    # Place cet onglet en premier — c'est la page d'accueil du classeur.
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(ws)))
    wb.active = 0
    return ws


# ══════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(
        description='Génère le classeur Excel des tournées + les 4 documents PDF BAI 38, '
                     'directement depuis liste-vehicule.xlsx et liste-magasins.xlsx.')
    parser.add_argument('--magasins', default='liste-magasins.xlsx',
                        help="Référentiel magasins (adresses, horaires, VIF, État)")
    parser.add_argument('--vehicules', default='liste-vehicule.xlsx',
                        help="Affectations camion/magasin/demi-journée + Quai / consignes / équipiers")
    parser.add_argument('--cagettes', default='Cagettes_magasins.xlsx',
                        help="Historique du nb de cagettes par magasin/demi-journée "
                             "(colonnes 'Code VIF','jour','cag1'..'cag4')")
    parser.add_argument('--annee', default=2026, type=int)
    parser.add_argument('--camion', default=None,
                        help="Limite le document 1 (fiches de collecte) à ce seul camion "
                             "(ex: V003). N'affecte pas les documents 2/3/4.")
    parser.add_argument('--date-jeudi', default=None,
                        help="Date du jeudi de la collecte (JJ/MM/AAAA), pour afficher la date "
                             "de chaque demi-journée en en-tête du document 3. Omis si non fourni.")

    parser.add_argument('--output-excel', default=None,
                        help="Chemin du classeur Excel unique — tournées (VIF + secteurs) + "
                             "magasins + contrôles manquants (défaut : "
                             "Tournées definitives 2026/Tournees_BAI38_{annee}_GOTW.xlsx)")
    parser.add_argument('--output-fiches', default=None,
                        help="Chemin du PDF des fiches de collecte (défaut : "
                             "Tournées definitives 2026/fiches_jour_vehicule_magasin_{annee}.pdf)")
    parser.add_argument('--output-pointage', default=None,
                        help="Chemin du PDF de pointage (défaut : Tournées definitives 2026/pointage_vehicules_{annee}.pdf)")
    parser.add_argument('--output-equipier', default=None,
                        help="Chemin du PDF jour/véhicule/magasin/équipier (défaut : "
                             "Tournées definitives 2026/fiches_jour_vehicule_magasin_equipier_{annee}.pdf)")
    parser.add_argument('--output-index', default=None,
                        help="Chemin du PDF index alphabétique des équipiers (défaut : "
                             "Tournées definitives 2026/fiches_equipier_jour_vehicule_{annee}.pdf)")

    # Rétro-compatibilité : anciens noms d'arguments des scripts d'origine
    parser.add_argument('--output', default=None, help=argparse.SUPPRESS)
    parser.add_argument('--excel', default=None, help=argparse.SUPPRESS)  # ignoré : plus de fichier de tournées externe
    parser.add_argument('--rapport', default=None, help=argparse.SUPPRESS)  # ignoré : fusionné dans --output-excel

    args = parser.parse_args()
    if args.output and not args.output_fiches:
        args.output_fiches = args.output
    if args.excel:
        print("INFO : --excel est ignoré par ce script (les tournées sont reconstruites "
              "depuis --vehicules) — utilisez Generer_documents_bai38.py si vous partez "
              "d'un fichier de tournées déjà calculé.")
    if args.rapport:
        print("INFO : --rapport est ignoré — tous les onglets (tournées, secteurs, "
              "contrôles manquants) sont désormais dans le même classeur (--output-excel).")

    # Suffixe ajouté à tous les fichiers de sortie pour rappeler leur origine :
    # ces documents sont issus des exports de l'outil "Go on the Web" (GOTW)
    # (liste-vehicule.xlsx / liste-magasins.xlsx), PAS du calcul d'optimisation
    # des tournées (generer_tournees_bai_v2.py + PDF de l'année précédente).
    ORIGINE = 'GOTW'

    # out_dir sert uniquement de dossier par défaut pour les sorties non
    # explicitement précisées, ET comme dossier de sortie de la carte HTML
    # (generer_carte_tournees ne prend pas de --output dédié). Quand
    # --output-excel est fourni (cas de l'appli Basilic), on aligne out_dir
    # dessus pour que la carte atterrisse au même endroit que les autres
    # documents plutôt que sur le chemin Windows codé en dur par défaut.
    if args.output_excel:
        out_dir = os.path.dirname(os.path.abspath(args.output_excel)) or '.'
    else:
        out_dir = r'G:\Drive partagés\BA380 - COLLECTE\Collecte 2026\Camions 2026\Tournées definitives 2026'
    os.makedirs(out_dir, exist_ok=True)
    if args.output_excel is None:
        args.output_excel = os.path.join(out_dir, f'Tournees_BAI38_{args.annee}_{ORIGINE}.xlsx')
    if args.output_fiches is None:
        args.output_fiches = os.path.join(out_dir, f'fiches_jour_vehicule_magasin_{args.annee}_{ORIGINE}.pdf')
    if args.output_pointage is None:
        args.output_pointage = os.path.join(out_dir, f'pointage_vehicules_{args.annee}_{ORIGINE}.pdf')
    if args.output_equipier is None:
        args.output_equipier = os.path.join(out_dir, f'fiches_jour_vehicule_magasin_equipier_{args.annee}_{ORIGINE}.pdf')
    if args.output_index is None:
        args.output_index = os.path.join(out_dir, f'fiches_equipier_jour_vehicule_{args.annee}_{ORIGINE}.pdf')

    print(f"Vehicules : {args.vehicules}")
    print(f"Ref       : {args.magasins}")

    # ── Chargement des tournées : reconstruites depuis liste-vehicule.xlsx ─
    # (plus de fichier de tournées externe — voir construire_tournees_depuis_vehicules)
    print("\nReconstruction des tournées depuis liste-vehicule.xlsx...")
    ctx = construire_tournees_depuis_vehicules(args)
    df, mag_cols, vif_cols, use_vif = ctx['df'], ctx['mag_cols'], ctx['vif_cols'], ctx['use_vif']
    quai_par_camion         = ctx['quai_par_camion']
    consignes_par_camion_dj = ctx['consignes_par_camion_dj']
    equipiers_par_camion_dj = ctx['equipiers_par_camion_dj']
    codes_veh               = ctx['codes_veh']
    dim_set                 = ctx['dim_set']
    df_veh                  = ctx['df_veh']
    nom_camion_par_code     = ctx['nom_camion_par_code']
    magasins_sans_camion    = ctx['magasins_sans_camion']

    if df.empty:
        print("ERREUR : aucune tournée reconstruite — vérifiez les colonnes de "
              f"{args.vehicules} ('Code', 'Tournée', 'Début', 'Code VIF', 'Magasin').")
        return

    # Référentiel magasins (documents 1 uniquement)
    df_ref = pd.read_excel(args.magasins)
    ref_dict = {}
    for _, r in df_ref.iterrows():
        nom = str(r.get('Nom', '')).strip()
        if nom and nom != 'nan':
            ref_dict[nom] = dict(r)


    # ── Classeur Excel unique : Tournees VIF / Tournees (+ secteurs) /
    # Magasins — sauvegardé une seule fois, à la fin, une fois les onglets de
    # contrôle (Quai manquant, Cagettes manquantes...) ajoutés eux aussi.
    print()
    wb = construire_classeur_tournees(df, mag_cols, vif_cols, df_ref)

    # ── Magasins actifs du référentiel jamais intégrés dans liste-vehicule.xlsx
    non_planifies = []
    if 'État' in df_ref.columns:
        actifs_ref = df_ref[df_ref['État'].astype(str).str.strip() == 'Collecté par la BAI']
        # NB : on part de df (tournées réellement reconstruites, une ligne par
        # camion x demi-journée) et non de df_veh brut — une ligne de
        # liste-vehicule.xlsx peut avoir un Code VIF renseigné mais un Code
        # (camion) ou une Tournée/Début vide, auquel cas elle est ignorée par
        # construire_tournees_depuis_vehicules (voir 'continue' plus haut) :
        # le magasin n'a alors AUCUN camion réellement affecté, même si son
        # VIF apparaît quelque part dans le fichier brut.
        vifs_planifies = set()
        for vc in vif_cols:
            if vc in df.columns:
                vifs_planifies |= set(vif_fmt(v) for v in df[vc])
        vifs_planifies.discard('')
        # Les magasins dont la ligne existe dans liste-vehicule.xlsx mais sans
        # camion (Code) renseigné sont signalés séparément (onglet 'Camion
        # non défini') plutôt que comme "non planifiés" : ce n'est pas qu'ils
        # sont absents du fichier, c'est qu'il leur manque un camion.
        vifs_sans_camion = set(v for v in magasins_sans_camion if v)
        for _, r in actifs_ref.iterrows():
            vif_r = vif_fmt(r.get('Code VIF', ''))
            nom_r = str(r.get('Nom', '')).strip()
            if not vif_r or not nom_r or nom_r == 'nan':
                continue
            if vif_r in vifs_sans_camion:
                continue
            if vif_r not in vifs_planifies:
                non_planifies.append((vif_r, nom_r))

        if non_planifies:
            print(f"ATTENTION : {len(non_planifies)} magasin(s) actif(s) du référentiel absent(s) "
                  f"de {args.vehicules} !")
        else:
            print(f"Planification : OK, tous les magasins actifs sont dans {args.vehicules}.")

        if magasins_sans_camion:
            print(f"ATTENTION : {len(magasins_sans_camion)} magasin(s) présent(s) dans "
                  f"{args.vehicules} mais sans camion (colonne Code) renseigné !")

    # ── Créneaux incomplets : magasin collecté un jour donné (Vendredi ou
    # Samedi) sur UNE seule des deux demi-journées de ce jour, alors qu'un
    # camion est prévu sur l'autre — signale un oubli de planification très
    # probable (ex. camion Vendredi après-midi + Samedi, mais aucun camion
    # Vendredi matin). Un magasin qui n'est simplement PAS planifié ce jour-là
    # (0 des 2 demi-journées, ex. magasin collecté uniquement le samedi) n'est
    # PAS signalé : ce n'est pas une anomalie, juste un jour de collecte
    # différent d'un magasin à l'autre. Ne redouble pas avec 'Magasins non
    # planifiés' (qui couvre le cas où le magasin n'apparaît nulle part dans
    # liste-vehicule.xlsx).
    DJ_JOURS_CTRL = {
        'Vendredi': ('Vendredi Matin', 'Vendredi Apres Midi'),
        'Samedi':   ('Samedi Matin',   'Samedi Apres Midi'),
    }
    creneaux_incomplets = []
    magasins_dj_non_couverts = []
    if 'État' in df_ref.columns:
        # presence_par_dj[dj] = {nom_magasin: camion} — sur TOUTES les
        # demi-journées (pas seulement Vendredi/Samedi), pour servir à la
        # fois au contrôle heuristique ci-dessous et au contrôle
        # 'Magasins/DJ non couverts' basé sur les créneaux déclarés.
        presence_par_dj = {dj: {} for dj in DJ_ORDER}
        for _, trow in df.iterrows():
            dj_t = str(trow.get('Demi-journee', '')).strip()
            if dj_t not in presence_par_dj:
                continue
            cam_t = str(trow.get('Camion', '')).strip()
            for mc in mag_cols:
                nom_t = str(trow.get(mc, '')).strip()
                if nom_t and nom_t != 'nan':
                    presence_par_dj[dj_t].setdefault(nom_t, cam_t)

        for _, r in actifs_ref.iterrows():
            nom_r = str(r.get('Nom', '')).strip()
            if not nom_r or nom_r == 'nan':
                continue
            for jour, (dj_matin, dj_am) in DJ_JOURS_CTRL.items():
                couvert_matin = nom_r in presence_par_dj[dj_matin]
                couvert_am = nom_r in presence_par_dj[dj_am]
                # Seul le cas « une seule des deux demi-journées couverte » est
                # une anomalie ; 0 ou 2 demi-journées couvertes sont normaux.
                if couvert_matin == couvert_am:
                    continue
                if couvert_matin:
                    dj_ok, dj_ko = dj_matin, dj_am
                else:
                    dj_ok, dj_ko = dj_am, dj_matin
                camion_ok = presence_par_dj[dj_ok][nom_r]
                # Camion à tournée figée/spéciale (ex. V026, V037) : son planning
                # n'a normalement pas à couvrir les deux demi-journées du jour,
                # donc pas d'alerte dans ce cas précis.
                if camion_ok in VEHICULES_FIGES:
                    continue
                vif_r = vif_fmt(r.get('Code VIF', ''))
                ville_r = str(r.get('Ville', '')).strip()
                txt_couvert = f"{DJ_LABELS.get(dj_ok, dj_ok)} ({camion_ok})"
                txt_manquant = DJ_LABELS.get(dj_ko, dj_ko)
                creneaux_incomplets.append((vif_r, nom_r, ville_r, jour, txt_couvert, txt_manquant))

        if creneaux_incomplets:
            print(f"ATTENTION : {len(creneaux_incomplets)} créneau(x) incomplet(s) détecté(s) "
                  f"(magasin collecté sur une demi-journée d'un jour mais pas l'autre) !")

        # ── Magasins/DJ non couverts : à partir des créneaux RÉELLEMENT
        # déclarés (colonne 'Créneaux' du référentiel, ex. 'vendredi
        # 09h30-11h30'), la liste précise des couples (magasin, demi-journée)
        # attendus mais pour lesquels AUCUN camion n'est prévu dans
        # liste-vehicule.xlsx. Plus direct et plus complet que 'Créneaux
        # incomplets' (qui ne fait que déduire une anomalie probable de la
        # symétrie Vendredi/Samedi) : ici on part de ce que le magasin est
        # censé faire, jour par jour. Un magasin sans colonne 'Créneaux'
        # renseignée n'est pas contrôlé ici (rien à comparer) — il reste
        # couvert par 'Magasins non planifiés' s'il est totalement absent.
        for _, r in actifs_ref.iterrows():
            nom_r = str(r.get('Nom', '')).strip()
            if not nom_r or nom_r == 'nan':
                continue
            djs_attendues = parse_creneaux(r.get('Créneaux', ''))
            if not djs_attendues:
                continue
            for dj_att in sorted(djs_attendues, key=lambda d: DJ_ORDER.index(d) if d in DJ_ORDER else 99):
                if nom_r in presence_par_dj.get(dj_att, {}):
                    continue
                vif_r = vif_fmt(r.get('Code VIF', ''))
                ville_r = str(r.get('Ville', '')).strip()
                magasins_dj_non_couverts.append((vif_r, nom_r, ville_r, DJ_LABELS.get(dj_att, dj_att)))

        if magasins_dj_non_couverts:
            print(f"ATTENTION : {len(magasins_dj_non_couverts)} couple(s) magasin/demi-journée "
                  f"attendu(s) d'après les créneaux mais sans camion dans {args.vehicules} !")

    # ── Cagettes (document 1 uniquement) ────────────────────────────────────
    cag_par_vif_dj = {}
    try:
        df_cag = pd.read_excel(args.cagettes, sheet_name='Feuil1')
        for _, r in df_cag.iterrows():
            vif = vif_fmt(r.get('Code VIF', ''))
            dj_n = normaliser_dj(r.get('jour', ''))
            if not vif or not dj_n:
                continue
            vals = []
            for k in ('cag1', 'cag2', 'cag3', 'cag4'):
                v = r.get(k, '')
                v = str(v).strip().split('.')[0] if str(v).strip() not in ('', 'nan') else ''
                vals.append(v)
            cag_par_vif_dj[(vif, dj_n)] = vals
        print(f"Cagettes: {args.cagettes} ({len(cag_par_vif_dj)} entrées magasin/demi-journée)")
    except Exception as e:
        print(f"AVERTISSEMENT : Cagettes non chargées ({args.cagettes} : {e})")

    # ═══════════════════════════════════════════════════════════════════════
    # DOCUMENT 1 — Fiches de collecte + rapport des manquants
    # ═══════════════════════════════════════════════════════════════════════
    df_work = df.copy()
    if args.camion:
        df_work = df_work[df_work['Camion'] == args.camion]
        if len(df_work) == 0:
            print(f"ERREUR : camion '{args.camion}' non trouvé")
            print("Disponibles :", sorted(df['Camion'].unique()))
            return

    dj_ord = {dj: i for i, dj in enumerate(DJ_ORDER)}
    df_work = df_work.copy()
    df_work['_o'] = df_work['Demi-journee'].map(dj_ord).fillna(99)
    df_work = df_work.sort_values(['Camion', '_o']).drop(columns=['_o'])

    print(f"\nSortie 1 (fiches)  : {args.output_fiches}")
    cv = rl_canvas.Canvas(args.output_fiches, pagesize=landscape(A4))

    manquants_quai      = {}
    camions_non_definis = {}
    manquants_cag        = {}

    for _, row in df_work.iterrows():
        dj      = str(row.get('Demi-journee', '')).strip()
        camion  = str(row.get('Camion', '')).strip()
        nom_cam = str(row.get('Nom camion', '')).strip()

        dj_n = normaliser_dj(dj)
        magasins = []
        if use_vif:
            pairs = list(zip(vif_cols, mag_cols))
        else:
            pairs = [(None, mc) for mc in mag_cols]

        for vc, mc in pairs:
            nom = str(row.get(mc, '')).strip()
            if not nom or nom == 'nan':
                continue
            r_ref = ref_dict.get(nom, {})
            vif = vif_fmt(row.get(vc, '')) if vc else ''
            if not vif:
                vif = vif_fmt(r_ref.get('Code VIF', ''))
            cag = cag_par_vif_dj.get((vif, dj_n), ['', '', '', ''])
            if (vif, dj_n) not in cag_par_vif_dj:
                manquants_cag[(vif, dj_n)] = {'nom': nom, 'dj_label': DJ_LABELS.get(dj_n, dj_n)}
            magasins.append({
                'nom':         nom,
                'vif':         vif,
                'adresse':     adresse_fmt(r_ref),
                'horaires':    horaires_fmt(r_ref),
                'en_dimanche': nom in dim_set,
                'cag':         cag,
            })

        if not magasins:
            continue

        quai = quai_par_camion.get(camion, '')
        if not quai:
            if camion not in codes_veh:
                camions_non_definis[camion] = nom_cam
            else:
                manquants_quai[camion] = nom_cam
        cons = consignes_par_camion_dj.get((camion, dj_n), {'c1': '', 'c2': ''})
        print(f"  {camion} — {dj} ({len(magasins)} magasins)")
        draw_fiche(cv, dj, camion, nom_cam, magasins, args.annee, quai=quai,
                   consigne1=cons['c1'], consigne2=cons['c2'])
        cv.showPage()

    args.output_fiches = sauver_pdf_avec_repli(cv, args.output_fiches)
    print(f"PDF sauvegardé : {args.output_fiches}")
    print("Quai et nb de cagettes pré-remplis quand disponibles dans les fichiers sources ;")
    print("les cases restées vides sont à compléter manuellement.")

    # ── Camions absents : codes camion présents dans liste-vehicule.xlsx
    # (une ligne au moins, ex. avec un Quai renseigné) mais qui n'ont
    # finalement aucune tournée (Code VIF/Magasin) qui leur soit affectée —
    # camion connu du fichier mais avec un planning vide, à vérifier.
    camions_avec_tournee = set(str(c).strip() for c in df['Camion']) if len(df) else set()
    camions_absents = {
        code: (nom_camion_par_code.get(code) or NOMS_CAMIONS_CONNUS.get(code, ''))
        for code in codes_veh if code not in camions_avec_tournee
    }
    if camions_absents:
        print(f"ATTENTION : {len(camions_absents)} camion(s) présent(s) dans {args.vehicules} "
              f"sans aucune tournée affectée : {', '.join(sorted(camions_absents))}")

    # ── Onglets de contrôle (Quai manquant, Cagettes manquantes...), ajoutés
    # dans LE MÊME classeur que Tournees VIF / Tournees / Magasins — un seul
    # fichier Excel avec tous les onglets, sauvegardé une seule fois ci-dessous.
    if camions_absents:
        ws_ca = wb.create_sheet('Camions absents')
        ws_ca.append(['Code', 'Nom camion'])
        for code, nom_cam in sorted(camions_absents.items()):
            ws_ca.append([code, nom_cam])

    if magasins_sans_camion:
        ws_msc = wb.create_sheet('Camion non défini')
        ws_msc.append(['Code VIF', 'Nom'])
        for cle_sc, nom_sc in sorted(magasins_sans_camion.items(), key=lambda kv: kv[1]):
            # La clé est le Code VIF normalisé, sauf repli sur le nom du
            # magasin quand la ligne n'avait pas de Code VIF renseigné non
            # plus (voir construire_tournees_depuis_vehicules) — dans ce cas
            # la colonne Code VIF reste vide.
            vif_sc = '' if cle_sc == nom_sc else cle_sc
            ws_msc.append([vif_sc, nom_sc])

    if manquants_quai:
        ws_q = wb.create_sheet('Quai manquant')
        ws_q.append(['Code', 'Nom camion', 'Quai'])
        for code, nom_cam in sorted(manquants_quai.items()):
            ws_q.append([code, nom_cam, ''])

    if camions_non_definis:
        ws_v = wb.create_sheet('Camions non définis')
        ws_v.append(['Code', 'Nom camion', 'Tournée (jour)', 'Début', 'Fin', 'Quai',
                      'Code VIF', 'Magasin', 'consigne1', 'Consigne2'])
        for code, nom_cam in sorted(camions_non_definis.items()):
            ws_v.append([code, nom_cam, '', '', '', '', '', '', '', ''])

    if manquants_cag:
        ws_c = wb.create_sheet('Cagettes manquantes')
        ws_c.append(['Code VIF', 'Nom', 'jour', 'cag1', 'cag2', 'cag3', 'cag4'])
        for (vif, dj_n), info in sorted(manquants_cag.items(), key=lambda kv: (kv[1]['nom'], kv[0][1])):
            ws_c.append([vif, info['nom'], info['dj_label'], '', '', '', ''])

    if non_planifies:
        ws_n = wb.create_sheet('Magasins non planifiés')
        ws_n.append(['Code VIF', 'Nom'])
        for vif_r, nom_r in sorted(non_planifies, key=lambda t: t[1]):
            ws_n.append([vif_r, nom_r])

    if creneaux_incomplets:
        from openpyxl.styles import PatternFill as _PF, Font as _Ft
        ws_ci = wb.create_sheet('Créneaux incomplets')
        entetes_ci = ['Code VIF', 'Nom', 'Ville', 'Jour', 'Créneau couvert (camion)', 'Créneau SANS camion']
        ws_ci.append(entetes_ci)
        for cell in ws_ci[1]:
            cell.font = _Ft(name='Calibri', bold=True, color='FFFFFF', size=10)
            cell.fill = _PF("solid", fgColor="C00000")
        for vif_r, nom_r, ville_r, jour_r, txt_couvert, txt_manquant in sorted(
                creneaux_incomplets, key=lambda t: (t[1], t[3])):
            ws_ci.append([vif_r, nom_r, ville_r, jour_r, txt_couvert, txt_manquant])
            ws_ci.cell(row=ws_ci.max_row, column=6).font = _Ft(bold=True, color='C00000')
        for col, w in zip('ABCDEF', [12, 30, 18, 12, 30, 22]):
            ws_ci.column_dimensions[col].width = w
        ws_ci.freeze_panes = 'A2'
        ws_ci.auto_filter.ref = ws_ci.dimensions

    if magasins_dj_non_couverts:
        from openpyxl.styles import PatternFill as _PF2, Font as _Ft2
        ws_dj = wb.create_sheet('Magasins-DJ non couverts')
        entetes_dj = ['Code VIF', 'Nom', 'Ville', 'Demi-journée attendue (créneaux)']
        ws_dj.append(entetes_dj)
        for cell in ws_dj[1]:
            cell.font = _Ft2(name='Calibri', bold=True, color='FFFFFF', size=10)
            cell.fill = _PF2("solid", fgColor="C00000")
        for vif_r, nom_r, ville_r, dj_label in sorted(
                magasins_dj_non_couverts, key=lambda t: (t[1], t[3])):
            ws_dj.append([vif_r, nom_r, ville_r, dj_label])
            ws_dj.cell(row=ws_dj.max_row, column=4).font = _Ft2(bold=True, color='C00000')
        for col, w in zip('ABCD', [12, 30, 18, 28]):
            ws_dj.column_dimensions[col].width = w
        ws_dj.freeze_panes = 'A2'
        ws_dj.auto_filter.ref = ws_dj.dimensions

    # ── Onglet 'Explications' (mode d'emploi du classeur) — ajouté en dernier
    # pour pouvoir documenter tous les onglets de contrôle réellement présents,
    # puis placé en première position comme page d'accueil.
    ajouter_onglet_explications(wb)

    args.output_excel = sauver_xlsx_avec_repli(wb, args.output_excel)
    if (manquants_quai or camions_non_definis or camions_absents or magasins_sans_camion
            or manquants_cag or non_planifies or creneaux_incomplets or magasins_dj_non_couverts):
        print(f"\nManquants : {len(manquants_quai)} camion(s) sans Quai, "
              f"{len(camions_non_definis)} camion(s) totalement absent(s) de {args.vehicules}, "
              f"{len(camions_absents)} camion(s) présent(s) dans {args.vehicules} mais sans "
              f"tournée affectée, "
              f"{len(magasins_sans_camion)} magasin(s) présent(s) dans {args.vehicules} sans "
              f"camion (Code) renseigné, "
              f"{len(manquants_cag)} entrée(s) magasin/demi-journée sans nb de cagettes, "
              f"{len(non_planifies)} magasin(s) jamais planifié(s), "
              f"{len(creneaux_incomplets)} créneau(x) incomplet(s) (jour couvert sur une seule "
              f"demi-journée sur les deux), "
              f"{len(magasins_dj_non_couverts)} couple(s) magasin/demi-journée attendu(s) "
              f"d'après les créneaux mais sans camion")
        print(f"  (Quai/Camions non définis/Camions absents/Camion non défini/Cagettes/"
              f"Magasins non planifiés/Créneaux incomplets/Magasins-DJ non couverts → à "
              f"compléter puis réinjecter dans {args.vehicules} / {args.cagettes}, avant "
              f"régénération)")
    else:
        print("\nAucun manquant : Quai, cagettes et planification sont complets.")
    print(f"→ Classeur Excel (tous onglets) : {args.output_excel}")

    # ═══════════════════════════════════════════════════════════════════════
    # DOCUMENTS 2, 3, 4 — Pointage / Équipier / Index (toujours tous camions,
    # indépendamment du filtre --camion qui ne concerne que le document 1)
    # ═══════════════════════════════════════════════════════════════════════
    par_dj = {}           # dj -> {code: nom_camion}
    magasins_par_dj = {}  # (dj, code) -> [magasins]
    for _, r in df.iterrows():
        dj = str(r.get('Demi-journee', '')).strip()
        code = str(r.get('Camion', '')).strip()
        nom = str(r.get('Nom camion', '')).strip()
        if nom == 'nan':
            nom = ''
        if not dj or dj == 'nan' or not code or code == 'nan':
            continue
        d = par_dj.setdefault(dj, {})
        if code not in d or not d[code]:
            d[code] = nom
        mags = [str(r.get(mc, '')).strip() for mc in mag_cols]
        mags = [m for m in mags if m and m != 'nan']
        magasins_par_dj[(dj, code)] = mags

    if not par_dj:
        print("\nERREUR : aucune donnée de tournée trouvée pour les documents 2/3/4.")
        return

    date_gen = date_fr(datetime.datetime.now())

    dates_par_dj = {}
    if args.date_jeudi:
        try:
            d_jeudi = datetime.datetime.strptime(args.date_jeudi, '%d/%m/%Y').date()
            offsets = {'Jeudi': 0, 'Vendredi': 1, 'Samedi': 2, 'Dimanche': 3}
            for dj_key in DJ_ORDER:
                jour = dj_key.split(' ')[0]
                d = d_jeudi + datetime.timedelta(days=offsets.get(jour, 0))
                dates_par_dj[dj_key] = d.strftime('%d/%m/%Y')
        except ValueError:
            print(f"AVERTISSEMENT : --date-jeudi '{args.date_jeudi}' invalide (attendu JJ/MM/AAAA), ignoré")

    # ── Document 2 : pointage Départ/Retour ─────────────────────────────────
    print(f"\nSortie 2 (pointage): {args.output_pointage}")
    cv2 = rl_canvas.Canvas(args.output_pointage, pagesize=A4)
    for dj_key in DJ_ORDER:
        camions_dict = par_dj.get(dj_key)
        if not camions_dict:
            continue
        camions = sorted(camions_dict.items(), key=lambda kv: camion_sort_key(kv[0]))
        print(f"  Pointage — {DJ_LABELS[dj_key]} ({len(camions)} camions)")
        draw_page(cv2, DJ_LABELS[dj_key], args.annee, camions)
        cv2.showPage()
    args.output_pointage = sauver_pdf_avec_repli(cv2, args.output_pointage)
    print(f"Sortie 3 (équipier): {args.output_equipier}")
    cv3 = rl_canvas.Canvas(args.output_equipier, pagesize=A4)
    page_num = 1
    for dj_key in DJ_ORDER:
        camions_dict = par_dj.get(dj_key)
        if not camions_dict:
            continue
        camions = sorted(camions_dict.items(), key=lambda kv: camion_sort_key(kv[0]))
        rows = []
        for code, nom in camions:
            magasins = magasins_par_dj.get((dj_key, code), [])
            equipiers = equipiers_par_camion_dj.get((code, dj_key), [])
            rows.append((code, nom, magasins, equipiers))
        print(f"  Équipier  — {DJ_LABELS[dj_key]} ({len(rows)} camions)")
        page_num = draw_equipier_page(cv3, DJ_LABELS[dj_key], rows, dates_par_dj.get(dj_key, ''), date_gen, page_num)
        cv3.showPage()
        page_num += 1
    args.output_equipier = sauver_pdf_avec_repli(cv3, args.output_equipier)
    rows_par_dj = {}
    for dj_key in DJ_ORDER:
        camions_dict = par_dj.get(dj_key)
        if not camions_dict:
            continue
        rows = []
        for (cam_code, dj_e), eq_list in equipiers_par_camion_dj.items():
            if dj_e != dj_key:
                continue
            nom_cam = camions_dict.get(cam_code, '')
            for nom_eq, tel in eq_list:
                rows.append((nom_eq, cam_code, nom_cam))
        if rows:
            rows.sort(key=lambda r: r[0].casefold())
            rows_par_dj[dj_key] = rows

    print(f"Sortie 4 (index)   : {args.output_index}")
    cv4 = rl_canvas.Canvas(args.output_index, pagesize=A4)
    for dj_key in DJ_ORDER:
        rows = rows_par_dj.get(dj_key)
        if not rows:
            continue
        print(f"  Index — {DJ_LABELS[dj_key]} ({len(rows)} équipiers)")
        draw_index_page(cv4, DJ_LABELS[dj_key], rows, 1, PW, PH, col_nom=70*mm, ncols=1, boxed=True)
        cv4.showPage()
    args.output_index = sauver_pdf_avec_repli(cv4, args.output_index)

    # ── Carte HTML interactive des tournées (comme dans les simulations de
    # generer_tournees_bai_v2.py) — recalcule les mêmes lignes 'Secteur' que
    # l'onglet 'Tournees' du classeur (construire_lignes_tournees_secteur),
    # pour disposer d'un Secteur par tournée même si liste-vehicule.xlsx n'en
    # contient pas directement.
    print()
    print('  Génération de la carte HTML des tournées...')
    lignes_carte = construire_lignes_tournees_secteur(df, mag_cols, df_ref)
    df_carte = pd.DataFrame(lignes_carte)
    generer_carte_tournees(df_carte, df_ref, mag_cols, out_dir)

    print(f"\nLes 4 documents et la carte HTML ont été générés dans : {out_dir}")


if __name__ == '__main__':
    main()
